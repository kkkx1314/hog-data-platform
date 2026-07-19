#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生猪期货基本面建模数据构建脚本
==============================================
功能：
  1. 从 akshare 获取 DCE 生猪期货（LH）所有合约历史日线 + 主力连续
  2. 读取月度基本面数据（能繁母猪、新生仔猪、利润）
  3. 读取涌益咨询日度数据（现货价格、肥标价差等）
  4. 读取涌益咨询周度数据（体重、屠宰量、鲜销率等）
  5. 合并为建模用宽表 → final_modeling_data.csv

使用方式：
  python build_futures_modeling_data.py
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import warnings
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════
# 0. 配置区（用户自行修改路径）
# ═══════════════════════════════════════════════════════════════

# 月度基本面数据
MONTHLY_DATA_PATH = r"D:\CC\Desktop\价格能繁母猪新生仔猪.xlsx"
MONTHLY_SHEET = "Sheet1"

# 涌益咨询日度数据
DAILY_DATA_PATH = r"D:\CC\Desktop\平台数据\2026年7月17日涌益咨询日度数据.xlsx"

# 涌益咨询周度数据
WEEKLY_DATA_PATH = r"D:\CC\Desktop\平台数据\2026.7.10-2026.7.16涌益咨询 周度数据.xlsx"

# 期货数据本地备用文件
FUTURES_ALL_PATH = "futures_all_contracts.csv"
FUTURES_MAIN_PATH = "futures_main.csv"

# 输出文件
OUTPUT_PATH = "final_modeling_data.csv"

# 期货数据时间范围
FUTURES_START_DATE = "2021-01-01"
FUTURES_END_DATE = datetime.today().strftime("%Y-%m-%d")

# 列名映射（可配置）
MONTHLY_COL_NAMES = {
    "date_num": 0,       # Excel日期序列号
    "price": 1,          # 月度现货价格
    "sow_10m": 2,        # 10个月前能繁母猪
    "piglet_6m": 3,      # 6个月前新生仔猪
    "profit_self": 4,    # 10个月前自繁自养利润
    "profit_outsource": 5,  # 10个月前外购仔猪利润
    "profit_piglet": 6,  # 10个月前仔猪销售利润
}

# ═══════════════════════════════════════════════════════════════
# 1. 工具函数
# ═══════════════════════════════════════════════════════════════

def text_of(value: Any) -> str:
    """安全转换为字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    return str(value).strip()


def to_float(value: Any) -> float | None:
    """安全转换为浮点数。"""
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        return float(value)
    t = text_of(value)
    if not t:
        return None
    try:
        return float(t.replace(",", "").replace("，", ""))
    except ValueError:
        return None


def parse_date(value: Any) -> pd.Timestamp | None:
    """通用日期解析。"""
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, datetime):
        return pd.Timestamp(value).normalize()
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        num = float(value)
        # Excel 日期序列号范围
        if 42000 <= num <= 52000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(num, unit="D")).normalize()
            except Exception:
                return None
        return None
    t = text_of(value)
    if not t:
        return None
    parsed = pd.to_datetime(t, errors="coerce")
    if not pd.isna(parsed):
        return parsed.normalize()
    # 尝试 "YYYY-MM-DD" 等常见格式
    m = re.search(r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})", t)
    if m:
        try:
            return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=int(m.group(3)))
        except Exception:
            return None
    return None


def print_progress(msg: str) -> None:
    """打印进度。"""
    print(f"  [{datetime.now().strftime('%H:%M:%S')}] {msg}")


# ═══════════════════════════════════════════════════════════════
# 2. 期货数据获取
# ═══════════════════════════════════════════════════════════════

def fetch_futures_akshare() -> pd.DataFrame:
    """从 akshare 获取所有 LH 合约日线数据。"""
    import akshare as ak
    all_rows = []
    contracts_seen = set()

    try:
        # 获取 LH 合约列表
        print_progress("正在从 akshare 获取生猪期货合约列表...")
        symbol_df = ak.futures_dce_warehouse_receipt(symbol="生猪")
        # 备用方法：直接尝试已知合约范围
    except Exception:
        pass

    # 遍历可能的合约月份（2021年1月上市 ~ 当前+1年）
    current = datetime.today()
    start_year = 2021
    end_year = current.year + 1

    for year in range(start_year, end_year + 1):
        for month in [1, 3, 5, 7, 9, 11]:
            if year == start_year and month < 1:
                continue
            if year == end_year and month > current.month + 2:
                break
            contract = f"LH{str(year)[-2:]}{month:02d}"
            if contract in contracts_seen:
                continue
            contracts_seen.add(contract)

            try:
                print_progress(f"  获取 {contract}...")
                df = ak.futures_main_sina(symbol=contract)
                if df is None or df.empty:
                    continue
                df["contract"] = contract
                all_rows.append(df)
            except Exception:
                continue

    if all_rows:
        result = pd.concat(all_rows, ignore_index=True)
        result = result.rename(columns={
            "日期": "date", "date": "date",
            "收盘价": "close", "close": "close",
            "成交量": "volume", "volume": "volume",
            "持仓量": "open_interest", "open_interest": "open_interest",
            "开盘价": "open", "open": "open",
            "最高价": "high", "high": "high",
            "最低价": "low", "low": "low",
        })
        return result
    return pd.DataFrame()


def fetch_main_contract_akshare() -> pd.DataFrame:
    """从 akshare 获取生猪主力连续合约（LH0）日线。"""
    try:
        import akshare as ak
        print_progress("获取主力连续合约 LH0...")
        df = ak.futures_main_sina(symbol="LH0")
        if df is not None and not df.empty:
            df = df.rename(columns={
                "日期": "date", "date": "date",
                "收盘价": "main_close", "close": "main_close",
            })
            return df[["date", "main_close"]].copy()
    except Exception:
        pass
    return pd.DataFrame()


def load_futures_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """加载期货数据（akshare 优先，失败则本地备用）。"""
    print("=" * 60)
    print(" 1. 获取期货数据")
    print("=" * 60)

    contracts_df = pd.DataFrame()
    main_df = pd.DataFrame()

    try:
        contracts_df = fetch_futures_akshare()
        main_df = fetch_main_contract_akshare()
    except Exception as e:
        print(f"  ️  akshare 获取失败: {e}")

    # 如果 akshare 失败或结果为空，使用本地备用文件
    if contracts_df.empty or main_df.empty:
        print("  ️  切换到本地备用文件...")
        if Path(FUTURES_ALL_PATH).exists():
            contracts_df = pd.read_csv(FUTURES_ALL_PATH)
            print(f"   从 {FUTURES_ALL_PATH} 读取合约数据: {len(contracts_df)} 行")
        else:
            print(f"   未找到 {FUTURES_ALL_PATH}")

        if Path(FUTURES_MAIN_PATH).exists():
            main_df = pd.read_csv(FUTURES_MAIN_PATH)
            print(f"   从 {FUTURES_MAIN_PATH} 读取主力数据: {len(main_df)} 行")
        else:
            print(f"   未找到 {FUTURES_MAIN_PATH}")

    # 标准化日期列
    for df_name, df in [("合约数据", contracts_df), ("主力数据", main_df)]:
        if df.empty:
            continue
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
            df = df.dropna(subset=["date"])
        # 过滤日期范围
        if "date" in df.columns:
            start_dt = pd.Timestamp(FUTURES_START_DATE)
            end_dt = pd.Timestamp(FUTURES_END_DATE)
            df = df[(df["date"] >= start_dt) & (df["date"] <= end_dt)]

    print(f"   合约行情: {len(contracts_df)} 行 · 主力连续: {len(main_df)} 行")
    return contracts_df, main_df


# ═══════════════════════════════════════════════════════════════
# 3. 月度数据处理
# ═══════════════════════════════════════════════════════════════

def load_monthly_data(path: str) -> pd.DataFrame:
    """读取月度基本面数据并清洗。"""
    print("\n" + "=" * 60)
    print(" 2. 读取月度基本面数据")
    print("=" * 60)

    if not Path(path).exists():
        print(f"   文件不存在: {path}")
        return pd.DataFrame()

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[MONTHLY_SHEET]
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        while vals and vals[-1] is None:
            vals.pop()
        if vals:
            rows.append(vals)
    wb.close()

    if len(rows) < 2:
        print("   数据行数不足")
        return pd.DataFrame()

    # 第一行是标题，跳过
    data_rows = rows[1:]
    print(f"  原始数据: {len(data_rows)} 行")

    records = []
    for row in data_rows:
        if len(row) <= max(MONTHLY_COL_NAMES.values()):
            continue
        rec = {}
        for key, idx in MONTHLY_COL_NAMES.items():
            if idx < len(row):
                rec[key] = row[idx]
            else:
                rec[key] = None
        records.append(rec)

    df = pd.DataFrame(records)

    # 转换 date_num → month_date
    df["month_date"] = df["date_num"].apply(parse_date)
    df = df.dropna(subset=["month_date"]).copy()
    # 确保是每月第一天
    df["month_date"] = df["month_date"].apply(lambda x: x.replace(day=1))

    # 数值列
    numeric_cols = ["price", "sow_10m", "piglet_6m",
                    "profit_self", "profit_outsource", "profit_piglet"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(to_float)

    # 排序 + 前向填充
    df = df.sort_values("month_date").reset_index(drop=True)
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].ffill()

    df = df.drop(columns=["date_num"])
    print(f"  清洗后: {len(df)} 行, {df['month_date'].min().date()} ~ {df['month_date'].max().date()}")
    print(f"  列: {list(df.columns)}")
    return df


# ═══════════════════════════════════════════════════════════════
# 4. 涌益日度数据解析
# ═══════════════════════════════════════════════════════════════

def _filled_right(values: list) -> list:
    """右向填充——用于展开合并单元格。"""
    result = []
    current = None
    for v in values:
        t = text_of(v)
        if t:
            current = v
        result.append(current)
    return result


def _parse_date_rows_province_columns(rows, sheet_name, metric_prefix):
    """日期为行、省份为列的标准表格。"""
    records = []
    if len(rows) < 2:
        return records

    # 第一行：日期列名 + 省份名
    header_row = rows[0]
    provinces = [text_of(c) for c in header_row[1:]]

    for row in rows[1:]:
        if not row:
            continue
        date_val = parse_date(row[0])
        if date_val is None:
            continue
        for i, province in enumerate(provinces):
            if i + 1 >= len(row):
                break
            if not province:
                continue
            val = to_float(row[i + 1])
            if val is None:
                continue
            records.append({
                "date": date_val,
                "province": province,
                "metric": metric_prefix,
                "value": val,
            })
    return records


def _parse_simple_metric_table(rows, sheet_name, date_col=1, metric_start_col=2):
    """单日期列 + 多指标列的简单表格。"""
    records = []
    if len(rows) < 2:
        return records
    header = rows[0]
    for row in rows[1:]:
        if not row:
            continue
        date_val = parse_date(row[date_col]) if date_col < len(row) else None
        if date_val is None:
            continue
        for i in range(metric_start_col, min(len(row), len(header))):
            metric_name = text_of(header[i])
            if not metric_name:
                continue
            val = to_float(row[i])
            if val is None:
                continue
            records.append({
                "date": date_val,
                "metric": metric_name,
                "value": val,
            })
    return records


def _parse_repeated_date_blocks(rows, sheet_name, start_col=2, province_col=0):
    """双层表头、日期分块横向展开。"""
    records = []
    if len(rows) < 3:
        return records

    max_len = max(len(r) for r in rows[:min(len(rows), 5)])
    date_row = list(rows[0]) + [None] * (max_len - len(rows[0]))
    metric_row = list(rows[1]) + [None] * (max_len - len(rows[1]))

    dates = _filled_right([parse_date(x) for x in date_row])
    metrics = _filled_right(metric_row)

    for row in rows[2:]:
        padded = list(row) + [None] * (max_len - len(row))
        province = text_of(padded[province_col])
        if not province or province in {"日期", "省份"}:
            continue
        for col in range(start_col, max_len):
            date_val = dates[col]
            metric_val = text_of(metrics[col])
            if date_val is None or not metric_val:
                continue
            val = to_float(padded[col])
            if val is None:
                continue
            records.append({
                "date": date_val,
                "province": province,
                "metric": metric_val,
                "value": val,
            })
    return records


def _parse_province_rows_date_columns(rows, sheet_name, metric_name="日度屠宰量"):
    """省份为行、日期为列的表格。Row 0 = 日期行，Row 1+ = 各省份数据。
    第一列为省份名，其余列为各日期的数据。"""
    records = []
    if len(rows) < 2:
        return records

    # Row 0: 省份名 + 日期列
    header = rows[0]
    dates = []
    for c in header[1:]:
        d = parse_date(c)
        dates.append(d)

    for row in rows[1:]:
        if not row:
            continue
        province = text_of(row[0])
        if not province or province in {"日期", "省份", "省"}:
            continue
        for i, date_val in enumerate(dates):
            if date_val is None or i + 1 >= len(row):
                continue
            val = to_float(row[i + 1])
            if val is None:
                continue
            records.append({
                "date": date_val,
                "province": province,
                "metric": metric_name,
                "value": val,
            })
    return records


# 要保留的日度指标
DAILY_TARGET_METRICS = {
    # 价格+宰量 sheet 中需要的指标
    "全国均价", "均价", "出栏均价",
    "标猪价", "肥猪价",
    "标猪均价", "肥猪均价",
    "宰量", "屠宰量", "日度屠宰量",
}

DAILY_SHEET_CONFIG = {
    "价格+宰量": {"mode": "simple", "date_col": 1, "metric_start_col": 2},
    "各省份均价": {"mode": "date_rows_province_cols", "metric": "均价"},
    "屠宰企业日度屠宰量": {"mode": "province_rows_date_cols", "metric": "日度屠宰量"},
    "市场主流标猪肥猪均价方便作图": {"mode": "simple", "date_col": 1, "metric_start_col": 2},
    "出栏价": {"mode": "repeated_blocks", "start_col": 2},
    "散户标肥价差": {"mode": "repeated_blocks", "start_col": 1},
    "市场主流标猪肥猪价格": {"mode": "repeated_blocks", "start_col": 2},
}


def _workbook_to_rows(workbook) -> dict:
    """将 openpyxl workbook 转为 {sheet_name: [rows]}."""
    sheets = {}
    for ws in workbook.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            while vals and vals[-1] is None:
                vals.pop()
            rows.append(vals)
        sheets[ws.title] = rows
    return sheets


def load_daily_data(path: str) -> pd.DataFrame:
    """解析涌益日度数据，提取现货价格和肥标价差。"""
    print("\n" + "=" * 60)
    print(" 3. 读取涌益日度数据")
    print("=" * 60)

    if not Path(path).exists():
        print(f"  ️  文件不存在: {path}，日度数据为空")
        return pd.DataFrame()

    print_progress(f"正在解析: {Path(path).name}")
    wb = load_workbook(path, data_only=True, read_only=True)
    all_sheets = _workbook_to_rows(wb)
    wb.close()

    all_records = []
    for sheet_name, rows in all_sheets.items():
        if sheet_name not in DAILY_SHEET_CONFIG:
            continue
        cfg = DAILY_SHEET_CONFIG[sheet_name]
        mode = cfg["mode"]

        if mode == "simple":
            records = _parse_simple_metric_table(rows, sheet_name, cfg["date_col"], cfg["metric_start_col"])
        elif mode == "date_rows_province_cols":
            records = _parse_date_rows_province_columns(rows, sheet_name, cfg["metric"])
        elif mode == "repeated_blocks":
            records = _parse_repeated_date_blocks(
                rows, sheet_name,
                start_col=cfg.get("start_col", 2),
                province_col=cfg.get("province_col", 0),
            )
        elif mode == "province_rows_date_cols":
            records = _parse_province_rows_date_columns(
                rows, sheet_name,
                metric_name=cfg.get("metric", "日度屠宰量"),
            )
        else:
            continue

        if records:
            print_progress(f"  {sheet_name}: {len(records)} 条")
            all_records.extend(records)

    if not all_records:
        print("  ️  未从日度数据中解析到任何记录")
        return pd.DataFrame()

    df = pd.DataFrame(all_records)
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    print(f"   总计: {len(df)} 条, {df['date'].min().date()} ~ {df['date'].max().date()}")
    print(f"  指标: {sorted(df['metric'].unique())}")
    return df


# ═══════════════════════════════════════════════════════════════
# 5. 涌益周度数据解析
# ═══════════════════════════════════════════════════════════════

# 周度数据中要跳过的 sheet
def _parse_weight_sheet(rows, sheet_name):
    """解析出栏均重 sheet：Row0=标题, Row1-2=多级表头, Row3+=数据。
    格式：[开始日期, 全国均重, 规模, 散户, 规模权重, 散户权重] 或
          [开始日期, 结束日期, 全国均重, 规模, 散户, 规模权重, 散户权重]
    提取 全国均重 作为 weight。"""
    records = []
    if len(rows) < 4:
        return records
    for row in rows[3:]:
        if not row or len(row) < 2:
            continue
        start_date = parse_date(row[0])
        if start_date is None:
            continue
        # 全国均重可能在 col 1 或 col 2
        val = to_float(row[1])
        if val is None and len(row) >= 3:
            val = to_float(row[2])
        if val is None:
            continue
        records.append({"date": start_date, "metric": "weight", "value": val})
    return records


def _parse_slaughter_sheet(rows, sheet_name):
    """解析「周度-规模屠宰2022.10.28」sheet：每行 [开始日期, 结束日期, 规模厂屠宰量, ...]。
    提取 规模厂屠宰量 作为 slaughter。"""
    records = []
    if len(rows) < 2:
        return records
    for row in rows[1:]:
        if not row or len(row) < 3:
            continue
        start_date = parse_date(row[0])
        if start_date is None:
            continue
        val = to_float(row[2])
        if val is None:
            continue
        records.append({"date": start_date, "metric": "slaughter", "value": val})
    return records


def _parse_fresh_sale_rate_sheet(rows, sheet_name):
    """解析「鲜销率」sheet：Row0=空, Row1=多级表头, Row2=省份表头, Row3+=数据。
    提取全国均值（col 2，列名含「全国」）作为 fresh_sale_rate。"""
    records = []
    if len(rows) < 4:
        return records
    # Row 1: [None, None, '鲜销率', None, None, '鲜销率', ...]
    # Row 2: [None, None, '辽宁', '山东', '河北', '河南', ...]
    # Row 3+: [date, date, val, val, val, ...]
    for row in rows[3:]:
        if not row or len(row) < 3:
            continue
        start_date = parse_date(row[0])
        if start_date is None:
            continue
        # 全国均值在 col 2（但也可能是省份数据，取第一个有效数值列）
        for i in range(2, min(len(row), 10)):
            val = to_float(row[i])
            if val is not None:
                records.append({"date": start_date, "metric": "fresh_sale_rate", "value": val})
                break
    return records


def _parse_price_sheet(rows, sheet_name):
    """解析「周度-商品猪出栏价」sheet：Row0=标题, Row1=表头, Row2+=数据。
    提取全国均价（col 2，列名含「全国」）作为 weekly_price。"""
    records = []
    if len(rows) < 3:
        return records
    for row in rows[2:]:
        if not row or len(row) < 3:
            continue
        start_date = parse_date(row[0])
        if start_date is None:
            continue
        val = to_float(row[2])
        if val is None:
            continue
        records.append({"date": start_date, "metric": "weekly_price", "value": val})
    return records


def _match_sheet(sheet_name: str) -> str | None:
    """按关键词匹配 sheet 名，返回对应的解析器 key。"""
    name = sheet_name.strip()
    if "均重" in name or "体重" in name:
        return "weight"
    if "屠宰" in name and ("规模" in name or "2022" in name):
        return "slaughter"
    if "鲜销率" in name or name == "鲜销率":
        return "fresh_sale"
    if "商品猪出栏价" in name or ("出栏价" in name and "商品" in name):
        return "price"
    return None


# sheet 类型 → 解析函数
WEEKLY_PARSERS = {
    "weight": _parse_weight_sheet,
    "slaughter": _parse_slaughter_sheet,
    "fresh_sale": _parse_fresh_sale_rate_sheet,
    "price": _parse_price_sheet,
}


def load_weekly_data(path: str) -> pd.DataFrame:
    """解析涌益周度数据，提取体重、屠宰量、鲜销率、周度价格等。"""
    print("\n" + "=" * 60)
    print(" 4. 读取涌益周度数据")
    print("=" * 60)

    if not Path(path).exists():
        print(f"[WARN] 文件不存在: {path}，周度数据为空")
        return pd.DataFrame()

    print_progress(f"正在解析: {Path(path).name}")
    wb = load_workbook(path, data_only=True, read_only=True)
    all_sheets = _workbook_to_rows(wb)
    wb.close()

    all_records = []
    for sheet_name, rows in all_sheets.items():
        key = _match_sheet(sheet_name)
        if key is None:
            continue
        parser = WEEKLY_PARSERS[key]
        try:
            records = parser(rows, sheet_name)
            if records:
                print_progress(f"  {sheet_name}: {len(records)} 条")
                all_records.extend(records)
        except Exception as e:
            print(f"  [WARN] {sheet_name} 解析失败: {e}")

    if not all_records:
        print("[WARN] 未从周度数据中解析到目标指标")
        return pd.DataFrame()

    df = pd.DataFrame(all_records)
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    print(f"[OK] 总计: {len(df)} 条, {df['date'].min().date()} ~ {df['date'].max().date()}")
    print(f"  指标: {sorted(df['metric'].unique())}")
    return df


# ═══════════════════════════════════════════════════════════════
# 6. 核心合并逻辑
# ═══════════════════════════════════════════════════════════════

def parse_contract_info(contract_str: str) -> dict:
    """从合约代码解析年份和月份。例如 LH2409 → year=2024, month=9."""
    m = re.search(r"LH(\d{2})(\d{2})", str(contract_str))
    if m:
        yy = int(m.group(1))
        mm = int(m.group(2))
        year = 2000 + yy if yy <= 50 else 1900 + yy  # 假设在 2000-2050 之间
        return {"contract_year": year, "contract_month": mm, "delivery_date": pd.Timestamp(year=year, month=mm, day=1)}
    return {"contract_year": None, "contract_month": None, "delivery_date": None}


def build_modeling_dataset(
    contracts_df: pd.DataFrame,
    main_df: pd.DataFrame,
    monthly_df: pd.DataFrame,
    daily_df: pd.DataFrame,
    weekly_df: pd.DataFrame,
) -> pd.DataFrame:
    """合并所有数据源，构建最终宽表。"""
    print("\n" + "=" * 60)
    print(" 5. 构建建模数据集")
    print("=" * 60)

    if contracts_df.empty:
        print("   无期货数据，无法继续")
        return pd.DataFrame()

    # ── 5.1 确保列名标准化 ──
    contracts_df = contracts_df.copy()
    # 自动检测列名并标准化
    col_map = {}
    for c in contracts_df.columns:
        cl = c.lower().replace(" ", "_")
        if "date" in cl or "日期" in c:
            col_map[c] = "date"
        elif "close" in cl or "收盘" in c:
            col_map[c] = "close"
        elif "volume" in cl or "成交" in c:
            col_map[c] = "volume"
        elif "open_interest" in cl or "持仓" in c:
            col_map[c] = "open_interest"
        elif "open" in cl or "开盘" in c:
            col_map[c] = "open"
        elif "high" in cl or "最高" in c:
            col_map[c] = "high"
        elif "low" in cl or "最低" in c:
            col_map[c] = "low"
        elif "contract" in cl or "合约" in c:
            col_map[c] = "contract"

    contracts_df = contracts_df.rename(columns=col_map)
    if "date" in contracts_df.columns:
        contracts_df["date"] = pd.to_datetime(contracts_df["date"]).dt.normalize()

    # 确保有 contract 列
    if "contract" not in contracts_df.columns:
        # 尝试从已解析的合约名中提取
        print("  ️  未找到 contract 列，尝试从数据推断...")

    # 确保必要列存在
    for col in ["close", "volume", "open_interest"]:
        if col not in contracts_df.columns:
            contracts_df[col] = np.nan
            print(f"  ️  缺少列: {col}，填充 NaN")

    print(f"  合约行情: {len(contracts_df)} 行, {contracts_df['contract'].nunique() if 'contract' in contracts_df.columns else 'N/A'} 个合约")

    # ── 5.2 主力价差 ──
    if not main_df.empty and "date" in main_df.columns and "main_close" in main_df.columns:
        main_df["date"] = pd.to_datetime(main_df["date"]).dt.normalize()
        contracts_df = contracts_df.merge(main_df[["date", "main_close"]], on="date", how="left")
        contracts_df["spread_vs_main"] = contracts_df["close"] - contracts_df["main_close"]
        print(f"   主力价差已计算")
    else:
        contracts_df["main_close"] = np.nan
        contracts_df["spread_vs_main"] = np.nan
        print(f"  ️  无主力连续数据，spread_vs_main 为空")

    # ── 5.3 匹配月度静态特征 ──
    if not monthly_df.empty:
        print_progress("匹配月度基本面特征...")

        # 为每行生成交割月份
        contracts_df["delivery_date"] = pd.NaT
        if "contract" in contracts_df.columns:
            contract_infos = contracts_df["contract"].apply(parse_contract_info)
            contracts_df["delivery_date"] = contract_infos.apply(lambda x: x["delivery_date"])
            contracts_df["contract_year"] = contract_infos.apply(lambda x: x["contract_year"])
            contracts_df["contract_month"] = contract_infos.apply(lambda x: x["contract_month"])

        # 使用 merge_asof 进行前向匹配
        monthly_sorted = monthly_df.sort_values("month_date").copy()

        for feature_col in ["sow_10m", "piglet_6m", "profit_self", "profit_outsource", "profit_piglet"]:
            if feature_col not in monthly_sorted.columns:
                contracts_df[feature_col] = np.nan
                continue

            monthly_features = monthly_sorted[["month_date", feature_col]].dropna(subset=[feature_col]).copy()
            monthly_features = monthly_features.rename(columns={"month_date": "delivery_date"})

            contracts_df = contracts_df.drop(columns=[feature_col], errors="ignore")

            contracts_valid = contracts_df.dropna(subset=["delivery_date"]).sort_values("delivery_date")
            contracts_na = contracts_df[contracts_df["delivery_date"].isna()]

            if not contracts_valid.empty:
                contracts_valid = pd.merge_asof(
                    contracts_valid.sort_values("delivery_date"),
                    monthly_features.sort_values("delivery_date"),
                    on="delivery_date",
                    direction="backward",
                )

            contracts_df = pd.concat([contracts_valid, contracts_na], ignore_index=True)

            filled_count = contracts_df[feature_col].notna().sum()
            print(f"  {feature_col}: {filled_count}/{len(contracts_df)} 非空")

        print("   月度特征匹配完成")
    else:
        for col in ["sow_10m", "piglet_6m", "profit_self", "profit_outsource", "profit_piglet"]:
            contracts_df[col] = np.nan
        print("  ️  无月度数据，基本面特征为空")

    # ── 5.4 日度动态变量 ──
    if not daily_df.empty and "date" in daily_df.columns:
        print_progress("合并日度动态变量...")
        daily_wide = _pivot_daily_data(daily_df)
        if not daily_wide.empty:
            # 按日期全局排序后 ffill
            daily_wide = daily_wide.sort_values("date")
            date_range = pd.DataFrame({
                "date": pd.date_range(
                    start=contracts_df["date"].min(),
                    end=contracts_df["date"].max(),
                    freq="D",
                )
            })
            daily_wide = date_range.merge(daily_wide, on="date", how="left")
            # ffill 除 date 外的所有列
            for col in daily_wide.columns:
                if col != "date":
                    daily_wide[col] = daily_wide[col].ffill()

            contracts_df = contracts_df.merge(daily_wide, on="date", how="left")
            # 对刚合并的日度列做分组 ffill
            daily_cols = [c for c in daily_wide.columns if c != "date"]
            for col in daily_cols:
                if col in contracts_df.columns:
                    contracts_df[col] = contracts_df.groupby("contract")[col].ffill()

            print(f"   日度变量已合并: {daily_cols}")
    else:
        print("  ️  无日度数据")

    # ── 5.5 周度动态变量 ──
    if not weekly_df.empty and "date" in weekly_df.columns:
        print_progress("合并周度动态变量...")
        weekly_wide = _pivot_weekly_data(weekly_df)
        if not weekly_wide.empty:
            weekly_wide = weekly_wide.sort_values("date")
            # 扩展到日频：每天使用当周数据
            date_range = pd.DataFrame({
                "date": pd.date_range(
                    start=contracts_df["date"].min(),
                    end=contracts_df["date"].max(),
                    freq="D",
                )
            })
            weekly_wide = date_range.merge(weekly_wide, on="date", how="left")
            for col in weekly_wide.columns:
                if col != "date":
                    weekly_wide[col] = weekly_wide[col].ffill()

            contracts_df = contracts_df.merge(weekly_wide, on="date", how="left")
            weekly_cols = [c for c in weekly_wide.columns if c != "date"]
            for col in weekly_cols:
                if col in contracts_df.columns:
                    contracts_df[col] = contracts_df.groupby("contract")[col].ffill()

            print(f"   周度变量已合并: {weekly_cols}")
    else:
        print("  ️  无周度数据")

    # ── 5.6 整理最终列顺序 ──
    priority_cols = [
        "date", "contract", "close", "volume", "open_interest",
        "spread_vs_main", "main_close",
        "contract_year", "contract_month", "delivery_date",
        "sow_10m", "piglet_6m",
        "profit_self", "profit_outsource", "profit_piglet",
    ]
    other_cols = [c for c in contracts_df.columns if c not in priority_cols]
    final_cols = [c for c in priority_cols if c in contracts_df.columns] + other_cols
    contracts_df = contracts_df[final_cols]

    # 按合约+日期排序
    sort_cols = ["contract", "date"] if "contract" in contracts_df.columns else ["date"]
    contracts_df = contracts_df.sort_values(sort_cols).reset_index(drop=True)

    print(f"\n   最终宽表: {len(contracts_df)} 行 × {len(contracts_df.columns)} 列")
    return contracts_df


def _pivot_daily_data(daily_df: pd.DataFrame) -> pd.DataFrame:
    """将日度长表转为宽表（每个指标一列）。"""
    if daily_df.empty or "metric" not in daily_df.columns:
        return pd.DataFrame()

    # 优先取全国数据；若无则取均值
    has_province = "province" in daily_df.columns
    if has_province:
        # 全国优先
        national = daily_df[daily_df["province"].str.contains("全国|均价", na=False)]
        if not national.empty:
            daily_df = national
        # 否则按 date+metric 取均值

    pivot = daily_df.pivot_table(
        index="date", columns="metric", values="value", aggfunc="mean"
    ).reset_index()

    # 清理列名（去掉特殊字符）
    pivot.columns = [
        re.sub(r"[^\w一-鿿]+", "_", str(c)).strip("_")
        for c in pivot.columns
    ]
    # 重命名关键列
    rename_map = {}
    for c in pivot.columns:
        if "均价" in c or "价格" in c:
            rename_map[c] = "spot_price"
        elif "肥标" in c or "价差" in c:
            rename_map[c] = "fat_lean_spread"
    pivot = pivot.rename(columns=rename_map)
    return pivot


def _pivot_weekly_data(weekly_df: pd.DataFrame) -> pd.DataFrame:
    """将周度长表转为宽表。"""
    if weekly_df.empty or "metric" not in weekly_df.columns:
        return pd.DataFrame()

    pivot = weekly_df.pivot_table(
        index="date", columns="metric", values="value", aggfunc="mean"
    ).reset_index()

    pivot.columns = [
        re.sub(r"[^\w一-鿿]+", "_", str(c)).strip("_")
        for c in pivot.columns
    ]
    # 重命名关键列
    rename_map = {}
    for c in pivot.columns:
        if "均重" in c or "体重" in c or "weight" in c.lower():
            rename_map[c] = "weight"
        elif "屠宰" in c or "宰量" in c or "slaughter" in c.lower():
            rename_map[c] = "slaughter"
        elif "鲜销" in c or "fresh" in c.lower():
            rename_map[c] = "fresh_sale_rate"
    pivot = pivot.rename(columns=rename_map)
    return pivot


# ═══════════════════════════════════════════════════════════════
# 7. 数据质量报告
# ═══════════════════════════════════════════════════════════════

def generate_quality_report(df: pd.DataFrame) -> None:
    """生成数据质量报告。"""
    print("\n" + "=" * 60)
    print(" 6. 数据质量报告")
    print("=" * 60)

    if df.empty:
        print("   数据集为空")
        return

    # 每个合约的概况
    if "contract" in df.columns:
        print(f"\n{'合约':<12} {'行数':>8} {'起始日期':>12} {'结束日期':>12} {'缺失率':>8}")
        print("-" * 56)
        for contract, grp in df.groupby("contract"):
            missing_rate = grp.isnull().mean().mean() * 100
            d_min = grp["date"].min().strftime("%Y-%m-%d") if pd.notna(grp["date"].min()) else "N/A"
            d_max = grp["date"].max().strftime("%Y-%m-%d") if pd.notna(grp["date"].max()) else "N/A"
            print(f"{contract:<12} {len(grp):>8} {d_min:>12} {d_max:>12} {missing_rate:>7.1f}%")
    else:
        print(f"\n  总行数: {len(df)}")
        print(f"  日期范围: {df['date'].min()} ~ {df['date'].max()}")

    # 各列缺失比例
    print(f"\n{'列名':<35} {'缺失比例':>10} {'非空数':>10}")
    print("-" * 58)
    for col in df.columns:
        missing_pct = df[col].isnull().mean() * 100
        non_null = df[col].notna().sum()
        bar = "#" * int(missing_pct / 5) if missing_pct > 0 else "—"
        print(f"{col:<35} {missing_pct:>6.1f}% {bar} {non_null:>8}")

    print(f"\n  总列数: {len(df.columns)}")
    print(f"  总行数: {len(df)}")


# ═══════════════════════════════════════════════════════════════
# 8. 主流程
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print(" 生猪期货基本面建模数据构建")
    print(f"   时间范围: {FUTURES_START_DATE} ~ {FUTURES_END_DATE}")
    print("=" * 60)

    # 1. 期货数据
    contracts_df, main_df = load_futures_data()

    # 2. 月度基本面
    monthly_df = load_monthly_data(MONTHLY_DATA_PATH)

    # 3. 日度数据
    daily_df = load_daily_data(DAILY_DATA_PATH)

    # 4. 周度数据
    weekly_df = load_weekly_data(WEEKLY_DATA_PATH)

    # 5. 合并构建宽表
    final_df = build_modeling_dataset(
        contracts_df, main_df, monthly_df, daily_df, weekly_df,
    )

    if final_df.empty:
        print("\n 无法生成建模数据，请检查数据源。")
        return

    # 6. 质量报告
    generate_quality_report(final_df)

    # 7. 保存
    final_df.to_csv(OUTPUT_PATH, index=False, encoding="utf-8-sig")
    print(f"\n 已保存: {OUTPUT_PATH}")
    print(f"   文件大小: {Path(OUTPUT_PATH).stat().st_size / 1024 / 1024:.1f} MB")


if __name__ == "__main__":
    main()
