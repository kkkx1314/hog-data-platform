from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any
import base64
import hashlib
import re

import plotly.io as pio

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

try:
    from lunardate import LunarDate
except Exception:
    LunarDate = None


# ── 数据目录：优先本地平台数据目录，其次项目内data目录（GitHub同步用）──
_LOCAL_DATA = Path(r"D:\CC\Desktop\平台数据")
_REPO_DATA = Path(__file__).parent / "data"
_REPO_DATA.mkdir(exist_ok=True)
PLATFORM_DATA_DIR = _LOCAL_DATA if _LOCAL_DATA.exists() else _REPO_DATA

# ── 自动识别最新数据文件 ──
def _find_latest_file(pattern: str, dir_path: Path = None) -> Path | None:
    """在数据目录中按文件名模式匹配最新文件"""
    search_dir = dir_path or PLATFORM_DATA_DIR
    if not search_dir.exists():
        return None
    candidates = []
    for f in search_dir.glob("*.xlsx"):
        if re.search(pattern, f.name, re.IGNORECASE):
            candidates.append((f.stat().st_mtime, f))
    if not candidates:
        for f in search_dir.glob("*.xls"):
            if re.search(pattern, f.name, re.IGNORECASE):
                candidates.append((f.stat().st_mtime, f))
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]
    return None


def _find_yongyi_daily() -> str:
    """自动识别涌益咨询日度数据"""
    f = _find_latest_file(r"涌益.*日度|日度.*涌益")
    if f: return str(f)
    # 兜底：任何包含"日度"的xlsx
    f = _find_latest_file(r"日度数据")
    if f: return str(f)
    return ""


def _find_yongyi_weekly() -> str:
    """自动识别涌益咨询周度数据"""
    f = _find_latest_file(r"涌益.*周度|周度.*涌益")
    if f: return str(f)
    f = _find_latest_file(r"周度数据")
    if f: return str(f)
    return ""


def _find_transport() -> str:
    """自动识别猪只调运数据"""
    f = _find_latest_file(r"猪只调运|调运.*分析")
    if f: return str(f)
    f = _find_latest_file(r"调运")
    if f: return str(f)
    return ""


def _find_fresh_frozen() -> str:
    """自动识别鲜品冻品/神农肉业数据"""
    f = _find_latest_file(r"鲜品|神农肉业|冻品")
    if f: return str(f)
    return ""


DEFAULT_YONGYI_PATH = _find_yongyi_daily()
DEFAULT_WEEKLY_PATH = _find_yongyi_weekly()
DEFAULT_TRANSPORT_PATH = _find_transport()
DEFAULT_FRESH_FROZEN_PATH = _find_fresh_frozen()
DEFAULT_FUTURES_PATH = ""  # 已移除期货分析模块，保留变量避免引用报错

# 周度数据：跳过这些 sheet（说明/目录/静态/辅助/停更）
WEEKLY_SKIP_SHEETS: set[str] = {
    "样本点选取", "生猪数据库·目录", "数据说明",
    "不同规模市占率（按生猪出栏划分）", "不同规模市占率（按母猪存栏划分）",
    "育肥头均费用", "2025.4.24料肉比",
    "周度-成本计算附件", "二育销量", "各存栏规模", "月度猪肉供应占比",
    "历史出栏体重", "运费", "减产能较2018年7月30日",
    "新增猪场", "周&月度-三元母猪价格", "MSY",
    "进口肉", "猪料原料占比",
}

WEEKLY_FORCE_SHOW_SHEETS: set[str] = {
    "历史猪价",
    "周度-商品猪出栏价",
    "周度-宰后结算价",
    "周度-毛白价差",
    "周度-50公斤二元母猪价格",
    "周度-规模场15公斤仔猪出栏价",
    "二育栏舍利用率",
    "周度-淘汰母猪价格",
    "周度-高胎淘母折扣价",
    "周度-低胎母猪折扣价",
    "周度-各体重段价差",
    "周度-养殖利润最新",
    "周度-猪肉价（前三等级白条均价）",
    "周度-猪肉产品价格",
    "华东冻品价格",
    "国产冻品2-4号肉价格",
    "育肥全价料价格",
    "月度-猪料销量",
}

WEEKLY_BOARD_SHEETS: dict[str, list[str]] = {
    "价格端": [
        "历史猪价",
        "周度-商品猪出栏价",
        "周度-50公斤二元母猪价格",
        "周度-规模场15公斤仔猪出栏价",
        "高频仔猪、母猪",
        "周度-淘汰母猪价格",
        "周度-高胎淘母折扣价",
        "周度-低胎母猪折扣价",
        "周度-宰后结算价",
        "周度-各体重段价差",
    ],
    "供给端": [
        "二育栏舍利用率",
        "月度-生产指标（2021.5.7新增）",
        "月度-生产指标2",
        "月度-二元三元能繁比例",
        "月度-能繁母猪存栏（2020年2月新增）",
        "月度-小猪存栏（2020年5月新增）",
        "月度-大猪存栏（2020年5月新增）",
        "月度-商品猪出栏量",
        "月度-小猪（50公斤以下）存栏",
    ],
    "体重端": [
        "周度-体重",
        "周度-体重拆分",
        "周度-屠宰厂宰前活猪重",
    ],
    "利润端": [
        "仔猪与商品猪利润对比",
        "周度-养殖利润最新",
    ],
    "屠宰端": [
        "周度-毛白价差",
        "周度-河南屠宰白条成本",
        "周度-冻品库存",
        "周度-冻品库存多样本",
        "鲜销率",
        "周度-屠宰企业日度屠宰量",
        "周度-屠宰新2022.10.28",
        "月度-淘汰母猪屠宰厂宰杀量",
    ],
    "猪肉产品": [
        "周度-猪肉价（前三等级白条均价）",
        "周度-猪肉产品价格",
        "华东冻品价格",
        "国产冻品2-4号肉价格",
    ],
    "其他": [
        "育肥全价料价格",
        "月度-猪料销量",
    ],
}

WEEKLY_FEED_SHEETS: set[str] = {"育肥全价料价格", "月度-猪料销量"}
WEEKLY_UNIT_CONVERSION_SHEETS: set[str] = {"周度-各体重段价差", "周度-淘汰母猪价格"}
WEEKLY_ALWAYS_SHOW_SHEETS: set[str] = WEEKLY_FORCE_SHOW_SHEETS | {sheet for sheets in WEEKLY_BOARD_SHEETS.values() for sheet in sheets}

# 特定 sheet 的指标排除列表（这些名称不作为可选指标）
WEEKLY_SHEET_METRIC_EXCLUDE: dict[str, set[str]] = {
    "月度-生产指标2": {"日期", "月份", "时间", "生产指标"},
}

# 特定 sheet 的指标包含列表（只显示这些指标）
WEEKLY_SHEET_METRIC_INCLUDE: dict[str, set[str]] = {
    "月度-生产指标2": {"基础母猪存栏", "后备母猪数", "配种数", "分娩母猪窝数", "窝均健仔数", "产房存活率", "配种分娩率", "断奶成活率", "育肥出栏成活率"},
}

# 供给端sheet只读取全国列
SUPPLY_SIDE_SHEETS: set[str] = {
    "月度-能繁母猪存栏（2020年2月新增）",
    "月度-小猪存栏（2020年5月新增）",
    "月度-大猪存栏（2020年5月新增）",
    "月度-商品猪出栏量",
    "月度-小猪（50公斤以下）存栏",
}


PROVINCE_REPLACEMENTS = {
    "内蒙古自治区": "内蒙古",
    "广西壮族自治区": "广西",
    "西藏自治区": "西藏",
    "宁夏回族自治区": "宁夏",
    "新疆维吾尔自治区": "新疆",
    "香港特别行政区": "香港",
    "澳门特别行政区": "澳门",
}

PROVINCE_COORDS = {
    "北京": (116.40, 39.90),
    "天津": (117.20, 39.12),
    "上海": (121.47, 31.23),
    "重庆": (106.55, 29.56),
    "河北": (114.48, 38.03),
    "山西": (112.55, 37.87),
    "辽宁": (123.43, 41.80),
    "吉林": (125.32, 43.90),
    "黑龙江": (126.64, 45.76),
    "江苏": (118.78, 32.04),
    "浙江": (120.15, 30.28),
    "安徽": (117.27, 31.86),
    "福建": (119.30, 26.08),
    "江西": (115.89, 28.68),
    "山东": (117.00, 36.67),
    "河南": (113.62, 34.75),
    "湖北": (114.31, 30.52),
    "湖南": (112.94, 28.23),
    "广东": (113.27, 23.13),
    "海南": (110.35, 20.02),
    "四川": (104.07, 30.67),
    "贵州": (106.71, 26.58),
    "云南": (102.71, 25.04),
    "陕西": (108.95, 34.27),
    "甘肃": (103.82, 36.06),
    "青海": (101.78, 36.62),
    "台湾": (121.52, 25.03),
    "内蒙古": (111.67, 40.82),
    "广西": (108.37, 22.82),
    "西藏": (91.11, 29.65),
    "宁夏": (106.28, 38.47),
    "新疆": (87.62, 43.82),
    "香港": (114.17, 22.28),
    "澳门": (113.54, 22.19),
}
VALID_PROVINCES = set(PROVINCE_COORDS)

LUNAR_MONTHS = ["正", "二", "三", "四", "五", "六", "七", "八", "九", "十", "冬", "腊"]
LUNAR_DAYS = [
    "初一", "初二", "初三", "初四", "初五", "初六", "初七", "初八", "初九", "初十",
    "十一", "十二", "十三", "十四", "十五", "十六", "十七", "十八", "十九", "二十",
    "廿一", "廿二", "廿三", "廿四", "廿五", "廿六", "廿七", "廿八", "廿九", "三十",
]
YEAR_COLOR_MAP = {
    "2026年": "#d62728",
    "2025年": "#2ca02c",
    "2024年": "#000000",
    "2023年": "#1f77b4",
    "2022年": "#ff7f0e",
    "2021年": "#9467bd",
}
FALLBACK_YEAR_COLORS = ["#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf", "#393b79", "#637939", "#8c6d31"]
REGION_DIMENSION_LABELS = {
    "none": "全国",
    "province": "地区",
    "city": "城市",
}
FREQUENCY_OPTIONS = ["日度数据", "周度数据（均值）", "月度数据（均值）"]
YONGYI_PRICE_SLAUGHTER_ALLOWED_METRICS = ["全国均价", "日屠宰量合计1", "日度屠宰量合计2"]
YONGYI_PRICE_METRIC = "全国均价"
YONGYI_NO_REGION_CHART_SHEETS = {"价格+宰量", "市场主流标猪肥猪均价方便作图"}
OUT_PRICE_COMPARE_OPTIONS = ["规模场", "小散户", "均价"]
MAINSTREAM_PRICE_COMPARE_OPTIONS = ["90-100kg均价", "130-140kg均价", "150kg左右均价", "标猪均价"]
YONGYI_SHEET_ORDER = ["各省份均价", "价格+宰量", "屠宰企业日度屠宰量", "出栏价", "市场主流标猪肥猪价格", "散户标肥价差", "交割地市出栏价"]
YONGYI_SHEET_DISPLAY_NAMES = {"出栏价": "企散价差"}


# -----------------------------
# 基础工具
# -----------------------------
def inject_css() -> None:
    st.markdown(
        """
        <style>
        /* ===== 整体背景 ===== */
        .stApp {
            background: linear-gradient(160deg, #f4f7fb 0%, #e9eef6 100%);
        }
        .block-container {
            padding-top: 1rem;
            padding-bottom: 2rem;
        }
        /* ===== 指标卡片 ===== */
        .card-wrap {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
            gap: 14px;
            margin: 0.8rem 0 1.2rem 0;
        }
        .metric-card {
            background: #ffffff;
            border-radius: 14px;
            padding: 16px 20px 14px 20px;
            border-top: 3px solid #2563eb;
            border-left: 1px solid #e2e8f0;
            border-right: 1px solid #e2e8f0;
            border-bottom: 1px solid #e2e8f0;
            box-shadow: 0 2px 8px rgba(37,99,235,0.06), 0 1px 3px rgba(0,0,0,0.04);
        }
        .metric-label {
            font-size: 12px;
            font-weight: 600;
            letter-spacing: 0.05em;
            color: #64748b;
            text-transform: uppercase;
            margin-bottom: 8px;
        }
        .metric-value {
            font-size: 22px;
            font-weight: 800;
            color: #0f172a;
            line-height: 1.25;
            word-break: break-word;
        }
        .metric-extra {
            margin-top: 8px;
            font-size: 12px;
            color: #94a3b8;
            line-height: 1.5;
        }
        /* ===== 涨跌颜色 ===== */
        .up   { color: #dc2626; font-weight: 700; }
        .down { color: #16a34a; font-weight: 700; }
        .flat { color: #64748b; font-weight: 700; }
        /* ===== 综合摘要框 ===== */
        .summary-box {
            background: #ffffff;
            border-radius: 14px;
            padding: 20px 24px;
            border: 1px solid #e2e8f0;
            border-left: 4px solid #2563eb;
            box-shadow: 0 2px 8px rgba(37,99,235,0.05);
            margin-top: 12px;
        }
        .summary-title {
            font-size: 20px;
            font-weight: 800;
            color: #0f172a;
            margin-bottom: 8px;
            letter-spacing: -0.01em;
        }
        .summary-subtitle {
            color: #64748b;
            margin-bottom: 10px;
            font-size: 13px;
        }
        .bullet-list {
            margin: 0;
            padding-left: 18px;
            color: #1e293b;
            line-height: 1.9;
            font-size: 14px;
        }
        .mini-note {
            color: #94a3b8;
            font-size: 12px;
            margin-top: 6px;
        }
        /* ===== 异常/机会提示框 ===== */
        .alert-box {
            background: #fffbeb;
            border-radius: 12px;
            padding: 14px 18px;
            border: 1px solid #fde68a;
            border-left: 4px solid #f59e0b;
            margin-top: 10px;
        }
        .alert-title {
            font-size: 15px;
            font-weight: 700;
            color: #78350f;
            margin-bottom: 8px;
        }
        /* ===== 板块分区标题 ===== */
        .section-header {
            font-size: 16px;
            font-weight: 700;
            color: #1e40af;
            padding: 8px 0 4px 0;
            border-bottom: 2px solid #bfdbfe;
            margin: 16px 0 10px 0;
        }
        /* ===== 模块导语 ===== */
        .module-lead {
            background: #eff6ff;
            border-radius: 10px;
            padding: 12px 16px;
            border: 1px solid #bfdbfe;
            font-size: 14px;
            color: #1e40af;
            margin-bottom: 14px;
        }
        .highlight-control-box {
            background: #fff7ed;
            border: 1px solid #fdba74;
            border-left: 4px solid #f97316;
            border-radius: 12px;
            padding: 10px 12px 4px 12px;
            margin: 8px 0 12px 0;
        }
        .highlight-control-title {
            font-size: 14px;
            font-weight: 800;
            color: #9a3412;
            margin-bottom: 6px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def resolve_excel_path(raw_path: str) -> Path:
    raw = (raw_path or "").strip().strip('"').strip("'")
    base = Path(raw)
    candidates: list[Path] = []
    if base.suffix:
        candidates.append(base)
    else:
        candidates.extend([base, Path(f"{raw}.xlsx"), Path(f"{raw}.xlsm"), Path(f"{raw}.xls")])
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    if base.parent.exists():
        fuzzy = sorted(
            p for p in base.parent.glob(f"{base.name}*") if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
        )
        if fuzzy:
            return fuzzy[0]
    raise FileNotFoundError(f"未找到Excel文件：{raw_path}")


def text_of(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    return str(value).strip()


def parse_date(value: Any) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        num = float(value)
        # 仅当数值在合理 Excel 日期序列号范围内（~2014~2040）才转换为日期，
        # 避免将屠宰头数等业务数据（2~6万）误判为日期
        if 42000 <= num <= 52000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(num, unit="D")).normalize()
            except Exception:
                return None
        return None
    text = text_of(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        match = re.search(r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})", text)
        if not match:
            return None
        try:
            parsed = pd.Timestamp(year=int(match.group(1)), month=int(match.group(2)), day=int(match.group(3)))
        except Exception:
            return None
    return pd.Timestamp(parsed).normalize()


def parse_date_range_text(value: Any) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    text = text_of(value)
    if not text:
        return None, None
    compact = text.replace(" ", "").replace("（", "").replace("）", "")
    full = re.search(
        r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})[日号]?\s*[-~—至]+\s*(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})",
        compact,
    )
    if full:
        start = pd.Timestamp(year=int(full.group(1)), month=int(full.group(2)), day=int(full.group(3)))
        end = pd.Timestamp(year=int(full.group(4)), month=int(full.group(5)), day=int(full.group(6)))
        return start.normalize(), end.normalize()
    same_year = re.search(
        r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})[日号]?\s*[-~—至]+\s*(\d{1,2})[./月-](\d{1,2})",
        compact,
    )
    if same_year:
        year = int(same_year.group(1))
        start = pd.Timestamp(year=year, month=int(same_year.group(2)), day=int(same_year.group(3)))
        end = pd.Timestamp(year=year, month=int(same_year.group(4)), day=int(same_year.group(5)))
        return start.normalize(), end.normalize()
    single = parse_date(text)
    return single, single


def parse_month_range_text(value: Any) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    text = text_of(value).replace(" ", "")
    if not text:
        return None, None
    full = re.search(r"(20\d{2})年?(\d{1,2})[-~—至](\d{1,2})月", text)
    if full:
        year = int(full.group(1))
        start_month = int(full.group(2))
        end_month = int(full.group(3))
        try:
            start = pd.Timestamp(year=year, month=start_month, day=1)
            end = pd.Timestamp(year=year, month=end_month, day=1)
            return start.normalize(), end.normalize()
        except Exception:
            return None, None
    single = re.search(r"(20\d{2})年?(\d{1,2})月", text)
    if single:
        try:
            ts = pd.Timestamp(year=int(single.group(1)), month=int(single.group(2)), day=1)
            return ts.normalize(), ts.normalize()
        except Exception:
            return None, None
    return None, None


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        return float(value)
    text = (
        str(value)
        .replace(",", "")
        .replace("%", "")
        .replace("—", "")
        .replace("--", "")
        .replace(" ", "")
        .strip()
    )
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def canonicalize_province(name: str) -> str:
    text = text_of(name)
    if not text:
        return "全部"
    if text in PROVINCE_REPLACEMENTS:
        return PROVINCE_REPLACEMENTS[text]
    for suffix in ["省", "市", "自治区", "特别行政区"]:
        if text.endswith(suffix) and len(text) > len(suffix):
            text = text[: -len(suffix)]
            break
    return text or "全部"


def clean_city_name(name: str) -> str:
    text = text_of(name)
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    for suffix in ["市", "地区", "自治州", "盟"]:
        if text.endswith(suffix) and len(text) > len(suffix):
            text = text[: -len(suffix)]
            break
    return text


def is_valid_city_name(name: str) -> bool:
    text = clean_city_name(name)
    if len(text) < 2:
        return False
    if any(flag in text for flag in ["或", "、", "/", "→", "到", "装", "明天", "今天", "半挂"]):
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fff]{2,12}", text))


def parse_price_band(value: Any) -> tuple[float | None, float | None, float | None, str]:
    raw = text_of(value)
    if not raw:
        return None, None, None, ""
    text = raw.replace("～", "-").replace("~", "-").replace("至", "-").replace("—", "-").replace("－", "-")
    nums = re.findall(r"-?\d+(?:\.\d+)?", text)
    if not nums:
        return None, None, None, raw
    if len(nums) == 1:
        low = high = float(nums[0])
    else:
        low, high = float(nums[0]), float(nums[1])
        if high < low:
            low, high = high, low
    return low, high, (low + high) / 2, raw


def lunar_label(ts: pd.Timestamp) -> tuple[str | None, float | None, bool, int | None]:
    if LunarDate is None or pd.isna(ts):
        return None, None, False, None
    try:
        lunar = LunarDate.fromSolarDate(ts.year, ts.month, ts.day)
        is_leap = bool(getattr(lunar, "isLeapMonth", False))
        month_name = LUNAR_MONTHS[lunar.month - 1] if 1 <= lunar.month <= 12 else str(lunar.month)
        day_name = LUNAR_DAYS[lunar.day - 1] if 1 <= lunar.day <= 30 else str(lunar.day)
        label = f"{month_name}月{day_name}"
        order = float(lunar.month * 100 + lunar.day)
        lunar_year = int(getattr(lunar, "year", ts.year))
        return label, order, is_leap, lunar_year
    except Exception:
        return None, None, False, None


def enrich_date_features(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    temp = df.copy()
    temp["date"] = pd.to_datetime(temp["date"]).dt.normalize()
    temp["year"] = temp["date"].dt.year
    temp["month"] = temp["date"].dt.month
    temp["week"] = temp["date"].dt.isocalendar().week.astype(int)
    temp["month_day"] = temp["date"].dt.strftime("%m-%d")
    lunar = temp["date"].map(lunar_label)
    temp["lunar_label"] = lunar.map(lambda x: x[0])
    temp["lunar_order"] = lunar.map(lambda x: x[1])
    temp["is_leap_lunar"] = lunar.map(lambda x: x[2])
    temp["lunar_year"] = lunar.map(lambda x: x[3])
    return temp


def get_year_color_map(year_labels: list[str]) -> dict[str, str]:
    year_numbers: list[int] = []
    for label in year_labels:
        text = str(label).replace("年", "")
        if text.isdigit():
            year_numbers.append(int(text))
    fallback_years = sorted({year for year in year_numbers if f"{year}年" not in YEAR_COLOR_MAP}, reverse=True)
    color_map = dict(YEAR_COLOR_MAP)
    for idx, year in enumerate(fallback_years):
        color_map[f"{year}年"] = FALLBACK_YEAR_COLORS[idx % len(FALLBACK_YEAR_COLORS)]
    return color_map


def is_national_scope(name: str) -> bool:
    text = text_of(name)
    return text == "全部" or any(flag in text for flag in ["全国", "全国均值", "全国均价"])


def build_plotly_key(*parts: Any) -> str:
    raw = "|".join(text_of(part) for part in parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def get_series_dimension(df: pd.DataFrame) -> str:
    if "city" in df.columns and df["city"].fillna("").ne("").any():
        return "city"
    unique_provinces = [p for p in df.get("province", pd.Series(dtype=object)).dropna().unique() if p != "全部"]
    return "province" if unique_provinces else "none"


def get_group_field(df: pd.DataFrame) -> str | None:
    dimension = get_series_dimension(df)
    if dimension == "city":
        return "city"
    if dimension == "province":
        return "province"
    return None


def get_group_label(df: pd.DataFrame) -> str:
    return REGION_DIMENSION_LABELS[get_series_dimension(df)]


def has_group_dimension(df: pd.DataFrame) -> bool:
    return get_group_field(df) is not None


def supports_region_charts(df: pd.DataFrame, sheet: str) -> bool:
    if sheet in YONGYI_NO_REGION_CHART_SHEETS:
        return False
    return has_group_dimension(df)


def pick_preferred_scope_name(options: list[str]) -> str | None:
    cleaned = [text_of(x) for x in options if text_of(x)]
    if not cleaned:
        return None
    for name in cleaned:
        if name == "全部":
            return name
    for name in cleaned:
        if is_national_scope(name):
            return name
    if "云南" in cleaned:
        return "云南"
    return cleaned[0]


def get_default_scope_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "province" not in df.columns:
        return df.copy()
    preferred_name = pick_preferred_scope_name(df["province"].dropna().unique().tolist())
    if preferred_name is None:
        return df.copy()
    result = df[df["province"] == preferred_name].copy()
    if not result.empty:
        result["display_name"] = "全国" if is_national_scope(preferred_name) else preferred_name
    return result


def normalize_out_price_metric(metric: str) -> str | None:
    text = text_of(metric)
    if not text:
        return None
    if "规模场" in text:
        return "规模场"
    if "小散户" in text or ("散户" in text and "大" not in text):
        return "小散户"
    if "均价" in text:
        return "均价"
    return None


def get_available_group_options(df: pd.DataFrame, group_field: str | None) -> list[str]:
    if group_field is None or group_field not in df.columns:
        return []
    values = [text_of(x) for x in df[group_field].dropna().unique()]
    values = [x for x in values if x and x != "全部"]
    return sorted(values)


def format_number(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "暂无"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    return f"{value:,.2f}"


def format_date_cn(ts: pd.Timestamp | None) -> str:
    if ts is None or pd.isna(ts):
        return "暂无日期"
    ts = pd.Timestamp(ts)
    return f"{ts.year}年{ts.month}月{ts.day}日"


def format_month_day_cn(ts: pd.Timestamp | None) -> str:
    if ts is None or pd.isna(ts):
        return ""
    ts = pd.Timestamp(ts)
    return f"{ts.month}月{ts.day}日"


def format_week_cn(ts: pd.Timestamp | None) -> str:
    if ts is None or pd.isna(ts):
        return ""
    ts = pd.Timestamp(ts)
    iso = ts.isocalendar()
    return f"{int(iso.year)}年第{int(iso.week)}周"


def format_month_cn(ts: pd.Timestamp | None) -> str:
    if ts is None or pd.isna(ts):
        return ""
    ts = pd.Timestamp(ts)
    return f"{ts.year}年{ts.month}月"


def delta_percent(delta: float | None, base: float | None) -> str:
    if delta is None or pd.isna(delta) or base in (None, 0) or pd.isna(base):
        return ""
    return f"（{delta / base:+.1%}）"


def build_cn_date_range_selector(df: pd.DataFrame, key_prefix: str, title: str = "📅 图表日期范围") -> tuple[pd.Timestamp, pd.Timestamp, pd.DataFrame]:
    dates = df["date"].dropna()
    if dates.empty:
        today = pd.Timestamp.today().normalize()
        st.session_state[f"{key_prefix}_range_label"] = "全部数据"
        st.session_state[f"{key_prefix}_range_text"] = f"图表范围：{format_date_cn(today)} — {format_date_cn(today)}"
        return today, today, df.iloc[0:0].copy()
    dmin = pd.Timestamp(dates.min()).date()
    dmax = pd.Timestamp(dates.max()).date()

    range_options_map: dict[str, tuple] = {"全部数据": (dmin, dmax), "自定义日期": (dmin, dmax)}
    for yr in sorted(set(range(dmin.year, dmax.year + 1)), reverse=True):
        ys = max(dmin, __import__('datetime').date(yr, 1, 1))
        ye = min(dmax, __import__('datetime').date(yr, 12, 31))
        if ys <= ye:
            range_options_map[f"{yr}年"] = (ys, ye)
    for label, months in [("近1个月", 1), ("近3个月", 3), ("近6个月", 6), ("近1年", 12)]:
        ms = pd.Timestamp(dmax) - pd.DateOffset(months=months)
        ms_d = max(dmin, ms.date())
        if ms_d < dmax:
            range_options_map[label] = (ms_d, dmax)

    selected_label = st.selectbox(title, list(range_options_map.keys()), index=0, key=f"{key_prefix}_range_select")
    start_d, end_d = range_options_map[selected_label]
    if selected_label == "自定义日期":
        c1, c2 = st.columns(2)
        with c1:
            start_d = st.date_input("开始日期", value=dmin, min_value=dmin, max_value=dmax, key=f"{key_prefix}_start")
        with c2:
            end_d = st.date_input("结束日期", value=dmax, min_value=dmin, max_value=dmax, key=f"{key_prefix}_end")
    range_text = f"图表范围：{format_date_cn(pd.Timestamp(start_d))} — {format_date_cn(pd.Timestamp(end_d))}"
    st.session_state[f"{key_prefix}_range_label"] = selected_label
    st.session_state[f"{key_prefix}_range_text"] = range_text
    st.caption(range_text)
    filtered = df[(df["date"] >= pd.Timestamp(start_d)) & (df["date"] <= pd.Timestamp(end_d))].copy()
    return pd.Timestamp(start_d), pd.Timestamp(end_d), filtered


def delta_html(delta: float | None, base: float | None = None, unit_prefix: str = "") -> str:
    if delta is None or pd.isna(delta):
        return '<span class="flat">暂无可比数据</span>'
    if abs(delta) < 1e-12:
        return '<span class="flat">持平</span>'
    cls = "up" if delta > 0 else "down"
    direction = "上涨" if delta > 0 else "下跌"
    percent = delta_percent(delta, base)
    return f'<span class="{cls}">{direction} {unit_prefix}{format_number(abs(delta))}{percent}</span>'


def describe_delta(delta: float | None, label: str = "") -> str:
    if delta is None or pd.isna(delta):
        return f"{label}暂无可比数据" if label else "暂无可比数据"
    if abs(delta) < 1e-12:
        return f"{label}持平" if label else "持平"
    direction = "增加" if delta > 0 else "减少"
    return f"{label}{direction} {format_number(abs(delta))}"


def trend_position_label(current_value: float, history: pd.Series) -> str:
    valid = history.dropna()
    if len(valid) < 5:
        return "样本不足，暂不判断所处区间"
    q33 = valid.quantile(0.33)
    q67 = valid.quantile(0.67)
    if current_value >= q67:
        return "处于近阶段偏高区间"
    if current_value <= q33:
        return "处于近阶段偏低区间"
    return "处于近阶段中性区间"


def percentile_position_text(current_value: float, history: pd.Series) -> tuple[str, float | None]:
    valid = history.dropna()
    if len(valid) < 10:
        return "样本不足", None
    percentile = float((valid <= current_value).mean())
    if percentile >= 0.8:
        return "历史偏高", percentile
    if percentile <= 0.2:
        return "历史偏低", percentile
    return "历史中性", percentile


def same_day_last_year_match(dates: pd.Series, target_date: pd.Timestamp) -> pd.Timestamp | None:
    if dates.empty:
        return None
    target = pd.Timestamp(target_date) - pd.DateOffset(years=1)
    candidates = pd.Series(pd.to_datetime(dates).dropna().unique())
    if candidates.empty:
        return None
    year_candidates = candidates[candidates.dt.year == target.year]
    if year_candidates.empty:
        return None
    deltas = (year_candidates - target).abs()
    idx = deltas.argmin()
    matched = pd.Timestamp(year_candidates.iloc[idx]).normalize()
    if abs((matched - target).days) > 7:
        return None
    return matched


def ensure_single_choice(key: str, options: list[str], default: str | None = None) -> None:
    if not options:
        return
    if default is None:
        default = options[0]
    if key not in st.session_state or st.session_state[key] not in options:
        st.session_state[key] = default if default in options else options[0]


def ensure_multi_choice(key: str, options: list[str], default: list[str]) -> None:
    existing = st.session_state.get(key, default)
    valid = [item for item in existing if item in options]
    if not valid:
        valid = [item for item in default if item in options] or options[:1]
    st.session_state[key] = valid


def render_metric_cards(cards: list[dict[str, str]]) -> None:
    html = '<div class="card-wrap">'
    for card in cards:
        html += (
            '<div class="metric-card">'
            f'<div class="metric-label">{card["label"]}</div>'
            f'<div class="metric-value">{card["value"]}</div>'
            f'<div class="metric-extra">{card.get("extra", "")}</div>'
            '</div>'
        )
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def render_summary_box(title: str, subtitle: str, body: str) -> None:
    st.markdown(
        f"<div class='summary-box'><div class='summary-title'>{title}</div><div class='summary-subtitle'>{subtitle}</div><div style='font-size:18px;font-weight:700;color:#14213d;line-height:1.8'>{body}</div></div>",
        unsafe_allow_html=True,
    )


def render_signal_messages(title: str, messages: list[str], empty_message: str = "当前未发现明显异常。") -> None:
    if messages:
        bullets = "".join(f"<li>{msg}</li>" for msg in messages)
        html = (
            "<div class='alert-box'>"
            f"<div class='alert-title'>⚑ {title}</div>"
            f"<ul class='bullet-list'>{bullets}</ul>"
            "</div>"
        )
        st.markdown(html, unsafe_allow_html=True)
    else:
        st.info(f"✓ {empty_message}")


def render_market_signals(signals: list[dict]) -> None:
    """渲染业务预警信号卡片，level: danger/warning/info"""
    if not signals:
        st.info("✓ 当前未发现明显业务预警信号。")
        return
    level_color = {"danger": "#fef2f2", "warning": "#fffbeb", "info": "#eff6ff"}
    level_border = {"danger": "#fca5a5", "warning": "#fcd34d", "info": "#93c5fd"}
    level_icon = {"danger": "🔴", "warning": "🟡", "info": "🔵"}
    for sig in signals:
        lv = sig.get("level", "info")
        bg = level_color.get(lv, "#f8fafc")
        border = level_border.get(lv, "#cbd5e1")
        icon = level_icon.get(lv, "ℹ️")
        st.markdown(
            f"<div style='background:{bg};border-left:4px solid {border};"
            f"border-radius:8px;padding:12px 16px;margin-bottom:10px'>"
            f"<div style='font-weight:700;font-size:14px;margin-bottom:4px'>{icon} {sig['title']}</div>"
            f"<div style='font-size:13px;color:#374151;line-height:1.6'>{sig['body']}</div>"
            "</div>",
            unsafe_allow_html=True,
        )


def render_section_header(title: str) -> None:
    st.markdown(f"<div class='section-header'>{title}</div>", unsafe_allow_html=True)


def render_module_lead(text: str) -> None:
    st.markdown(f"<div class='module-lead'>{text}</div>", unsafe_allow_html=True)


def normalize_weekly_scope_label(label: str) -> str:
    text = text_of(label).replace("\n", "").replace(" ", "")
    if not text:
        return ""
    if text == "全国均重":
        return text
    text = text.replace("全国均值", "全国").replace("全国平均", "全国").replace("全国1", "全国").replace("全国2", "全国").replace("全国3", "全国")
    if text.startswith("全") and "国" in text:
        return "全国"
    return text


def infer_region_type(series_name: str, province: str = "") -> str:
    label = normalize_weekly_scope_label(series_name or province)
    if not label:
        return "single"
    if label == "全国":
        return "national"
    if label in VALID_PROVINCES:
        return "province"
    if any(flag in label for flag in ["华北", "华中", "华东", "华南", "东北", "西北", "西南"]):
        return "region"
    return "single"


def apply_topic_to_records(records: list[dict], topic: str) -> list[dict]:
    for record in records:
        if record.get("topic") in [None, "", np.nan]:
            record["topic"] = topic
    return records


def build_cross_sheet_spread_df(left_df: pd.DataFrame, right_df: pd.DataFrame, spread_name: str) -> pd.DataFrame:
    if left_df.empty or right_df.empty:
        return pd.DataFrame(columns=left_df.columns if not left_df.empty else right_df.columns)
    group_cols = ["date"]
    if "province" in left_df.columns and "province" in right_df.columns:
        group_cols.append("province")
    left = left_df[group_cols + ["value"]].rename(columns={"value": "left_value"})
    right = right_df[group_cols + ["value"]].rename(columns={"value": "right_value"})
    merged = left.merge(right, on=group_cols, how="inner")
    if merged.empty:
        return pd.DataFrame(columns=left_df.columns)
    result = pd.DataFrame({
        "date": merged["date"],
        "province": merged["province"] if "province" in merged.columns else "全部",
        "city": "",
        "metric": spread_name,
        "series_name": merged["province"] if "province" in merged.columns else spread_name,
        "display_name": merged["province"] if "province" in merged.columns else spread_name,
        "value": merged["right_value"] - merged["left_value"],
        "sheet": left_df["sheet"].iloc[0] if "sheet" in left_df.columns and not left_df.empty else spread_name,
    })
    for col in ["freq_type", "topic", "region", "region_type", "source_row_type"]:
        if col in left_df.columns:
            result[col] = left_df[col].dropna().iloc[0] if not left_df[col].dropna().empty else None
    return enrich_date_features(result)


def build_weekly_record(
    *,
    sheet: str,
    date: Any,
    metric: str,
    series_name: str,
    value: Any,
    province: str = "全部",
    city: str = "",
    display_name: str | None = None,
    freq_type: str = "周度",
    start_date: Any = None,
    end_date: Any = None,
    region: str = "",
    region_type: str = "",
    low_price: float | None = None,
    high_price: float | None = None,
    raw_range: str = "",
    raw_text: str = "",
    source_row_type: str = "normal",
    extra: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    start_ts = parse_date(start_date)
    end_ts = parse_date(end_date)
    ts = parse_date(date) or end_ts or start_ts
    if ts is None:
        return None
    number = to_float(value)
    if number is None:
        return None
    province_name = canonicalize_province(province)
    series_text = normalize_weekly_scope_label(series_name) or text_of(series_name) or province_name
    region_text = normalize_weekly_scope_label(region) or text_of(region)
    record = {
        "sheet": sheet,
        "freq_type": freq_type,
        "date": ts,
        "start_date": start_ts,
        "end_date": end_ts,
        "metric": text_of(metric) or "数值",
        "series_name": series_text,
        "display_name": display_name or series_text,
        "province": province_name,
        "city": text_of(city),
        "region": region_text,
        "region_type": region_type or infer_region_type(series_text, province_name),
        "value": float(number),
        "low_price": low_price,
        "high_price": high_price,
        "price_range_text": raw_range,
        "raw_range": raw_range,
        "raw_text": raw_text,
        "source_row_type": source_row_type,
    }
    if extra:
        record.update(extra)
    return record


def parse_latest_date_from_rows(rows: list[list[Any]]) -> pd.Timestamp | None:
    latest: pd.Timestamp | None = None
    for row in rows[:12]:
        for cell in row:
            ts = parse_date(cell)
            if ts is not None and (latest is None or ts > latest):
                latest = ts
            start_ts, end_ts = parse_date_range_text(cell)
            for candidate in [start_ts, end_ts]:
                if candidate is not None and (latest is None or candidate > latest):
                    latest = candidate
            month_start, month_end = parse_month_range_text(cell)
            for candidate in [month_start, month_end]:
                if candidate is not None and (latest is None or candidate > latest):
                    latest = candidate
    return latest


def infer_weekly_freq_type(sheet_name: str) -> str:
    text = text_of(sheet_name)
    if text in {"二育栏舍利用率"}:
        return "旬度"
    if text.startswith("月度") or "月度" in text or "月份" in text:
        return "月度"
    return "周度"


def normalize_weekly_header_label(label: Any) -> str:
    text = text_of(label).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip(" ·")
    return normalize_weekly_scope_label(text)


def is_weekly_auxiliary_column(label: str) -> bool:
    text = normalize_weekly_header_label(label)
    if not text:
        return True
    return any(flag in text for flag in ["环比", "同比", "较上周", "较去年", "增减", "变化", "备注", "说明", "注"])


def detect_weekly_scope(label: str) -> tuple[str, str, str]:
    cleaned = normalize_weekly_header_label(label)
    if not cleaned:
        return "全部", "", "single"
    if is_national_scope(cleaned):
        return "全部", "全国", "national"
    province = canonicalize_province(cleaned)
    if province in VALID_PROVINCES:
        return province, "", "province"
    if any(flag in cleaned for flag in ["华北", "华中", "华东", "华南", "东北", "西北", "西南"]):
        return "全部", cleaned, "region"
    return "全部", "", "single"


def build_weekly_series_name(base_label: str, metric_label: str = "", suffix: str = "") -> str:
    parts = [normalize_weekly_header_label(base_label), normalize_weekly_header_label(metric_label), normalize_weekly_header_label(suffix)]
    return "｜".join([part for part in parts if part]) or "数值"


def standardize_weekly_headers(merged_header: list[str]) -> list[str]:
    return [normalize_weekly_header_label(item) for item in merged_header]


def make_weekly_period_record(
    *,
    sheet_name: str,
    start_date: Any,
    end_date: Any,
    metric: str,
    series_name: str,
    value: Any,
    freq_type: str,
    source_row_type: str = "normal",
    raw_text: str = "",
    extra: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    low, high, mid, raw_range = parse_price_band(value)
    numeric_value = mid if mid is not None else value
    province, region, region_type = detect_weekly_scope(series_name)
    return build_weekly_record(
        sheet=sheet_name,
        date=end_date,
        start_date=start_date,
        end_date=end_date,
        metric=metric,
        series_name=series_name,
        value=numeric_value,
        province=province,
        region=region,
        region_type=region_type,
        low_price=low,
        high_price=high,
        raw_range=raw_range if raw_range != text_of(value) else "",
        raw_text=raw_text,
        source_row_type=source_row_type,
        freq_type=freq_type,
        extra=extra,
    )


def parse_weekly_text_period_cell(value: Any, fallback_year: int | None = None) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    start_ts, end_ts = parse_date_range_text(value)
    if start_ts is not None or end_ts is not None:
        return start_ts, end_ts
    text = text_of(value).replace(" ", "")
    if not text:
        return None, None
    match = re.search(r"(\d{1,2})[./月-](\d{1,2})[日号]?[-~—至]+(\d{1,2})[./月-](\d{1,2})", text)
    if match and fallback_year:
        try:
            start = pd.Timestamp(year=fallback_year, month=int(match.group(1)), day=int(match.group(2)))
            end = pd.Timestamp(year=fallback_year, month=int(match.group(3)), day=int(match.group(4)))
            if end < start:
                end = pd.Timestamp(year=fallback_year + 1, month=int(match.group(3)), day=int(match.group(4)))
            return start.normalize(), end.normalize()
        except Exception:
            return None, None
    return None, None


def parse_weekly_month_marker(value: Any, current_year: int | None) -> pd.Timestamp | None:
    direct = parse_date(value)
    if direct is not None:
        return direct
    text = text_of(value).replace(" ", "")
    if not text:
        return None
    month_only = re.fullmatch(r"(\d{1,2})月", text)
    if month_only and current_year:
        try:
            return pd.Timestamp(year=current_year, month=int(month_only.group(1)), day=1).normalize()
        except Exception:
            return None
    return None


def parse_month_or_range_anchor(value: Any, current_year: int | None = None) -> tuple[pd.Timestamp | None, pd.Timestamp | None, pd.Timestamp | None, int | None]:
    start_ts, end_ts = parse_month_range_text(value)
    if start_ts is not None or end_ts is not None:
        anchor = end_ts or start_ts
        return start_ts, end_ts, anchor, int(anchor.year) if anchor is not None else current_year
    anchor = parse_weekly_month_marker(value, current_year)
    if anchor is not None:
        return anchor, anchor, anchor, int(anchor.year)
    return None, None, None, current_year


def parse_monthly_breeding_ratio_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 2:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[0]]
    current_year: int | None = None
    for row in rows[1:]:
        if not row:
            continue
        period_cell = row[0] if len(row) > 0 else None
        start_ts, end_ts, anchor, current_year = parse_month_or_range_anchor(period_cell, current_year)
        if anchor is None:
            continue
        period_text = text_of(period_cell)
        for col in range(1, len(headers)):
            metric = headers[col]
            if not metric or col >= len(row):
                continue
            record = make_weekly_period_record(
                sheet_name=sheet_name,
                start_date=start_ts,
                end_date=end_ts,
                metric=metric,
                series_name=metric,
                value=row[col],
                freq_type="月度",
                raw_text=period_text,
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_profit_compare_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 3:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[2]]
    for row in rows[3:]:
        if not row:
            continue
        start_ts, end_ts = parse_weekly_text_period_cell(row[0] if len(row) > 0 else None)
        if end_ts is None:
            continue
        for col in range(1, len(headers)):
            metric = headers[col]
            if not metric or col >= len(row):
                continue
            record = make_weekly_period_record(
                sheet_name=sheet_name,
                start_date=start_ts,
                end_date=end_ts,
                metric=metric,
                series_name=metric,
                value=row[col],
                freq_type="周度",
                raw_text=text_of(row[0]),
            )
            if record is not None:
                records.append(record)
    return records


# -----------------------------
# Excel 原始行读取
# -----------------------------
def workbook_to_rows(workbook) -> dict[str, list[list[Any]]]:
    sheets: dict[str, list[list[Any]]] = {}
    for ws in workbook.worksheets:
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        sheets[ws.title] = rows
    return sheets


def _get_file_version(path_str: str) -> str:
    """返回带文件修改时间的版本字符串，用于缓存键自动失效。"""
    try:
        path = resolve_excel_path(path_str)
        stat = path.stat()
        return f"{path_str}@@v={stat.st_mtime:.6f}_sz={stat.st_size}"
    except Exception:
        return path_str


@st.cache_data(show_spinner=False)
def _cached_read_workbook_rows(versioned_path: str) -> dict[str, list[list[Any]]]:
    """内部缓存函数：versioned_path 包含 mtime，文件更新时自动失效。"""
    # 提取真实路径（去掉 @@v=... 后缀）
    real_path = versioned_path.split("@@v=")[0]
    path = resolve_excel_path(real_path)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = workbook_to_rows(wb)
    wb.close()
    return sheets


def read_workbook_rows_from_path(path_str: str) -> dict[str, list[list[Any]]]:
    """读取 Excel 工作簿，缓存自动感知文件更新时间。"""
    return _cached_read_workbook_rows(_get_file_version(path_str))


@st.cache_data(show_spinner=False)
def read_workbook_rows_from_bytes(file_bytes: bytes) -> dict[str, list[list[Any]]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = workbook_to_rows(wb)
    wb.close()
    return sheets


# -----------------------------
# 涌益数据解析
# -----------------------------
def add_numeric(
    records: list[dict],
    sheet: str,
    date: pd.Timestamp | None,
    province: str,
    metric: str,
    value: Any,
    city: str = "",
    display_name: str | None = None,
) -> None:
    date = parse_date(date)
    metric = text_of(metric)
    number = to_float(value)
    city = text_of(city)
    if date is None or not metric or number is None:
        return
    province_name = canonicalize_province(province)
    records.append(
        {
            "sheet": sheet,
            "date": date,
            "province": province_name,
            "city": city,
            "metric": metric,
            "display_name": display_name or city or province_name,
            "value": number,
        }
    )


def filled_right(values: list[Any]) -> list[Any]:
    result = []
    current = None
    for value in values:
        if text_of(value):
            current = value
        result.append(current)
    return result


def parse_repeated_date_blocks(rows: list[list[Any]], sheet: str, start_col: int, province_col: int = 0, data_start_row: int = 2) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= data_start_row:
        return records
    max_len = max(len(row) for row in rows[: min(len(rows), data_start_row + 4)])
    date_row = rows[0] + [None] * (max_len - len(rows[0]))
    metric_row = rows[1] + [None] * (max_len - len(rows[1]))
    dates = filled_right([parse_date(x) for x in date_row])

    for row in rows[data_start_row:]:
        padded = row + [None] * (max_len - len(row))
        province = text_of(padded[province_col])
        if not province or province in {"日期", "省份"}:
            continue
        for col in range(start_col, max_len):
            date = dates[col]
            metric = text_of(metric_row[col])
            value = padded[col]
            if date is None or not metric:
                continue
            normalized_metric = normalize_out_price_metric(metric) if sheet == "出栏价" else metric
            if sheet == "出栏价" and normalized_metric is None:
                continue
            add_numeric(records, sheet, date, province, normalized_metric, value)
    return records


def parse_simple_metric_table(rows: list[list[Any]], sheet: str, date_col: int = 0, metric_start_col: int | None = None) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= 1:
        return records
    metrics = [text_of(x) for x in rows[0]]
    metric_start = metric_start_col if metric_start_col is not None else date_col + 1
    for row in rows[1:]:
        if not row:
            continue
        date = parse_date(row[date_col] if len(row) > date_col else None)
        if date is None:
            continue
        for col in range(metric_start, len(metrics)):
            metric = metrics[col]
            if not metric:
                continue
            value = row[col] if col < len(row) else None
            add_numeric(records, sheet, date, "全部", metric, value)
    return records


def parse_date_rows_province_columns(rows: list[list[Any]], sheet: str, metric_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= 1:
        return records
    provinces = [text_of(x) for x in rows[0]]
    for row in rows[1:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for col in range(1, len(provinces)):
            province = provinces[col]
            if not province:
                continue
            value = row[col] if col < len(row) else None
            add_numeric(records, sheet, date, province, metric_name, value)
    return records


def parse_province_rows_date_columns(rows: list[list[Any]], sheet: str, metric_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= 1:
        return records
    dates = [parse_date(x) for x in rows[0]]
    for row in rows[1:]:
        if not row:
            continue
        province = text_of(row[0] if len(row) > 0 else None)
        if not province or province == "省份":
            continue
        for col in range(1, len(dates)):
            date = dates[col]
            if date is None:
                continue
            value = row[col] if col < len(row) else None
            add_numeric(records, sheet, date, province, metric_name, value)
    return records


def normalize_delivery_meta_metric(metric: Any) -> str:
    text = text_of(metric).replace(" ", "")
    if not text:
        return ""
    if "LH2505及以后" in text:
        return "premium_lh2505_plus"
    if "LH2409-LH2503" in text:
        return "premium_lh2409_to_lh2503"
    if "交易均重" in text:
        return "交易均重"
    return text_of(metric).strip()


def build_delivery_premium_df(metadata_df: pd.DataFrame) -> pd.DataFrame:
    if metadata_df.empty:
        return pd.DataFrame(columns=["province", "city", "premium_lh2505_plus", "premium_lh2409_to_lh2503"])
    temp = metadata_df.copy()
    temp["meta_metric_normalized"] = temp["meta_metric"].map(normalize_delivery_meta_metric)
    temp = temp[temp["meta_metric_normalized"].isin(["premium_lh2505_plus", "premium_lh2409_to_lh2503"])].copy()
    if temp.empty:
        return pd.DataFrame(columns=["province", "city", "premium_lh2505_plus", "premium_lh2409_to_lh2503"])
    temp["meta_value"] = temp["meta_value"].map(to_float)
    pivot = temp.pivot_table(index=["province", "city"], columns="meta_metric_normalized", values="meta_value", aggfunc="first").reset_index()
    for col in ["premium_lh2505_plus", "premium_lh2409_to_lh2503"]:
        if col not in pivot.columns:
            pivot[col] = np.nan
    return pivot


def parse_delivery_city_sheet(rows: list[list[Any]], sheet: str) -> tuple[list[dict], list[dict]]:
    numeric_records: list[dict] = []
    metadata_records: list[dict] = []
    if len(rows) <= 4:
        return numeric_records, metadata_records
    province_row = rows[0]
    city_row = rows[4] if len(rows) > 4 else []
    provinces = filled_right([text_of(x) for x in province_row])

    for meta_row_idx in [1, 2]:
        if meta_row_idx >= len(rows):
            continue
        meta_row = rows[meta_row_idx]
        metric = text_of(meta_row[0] if len(meta_row) > 0 else None)
        if not metric:
            continue
        clean_metric = metric.split("（")[0].split("(")[0].strip()
        for col in range(1, min(len(meta_row), len(provinces), len(city_row))):
            city = text_of(city_row[col] if col < len(city_row) else None)
            province = provinces[col]
            value = meta_row[col]
            if city and text_of(value) and clean_metric != "交易均重":
                metadata_records.append(
                    {
                        "sheet": sheet,
                        "province": canonicalize_province(province),
                        "city": city,
                        "meta_metric": clean_metric,
                        "meta_metric_normalized": normalize_delivery_meta_metric(clean_metric),
                        "meta_value": text_of(value),
                    }
                )

    for row in rows[5:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for col in range(1, min(len(row), len(provinces), len(city_row))):
            province = provinces[col]
            city = text_of(city_row[col] if col < len(city_row) else None)
            if not province or not city:
                continue
            value = row[col]
            add_numeric(numeric_records, sheet, date, province, "交割地出栏价", value, city=city)
    return numeric_records, metadata_records


def finalize_numeric_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        cols = ["sheet", "date", "province", "city", "display_name", "metric", "value", "year", "month", "week", "month_day", "lunar_label", "lunar_order", "is_leap_lunar"]
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(records)
    if "city" not in df.columns:
        df["city"] = ""
    if "display_name" not in df.columns:
        df["display_name"] = df["city"].where(df["city"].fillna("").ne(""), df["province"])
    df = enrich_date_features(df)
    return df.sort_values(["sheet", "metric", "province", "city", "date"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def build_yongyi_dataset_from_rows(sheets: dict[str, list[list[Any]]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    numeric_records: list[dict] = []
    metadata_records: list[dict] = []
    logs: list[dict] = []
    for sheet, rows in sheets.items():
        numeric_before = len(numeric_records)
        metadata_before = len(metadata_records)
        if sheet == "价格+宰量":
            price_slaughter_records = parse_simple_metric_table(rows, sheet, date_col=1, metric_start_col=2)
            numeric_records.extend(
                record for record in price_slaughter_records if record["metric"] in YONGYI_PRICE_SLAUGHTER_ALLOWED_METRICS
            )
            mode = "单日期列+多指标（限定价格/宰量口径）"
        elif sheet == "各省份均价":
            numeric_records.extend(parse_date_rows_province_columns(rows, sheet, "均价"))
            mode = "日期行为主、各省份为列"
        elif sheet == "屠宰企业日度屠宰量":
            numeric_records.extend(parse_province_rows_date_columns(rows, sheet, "日度屠宰量"))
            mode = "省份行为主、日期为列"
        elif sheet == "市场主流标猪肥猪均价方便作图":
            numeric_records.extend(parse_simple_metric_table(rows, sheet))
            mode = "单日期列+多指标"
        elif sheet in {"出栏价", "散户标肥价差", "市场主流标猪肥猪价格"}:
            start_col = 2 if sheet in {"出栏价", "市场主流标猪肥猪价格"} else 1
            numeric_records.extend(parse_repeated_date_blocks(rows, sheet, start_col=start_col))
            mode = "双层表头、日期分块横向展开"
        elif sheet == "交割地市出栏价":
            n, m = parse_delivery_city_sheet(rows, sheet)
            numeric_records.extend(n)
            metadata_records.extend(m)
            mode = "省份分组+地市列"
        else:
            mode = "未适配"
        logs.append(
            {
                "sheet": sheet,
                "mode": mode,
                "numeric_rows": len(numeric_records) - numeric_before,
                "meta_rows": len(metadata_records) - metadata_before,
            }
        )
    return finalize_numeric_df(numeric_records), pd.DataFrame(metadata_records), pd.DataFrame(logs)


def build_yongyi_dataset_from_path(path_str: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    sheets = read_workbook_rows_from_path(path_str)
    return build_yongyi_dataset_from_rows(sheets)


# -----------------------------
# 通用聚合与图表
# -----------------------------
def build_native_weekly_anchor(df: pd.DataFrame) -> pd.Series:
    anchor = pd.to_datetime(df["end_date"], errors="coerce") if "end_date" in df.columns else pd.Series(pd.NaT, index=df.index)
    if not isinstance(anchor, pd.Series):
        anchor = pd.Series(anchor, index=df.index)
    if "date" in df.columns:
        anchor = anchor.combine_first(pd.to_datetime(df["date"], errors="coerce"))
    return anchor.dt.normalize()


def is_weekly_feed_sheet(sheet_name: str) -> bool:
    return text_of(sheet_name) in WEEKLY_FEED_SHEETS


def should_show_weekly_comparison(sheet_name: str) -> bool:
    return is_weekly_feed_sheet(sheet_name)


def format_tenday_label(value: Any) -> str:
    ts = parse_date(value)
    if ts is None:
        return text_of(value)
    day = int(ts.day)
    if day <= 10:
        part = "上旬"
    elif day <= 20:
        part = "中旬"
    else:
        part = "下旬"
    return f"{ts.year}年{ts.month}月{part}"


def apply_weekly_unit_conversion(records: list[dict], sheet_name: str) -> list[dict]:
    if sheet_name not in WEEKLY_UNIT_CONVERSION_SHEETS:
        return records
    for record in records:
        if record.get("value") is not None:
            record["value"] = float(record["value"]) * 2
        low_price = record.get("low_price")
        high_price = record.get("high_price")
        if low_price is not None:
            record["low_price"] = float(low_price) * 2
        if high_price is not None:
            record["high_price"] = float(high_price) * 2
        if record.get("low_price") is not None and record.get("high_price") is not None:
            range_text = f"{format_number(record['low_price'])} - {format_number(record['high_price'])}"
            record["raw_range"] = range_text
            record["price_range_text"] = range_text
    return records


def aggregate_regular(df: pd.DataFrame, frequency: str, use_lunar: bool = False) -> pd.DataFrame:
    temp = df.copy()
    group_field = get_group_field(temp)
    if use_lunar:
        for col in ["lunar_label", "lunar_order", "is_leap_lunar", "lunar_year"]:
            if col not in temp.columns:
                temp = enrich_date_features(temp)
                break
    if frequency == "日度数据" and use_lunar and "lunar_label" in temp.columns and temp["lunar_label"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        temp["period"] = temp["date"]
        temp["period_label"] = temp.apply(lambda r: f"{int(r['lunar_year'])}年{r['lunar_label']}" if pd.notna(r.get('lunar_year')) and text_of(r.get('lunar_label')) else format_date_cn(pd.Timestamp(r['date'])), axis=1)
        temp["_sort_key"] = temp["date"]
    elif frequency == "日度数据":
        temp["period"] = temp["date"]
        temp["period_label"] = temp["date"].map(format_date_cn)
        temp["_sort_key"] = temp["date"]
    elif frequency == "周度数据（均值）" and use_lunar and "lunar_order" in temp.columns and temp["lunar_order"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
        temp["_lday"] = temp["lunar_order"].fillna(0).astype(int) % 100
        temp["_lweek_seg"] = temp["_lday"].apply(lambda d: min((max(d, 1) - 1) // 7 + 1, 5))
        temp["period"] = temp["lunar_year"].astype(str) + "-" + temp["_lmonth"].astype(str).str.zfill(2) + "-W" + temp["_lweek_seg"].astype(str)
        temp["period_label"] = temp.apply(lambda r: f"{int(r['lunar_year'])}年{LUNAR_MONTHS[int(r['_lmonth'])-1]}月第{int(r['_lweek_seg'])}周" if pd.notna(r['lunar_year']) and 1 <= int(r['_lmonth']) <= 12 else "", axis=1)
        temp["_sort_key"] = temp["lunar_year"].astype(int) * 1000 + temp["_lmonth"] * 10 + temp["_lweek_seg"]
    elif frequency == "周度数据（均值）":
        native_weekly_mask = temp["freq_type"].eq("周度") if "freq_type" in temp.columns else pd.Series(False, index=temp.index)
        if native_weekly_mask.any():
            temp.loc[native_weekly_mask, "period"] = build_native_weekly_anchor(temp.loc[native_weekly_mask])
            temp.loc[native_weekly_mask, "period_label"] = temp.loc[native_weekly_mask, "period"].map(format_week_cn)
            temp.loc[native_weekly_mask, "_sort_key"] = temp.loc[native_weekly_mask, "period"]
        regular_mask = ~native_weekly_mask
        if regular_mask.any():
            temp.loc[regular_mask, "period"] = temp.loc[regular_mask, "date"].dt.to_period("W-MON").dt.start_time
            temp.loc[regular_mask, "period_label"] = temp.loc[regular_mask, "period"].map(format_week_cn)
            temp.loc[regular_mask, "_sort_key"] = temp.loc[regular_mask, "period"]
    elif frequency == "旬度数据":
        temp["period"] = temp["date"]
        temp["period_label"] = temp["date"].map(format_tenday_label)
        temp["_sort_key"] = temp["date"]
    elif frequency == "月度数据（均值）" and use_lunar and "lunar_order" in temp.columns and temp["lunar_order"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
        temp["period"] = temp["lunar_year"].astype(str) + "-" + temp["_lmonth"].astype(str).str.zfill(2)
        temp["period_label"] = temp.apply(lambda r: f"{int(r['lunar_year'])}年{LUNAR_MONTHS[int(r['_lmonth'])-1]}月" if pd.notna(r['lunar_year']) and 1 <= int(r['_lmonth']) <= 12 else "", axis=1)
        temp["_sort_key"] = temp["lunar_year"].astype(int) * 100 + temp["_lmonth"]
    else:
        temp["period"] = temp["date"].dt.to_period("M").dt.start_time
        temp["period_label"] = temp["period"].map(format_month_cn)
        temp["_sort_key"] = temp["period"]
    group_cols = ["period", "period_label"] + ([group_field] if group_field else [])
    result = temp.groupby(group_cols, as_index=False).agg({"value": "mean", "_sort_key": "min"}).sort_values(["_sort_key"] + ([group_field] if group_field else []))
    if group_field:
        result["series_name"] = result[group_field]
    else:
        display_values = temp["display_name"].dropna() if "display_name" in temp.columns else pd.Series(dtype=object)
        fallback_name = text_of(display_values.iloc[0]) if not display_values.empty else ""
        default_name = "全国" if "province" in temp.columns and any(is_national_scope(x) for x in temp["province"].dropna().unique()) else fallback_name
        result["series_name"] = default_name or "云南"
    return result


def aggregate_seasonal(df: pd.DataFrame, use_lunar: bool, frequency: str = "日度数据") -> pd.DataFrame:
    """季节性聚合。
    - 公历模式：按月日（日度）/ 周号（周度）/ 月（月度）聚合，x 轴为公历标签。
    - 农历模式：先把每条日度记录映射到农历日/周/月标签，再按农历标签 + 年份聚合均值。
      周度/月度在农历模式下的逻辑：先把每一天映射到农历月+农历周段，再以年维度求均值。
    """
    temp = df.copy()
    # 补齐日期特征
    if "year" not in temp.columns or temp["year"].isna().all():
        temp = enrich_date_features(temp)
    for col in ["lunar_label", "lunar_order", "is_leap_lunar", "lunar_year"]:
        if col not in temp.columns:
            temp = enrich_date_features(temp)
            break

    if use_lunar and "lunar_label" in temp.columns and temp["lunar_label"].notna().any():
        # 过滤闰月
        temp = temp[temp["lunar_label"].notna() & (~temp["is_leap_lunar"].fillna(False))].copy()
        year_col = temp["lunar_year"] if "lunar_year" in temp.columns else pd.Series(dtype=float)
        temp["season_year"] = year_col.fillna(temp["year"]).astype(int)

        if frequency == "周度数据（均值）":
            # 农历周：以农历月 + 该月内第几个7天段作为标签
            # lunar_order = month*100+day，取 day//7 作周段
            temp["_lday"] = temp["lunar_order"].fillna(0).astype(int) % 100
            temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
            temp["_week_seg"] = temp["_lday"].apply(lambda d: min((d - 1) // 7 + 1, 4))  # 1-4段
            temp["season_label"] = temp.apply(lambda r: f"{LUNAR_MONTHS[int(r['_lmonth'])-1]}月第{r['_week_seg']}周" if 1 <= int(r['_lmonth']) <= 12 else "", axis=1)
            temp["season_order"] = temp["_lmonth"] * 100 + temp["_week_seg"] * 10
        elif frequency == "月度数据（均值）":
            # 农历月
            temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
            temp["season_label"] = temp["_lmonth"].apply(lambda m: f"{LUNAR_MONTHS[m-1]}月" if 1 <= m <= 12 else "")
            temp["season_order"] = temp["_lmonth"] * 100
        else:
            # 日度：直接用农历日标签
            temp["season_label"] = temp["lunar_label"]
            temp["season_order"] = temp["lunar_order"]
    else:
        # 公历模式
        if frequency == "周度数据（均值）":
            # ISO 周号
            temp["_iso_week"] = temp["date"].dt.isocalendar().week.astype(int)
            temp["_iso_year_w"] = temp["date"].dt.isocalendar().year.astype(int)
            temp["season_label"] = temp["_iso_week"].apply(lambda w: f"第{w:02d}周")
            temp["season_order"] = temp["_iso_week"]
            temp["season_year"] = temp["_iso_year_w"]
        elif frequency == "旬度数据":
            temp["_tenday"] = temp["date"].dt.day.apply(lambda d: 1 if d <= 10 else (2 if d <= 20 else 3))
            temp["season_label"] = temp.apply(lambda r: f"{r['date'].month}月{'上旬' if r['_tenday']==1 else ('中旬' if r['_tenday']==2 else '下旬')}", axis=1)
            temp["season_order"] = temp["date"].dt.month * 100 + temp["_tenday"] * 10
            temp["season_year"] = temp["year"].astype(int)
        elif frequency == "月度数据（均值）":
            temp["season_label"] = temp["date"].dt.month.apply(lambda m: f"{m}月")
            temp["season_order"] = temp["date"].dt.month
            temp["season_year"] = temp["year"].astype(int)
        else:
            temp["season_label"] = temp["date"].map(format_month_day_cn)
            temp["season_order"] = temp["date"].dt.month * 100 + temp["date"].dt.day
            temp["season_year"] = temp["year"].astype(int)

    # 过滤掉空标签
    temp = temp[temp["season_label"].fillna("").ne("")].copy()

    result = (
        temp.groupby(["season_year", "season_label", "season_order"], as_index=False)["value"]
        .mean()
        .sort_values(["season_order", "season_year"])
    )
    result["年份"] = result["season_year"].astype(str) + "年"
    return result


def aggregate_series_frequency(df: pd.DataFrame, frequency: str, value_col: str = "value", use_lunar: bool = False) -> pd.DataFrame:
    temp = df.copy()
    if use_lunar:
        for col in ["lunar_label", "lunar_order", "is_leap_lunar", "lunar_year"]:
            if col not in temp.columns:
                temp = enrich_date_features(temp)
                break
    if frequency == "日度数据" and use_lunar and "lunar_label" in temp.columns and temp["lunar_label"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        lunar_year = temp["lunar_year"] if "lunar_year" in temp.columns else pd.Series(index=temp.index, dtype=float)
        temp["period"] = temp["date"]
        temp["period_label"] = [f"{int(y)}年{label}" if pd.notna(y) and text_of(label) else format_date_cn(pd.Timestamp(d)) for y, label, d in zip(lunar_year, temp["lunar_label"], temp["date"])]
        temp["_lunar_sort"] = temp["date"]
    elif frequency == "日度数据":
        temp["period"] = temp["date"]
        temp["period_label"] = temp["date"].map(format_date_cn)
        temp["_lunar_sort"] = temp["date"]
    elif frequency == "周度数据（均值）" and use_lunar and "lunar_order" in temp.columns and temp["lunar_order"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
        temp["_lday"] = temp["lunar_order"].fillna(0).astype(int) % 100
        temp["_lweek_seg"] = temp["_lday"].apply(lambda d: min((max(d, 1) - 1) // 7 + 1, 5))
        temp["period"] = temp["lunar_year"].astype(str) + "-" + temp["_lmonth"].astype(str).str.zfill(2) + "-W" + temp["_lweek_seg"].astype(str)
        temp["period_label"] = temp.apply(lambda r: f"{LUNAR_MONTHS[int(r['_lmonth'])-1]}月第{int(r['_lweek_seg'])}周" if 1 <= int(r['_lmonth']) <= 12 else "", axis=1)
        temp["_lunar_sort"] = temp["lunar_year"].astype(int) * 1000 + temp["_lmonth"] * 10 + temp["_lweek_seg"]
    elif frequency == "周度数据（均值）":
        native_weekly_mask = temp["freq_type"].eq("周度") if "freq_type" in temp.columns else pd.Series(False, index=temp.index)
        if native_weekly_mask.any():
            period_series = build_native_weekly_anchor(temp.loc[native_weekly_mask])
            temp.loc[native_weekly_mask, "period"] = period_series.values
            period_sheet_names = temp.loc[native_weekly_mask, "sheet"] if "sheet" in temp.columns else pd.Series("", index=temp.loc[native_weekly_mask].index)
            period_labels = [
                format_tenday_label(period) if sheet_name == "二育栏舍利用率" and pd.notna(period) else (format_week_cn(period) if pd.notna(period) else "")
                for period, sheet_name in zip(period_series, period_sheet_names)
            ]
            temp.loc[native_weekly_mask, "period_label"] = period_labels
            temp.loc[native_weekly_mask, "_lunar_sort"] = period_series.values
        regular_mask = ~native_weekly_mask
        if regular_mask.any():
            regular_period = temp.loc[regular_mask, "date"].dt.to_period("W-MON").dt.start_time
            temp.loc[regular_mask, "period"] = regular_period.values
            temp.loc[regular_mask, "period_label"] = regular_period.map(format_week_cn).tolist()
            temp.loc[regular_mask, "_lunar_sort"] = regular_period.values
    elif frequency == "旬度数据":
        temp["period"] = temp["date"]
        temp["period_label"] = temp["date"].map(format_tenday_label)
        temp["_lunar_sort"] = temp["date"]
    elif frequency == "月度数据（均值）" and use_lunar and "lunar_order" in temp.columns and temp["lunar_order"].notna().any():
        temp = temp[~temp["is_leap_lunar"].fillna(False)].copy()
        temp["_lmonth"] = (temp["lunar_order"].fillna(0).astype(int) // 100)
        temp["period"] = temp["lunar_year"].astype(str) + "-" + temp["_lmonth"].astype(str).str.zfill(2)
        temp["period_label"] = temp["_lmonth"].apply(lambda m: f"{LUNAR_MONTHS[int(m)-1]}月" if 1 <= int(m) <= 12 else "")
        temp["_lunar_sort"] = temp["lunar_year"].astype(int) * 100 + temp["_lmonth"]
    else:
        temp["period"] = temp["date"].dt.to_period("M").dt.start_time
        temp["period_label"] = temp["period"].map(format_month_cn)
        temp["_lunar_sort"] = temp["period"]
    if "period" not in temp.columns:
        temp["period"] = temp["date"]
    if "period_label" not in temp.columns:
        temp["period_label"] = temp["date"].map(format_date_cn)
    if "_lunar_sort" not in temp.columns:
        temp["_lunar_sort"] = temp["date"]
    if "series_name" not in temp.columns:
        group_field = get_group_field(temp)
        if group_field and group_field in temp.columns:
            temp["series_name"] = temp[group_field]
        else:
            display_values = temp["display_name"].dropna() if "display_name" in temp.columns else pd.Series(dtype=object)
            fallback = text_of(display_values.iloc[0]) if not display_values.empty else ""
            temp["series_name"] = "全国" if "province" in temp.columns and any(is_national_scope(x) for x in temp["province"].dropna().unique()) else (fallback or "数值")
    agg_map: dict[str, str] = {value_col: "mean"}
    if "low_price" in temp.columns:
        agg_map["low_price"] = "mean"
    if "high_price" in temp.columns:
        agg_map["high_price"] = "mean"
    agg_map_with_sort = dict(agg_map)
    if "_lunar_sort" in temp.columns:
        agg_map_with_sort["_lunar_sort"] = "min"
    result = temp.groupby(["period", "period_label", "series_name"], as_index=False).agg(agg_map_with_sort).sort_values(["_lunar_sort", "series_name"])
    if "low_price" in result.columns and "high_price" in result.columns:
        result["range_label"] = result.apply(lambda row: f"{format_number(row['low_price'])} - {format_number(row['high_price'])}", axis=1)
    return result


def get_snapshot_date(df: pd.DataFrame, target_date: pd.Timestamp) -> pd.Timestamp | None:
    dates = pd.Series(pd.to_datetime(df["date"].dropna().unique())).sort_values().reset_index(drop=True)
    if dates.empty:
        return None
    target_date = pd.Timestamp(target_date).normalize()
    exact = dates[dates == target_date]
    if not exact.empty:
        return pd.Timestamp(exact.iloc[0]).normalize()
    earlier = dates[dates <= target_date]
    if not earlier.empty:
        return pd.Timestamp(earlier.iloc[-1]).normalize()
    return pd.Timestamp(dates.iloc[0]).normalize()


def build_rank_df(df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[pd.DataFrame, pd.Timestamp | None]:
    group_field = get_group_field(df)
    if group_field is None:
        return pd.DataFrame(), None
    snap_date = get_snapshot_date(df, target_date)
    if snap_date is None:
        return pd.DataFrame(), None
    snap_df = df[df["date"] == snap_date].copy()
    if group_field == "province":
        snap_df = snap_df[snap_df["province"] != "全部"].copy()
    else:
        snap_df = snap_df[snap_df["city"].fillna("").ne("")].copy()
    if snap_df.empty:
        return pd.DataFrame(), snap_date
    rank_df = snap_df.groupby(group_field, as_index=False)["value"].mean().sort_values("value", ascending=False)
    rank_df["series_name"] = rank_df[group_field]
    rank_df["日期标签"] = format_date_cn(snap_date)
    return rank_df, snap_date


def build_map_df(rank_df: pd.DataFrame) -> pd.DataFrame:
    if rank_df.empty or "province" not in rank_df.columns:
        return pd.DataFrame()
    temp = rank_df.copy()
    temp = temp[temp["province"].isin(PROVINCE_COORDS)].copy()
    if temp.empty:
        return temp
    temp["lon"] = temp["province"].map(lambda x: PROVINCE_COORDS[x][0])
    temp["lat"] = temp["province"].map(lambda x: PROVINCE_COORDS[x][1])
    return temp


def build_regular_heatmap_source(df: pd.DataFrame, target_date: pd.Timestamp, last_n: int = 30) -> pd.DataFrame:
    group_field = get_group_field(df)
    if group_field is None:
        return pd.DataFrame()
    temp = df.copy()
    if group_field == "province":
        temp = temp[temp["province"] != "全部"].copy()
    else:
        temp = temp[temp["city"].fillna("").ne("")].copy()
    if temp.empty:
        return pd.DataFrame()
    dates = sorted(temp["date"].dropna().unique())
    snap_date = get_snapshot_date(temp, target_date)
    if snap_date is None:
        return pd.DataFrame()
    recent_dates = [pd.Timestamp(x) for x in dates if pd.Timestamp(x) <= snap_date][-last_n:]
    temp = temp[temp["date"].isin(recent_dates)].copy()
    if temp.empty:
        return pd.DataFrame()
    temp["日期标签"] = temp["date"].map(format_date_cn)
    result = temp.groupby([group_field, "日期标签", "date"], as_index=False)["value"].mean()
    result["series_name"] = result[group_field]
    return result.sort_values([group_field, "date"])


def filter_df_by_date_window(df: pd.DataFrame, start_date: Any, end_date: Any) -> pd.DataFrame:
    if df.empty or "date" not in df.columns:
        return df.copy()
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    return df[(df["date"] >= start_ts) & (df["date"] <= end_ts)].copy()


def build_multi_series_line_chart(df: pd.DataFrame, frequency: str, title: str, area: bool = False, hover_title: str = "数值", use_lunar: bool = False) -> go.Figure:
    chart_df = aggregate_series_frequency(df, frequency, use_lunar=use_lunar)
    order = chart_df["period_label"].drop_duplicates().tolist()
    x_label = "农历日期" if (use_lunar and frequency == "日度数据") else "日期"
    fig_func = px.area if area else px.line
    fig = fig_func(chart_df, x="period_label", y="value", color="series_name", category_orders={"period_label": order}, markers=False)
    fig.update_traces(hovertemplate=f"日期：%{{x}}<br>序列：%{{fullData.name}}<br>{hover_title}：%{{y:.2f}}<extra></extra>", line=dict(width=1.5))
    fig.update_layout(title=title, xaxis_title=x_label, yaxis_title=hover_title, template="plotly_white", hovermode="x unified", xaxis=dict(type="category", categoryorder="array", categoryarray=order))
    if chart_df["series_name"].nunique() <= 1:
        fig.update_layout(showlegend=False)
    return fig


def build_product_price_chart(df: pd.DataFrame, frequency: str, title: str) -> go.Figure:
    chart_df = aggregate_series_frequency(df, frequency)
    order = chart_df["period_label"].drop_duplicates().tolist()
    fig = px.line(chart_df, x="period_label", y="value", color="series_name", category_orders={"period_label": order}, markers=False)
    if "range_label" in chart_df.columns:
        fig.update_traces(customdata=np.stack([chart_df["range_label"]], axis=-1), hovertemplate="日期：%{x}<br>品类：%{fullData.name}<br>均值：%{y:.2f}<br>参考区间：%{customdata[0]}<extra></extra>")
    else:
        fig.update_traces(hovertemplate="日期：%{x}<br>品类：%{fullData.name}<br>均值：%{y:.2f}<extra></extra>")
    fig.update_layout(title=title, xaxis_title="日期", yaxis_title="价格", template="plotly_white", hovermode="x unified")
    if chart_df["series_name"].nunique() <= 1:
        fig.update_layout(showlegend=False)
    return fig


def build_trend_line_chart(df: pd.DataFrame, frequency: str, scope_title: str = "", use_lunar: bool = False) -> go.Figure:
    chart_df = aggregate_series_frequency(df, frequency, use_lunar=use_lunar)
    chart_df = chart_df.sort_values(["_lunar_sort", "series_name"] if "_lunar_sort" in chart_df.columns else ["period", "series_name"])
    order = chart_df["period_label"].drop_duplicates().tolist()
    group_label = get_group_label(df)
    x_label = "农历日期" if (use_lunar and frequency == "日度数据") else "日期"
    title = ("趋势折线图" + (f"｜{scope_title}" if scope_title else ""))
    if chart_df["series_name"].nunique() > 1:
        fig = px.line(chart_df, x="period_label", y="value", color="series_name", category_orders={"period_label": order}, markers=False)
        fig.update_traces(hovertemplate=f"日期：%{{x}}<br>{group_label}：%{{fullData.name}}<br>数值：%{{y:.2f}}<extra></extra>", line=dict(width=1.5))
        fig.update_layout(legend_title=group_label)
    else:
        fig = px.line(chart_df, x="period_label", y="value", category_orders={"period_label": order}, markers=False)
        fig.update_traces(hovertemplate="日期：%{x}<br>数值：%{y:.2f}<extra></extra>", line=dict(width=1.5))
        fig.update_layout(showlegend=False)
    fig.update_layout(title=title, xaxis=dict(title=x_label, type="category", categoryorder="array", categoryarray=order), yaxis_title="数值", template="plotly_white", hovermode="x unified")
    return fig


def build_trend_line_chart_with_spread(main_df: pd.DataFrame, spread_df: pd.DataFrame, frequency: str, scope_title: str = "", use_lunar: bool = False) -> go.Figure:
    """主指标走势（左轴）+ 价差走势（右轴，副轴）的双轴折线图。价格冷色系，价差高对比暖色/紫红系。"""
    main_agg = aggregate_series_frequency(main_df, frequency, use_lunar=use_lunar)
    spread_agg = aggregate_series_frequency(spread_df, frequency, use_lunar=use_lunar)
    sort_col = "_lunar_sort" if "_lunar_sort" in main_agg.columns else "period"
    order = main_agg.sort_values(sort_col)["period_label"].drop_duplicates().tolist()
    if not order and not spread_agg.empty:
        sort_col2 = "_lunar_sort" if "_lunar_sort" in spread_agg.columns else "period"
        order = spread_agg.sort_values(sort_col2)["period_label"].drop_duplicates().tolist()
    else:
        sort_col2 = "_lunar_sort" if "_lunar_sort" in spread_agg.columns else "period"
        spread_order = spread_agg.sort_values(sort_col2)["period_label"].drop_duplicates().tolist() if not spread_agg.empty else []
        order += [label for label in spread_order if label not in order]
    x_label = "农历日期" if (use_lunar and frequency == "日度数据") else "日期"
    title = "趋势折线图（含价差副轴）" + (f"｜{scope_title}" if scope_title else "")
    fig = go.Figure()
    price_colors = ["#1a56db", "#0e9f6e", "#1d6fa4", "#057a55", "#5145cd", "#03543f"]
    for i, sname in enumerate(sorted(main_agg["series_name"].dropna().unique())):
        sub = main_agg[main_agg["series_name"] == sname]
        fig.add_trace(go.Scatter(
            x=sub["period_label"], y=sub["value"],
            mode="lines", name=sname,
            yaxis="y",
            line=dict(color=price_colors[i % len(price_colors)], width=2),
            marker=dict(size=5),
            hovertemplate=f"日期：%{{x}}<br>价格 {sname}：%{{y:.2f}}<extra></extra>",
        ))
    spread_colors = ["#d946ef", "#ef4444", "#f97316", "#db2777"]
    for i, sname in enumerate(sorted(spread_agg["series_name"].dropna().unique())):
        sub = spread_agg[spread_agg["series_name"] == sname]
        fig.add_trace(go.Scatter(
            x=sub["period_label"], y=sub["value"],
            mode="lines", name=sname,
            yaxis="y2",
            line=dict(color=spread_colors[i % len(spread_colors)], width=2.8, dash="dash"),
            marker=dict(size=5),
            hovertemplate=f"日期：%{{x}}<br>价差 {sname}：%{{y:.2f}}<extra></extra>",
        ))
    fig.update_layout(
        title=title,
        xaxis=dict(title=x_label, type="category", categoryorder="array", categoryarray=order),
        yaxis=dict(title="价格"),
        yaxis2=dict(title="价差", overlaying="y", side="right", showgrid=False),
        template="plotly_white",
        hovermode="x unified",
        legend_title="序列",
    )
    return fig


def build_region_spread_df(df: pd.DataFrame, region_a: str, region_b: str, label: str | None = None) -> pd.DataFrame:
    if not region_a or not region_b or region_a == region_b:
        return pd.DataFrame(columns=df.columns)
    left = df[df["province"] == region_a].copy()
    right = df[df["province"] == region_b].copy()
    if left.empty or right.empty:
        return pd.DataFrame(columns=df.columns)
    merged = left[["date", "metric", "province", "city", "value"]].rename(columns={"value": "value_a"}).merge(
        right[["date", "metric", "province", "city", "value"]].rename(columns={"value": "value_b"}),
        on=["date", "metric"],
        how="inner",
    )
    if merged.empty:
        return pd.DataFrame(columns=df.columns)
    result = pd.DataFrame({
        "date": merged["date"],
        "metric": label or f"{region_a}-{region_b}价差",
        "value": merged["value_a"] - merged["value_b"],
        "province": "全部",
        "city": "",
        "display_name": label or f"{region_a}-{region_b}价差",
        "series_name": label or f"{region_a}-{region_b}价差",
        "sheet": left["sheet"].iloc[0] if "sheet" in left.columns and not left.empty else "区域价差",
    })
    return enrich_date_features(result)


def build_scattered_fat_price_views(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return pd.DataFrame(columns=df.columns), pd.DataFrame(columns=df.columns)
    metric_names = df["metric"].dropna().astype(str).unique().tolist()
    base_metric = next((m for m in metric_names if "市场" in m and "散户" in m), "")
    if not base_metric:
        return pd.DataFrame(columns=df.columns), pd.DataFrame(columns=df.columns)
    base_df = df[df["metric"] == base_metric].copy()
    # 用 date+province 聚合均值，避免同一天同省有多行导致 merge 膨胀
    base_agg = base_df.groupby(["date", "province"], as_index=False)["value"].mean()
    derived_frames: list[pd.DataFrame] = []
    spread_frames: list[pd.DataFrame] = []
    for spread_metric_kw, derived_metric in [("150公斤左右较标猪", "150公斤猪价格"), ("175公斤左右较标猪", "175公斤猪价格")]:
        pattern = spread_metric_kw.replace("左右", ".*")
        spread_df = df[df["metric"].astype(str).str.contains(pattern, regex=True, na=False)].copy()
        if spread_df.empty:
            continue
        spread_agg = spread_df.groupby(["date", "province"], as_index=False)["value"].mean()
        spread_agg = spread_agg.rename(columns={"value": "spread_value"})
        merged = base_agg.merge(spread_agg, on=["date", "province"], how="inner")
        if merged.empty:
            continue
        derived = merged[["date", "province"]].copy()
        derived["metric"] = derived_metric
        derived["sheet"] = base_df["sheet"].iloc[0] if not base_df.empty else df["sheet"].iloc[0]
        derived["city"] = ""
        derived["display_name"] = derived_metric
        derived["series_name"] = derived_metric
        derived["value"] = merged["value"] + merged["spread_value"]
        derived = enrich_date_features(derived)
        derived_frames.append(derived)
        spread_std = merged[["date", "province"]].copy()
        actual_spread_metric = spread_df["metric"].dropna().iloc[0] if not spread_df["metric"].dropna().empty else spread_metric_kw
        spread_std["metric"] = actual_spread_metric
        spread_std["series_name"] = actual_spread_metric
        spread_std["sheet"] = base_df["sheet"].iloc[0] if not base_df.empty else df["sheet"].iloc[0]
        spread_std["city"] = ""
        spread_std["display_name"] = actual_spread_metric
        spread_std["value"] = merged["spread_value"]
        spread_std = enrich_date_features(spread_std)
        spread_frames.append(spread_std)
    # base_df as standardized frame
    base_std = base_df.copy()
    base_std["series_name"] = base_metric
    price_df = pd.concat([base_std] + derived_frames, ignore_index=True) if derived_frames else base_std.copy()
    spread_out = pd.concat(spread_frames, ignore_index=True) if spread_frames else pd.DataFrame(columns=df.columns)
    return price_df, spread_out


def build_bar_chart(df: pd.DataFrame, x: str, y: str, title: str, orientation: str = "v", text_col: str | None = None, color_col: str | None = None) -> go.Figure:
    if df.empty:
        fig = go.Figure()
        fig.update_layout(template="plotly_white", title=title)
        return fig
    fig = px.bar(df, x=x, y=y, orientation=orientation, text=text_col, color=color_col or y)
    fig.update_layout(title=title, template="plotly_white", coloraxis_showscale=False)
    fig.update_traces(textposition="outside")
    return fig


def build_dual_axis_line_chart(price_df: pd.DataFrame, volume_df: pd.DataFrame, frequency: str, left_title: str, right_title: str, use_lunar: bool = False) -> go.Figure:
    price_chart = aggregate_regular(price_df, frequency, use_lunar=use_lunar)
    volume_chart = aggregate_regular(volume_df, frequency, use_lunar=use_lunar)
    sort_col = "_sort_key" if "_sort_key" in price_chart.columns else "period"
    price_chart = price_chart.sort_values(sort_col)
    volume_chart_sorted = volume_chart.sort_values("_sort_key" if "_sort_key" in volume_chart.columns else "period")
    order = list(dict.fromkeys(price_chart["period_label"].tolist() + volume_chart_sorted["period_label"].tolist()))
    x_label = "农历日期" if (use_lunar and frequency == "日度数据") else "日期"
    fig = go.Figure()
    if not price_chart.empty:
        fig.add_trace(
            go.Scatter(
                x=price_chart["period_label"],
                y=price_chart["value"],
                mode="lines",
                name=left_title,
                yaxis="y",
                hovertemplate="日期：%{x}<br>指标：" + left_title + "<br>数值：%{y:.2f}<extra></extra>",
            )
        )
    if not volume_chart.empty:
        fig.add_trace(
            go.Bar(
                x=volume_chart_sorted["period_label"],
                y=volume_chart_sorted["value"],
                name=right_title,
                yaxis="y2",
                opacity=0.45,
                hovertemplate="日期：%{x}<br>指标：" + right_title + "<br>数值：%{y:.2f}<extra></extra>",
            )
        )
    fig.update_layout(
        title="价格与宰量双轴图",
        xaxis=dict(title=x_label, categoryorder="array", categoryarray=order),
        yaxis=dict(title=left_title),
        yaxis2=dict(title=right_title, overlaying="y", side="right"),
        template="plotly_white",
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        barmode="overlay",
    )
    return fig


def build_metric_spread_df(df: pd.DataFrame, left_metric: str, right_metric: str, spread_name: str) -> pd.DataFrame:
    subset = df[df["metric"].isin([left_metric, right_metric])].copy()
    if subset.empty:
        return subset
    group_cols = ["sheet", "date", "province", "city"]
    extra_cols = [col for col in ["display_name", "series_name"] if col in subset.columns]
    pivot = subset.pivot_table(index=group_cols, columns="metric", values="value", aggfunc="mean").reset_index()
    if left_metric not in pivot.columns or right_metric not in pivot.columns:
        return pd.DataFrame(columns=df.columns)
    result = pivot[group_cols].copy()
    if "sheet" in result.columns:
        result["sheet"] = result["sheet"].fillna(subset["sheet"].iloc[0] if "sheet" in subset.columns and not subset.empty else "价差")
    result["metric"] = spread_name
    result["display_name"] = spread_name
    result["series_name"] = spread_name
    result["value"] = pivot[left_metric] - pivot[right_metric]
    result = result.dropna(subset=["value"])
    return enrich_date_features(result)


def build_spread_label(left_metric: str, right_metric: str) -> str:
    return f"{left_metric}-{right_metric}价差"


def build_optional_metric_spread_df(df: pd.DataFrame, left_metric: str, right_metric: str) -> pd.DataFrame:
    if not left_metric or not right_metric or left_metric == right_metric:
        return pd.DataFrame(columns=df.columns)
    return build_metric_spread_df(df, left_metric, right_metric, build_spread_label(left_metric, right_metric))


def build_trend_area_chart(df: pd.DataFrame, frequency: str, use_lunar: bool = False) -> go.Figure:
    chart_df = aggregate_regular(df, frequency, use_lunar=use_lunar)
    sort_col = "_sort_key" if "_sort_key" in chart_df.columns else "period"
    chart_df = chart_df.sort_values([sort_col, "series_name"])
    order = chart_df["period_label"].drop_duplicates().tolist()
    group_label = get_group_label(df)
    x_label = "农历日期" if (use_lunar and frequency == "日度数据") else "日期"
    if chart_df["series_name"].nunique() > 1:
        fig = px.area(chart_df, x="period_label", y="value", color="series_name", category_orders={"period_label": order})
        fig.update_traces(hovertemplate=f"日期：%{{x}}<br>{group_label}：%{{fullData.name}}<br>数值：%{{y:.2f}}<extra></extra>")
        fig.update_layout(legend_title=group_label)
    else:
        fig = px.area(chart_df, x="period_label", y="value", category_orders={"period_label": order})
        fig.update_traces(hovertemplate="日期：%{x}<br>数值：%{y:.2f}<extra></extra>")
        fig.update_layout(showlegend=False)
    fig.update_layout(title="趋势面积图", xaxis_title=x_label, yaxis_title="数值", template="plotly_white", hovermode="x unified")
    return fig


def build_rank_bar_chart(df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[go.Figure, str]:
    group_label = get_group_label(df)
    rank_df, snap_date = build_rank_df(df, target_date)
    fig = go.Figure()
    if rank_df.empty:
        fig.update_layout(template="plotly_white", title=f"{group_label}排行图")
        return fig, f"当前筛选下没有可用于排行的{group_label}数据。"
    group_field = get_group_field(df) or "series_name"
    fig = px.bar(rank_df.sort_values("value", ascending=True), x="value", y=group_field, orientation="h", color="value", color_continuous_scale="RdYlGn_r")
    fig.update_traces(hovertemplate=f"{group_label}：%{{y}}<br>数值：%{{x:.2f}}<extra></extra>")
    fig.update_layout(title=f"{group_label}排行图（{format_date_cn(snap_date)}）", xaxis_title="数值", yaxis_title=group_label, coloraxis_showscale=False, template="plotly_white")
    return fig, ""


def build_box_chart(df: pd.DataFrame) -> tuple[go.Figure, str]:
    group_field = get_group_field(df)
    group_label = get_group_label(df)
    fig = go.Figure()
    if group_field is None:
        fig.update_layout(template="plotly_white", title="分布箱线图")
        return fig, f"当前筛选下没有{group_label}维度数据。"
    temp = df.copy()
    if group_field == "province":
        temp = temp[temp["province"] != "全部"].copy()
    else:
        temp = temp[temp["city"].fillna("").ne("")].copy()
    if temp.empty:
        fig.update_layout(template="plotly_white", title="分布箱线图")
        return fig, f"当前筛选下没有{group_label}维度数据。"
    fig = px.box(temp, x=group_field, y="value", color=group_field)
    fig.update_traces(hovertemplate=f"{group_label}：%{{x}}<br>数值：%{{y:.2f}}<extra></extra>")
    fig.update_layout(title="分布箱线图", xaxis_title=group_label, yaxis_title="数值", showlegend=False, template="plotly_white")
    return fig, ""


def build_regular_heatmap(df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[go.Figure, str]:
    group_label = get_group_label(df)
    temp = build_regular_heatmap_source(df, target_date)
    fig = go.Figure()
    if temp.empty:
        fig.update_layout(template="plotly_white", title=f"{group_label}热力图")
        return fig, f"当前筛选下没有足够的{group_label}时间序列数据。"
    pivot = temp.pivot_table(index="series_name", columns="日期标签", values="value", aggfunc="mean")
    fig = go.Figure(data=go.Heatmap(z=pivot.values, x=list(pivot.columns), y=list(pivot.index), colorscale="RdYlGn_r", hovertemplate=f"日期：%{{x}}<br>{group_label}：%{{y}}<br>数值：%{{z:.2f}}<extra></extra>"))
    fig.update_layout(title=f"{group_label}热力图（近阶段）", xaxis_title="日期", yaxis_title=group_label, template="plotly_white")
    return fig, ""


def build_map_chart(df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[go.Figure, str]:
    if get_series_dimension(df) != "province":
        fig = go.Figure()
        fig.update_layout(template="plotly_white", title="地图气泡图")
        return fig, "当前筛选口径不适合绘制地图气泡图。"
    rank_df, snap_date = build_rank_df(df, target_date)
    map_df = build_map_df(rank_df)
    fig = go.Figure()
    if map_df.empty:
        fig.update_layout(template="plotly_white", title="地图气泡图")
        return fig, "当前筛选下没有可映射的省份数据。"
    max_value = map_df["value"].max() or 1
    sizes = 14 + (map_df["value"] / max_value) * 26
    fig = go.Figure(
        go.Scattergeo(
            lon=map_df["lon"],
            lat=map_df["lat"],
            text=map_df["province"],
            customdata=np.stack([map_df["province"], map_df["value"]], axis=-1),
            mode="markers+text",
            textposition="top center",
            marker=dict(size=sizes, color=map_df["value"], colorscale="RdYlGn_r", line=dict(width=0.8, color="white"), sizemode="diameter", opacity=0.88, colorbar=dict(title="数值")),
            hovertemplate="地区：%{customdata[0]}<br>日期：" + format_date_cn(snap_date) + "<br>数值：%{customdata[1]:.2f}<extra></extra>",
        )
    )
    fig.update_layout(
        title=f"地图气泡图（{format_date_cn(snap_date)}）",
        template="plotly_white",
        geo=dict(scope="asia", projection_type="natural earth", showland=True, landcolor="rgb(243,243,243)", showcountries=True, countrycolor="rgb(204,204,204)", lataxis=dict(range=[18, 54]), lonaxis=dict(range=[73, 136]), center=dict(lat=35, lon=104)),
        margin=dict(l=10, r=10, t=60, b=10),
    )
    return fig, ""


def build_seasonal_line_chart(df: pd.DataFrame, use_lunar: bool, frequency: str = "日度数据") -> tuple[go.Figure, str]:
    temp = aggregate_seasonal(df, use_lunar, frequency)
    fig = go.Figure()
    if temp.empty or temp["season_year"].nunique() < 2:
        fig.update_layout(template="plotly_white", title="季节性折线图")
        return fig, "当前筛选下历年样本不足，暂时无法做季节性对比。"
    order = temp.sort_values("season_order")["season_label"].drop_duplicates().tolist()
    color_map = get_year_color_map(temp["年份"].dropna().unique().tolist())
    fig = px.line(temp, x="season_label", y="value", color="年份", markers=False,
                  category_orders={"season_label": order}, color_discrete_map=color_map)
    freq_label = {"周度数据（均值）": "农历周" if use_lunar else "周", "月度数据（均值）": "农历月" if use_lunar else "月"}.get(frequency, "农历日期" if use_lunar else "月日")
    fig.update_traces(hovertemplate="日期：%{x}<br>年份：%{fullData.name}<br>数值：%{y:.2f}<extra></extra>",
                      line=dict(width=1.5))
    fig.update_layout(title="季节性折线图", xaxis_title=freq_label, yaxis_title="数值",
                      legend_title="年份", template="plotly_white", hovermode="x unified")
    return fig, ""


def build_seasonal_heatmap(df: pd.DataFrame, use_lunar: bool, frequency: str = "日度数据") -> tuple[go.Figure, str]:
    temp = aggregate_seasonal(df, use_lunar, frequency)
    fig = go.Figure()
    if temp.empty or temp["season_year"].nunique() < 2:
        fig.update_layout(template="plotly_white", title="季节性热力图")
        return fig, "当前筛选下历年样本不足，暂时无法做季节性热力图。"
    pivot = temp.pivot_table(index="年份", columns="season_label", values="value", aggfunc="mean")
    order = temp.sort_values("season_order")["season_label"].drop_duplicates().tolist()
    pivot = pivot.reindex(columns=order)
    freq_label = {"周度数据（均值）": "农历周" if use_lunar else "周", "月度数据（均值）": "农历月" if use_lunar else "月"}.get(frequency, "农历日期" if use_lunar else "月日")
    fig = go.Figure(data=go.Heatmap(
        z=pivot.values, x=list(pivot.columns), y=list(pivot.index),
        colorscale="RdYlGn_r",
        hovertemplate=f"{freq_label}：%{{x}}<br>年份：%{{y}}<br>数值：%{{z:.2f}}<extra></extra>"))
    fig.update_layout(title="季节性热力图", xaxis_title=freq_label, yaxis_title="年份", template="plotly_white")
    return fig, ""


def render_plotly(fig: go.Figure, *key_parts: Any) -> None:
    st.plotly_chart(fig, key=build_plotly_key(*key_parts), width='stretch')


def sort_date_label_frame(df: pd.DataFrame, label_col: str = "日期标签") -> pd.DataFrame:
    temp = df.copy()
    if "date" in temp.columns:
        temp = temp.sort_values(["date"] + ([label_col] if label_col in temp.columns else []))
    return temp


def build_summary_image_bytes(title: str, subtitle: str, body_lines: list[str], width: int = 1400, height: int = 520) -> bytes:
    fig = go.Figure()
    fig.update_layout(
        template="plotly_white",
        width=width,
        height=height,
        margin=dict(l=50, r=50, t=50, b=40),
        paper_bgcolor="white",
        plot_bgcolor="white",
        xaxis=dict(visible=False),
        yaxis=dict(visible=False),
        annotations=[
            dict(x=0, y=1.0, xref="paper", yref="paper", text=f"<b>{title}</b>", showarrow=False, align="left", xanchor="left", yanchor="top", font=dict(size=24, color="#0f172a")),
            dict(x=0, y=0.92, xref="paper", yref="paper", text=subtitle, showarrow=False, align="left", xanchor="left", yanchor="top", font=dict(size=14, color="#64748b")),
            dict(x=0, y=0.84, xref="paper", yref="paper", text="<br>".join(body_lines), showarrow=False, align="left", xanchor="left", yanchor="top", font=dict(size=16, color="#1e293b")),
        ],
    )
    return pio.to_image(fig, format="png", width=width, height=height, scale=2)


def combine_vertical_images(image_blocks: list[bytes]) -> bytes:
    valid = [block for block in image_blocks if block]
    if not valid:
        return b""
    if len(valid) == 1:
        return valid[0]
    try:
        from PIL import Image
    except Exception as exc:
        raise RuntimeError(f"缺少 Pillow 依赖：{exc}")
    images = [Image.open(BytesIO(block)).convert("RGB") for block in valid]
    target_width = max(img.width for img in images)
    resized: list[Any] = []
    total_height = 0
    for img in images:
        if img.width != target_width:
            new_height = int(img.height * target_width / img.width)
            img = img.resize((target_width, new_height))
        resized.append(img)
        total_height += img.height
    canvas = Image.new("RGB", (target_width, total_height), color="white")
    y = 0
    for img in resized:
        canvas.paste(img, (0, y))
        y += img.height
    out = BytesIO()
    canvas.save(out, format="PNG")
    return out.getvalue()


def build_futures_export_image(fig: go.Figure, contract: str, target_date: pd.Timestamp, advice: str, signals: list[str]) -> bytes:
    try:
        chart_bytes = pio.to_image(fig, format="png", width=1400, height=700, scale=2)
    except Exception as exc:
        raise RuntimeError(f"Plotly 图片导出失败，请确认 kaleido 可用：{exc}")
    summary_lines = [advice] + [f"• {text}" for text in signals[:6]]
    summary_bytes = build_summary_image_bytes(
        title=f"{contract} 技术分析总结",
        subtitle=f"分析日期：{format_date_cn(pd.Timestamp(target_date))}",
        body_lines=summary_lines,
        width=1400,
        height=560,
    )
    return combine_vertical_images([chart_bytes, summary_bytes])


def get_sheet_metric_snapshot(df: pd.DataFrame, sheet: str, metric_keywords: list[str], target_date: pd.Timestamp) -> dict[str, Any] | None:
    sheet_df = df[df["sheet"] == sheet].copy()
    if sheet_df.empty:
        return None
    # 强制全国口径优先：先找真正的全国行（province==全部或含"全国"）
    national_df = sheet_df[sheet_df["province"].apply(lambda p: is_national_scope(str(p)) if pd.notna(p) else False)].copy()
    if not national_df.empty:
        scoped_df = national_df
        scope_label = "全国"
    else:
        # fallback：按原有 pick_preferred_scope_name 逻辑取首选地区
        preferred_scope = pick_preferred_scope_name(sheet_df["province"].dropna().unique().tolist())
        if preferred_scope:
            scoped_df = sheet_df[sheet_df["province"] == preferred_scope].copy()
            scope_label = preferred_scope
        else:
            scoped_df = sheet_df.copy()
            scope_label = "云南"
    if scoped_df.empty:
        return None
    snap_date = get_snapshot_date(scoped_df, target_date)
    if snap_date is None:
        return None
    day_df = scoped_df[scoped_df["date"] == snap_date].copy()
    if day_df.empty:
        return None
    metric_df = pd.DataFrame()
    for keyword in metric_keywords:
        metric_df = day_df[day_df["metric"].astype(str).str.contains(keyword, na=False)].copy()
        if not metric_df.empty:
            break
    if metric_df.empty:
        metric_df = day_df.copy()
    grouped = metric_df.groupby("metric", as_index=False)["value"].mean().sort_values("metric")
    if grouped.empty:
        return None
    picked = grouped.iloc[0]
    prev_dates = scoped_df[scoped_df["date"] < snap_date]["date"]
    prev_date = prev_dates.max() if not prev_dates.empty else None
    prev_value = np.nan
    if prev_date is not None:
        prev_df = scoped_df[(scoped_df["date"] == prev_date) & (scoped_df["metric"] == picked["metric"])]
        if not prev_df.empty:
            prev_value = prev_df["value"].mean()
    return {
        "sheet": sheet,
        "metric": text_of(picked["metric"]),
        "value": float(picked["value"]),
        "date": snap_date,
        "prev_date": prev_date,
        "prev_value": prev_value,
        "scope": scope_label,
    }


def build_yongyi_global_summary(numeric_df: pd.DataFrame, target_date: pd.Timestamp) -> dict[str, Any]:
    if numeric_df.empty:
        return {"ok": False}
    target_date = pd.Timestamp(target_date).normalize()
    out_price = get_sheet_metric_snapshot(numeric_df, "出栏价", ["全国均价", "均价"], target_date)
    slaughter = get_sheet_metric_snapshot(numeric_df, "价格+宰量", ["屠宰量"], target_date)
    if slaughter is None:
        slaughter = get_sheet_metric_snapshot(numeric_df, "屠宰企业日度屠宰量", ["日度屠宰量"], target_date)
    # 150/175kg猪较标猪价差：直接从散户标肥价差读取原始价差
    fat_spread_150 = None
    fat_spread_175 = None
    fat_spread_sheet_df = numeric_df[numeric_df["sheet"] == "散户标肥价差"].copy()
    if not fat_spread_sheet_df.empty:
        _, spread_view_df = build_scattered_fat_price_views(fat_spread_sheet_df)
        if not spread_view_df.empty:
            # 取全国均价：筛选 province=="全部" 的行，若无则取全部省份均值
            if "province" in spread_view_df.columns:
                national = spread_view_df[spread_view_df["province"].apply(
                    lambda p: is_national_scope(str(p)) if pd.notna(p) else False
                )]
                scoped = national if not national.empty else spread_view_df
            else:
                scoped = spread_view_df
            snap = get_snapshot_date(scoped, target_date)
            if snap is not None:
                snap_data = scoped[scoped["date"] == snap]
                prev_date = scoped[scoped["date"] < snap]["date"].max()
                prev_data = scoped[scoped["date"] == prev_date] if pd.notna(prev_date) else pd.DataFrame()
                # 150kg较标猪价差：筛选含"较标猪"的原始价差
                s150 = snap_data[snap_data["metric"].astype(str).str.contains("150.*较标猪|较标猪.*150", na=False, regex=True)]
                if not s150.empty:
                    val_150 = s150["value"].mean()
                    prev_150 = np.nan
                    if not prev_data.empty:
                        p150 = prev_data[prev_data["metric"].astype(str).str.contains("150.*较标猪|较标猪.*150", na=False, regex=True)]
                        if not p150.empty:
                            prev_150 = p150["value"].mean()
                    fat_spread_150 = {"metric": "150kg猪较标猪价差", "value": float(val_150),
                                      "prev_value": prev_150, "date": snap}
                # 175kg较标猪价差（相同方式）
                s175 = snap_data[snap_data["metric"].astype(str).str.contains("175.*较标猪|较标猪.*175", na=False, regex=True)]
                if not s175.empty:
                    val_175 = s175["value"].mean()
                    prev_175 = np.nan
                    if not prev_data.empty:
                        p175 = prev_data[prev_data["metric"].astype(str).str.contains("175.*较标猪|较标猪.*175", na=False, regex=True)]
                        if not p175.empty:
                            prev_175 = p175["value"].mean()
                    fat_spread_175 = {"metric": "175kg猪较标猪价差", "value": float(val_175),
                                      "prev_value": prev_175, "date": snap}

    lines: list[str] = []
    cards: list[dict[str, str]] = []
    summary_date = target_date

    if out_price is not None:
        summary_date = out_price["date"]
        delta = out_price["value"] - out_price["prev_value"] if pd.notna(out_price["prev_value"]) else np.nan
        scope_name = "全国" if is_national_scope(out_price["scope"]) else out_price["scope"]
        direction = "偏强" if pd.notna(delta) and delta > 0 else ("偏弱" if pd.notna(delta) and delta < 0 else "平稳")
        lines.append(f"今日现货{direction}，{scope_name}出栏均价 {format_number(out_price['value'])}，{describe_delta(delta, '较前一日')}。")
        cards.append({"label": "现货均价", "value": format_number(out_price["value"]), "extra": f"口径：{scope_name}｜{describe_delta(delta, '日环比')}"})

    if slaughter is not None:
        delta = slaughter["value"] - slaughter["prev_value"] if pd.notna(slaughter["prev_value"]) else np.nan
        lines.append(f"屠宰量方面，当前口径为 {format_number(slaughter['value'])}，{describe_delta(delta, '较前一日')}。")
        cards.append({"label": "屠宰量", "value": format_number(slaughter["value"]), "extra": describe_delta(delta, "较前一日")})

    if fat_spread_150 is not None:
        delta = fat_spread_150["value"] - fat_spread_150["prev_value"] if pd.notna(fat_spread_150["prev_value"]) else np.nan
        spread_state = "走阔" if pd.notna(delta) and delta > 0 else ("收窄" if pd.notna(delta) and delta < 0 else "持平")
        lines.append(f"150kg猪较标猪价差 {spread_state}，当前参考值 {format_number(fat_spread_150['value'])}，{describe_delta(delta, '较前一日')}。")
        cards.append({"label": "150kg猪较标猪价差", "value": format_number(fat_spread_150["value"]), "extra": f"散户标肥价差｜{describe_delta(delta, '较前一日')}"})

    if fat_spread_175 is not None:
        delta = fat_spread_175["value"] - fat_spread_175["prev_value"] if pd.notna(fat_spread_175["prev_value"]) else np.nan
        spread_state = "走阔" if pd.notna(delta) and delta > 0 else ("收窄" if pd.notna(delta) and delta < 0 else "持平")
        lines.append(f"175kg猪较标猪价差 {spread_state}，当前参考值 {format_number(fat_spread_175['value'])}，{describe_delta(delta, '较前一日')}。")
        cards.append({"label": "175kg猪较标猪价差", "value": format_number(fat_spread_175["value"]), "extra": f"散户标肥价差｜{describe_delta(delta, '较前一日')}"})

    brief = " ".join(lines[:4]) if lines else "暂无可用于汇总的现货数据。"
    return {
        "ok": bool(lines),
        "date": summary_date,
        "title": "今日现货综合摘要",
        "brief": brief,
        "lines": lines,
        "cards": cards,
    }


def build_summary_payload(df: pd.DataFrame, target_date: pd.Timestamp, show_comparison: bool = True) -> dict[str, Any]:
    if df.empty:
        return {"ok": False, "message": "当前筛选下没有数据。"}
    target_date = pd.Timestamp(target_date).normalize()
    summary_date = get_snapshot_date(df, target_date)
    if summary_date is None:
        return {"ok": False, "message": "当前筛选下没有可总结日期。"}
    current_df = df[df["date"] == summary_date].copy()
    if current_df.empty:
        return {"ok": False, "message": "当前筛选下没有可总结数据。"}

    group_field = get_group_field(df)
    group_label = get_group_label(df)
    series_by_date = df.groupby("date", as_index=False)["value"].mean().sort_values("date")
    current_mean = current_df["value"].mean()
    range_high = current_df["value"].max()
    range_low = current_df["value"].min()
    range_span = range_high - range_low if pd.notna(range_high) and pd.notna(range_low) else np.nan

    rank_df = pd.DataFrame()
    top_rows = pd.DataFrame()
    bottom_rows = pd.DataFrame()
    if group_field is not None:
        rank_df = current_df.groupby(group_field, as_index=False)["value"].mean().sort_values("value", ascending=False)
        rank_df["series_name"] = rank_df[group_field]
        if len(rank_df) > 1:
            top_rows = rank_df.head(3)
            bottom_rows = rank_df.sort_values("value", ascending=True).head(3)

    leading_text = "；".join(f"{row.series_name} {format_number(row.value)}" for row in top_rows.itertuples()) if not top_rows.empty else f"暂无{group_label}排行"
    lagging_text = "；".join(f"{row.series_name} {format_number(row.value)}" for row in bottom_rows.itertuples()) if not bottom_rows.empty else f"暂无{group_label}排行"

    if not show_comparison:
        concise_summary = f"当前值 {format_number(current_mean)}。"
        cards = [
            {"label": "核心结论", "value": concise_summary, "extra": f"口径：{group_label if group_field is not None else '全国'}"},
            {"label": "分析日期", "value": format_date_cn(summary_date), "extra": f"若输入日期无数据，则自动回退到最近可用日期。{'' if summary_date == target_date else '当前已回退到最近数据日期。'}"},
            {"label": "当前值", "value": format_number(current_mean), "extra": ""},
            {"label": "当日波动区间", "value": format_number(range_span), "extra": f"最高 {format_number(range_high)} / 最低 {format_number(range_low)}"},
        ]
        summary_lines = [
            f"<b>当前值：</b>{format_number(current_mean)}",
            f"分析日期：<b>{format_date_cn(summary_date)}</b>。",
        ]
        if group_field is None:
            summary_lines.append("当前筛选为单一口径，仅展示当前具体值。")
        else:
            summary_lines.extend([
                f"当日高位{group_label}：<b>{leading_text}</b>。",
                f"当日低位{group_label}：<b>{lagging_text}</b>。",
            ])
        if "province" in current_df.columns and any(is_national_scope(x) for x in current_df["province"].dropna().unique()):
            summary_lines.append(f"当前默认展示的全国口径数值为 <b>{format_number(current_mean)}</b>。")
        return {
            "ok": True,
            "summary_date": summary_date,
            "cards": cards,
            "summary_lines": summary_lines,
            "brief_text": concise_summary,
            "top_rows": top_rows,
            "bottom_rows": bottom_rows,
            "range_df": rank_df,
            "rank_field": "series_name",
            "group_label": group_label,
        }

    prev_date_series = series_by_date[series_by_date["date"] < summary_date]
    prev_date = prev_date_series["date"].max() if not prev_date_series.empty else None
    prev_mean = prev_date_series.loc[prev_date_series["date"] == prev_date, "value"].mean() if prev_date is not None else np.nan

    yoy_match = same_day_last_year_match(series_by_date["date"], summary_date)
    yoy_mean = series_by_date.loc[series_by_date["date"] == yoy_match, "value"].mean() if yoy_match is not None else np.nan

    window_df = series_by_date[series_by_date["date"] <= summary_date].tail(30)
    momentum_text = trend_position_label(current_mean, window_df["value"]) if len(window_df) >= 3 else "近期样本不足"

    seasonal_text = "暂无历年同期对比样本"
    if yoy_match is not None:
        target_md = summary_date.strftime("%m-%d")
        same_day_samples = series_by_date[series_by_date["date"].dt.strftime("%m-%d") == target_md]
        if len(same_day_samples) >= 2:
            same_day_mean = same_day_samples["value"].mean()
            diff = current_mean - same_day_mean
            if diff > 0:
                seasonal_text = f"高于历年同期均值 {format_number(abs(diff))}"
            elif diff < 0:
                seasonal_text = f"低于历年同期均值 {format_number(abs(diff))}"
            else:
                seasonal_text = "与历年同期均值基本持平"

    if group_field is None:
        change_rank_text = "当前为单一口径，不涉及横向涨跌分化。"
    else:
        change_rank_text = f"暂无{group_label}环比变化"
        if prev_date is not None and not rank_df.empty:
            prev_group = df[df["date"] == prev_date].groupby(group_field, as_index=False)["value"].mean().rename(columns={"value": "prev_value"})
            merged = rank_df.merge(prev_group, on=group_field, how="inner")
            if not merged.empty:
                merged["delta"] = merged["value"] - merged["prev_value"]
                rise = merged.sort_values("delta", ascending=False).head(2)
                fall = merged.sort_values("delta", ascending=True).head(2)
                rise_text = "；".join(f"{getattr(row, group_field)} {format_number(row.delta)}" for row in rise.itertuples() if row.delta > 0)
                fall_text = "；".join(f"{getattr(row, group_field)} {format_number(abs(row.delta))}" for row in fall.itertuples() if row.delta < 0)
                parts = []
                if rise_text:
                    parts.append(f"领涨{group_label}：{rise_text}")
                if fall_text:
                    parts.append(f"回落{group_label}：{fall_text}")
                if parts:
                    change_rank_text = "；".join(parts) + "。"

    concise_summary = "，".join(
        part for part in [
            f"{momentum_text}，当前值 {format_number(current_mean)}",
            describe_delta(current_mean - prev_mean if pd.notna(prev_mean) else np.nan, "较前一日"),
            describe_delta(current_mean - yoy_mean if pd.notna(yoy_mean) else np.nan, "较去年同期"),
            seasonal_text,
        ] if part
    ) + "。"

    cards = [
        {"label": "核心结论", "value": concise_summary, "extra": f"口径：{group_label if group_field is not None else '全国'}"},
        {"label": "分析日期", "value": format_date_cn(summary_date), "extra": f"若输入日期无数据，则自动回退到最近可用日期。{'' if summary_date == target_date else '当前已回退到最近数据日期。'}"},
        {"label": "当日均值", "value": format_number(current_mean), "extra": momentum_text},
        {"label": "较前一日", "value": delta_html(current_mean - prev_mean if pd.notna(prev_mean) else np.nan, prev_mean), "extra": f"对比日期：{format_date_cn(prev_date) if prev_date is not None else '暂无'}"},
        {"label": "较去年同期", "value": delta_html(current_mean - yoy_mean if pd.notna(yoy_mean) else np.nan, yoy_mean), "extra": f"对比日期：{format_date_cn(yoy_match) if yoy_match is not None else '暂无'}"},
        {"label": "当日波动区间", "value": format_number(range_span), "extra": f"最高 {format_number(range_high)} / 最低 {format_number(range_low)}"},
        {"label": "历年同期位置", "value": seasonal_text, "extra": "基于同月同日样本判断。"},
    ]

    summary_lines = [
        f"<b>核心判断：</b>{concise_summary}",
        f"整体看，{format_date_cn(summary_date)} 当前口径均值为 <b>{format_number(current_mean)}</b>，{momentum_text}。",
        f"相较前一可比日期，{delta_html(current_mean - prev_mean if pd.notna(prev_mean) else np.nan, prev_mean)}；相较去年同期，{delta_html(current_mean - yoy_mean if pd.notna(yoy_mean) else np.nan, yoy_mean)}。",
    ]
    if group_field is None:
        summary_lines.append("当前筛选为单一口径，因此不展示横向排行对比。")
    else:
        summary_lines.extend([
            f"当日高位{group_label}主要集中在：<b>{leading_text}</b>。",
            f"当日低位{group_label}主要集中在：<b>{lagging_text}</b>。",
            change_rank_text,
        ])
    if "province" in current_df.columns and any(is_national_scope(x) for x in current_df["province"].dropna().unique()):
        summary_lines.append(f"当前默认展示的全国口径数值为 <b>{format_number(current_mean)}</b>。")

    return {
        "ok": True,
        "summary_date": summary_date,
        "cards": cards,
        "summary_lines": summary_lines,
        "brief_text": concise_summary,
        "top_rows": top_rows,
        "bottom_rows": bottom_rows,
        "range_df": rank_df,
        "rank_field": "series_name",
        "group_label": group_label,
    }


def render_summary(payload: dict[str, Any]) -> None:
    if not payload.get("ok"):
        st.warning(payload.get("message", "暂无总结结果。"))
        return
    render_metric_cards(payload["cards"])
    group_label = payload.get("group_label", "地区")
    brief_text = payload.get("brief_text", "")
    if brief_text:
        st.markdown(f"<div class='summary-box'><div class='summary-subtitle'>当日一句话结论</div><div style='font-size:20px;font-weight:800;color:#14213d;line-height:1.7'>{brief_text}</div></div>", unsafe_allow_html=True)
    bullets = "".join(f"<li>{line}</li>" for line in payload["summary_lines"])
    html = (
        '<div class="summary-box">'
        '<div class="summary-title">指定日期深度总结</div>'
        f'<div class="summary-subtitle">以下结论已结合当前已选的 Sheet、指标、{group_label}和日期口径自动生成。</div>'
        f'<ul class="bullet-list">{bullets}</ul>'
        '</div>'
    )
    st.markdown(html, unsafe_allow_html=True)
    top_rows = payload.get("top_rows")
    bottom_rows = payload.get("bottom_rows")
    rank_field = payload.get("rank_field", "series_name")
    if top_rows is not None and not top_rows.empty:
        col1, col2 = st.columns(2)
        with col1:
            fig = px.bar(top_rows.sort_values("value", ascending=True), x="value", y=rank_field, orientation="h", color="value", color_continuous_scale="Reds")
            fig.update_traces(hovertemplate=f"{group_label}：%{{y}}<br>数值：%{{x:.2f}}<extra></extra>")
            fig.update_layout(title=f"高位{group_label}", xaxis_title="数值", yaxis_title=group_label, coloraxis_showscale=False, template="plotly_white")
            render_plotly(fig, "summary", "top", group_label)
        with col2:
            if bottom_rows is not None and not bottom_rows.empty:
                fig = px.bar(bottom_rows.sort_values("value", ascending=False), x="value", y=rank_field, orientation="h", color="value", color_continuous_scale="Greens")
                fig.update_traces(hovertemplate=f"{group_label}：%{{y}}<br>数值：%{{x:.2f}}<extra></extra>")
                fig.update_layout(title=f"低位{group_label}", xaxis_title="数值", yaxis_title=group_label, coloraxis_showscale=False, template="plotly_white")
                render_plotly(fig, "summary", "bottom", group_label)


# -----------------------------
# 期货数据解析
# -----------------------------
def add_futures_record(records: list[dict], sheet: str, date: pd.Timestamp | None, metric: str, value: Any, contract: str, series_name: str, series_type: str = "daily") -> None:
    ts = parse_date(date)
    number = to_float(value)
    if ts is None or number is None:
        return
    contract_lh = resolve_lh_contract_code(contract, ts)
    records.append(
        {
            "sheet": sheet,
            "date": ts,
            "metric": metric,
            "value": number,
            "contract": contract_lh,
            "contract_month_code": contract,
            "contract_lh": contract_lh,
            "series_name": contract_lh if series_name == contract else series_name,
            "series_type": series_type,
            "province": "全部",
            "city": "",
            "display_name": contract_lh if series_name == contract else series_name,
        }
    )


def infer_futures_contract(title: str) -> str:
    text = text_of(title)
    match = re.search(r"(\d{2})合约", text)
    if match:
        return f"{match.group(1)}合约"
    if "主力合约" in text:
        return "主力合约"
    if "连续合约" in text:
        return "连续合约"
    if "注册仓单" in text:
        return "仓单"
    if "交割量" in text:
        return "交割量"
    if "单边交易" in text:
        return "全部合约"
    return text or "未识别"


def _sorted_contracts(contracts: list[str]) -> list[str]:
    def sort_key(contract: str) -> tuple[int, int, str]:
        text = text_of(contract).upper()
        match = re.fullmatch(r"LH(\d{2})(\d{2})", text)
        if match:
            return (0, int(match.group(1)) * 100 + int(match.group(2)), text)
        return (1, 0, text)
    return sorted([text_of(x) for x in contracts if text_of(x)], key=sort_key)


def get_active_lh_contracts(price_df: pd.DataFrame, target_date: Any, include_special: bool = False) -> list[str]:
    if price_df.empty or "contract" not in price_df.columns or "date" not in price_df.columns:
        return []
    target_ts = parse_date(target_date)
    if target_ts is None:
        return []
    temp = price_df.copy()
    temp = temp[temp["contract"].fillna("").astype(str).str.startswith("LH")].copy()
    if temp.empty:
        return []
    contract_ranges = temp.groupby("contract", as_index=False).agg(start_date=("date", "min"), end_date=("date", "max"))
    active = contract_ranges[(contract_ranges["start_date"] <= target_ts) & (contract_ranges["end_date"] >= target_ts)]["contract"].tolist()
    if include_special:
        special = [x for x in price_df["contract"].dropna().unique().tolist() if x in {"主力合约", "连续合约", "仓单", "交割量"}]
        active.extend(special)
    return _sorted_contracts(active)


def resolve_lh_contract_code(contract: str, date: Any) -> str:
    raw = text_of(contract)
    ts = parse_date(date)
    month_match = re.search(r"(\d{2})合约", raw)
    if not month_match or ts is None:
        return raw
    month = int(month_match.group(1))
    year = int(pd.Timestamp(ts).year)
    candidate = pd.Timestamp(year=year, month=month, day=1)
    if pd.Timestamp(ts).normalize() > (candidate + pd.DateOffset(months=1)):
        year += 1
    return f"LH{str(year)[-2:]}{month:02d}"


def classify_lh_contract_bucket(contract_lh: str) -> str | None:
    text = text_of(contract_lh).upper()
    match = re.fullmatch(r"LH(\d{2})(\d{2})", text)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    code = year * 100 + month
    if 2409 <= code <= 2503:
        return "premium_lh2409_to_lh2503"
    if code >= 2505:
        return "premium_lh2505_plus"
    return None


def infer_futures_metric(sheet: str, title: str) -> str:
    text = text_of(title)
    if sheet == "收盘价":
        return "收盘价"
    if " - " in text or "净持仓" in text:
        return "前20净持仓"
    if any(flag in text for flag in ["多单持仓合计", "持买单量合计"]):
        return "前20多单"
    if any(flag in text for flag in ["空单持仓合计", "持卖单量合计"]):
        return "前20空单"
    if "成交量" in text:
        return "成交量"
    if "持仓量" in text:
        return "持仓量"
    if "基差" in text:
        return "基差"
    if any(flag in text for flag in ["现货价", "市场价", "出栏均价"]):
        return "现货价"
    if "主力合约：收盘价" in text:
        return "主力收盘价"
    if "连续合约：收盘价" in text:
        return "连续收盘价"
    if "注册仓单" in text:
        return "注册仓单"
    if "交割量" in text:
        return "交割量"
    return text or sheet


def make_futures_series_name(metric: str, contract: str, title: str) -> str:
    text = text_of(title)
    if metric in {"收盘价", "持仓量", "成交量", "前20净持仓", "前20多单", "前20空单"}:
        return contract
    if metric == "基差":
        if "四川" in text and "连续" in text:
            return "四川-连续基差"
        if "四川" in text and "主力" in text:
            return "四川-主力基差"
        if "河南" in text and "连续" in text:
            return "河南-连续基差"
        if "河南" in text and "主力" in text:
            return "河南-主力基差"
        return text or "基差"
    if metric == "现货价":
        if "四川" in text:
            return "四川现货价"
        if "河南" in text:
            return "河南现货价"
        return text or "现货价"
    if metric == "注册仓单":
        return "注册仓单"
    if metric == "主力收盘价":
        return "主力收盘价"
    if metric == "连续收盘价":
        return "连续收盘价"
    if metric == "交割量":
        return "交割量"
    return text or f"{contract}｜{metric}"


def parse_year_from_label(label: str) -> int | None:
    text = text_of(label)
    match = re.search(r"(20\d{2})", text)
    return int(match.group(1)) if match else None


def make_date_from_period_label(label: str, year: int) -> pd.Timestamp | None:
    text = text_of(label)
    if not text:
        return None
    if re.fullmatch(r"\d{2}-\d{2}", text):
        month, day = text.split("-")
        try:
            return pd.Timestamp(year=year, month=int(month), day=int(day))
        except Exception:
            return None
    month_match = re.fullmatch(r"(\d{1,2})月", text)
    if month_match:
        try:
            return pd.Timestamp(year=year, month=int(month_match.group(1)), day=1)
        except Exception:
            return None
    return None


def parse_futures_standard_table(rows: list[list[Any]], sheet: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= 10:
        return records
    header = rows[1] if len(rows) > 1 else []
    for row in rows[10:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for col in range(1, len(header)):
            title = text_of(header[col] if col < len(header) else None)
            if not title:
                continue
            value = row[col] if col < len(row) else None
            contract = infer_futures_contract(title)
            metric = infer_futures_metric(sheet, title)
            series_name = make_futures_series_name(metric, contract, title)
            add_futures_record(records, sheet, date, metric, value, contract, series_name)
    return records


def get_block_starts(title_row: list[Any]) -> list[int]:
    starts = [idx for idx, value in enumerate(title_row) if text_of(value) == "指标名称"]
    return starts


def parse_futures_multi_blocks(rows: list[list[Any]], sheet: str, data_start_row: int = 10) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= data_start_row:
        return records
    max_len = max(len(row) for row in rows[: min(len(rows), 20)])
    title_row = (rows[1] if len(rows) > 1 else []) + [None] * (max_len - len(rows[1]) if len(rows) > 1 else max_len)
    starts = get_block_starts(title_row)
    if not starts:
        return records
    starts.append(max_len)
    for idx in range(len(starts) - 1):
        start = starts[idx]
        end = starts[idx + 1]
        for col in range(start + 1, end):
            title = text_of(title_row[col])
            if not title or title == "指标名称":
                continue
            contract = infer_futures_contract(title)
            metric = infer_futures_metric(sheet, title)
            series_name = make_futures_series_name(metric, contract, title)
            for row in rows[data_start_row:]:
                if not row:
                    continue
                date = parse_date(row[start] if len(row) > start else None)
                value = row[col] if len(row) > col else None
                add_futures_record(records, sheet, date, metric, value, contract, series_name)
    return records


def parse_futures_seasonal_blocks(rows: list[list[Any]], sheet: str, year_row_idx: int = 10, data_start_row: int = 11) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= data_start_row:
        return records
    max_len = max(len(row) for row in rows[: min(len(rows), data_start_row + 3)])
    title_row = (rows[1] if len(rows) > 1 else []) + [None] * (max_len - len(rows[1]) if len(rows) > 1 else max_len)
    year_row = (rows[year_row_idx] if len(rows) > year_row_idx else []) + [None] * (max_len - len(rows[year_row_idx]) if len(rows) > year_row_idx else max_len)
    starts = get_block_starts(title_row)
    if not starts:
        return records
    starts.append(max_len)
    for idx in range(len(starts) - 1):
        start = starts[idx]
        end = starts[idx + 1]
        block_title = text_of(title_row[start + 1] if start + 1 < len(title_row) else None)
        contract = infer_futures_contract(block_title)
        metric = infer_futures_metric(sheet, block_title)
        series_name = make_futures_series_name(metric, contract, block_title)
        for col in range(start + 1, end):
            year = parse_year_from_label(year_row[col] if col < len(year_row) else None)
            if year is None:
                continue
            for row in rows[data_start_row:]:
                if not row:
                    continue
                period_label = text_of(row[start] if len(row) > start else None)
                date = make_date_from_period_label(period_label, year)
                value = row[col] if len(row) > col else None
                add_futures_record(records, sheet, date, metric, value, contract, series_name, series_type="seasonal")
    return records


def finalize_futures_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        cols = ["sheet", "date", "metric", "value", "contract", "contract_month_code", "contract_lh", "series_name", "series_type", "province", "city", "display_name", "year", "month", "week", "month_day", "lunar_label", "lunar_order", "is_leap_lunar"]
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(records)
    for col in ["contract_month_code", "contract_lh"]:
        if col not in df.columns:
            df[col] = df["contract"] if "contract" in df.columns else None
    df = enrich_date_features(df)
    return df.sort_values(["sheet", "metric", "contract", "date"]).reset_index(drop=True)


def build_futures_dataset_from_path(path_str: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    return _cached_build_futures_dataset_from_path(_get_file_version(path_str))


@st.cache_data(show_spinner=False)
def _cached_build_futures_dataset_from_path(versioned_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    real_path = versioned_path.split("@@v=")[0]
    sheets = read_workbook_rows_from_path(real_path)
    records: list[dict] = []
    logs: list[dict] = []
    for sheet, rows in sheets.items():
        before = len(records)
        if sheet == "收盘价":
            records.extend(parse_futures_seasonal_blocks(rows, sheet))
            mode = "季节性分块表"
        elif sheet in {"持仓量", "成交量"}:
            records.extend(parse_futures_standard_table(rows, sheet))
            mode = "标准日度表"
        elif sheet in {"前20多空净持仓", "期现基差"}:
            records.extend(parse_futures_multi_blocks(rows, sheet))
            mode = "多指标并排分块表"
        elif sheet == "仓单、虚实盘、交割量":
            records.extend(parse_futures_multi_blocks(rows, sheet))
            records.extend(parse_futures_seasonal_blocks(rows, sheet))
            mode = "混合分块表"
        else:
            mode = "未适配"
        logs.append({"sheet": sheet, "mode": mode, "rows": len(records) - before})
    return finalize_futures_df(records), pd.DataFrame(logs)


# -----------------------------
# 调运数据解析
# -----------------------------
def build_transport_dataset_from_path(path_str: str) -> pd.DataFrame:
    return _cached_build_transport_dataset_from_path(_get_file_version(path_str))


@st.cache_data(show_spinner=False)
def _cached_build_transport_dataset_from_path(versioned_path: str) -> pd.DataFrame:
    real_path = versioned_path.split("@@v=")[0]
    path = resolve_excel_path(real_path)
    df = pd.read_excel(path)
    rename_map = {col: text_of(col) for col in df.columns}
    df = df.rename(columns=rename_map)
    needed_cols = ["调出省份", "调入省份", "调出城市", "调入城市", "日期"]
    for col in needed_cols:
        if col not in df.columns:
            raise ValueError(f"调运数据缺少字段：{col}")
    temp = df.copy()
    temp["date"] = pd.to_datetime(temp["日期"], errors="coerce").dt.normalize()
    temp["调出省份"] = temp["调出省份"].map(canonicalize_province)
    temp["调入省份"] = temp["调入省份"].map(canonicalize_province)
    temp["调出城市"] = temp["调出城市"].map(clean_city_name)
    temp["调入城市"] = temp["调入城市"].map(clean_city_name)
    temp = temp[temp["date"].notna()].copy()
    temp = temp[temp["调出省份"].isin(VALID_PROVINCES) & temp["调入省份"].isin(VALID_PROVINCES)].copy()
    temp = temp[temp["调出城市"].map(is_valid_city_name) & temp["调入城市"].map(is_valid_city_name)].copy()
    temp["省际路线"] = temp["调出省份"] + "→" + temp["调入省份"]
    temp["城市路线"] = temp["调出省份"] + "-" + temp["调出城市"] + " → " + temp["调入省份"] + "-" + temp["调入城市"]
    temp["count"] = 1
    return temp.sort_values(["date", "调出省份", "调入省份", "调出城市", "调入城市"]).reset_index(drop=True)


# -----------------------------
# 鲜冻品数据解析
# -----------------------------

# 预设年份（鲜品/冻品日期无年份信息，默认当年）
_CURRENT_YEAR = 2026


def _parse_month_day_float(value: Any) -> pd.Timestamp | None:
    """解析"月.日"浮点日期（如 5.16 → 2025-05-16），用于鲜品价格日报。"""
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        num = float(value)
        month = int(num)
        day = int(round((num - month) * 100))
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                return pd.Timestamp(year=_CURRENT_YEAR, month=month, day=day)
            except Exception:
                return None
    return None


def _parse_week_range_date(value: Any) -> pd.Timestamp | None:
    """解析周区间日期字符串（如"5月18日-5月24日"→起始日），用于冻品价格日报。"""
    text = text_of(value)
    if not text:
        return None
    text_clean = text.replace("\n", "").replace(" ", "")
    # 尝试中文周区间格式: "5月18日-5月24日"
    m = re.match(r"(\d{1,2})月(\d{1,2})日?[-~～](\d{1,2})月(\d{1,2})日?", text_clean)
    if m:
        try:
            return pd.Timestamp(year=_CURRENT_YEAR, month=int(m.group(1)), day=int(m.group(2)))
        except Exception:
            pass
    # 尝试小数点周区间格式: "5.18-5.24"
    m = re.match(r"(\d{1,2})\.(\d{1,2})[-~～](\d{1,2})\.(\d{1,2})", text_clean)
    if m:
        try:
            return pd.Timestamp(year=_CURRENT_YEAR, month=int(m.group(1)), day=int(m.group(2)))
        except Exception:
            pass
    # 尝试单个中文日期: "5月18日"
    m = re.match(r"(\d{1,2})月(\d{1,2})日?", text_clean)
    if m:
        try:
            return pd.Timestamp(year=_CURRENT_YEAR, month=int(m.group(1)), day=int(m.group(2)))
        except Exception:
            pass
    # 尝试浮点格式兜底
    return _parse_month_day_float(text)


def _is_price_sheet(sheet_name: str, rows: list[list[Any]]) -> bool:
    """判断一个 sheet 是否为价格日报（而非名称对照等非数据表）。
    价格日报特征：第 0 行含产品名列，第 1 行含供应商名，第 2 行起有数值。"""
    if len(rows) < 3:
        return False
    # 名称对照表：sheet 名含"名称"或"对照"，或第 0 行数据含"名称"
    if any(kw in sheet_name for kw in ("名称", "对照")):
        return False
    if any("名称" in str(c) or "对照" in str(c) for c in rows[0] if c):
        return False
    # 检查第 2 行（第一个数据行）col[0] 是否为日期（浮点或周区间）
    first_val = rows[2][0] if len(rows) > 2 and rows[2] else None
    if first_val is not None:
        if isinstance(first_val, (int, float)):
            month = int(float(first_val))
            if 1 <= month <= 12 and 0 < float(first_val) - month < 1:
                return True
        text = text_of(first_val)
        if text and re.search(r"\d{1,2}月\d{1,2}日", text):
            return True
    # 检查第 1 行是否包含供应商关键字（第 1 行 = 供应商行，对价格日报有效）
    supplier_keywords = ["神农", "双汇", "牧原", "千喜鹤", "众品"]
    row1_text = " ".join(str(c) for c in rows[1] if c)
    if any(kw in row1_text for kw in supplier_keywords):
        return True
    # 检查是否有明显数值列（第 2 行起有浮点价格）
    if len(rows) > 2:
        numeric_count = 0
        for col in range(1, min(len(rows[2]), 10)):
            v = rows[2][col] if col < len(rows[2]) else None
            if isinstance(v, (int, float)) and not pd.isna(v):
                numeric_count += 1
        if numeric_count >= 2:
            return True
    return False


def add_price_record(records: list[dict], sheet: str, category: str, product: str, supplier: str, date: pd.Timestamp | None, raw_value: Any) -> None:
    """添加一条价格记录。product = 产品名（如"去皮前段"），supplier = 供应商名（如"神农鲜品"）。"""
    low, high, mid, raw_text = parse_price_band(raw_value)
    if date is None or mid is None or not product:
        return
    records.append(
        {
            "sheet": sheet,
            "dataset_type": "鲜品" if "鲜品" in sheet else "冻品",
            "category": category or "未分组",
            "product": product,
            "supplier": supplier or "未知",
            "series_name": f"{product} · {supplier}" if supplier else product,
            "date": date,
            "value": mid,
            "mid_price": mid,
            "low_price": low,
            "high_price": high,
            "price_range_text": raw_text,
            "province": "全部",
            "city": "",
            "display_name": f"{product} · {supplier}" if supplier else product,
        }
    )


def build_fresh_frozen_dataset_from_path(path_str: str) -> pd.DataFrame:
    """从神农肉业鲜品/冻品 Excel 解析价格数据。缓存自动感知文件更新时间。"""
    return _cached_build_fresh_frozen_dataset_from_path(_get_file_version(path_str))


@st.cache_data(show_spinner=False)
def _cached_build_fresh_frozen_dataset_from_path(versioned_path: str) -> pd.DataFrame:
    """内部缓存函数：versioned_path 包含 mtime，文件更新时自动失效。"""
    real_path = versioned_path.split("@@v=")[0]
    sheets = read_workbook_rows_from_path(real_path)
    records: list[dict] = []
    sheets_processed = 0
    sheets_skipped = 0
    parse_diag: list[str] = []  # 诊断信息

    parse_diag.append(f"文件: {real_path}")
    parse_diag.append(f"共 {len(sheets)} 个 sheet: {list(sheets.keys())}")

    for sheet, rows in sheets.items():
        if len(rows) < 3:
            parse_diag.append(f"  [{sheet}] 跳过：行数不足 ({len(rows)} < 3)")
            sheets_skipped += 1
            continue

        # ── 尝试标准格式（Row 0=产品名, Row 1=供应商名）──
        standard_ok = _is_price_sheet(sheet, rows)
        wide_ok = False
        if not standard_ok:
            wide_ok = _is_wide_format_sheet(sheet, rows)

        if not standard_ok and not wide_ok:
            skip_reason = _get_skip_reason(sheet, rows)
            parse_diag.append(f"  [{sheet}] 跳过：{skip_reason}")
            sheets_skipped += 1
            continue

        # ── 选择解析路径 ──
        if standard_ok:
            sheet_records = _parse_standard_sheet(sheet, rows, records, parse_diag)
        else:
            sheet_records = _parse_wide_format_sheet(sheet, rows, records, parse_diag)

        if sheet_records > 0:
            sheets_processed += 1
        else:
            sheets_skipped += 1

    parse_diag.append(f"结果: {sheets_processed} 个sheet处理, {sheets_skipped} 个跳过, {len(records)} 条记录")

    if not records:
        cols = ["sheet", "dataset_type", "category", "product", "supplier", "series_name",
                "date", "value", "mid_price", "low_price", "high_price",
                "price_range_text", "province", "city", "display_name"]
        df = pd.DataFrame(columns=cols)
        df.attrs["sheets_processed"] = sheets_processed
        df.attrs["sheets_skipped"] = sheets_skipped
        df.attrs["total_sheets"] = sheets_processed + sheets_skipped
        df.attrs["parse_diag"] = parse_diag
        return df
    df = pd.DataFrame(records)
    df = enrich_date_features(df)
    df.attrs["parse_diag"] = parse_diag
    return df.sort_values(["dataset_type", "category", "product", "date"]).reset_index(drop=True)


def _get_skip_reason(sheet_name: str, rows: list[list[Any]]) -> str:
    """返回 sheet 被跳过的原因。"""
    if any(kw in sheet_name for kw in ("名称", "对照")):
        return f"sheet名含「名称/对照」"
    if any("名称" in str(c) or "对照" in str(c) for c in rows[0] if c):
        return f"第0行含「名称/对照」: {[c for c in rows[0][:5] if c]}"
    first_val = rows[2][0] if len(rows) > 2 and rows[2] else None
    if first_val is not None:
        if isinstance(first_val, (int, float)):
            return f"首行col0为数值但非日期格式 ({first_val})"
        t = text_of(first_val)
        if t:
            return f"首行col0不匹配日期格式: {t[:30]!r}"
    row1_text = " ".join(str(c) for c in rows[1][:5] if c)
    return f"第1行无供应商关键字，且非宽表日期格式: {row1_text[:50]!r}"


def _is_wide_format_sheet(sheet_name: str, rows: list[list[Any]]) -> bool:
    """检测是否为宽表格式（5.鲜品冻品价格数据库.xlsx 格式）。
    特征：Row 0=分类名, Row 1=产品名, Row 2 col 0='YYYY-MM-DD HH:MM:SS' 日期字符串。"""
    if len(rows) < 3:
        return False
    if any(kw in sheet_name for kw in ("名称", "对照")):
        return False
    supplier_keywords = ["神农", "双汇", "牧原", "千喜鹤", "众品"]
    row1_text = " ".join(str(c) for c in rows[1] if c)
    if any(kw in row1_text for kw in supplier_keywords):
        return False
    first_val = rows[2][0] if len(rows) > 2 and rows[2] else None
    if first_val is not None:
        t = text_of(first_val)
        if re.match(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", t):
            return True
    return False
    """检测是否为宽表格式（5.鲜品冻品价格数据库.xlsx 格式）。
    特征：Row 0=分类名, Row 1=产品名, Row 2 col 0='YYYY-MM-DD HH:MM:SS' 日期字符串。"""
    if len(rows) < 3:
        return False
    # 跳过名称对照
    if any(kw in sheet_name for kw in ("名称", "对照")):
        return False
    # Row 1 不应包含供应商关键字（否则是标准格式）
    supplier_keywords = ["神农", "双汇", "牧原", "千喜鹤", "众品"]
    row1_text = " ".join(str(c) for c in rows[1] if c)
    if any(kw in row1_text for kw in supplier_keywords):
        return False
    # Row 2 col 0 应为 YYYY-MM-DD 日期字符串
    first_val = rows[2][0] if len(rows) > 2 and rows[2] else None
    if first_val is not None:
        t = text_of(first_val)
        if re.match(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", t):
            return True
    return False


def _parse_standard_sheet(sheet: str, rows: list[list[Any]], records: list[dict],
                          parse_diag: list[str]) -> int:
    """解析标准格式（神农肉业-鲜品/冻品价格.xlsx）：Row 0=产品名, Row 1=供应商名, Row 2+=数据。"""
    product_row_raw = rows[0]
    supplier_row = rows[1]
    products_filled = filled_right([text_of(x) for x in product_row_raw])

    n_cols = min(len(supplier_row), len(products_filled))
    if n_cols <= 1:
        parse_diag.append(f"  [{sheet}] 跳过：列数不足 (n_cols={n_cols})")
        return 0

    first_data_row = rows[2]
    first_raw_date = first_data_row[0] if len(first_data_row) > 0 else None
    test_date = _parse_month_day_float(first_raw_date) or _parse_week_range_date(first_raw_date)

    sheet_records = 0
    for row_idx, row in enumerate(rows[2:], start=2):
        if not row:
            continue
        raw_date = row[0] if len(row) > 0 else None
        date = _parse_month_day_float(raw_date)
        if date is None:
            date = _parse_week_range_date(raw_date)
        if date is None:
            date = parse_date(raw_date)
        if date is None:
            date = _try_loose_date_parse(raw_date)
        if date is None:
            if sheet_records == 0:
                parse_diag.append(f"  [{sheet}] Row {row_idx}: 日期解析失败 → {raw_date!r}")
            continue

        for col in range(1, n_cols):
            product_name = text_of(products_filled[col] if col < len(products_filled) else None)
            supplier_name = text_of(supplier_row[col] if col < len(supplier_row) else None)
            value = row[col] if col < len(row) else None
            if value is not None and isinstance(value, str):
                cleaned = value.strip()
                if cleaned in ("暂无", "停宰", "缺货", "-", "--", "无", "—", "停宰 ", ""):
                    continue
            add_price_record(records, sheet, product_name, product_name, supplier_name, date, value)
            sheet_records += 1

    if sheet_records > 0:
        products_set = sorted(set(r["product"] for r in records[-sheet_records:]))
        suppliers_set = sorted(set(r["supplier"] for r in records[-sheet_records:]))
        parse_diag.append(f"  [{sheet}] ✅ 标准格式 {sheet_records}条 | 首日={test_date} | 产品={products_set[:5]}... | 供应商={suppliers_set}")
    else:
        parse_diag.append(f"  [{sheet}] ⚠️ 标准格式但0条记录 (首日={test_date}, n_cols={n_cols})")
    return sheet_records


def _parse_wide_format_sheet(sheet: str, rows: list[list[Any]], records: list[dict],
                              parse_diag: list[str]) -> int:
    """解析宽表格式（5.鲜品冻品价格数据库.xlsx）：Row 0=分类名, Row 1=产品名, Row 2+=YYYY-MM-DD 日期+价格。"""
    category_row = rows[0]
    product_row = rows[1]
    categories_filled = filled_right([text_of(x) for x in category_row])

    n_cols = min(len(product_row), len(categories_filled))
    if n_cols <= 1:
        parse_diag.append(f"  [{sheet}] 跳过(宽表)：列数不足 (n_cols={n_cols})")
        return 0

    first_raw_date = rows[2][0] if len(rows) > 2 and rows[2] else None
    try:
        test_date = pd.Timestamp(first_raw_date) if first_raw_date else None
    except Exception:
        test_date = None

    sheet_records = 0
    for row_idx, row in enumerate(rows[2:], start=2):
        if not row:
            continue
        raw_date = row[0] if len(row) > 0 else None
        try:
            date = pd.Timestamp(raw_date) if raw_date else None
        except Exception:
            date = None
        if date is None or pd.isna(date):
            if sheet_records == 0:
                parse_diag.append(f"  [{sheet}] Row {row_idx}: 宽表日期解析失败 → {raw_date!r}")
            continue

        for col in range(1, n_cols):
            product_name = text_of(product_row[col] if col < len(product_row) else None)
            category_name = text_of(categories_filled[col] if col < len(categories_filled) else None)
            value = row[col] if col < len(row) else None
            if value is not None and isinstance(value, str):
                cleaned = value.strip()
                if cleaned in ("暂无", "停宰", "缺货", "-", "--", "无", "—", ""):
                    continue
            # 宽表格式无供应商信息，supplier 留空
            add_price_record(records, sheet, category_name, product_name, "", date, value)
            sheet_records += 1

    if sheet_records > 0:
        products_set = sorted(set(r["product"] for r in records[-sheet_records:]))
        parse_diag.append(f"  [{sheet}] ✅ 宽表格式 {sheet_records}条 | 首日={test_date} | 产品={products_set[:5]}...")
    else:
        parse_diag.append(f"  [{sheet}] ⚠️ 宽表格式但0条记录 (首日={test_date}, n_cols={n_cols})")
    return sheet_records


def _try_loose_date_parse(value: Any) -> pd.Timestamp | None:
    """宽松日期解析：尝试多种格式。"""
    t = text_of(value)
    if not t:
        return None
    # 纯数字 -> 尝试月日浮点
    if re.match(r"^\d{1,2}\.\d{1,2}$", t):
        return _parse_month_day_float(float(t))
    # 2025.5.16 格式
    m = re.match(r"(\d{4})[./年](\d{1,2})[./月](\d{1,2})", t)
    if m:
        try:
            return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=int(m.group(3)))
        except Exception:
            pass
    # 5/16 格式
    m = re.match(r"(\d{1,2})/(\d{1,2})$", t)
    if m:
        try:
            return pd.Timestamp(year=_CURRENT_YEAR, month=int(m.group(1)), day=int(m.group(2)))
        except Exception:
            pass
    # 默认用 pandas 尝试
    try:
        parsed = pd.to_datetime(t, errors="coerce")
        if pd.notna(parsed):
            return pd.Timestamp(parsed)
    except Exception:
        pass
    return None


# -----------------------------
# 期货分析工具
# -----------------------------
def lookup_delivery_premium(metadata_df: pd.DataFrame, province: str, city: str, contract_lh: str) -> float | None:
    premium_df = build_delivery_premium_df(metadata_df)
    if premium_df.empty:
        return None
    bucket = classify_lh_contract_bucket(contract_lh)
    if bucket is None:
        return None
    matched = premium_df[(premium_df["province"] == canonicalize_province(province)) & (premium_df["city"] == text_of(city))].copy()
    if matched.empty:
        return None
    value = matched.iloc[0].get(bucket)
    return None if pd.isna(value) else float(value)


def build_yongyi_city_basis_all_df(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, contract_lh: str) -> pd.DataFrame:
    columns = ["date", "value", "series_name", "metric", "province", "city", "contract", "close_price", "spot_price", "premium", "display_name", "sheet"]
    price_df = futures_series(futures_df, "收盘价", "收盘价")
    price_df = price_df[price_df["contract"] == contract_lh].copy()
    spot_df = yongyi_numeric_df[(yongyi_numeric_df["sheet"] == "交割地市出栏价") & (yongyi_numeric_df["metric"] == "交割地出栏价")].copy()
    if price_df.empty or spot_df.empty:
        return pd.DataFrame(columns=columns)
    merged = spot_df[["date", "value", "province", "city"]].rename(columns={"value": "spot_price"}).merge(
        price_df[["date", "value", "contract"]].rename(columns={"value": "close_price"}),
        on="date",
        how="inner",
    )
    if merged.empty:
        return pd.DataFrame(columns=columns)
    premium_df = build_delivery_premium_df(yongyi_meta_df)
    bucket = classify_lh_contract_bucket(contract_lh)
    if bucket is not None and not premium_df.empty and bucket in premium_df.columns:
        merged = merged.merge(
            premium_df[["province", "city", bucket]].rename(columns={bucket: "premium"}),
            on=["province", "city"],
            how="left",
        )
    else:
        merged["premium"] = np.nan
    merged["value"] = merged["close_price"] - merged["spot_price"] * 1000 + merged["premium"]
    merged["metric"] = "区域基差（涌益）"
    merged["series_name"] = merged["province"].fillna("") + "｜" + merged["city"].fillna("")
    merged["display_name"] = merged["series_name"]
    merged["sheet"] = "区域基差（涌益）"
    merged = merged.dropna(subset=["value"])
    return enrich_date_features(merged)


def build_yongyi_region_basis_df(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, contract_lh: str, province: str = "") -> pd.DataFrame:
    columns = ["date", "value", "series_name", "metric", "province", "city", "contract", "close_price", "spot_price", "premium", "display_name", "sheet", "city_count"]
    city_basis_df = build_yongyi_city_basis_all_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh)
    if city_basis_df.empty:
        return pd.DataFrame(columns=columns)
    province_name = canonicalize_province(province) if text_of(province) else ""
    if province_name:
        city_basis_df = city_basis_df[city_basis_df["province"] == province_name].copy()
    if city_basis_df.empty:
        return pd.DataFrame(columns=columns)
    merged = city_basis_df.groupby(["date", "province", "contract"], as_index=False).agg(
        close_price=("close_price", "mean"),
        spot_price=("spot_price", "mean"),
        premium=("premium", "mean"),
        value=("value", "mean"),
        city_count=("city", "nunique"),
    )
    merged["city"] = ""
    merged["metric"] = "区域基差（涌益）"
    merged["series_name"] = merged["province"]
    merged["display_name"] = merged["province"]
    merged["sheet"] = "区域基差（涌益）"
    return enrich_date_features(merged)


def build_yongyi_basis_df(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, contract_lh: str, province: str, city: str) -> pd.DataFrame:
    columns = ["date", "value", "series_name", "metric", "province", "city", "contract", "close_price", "spot_price", "premium", "display_name", "sheet"]
    basis_df = build_yongyi_city_basis_all_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh)
    if basis_df.empty:
        return pd.DataFrame(columns=columns)
    basis_df = basis_df[(basis_df["province"] == canonicalize_province(province)) & (basis_df["city"] == text_of(city))].copy()
    if basis_df.empty:
        return pd.DataFrame(columns=columns)
    basis_df["metric"] = "基差（涌益版）"
    basis_df["series_name"] = f"{contract_lh}｜{text_of(city)}"
    basis_df["display_name"] = basis_df["series_name"]
    basis_df["sheet"] = "基差（涌益版）"
    return basis_df


def futures_series(df: pd.DataFrame, sheet: str, metric: str | None = None) -> pd.DataFrame:
    temp = df[df["sheet"] == sheet].copy()
    if metric is not None:
        temp = temp[temp["metric"] == metric].copy()
    return temp


def build_yongyi_basis_snapshot(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, contract_lh: str, province: str, city: str, target_date: pd.Timestamp) -> dict[str, Any]:
    basis_df = build_yongyi_basis_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh, province, city)
    if basis_df.empty:
        city_basis_df = build_yongyi_city_basis_all_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh)
        if city_basis_df.empty:
            return {"ok": False, "message": "当前合约暂无可计算的涌益基差数据。"}
        province_name = canonicalize_province(province)
        province_df = city_basis_df[city_basis_df["province"] == province_name].copy()
        if province_df.empty:
            return {"ok": False, "message": f"{contract_lh} 在 {province_name} 暂无可计算的交割地基差数据。"}
        city_name = text_of(city)
        city_df = province_df[province_df["city"] == city_name].copy()
        if city_df.empty:
            return {"ok": False, "message": f"{contract_lh} 在 {province_name} {city_name} 暂无可计算的交割地基差数据。"}
        return {"ok": False, "message": "当前城市基差记录为空，请检查升贴水映射或日期覆盖范围。"}
    snap_date = get_snapshot_date(basis_df, target_date)
    if snap_date is None:
        return {"ok": False, "message": "当前日期之前没有可用的基差快照数据。"}
    row = basis_df[basis_df["date"] == snap_date].sort_values("date").iloc[-1]
    close_price = row.get("close_price", np.nan)
    spot_price = row.get("spot_price", np.nan)
    premium = row.get("premium", np.nan)
    return {
        "ok": True,
        "date": snap_date,
        "basis_df": basis_df,
        "basis_value": float(row["value"]),
        "close_price": None if pd.isna(close_price) else float(close_price),
        "spot_price": None if pd.isna(spot_price) else float(spot_price),
        "premium": None if pd.isna(premium) else float(premium),
        "contract": contract_lh,
        "province": canonicalize_province(province),
        "city": text_of(city),
    }


def build_yongyi_region_basis_snapshot(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, contract_lh: str, province: str, target_date: pd.Timestamp) -> dict[str, Any]:
    basis_df = build_yongyi_region_basis_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh, province)
    if basis_df.empty:
        return {"ok": False, "message": "所选合约与区域没有重叠日期数据。"}
    snap_date = get_snapshot_date(basis_df, target_date)
    if snap_date is None:
        return {"ok": False, "message": "当前日期没有可用的区域基差数据。"}
    row = basis_df[basis_df["date"] == snap_date].sort_values("date").iloc[-1]
    close_price = row.get("close_price", np.nan)
    spot_price = row.get("spot_price", np.nan)
    premium = row.get("premium", np.nan)
    city_count = row.get("city_count", np.nan)
    return {
        "ok": True,
        "date": snap_date,
        "basis_df": basis_df,
        "basis_value": float(row["value"]),
        "close_price": None if pd.isna(close_price) else float(close_price),
        "spot_price": None if pd.isna(spot_price) else float(spot_price),
        "premium": None if pd.isna(premium) else float(premium),
        "city_count": None if pd.isna(city_count) else int(city_count),
        "contract": contract_lh,
        "province": canonicalize_province(province),
    }


def build_spread_df(price_df: pd.DataFrame, contract_a: str, contract_b: str) -> pd.DataFrame:
    left = price_df[price_df["contract"] == contract_a][["date", "value"]].rename(columns={"value": "a_value"})
    right = price_df[price_df["contract"] == contract_b][["date", "value"]].rename(columns={"value": "b_value"})
    merged = left.merge(right, on="date", how="inner")
    if merged.empty:
        return merged
    merged["value"] = merged["a_value"] - merged["b_value"]
    merged["series_name"] = f"{contract_a}-{contract_b}"
    merged["province"] = "全部"
    merged["city"] = ""
    merged["display_name"] = merged["series_name"]
    merged["sheet"] = "价差"
    merged["metric"] = "合约价差"
    return enrich_date_features(merged)


def get_latest_value(series_df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[pd.Timestamp | None, float | None, float | None]:
    snap_date = get_snapshot_date(series_df, target_date)
    if snap_date is None:
        return None, None, None
    current = series_df.loc[series_df["date"] == snap_date, "value"].mean()
    prev_dates = series_df[series_df["date"] < snap_date]["date"]
    prev_date = prev_dates.max() if not prev_dates.empty else None
    prev = series_df.loc[series_df["date"] == prev_date, "value"].mean() if prev_date is not None else np.nan
    return snap_date, current, prev


def build_yongyi_city_basis_snapshot(
    futures_df: pd.DataFrame,
    yongyi_numeric_df: pd.DataFrame | None,
    yongyi_meta_df: pd.DataFrame | None,
    target_date: pd.Timestamp,
    focus_contract: str | None = None,
) -> dict[str, Any] | None:
    if yongyi_numeric_df is None or yongyi_meta_df is None or yongyi_numeric_df.empty:
        return None
    price_df = futures_series(futures_df, "收盘价", "收盘价")
    contracts = get_active_lh_contracts(price_df, target_date)
    contract_lh = focus_contract if focus_contract in contracts else (contracts[0] if contracts else None)
    if not contract_lh:
        return None
    city_basis_df = build_yongyi_city_basis_all_df(futures_df, yongyi_numeric_df, yongyi_meta_df, contract_lh)
    if city_basis_df.empty:
        return None
    snap_date = get_snapshot_date(city_basis_df, target_date)
    if snap_date is None:
        return None
    candidates: list[dict[str, Any]] = []
    for name in sorted(city_basis_df["series_name"].dropna().unique()):
        series_df = city_basis_df[city_basis_df["series_name"] == name].copy()
        current = _latest_series_value(series_df, snap_date)
        if current is None:
            continue
        _, percentile = percentile_position_text(current, series_df[series_df["date"] <= snap_date]["value"])
        if percentile is None:
            continue
        latest_row = series_df[series_df["date"] == snap_date].sort_values("date").iloc[-1]
        candidates.append(
            {
                "name": name,
                "value": current,
                "percentile": percentile,
                "province": latest_row.get("province", ""),
                "city": latest_row.get("city", ""),
            }
        )
    if not candidates:
        return None
    high_candidates = [item for item in candidates if item["percentile"] >= 0.8]
    low_candidates = [item for item in candidates if item["percentile"] <= 0.2]
    high = max(high_candidates, key=lambda x: x["percentile"]) if high_candidates else None
    low = min(low_candidates, key=lambda x: x["percentile"]) if low_candidates else None
    extreme = max(candidates, key=lambda x: abs(x["percentile"] - 0.5))
    if high is not None and low is not None:
        message = f"{contract_lh} 市级基差分化明显，高位城市 {high['name']} {format_number(high['value'])}（{high['percentile']:.0%}），低位城市 {low['name']} {format_number(low['value'])}（{low['percentile']:.0%}）。"
    elif extreme["percentile"] >= 0.8 or extreme["percentile"] <= 0.2:
        direction = "偏高" if extreme["percentile"] >= 0.8 else "偏低"
        message = f"{contract_lh} 市级基差方面，{extreme['name']} 当前为 {format_number(extreme['value'])}，历史分位 {extreme['percentile']:.0%}，{direction}。"
    else:
        return None
    return {
        "contract": contract_lh,
        "date": snap_date,
        "high": high,
        "low": low,
        "extreme": extreme,
        "message": message,
    }


def build_yongyi_city_basis_signal(
    futures_df: pd.DataFrame,
    yongyi_numeric_df: pd.DataFrame | None,
    yongyi_meta_df: pd.DataFrame | None,
    target_date: pd.Timestamp,
    focus_contract: str | None = None,
) -> str | None:
    snapshot = build_yongyi_city_basis_snapshot(futures_df, yongyi_numeric_df, yongyi_meta_df, target_date, focus_contract)
    return snapshot.get("message") if snapshot else None


def build_futures_opportunity_messages(
    futures_df: pd.DataFrame,
    target_date: pd.Timestamp,
    yongyi_numeric_df: pd.DataFrame | None = None,
    yongyi_meta_df: pd.DataFrame | None = None,
    focus_contract: str | None = None,
    include_yongyi: bool = False,
) -> list[str]:
    messages: list[str] = []
    price_df = futures_series(futures_df, "收盘价", "收盘价")
    active_contracts = get_active_lh_contracts(price_df, target_date)
    if not price_df.empty:
        contracts = active_contracts
        spread_candidates: list[tuple[str, float, float]] = []
        for idx in range(len(contracts) - 1):
            spread_df = build_spread_df(price_df, contracts[idx], contracts[idx + 1])
            if spread_df.empty:
                continue
            snap_date, current, _ = get_latest_value(spread_df, target_date)
            if snap_date is None or current is None:
                continue
            position_text, percentile = percentile_position_text(current, spread_df[spread_df["date"] <= snap_date]["value"])
            if percentile is not None and (percentile >= 0.8 or percentile <= 0.2):
                spread_candidates.append((spread_df["series_name"].iloc[0], current, percentile))
        if spread_candidates:
            name, current, percentile = sorted(spread_candidates, key=lambda x: abs(x[2] - 0.5), reverse=True)[0]
            direction = "存在回归观察价值" if percentile >= 0.8 else "存在走阔观察价值"
            messages.append(f"价差方面，{name} 当前为 {format_number(current)}，处于历史分位 {percentile:.0%}，{direction}。")

    basis_df = futures_series(futures_df, "期现基差", "基差")
    if not basis_df.empty:
        candidates = []
        for name in sorted(basis_df["series_name"].dropna().unique()):
            series_df = basis_df[basis_df["series_name"] == name]
            snap_date, current, _ = get_latest_value(series_df, target_date)
            if snap_date is None or current is None:
                continue
            position_text, percentile = percentile_position_text(current, series_df[series_df["date"] <= snap_date]["value"])
            if percentile is not None:
                candidates.append((name, current, percentile))
        if candidates:
            name, current, percentile = sorted(candidates, key=lambda x: abs(x[2] - 0.5), reverse=True)[0]
            messages.append(f"基差方面，{name} 当前为 {format_number(current)}，历史分位 {percentile:.0%}，{'偏高' if percentile >= 0.8 else '偏低' if percentile <= 0.2 else '中性'}。")

    if include_yongyi:
        yongyi_message = build_yongyi_city_basis_signal(futures_df, yongyi_numeric_df, yongyi_meta_df, target_date, focus_contract)
        if yongyi_message:
            messages.append(yongyi_message)

    net_df = futures_series(futures_df, "前20多空净持仓", "前20净持仓")
    if not net_df.empty and not price_df.empty:
        for contract in active_contracts:
            net_series = net_df[net_df["contract"] == contract].sort_values("date")
            price_series = price_df[price_df["contract"] == contract].sort_values("date")
            merged = net_series[["date", "value"]].rename(columns={"value": "net"}).merge(price_series[["date", "value"]].rename(columns={"value": "price"}), on="date", how="inner")
            merged = merged[merged["date"] <= pd.Timestamp(target_date)].tail(5)
            if len(merged) >= 3:
                net_delta = merged["net"].iloc[-1] - merged["net"].iloc[0]
                price_delta = merged["price"].iloc[-1] - merged["price"].iloc[0]
                if net_delta > 0 and price_delta < 0:
                    messages.append(f"{contract} 近几日净持仓增加而价格回落，存在资金与价格背离。")
                    break
                if net_delta < 0 and price_delta > 0:
                    messages.append(f"{contract} 近几日净持仓回落而价格上行，盘面追涨需留意。")
                    break

    warehouse_df = futures_series(futures_df, "仓单、虚实盘、交割量", "注册仓单")
    if not warehouse_df.empty:
        snap_date, current, _ = get_latest_value(warehouse_df, target_date)
        if snap_date is not None and current is not None:
            hist = warehouse_df[warehouse_df["date"] <= snap_date].sort_values("date")
            ma20 = hist.tail(20)["value"].mean()
            if pd.notna(ma20) and ma20 != 0:
                diff_ratio = current / ma20 - 1
                if abs(diff_ratio) >= 0.2:
                    messages.append(f"仓单当前为 {format_number(current)}，较近20日均值{'偏高' if diff_ratio > 0 else '偏低'} {abs(diff_ratio):.0%}。")

    if not messages:
        messages.append("当前日期下，期货端主要指标大多处于中性区间，可继续跟踪价差、基差与净持仓是否出现新的共振。")
    return messages[:4]


# -----------------------------
# 调运分析工具
# -----------------------------
def transport_daily_count(df: pd.DataFrame, field: str, target: str) -> pd.DataFrame:
    temp = df[df[field] == target].groupby("date", as_index=False)["count"].sum().sort_values("date")
    temp["series_name"] = target
    temp["province"] = "全部"
    temp["city"] = ""
    temp["display_name"] = target
    return temp


def build_transport_anomaly_messages(df: pd.DataFrame, target_date: pd.Timestamp) -> list[str]:
    messages: list[str] = []
    series = df.groupby("date", as_index=False)["count"].sum().sort_values("date")
    snap_date = get_snapshot_date(series, target_date)
    if snap_date is None:
        return messages
    current = series.loc[series["date"] == snap_date, "count"].sum()
    prev7 = series[series["date"] < snap_date].tail(7)
    prev10 = series[series["date"] < snap_date].tail(10)
    hist_mean = series[series["date"] < snap_date]["count"].mean()
    if not prev7.empty:
        avg7 = prev7["count"].mean()
        if avg7 and current >= avg7 * 1.4:
            messages.append(f"{format_date_cn(snap_date)} 调运总量较近7日均值放大 {current / avg7 - 1:.0%}。")
        if avg7 and current <= avg7 * 0.7:
            messages.append(f"{format_date_cn(snap_date)} 调运总量较近7日均值回落 {1 - current / avg7:.0%}。")
    if not prev10.empty:
        avg10 = prev10["count"].mean()
        if avg10 and current < avg10 * 0.8:
            messages.append(f"当前总量低于近10日均值 {1 - current / avg10:.0%}。")
    if pd.notna(hist_mean) and hist_mean and current < hist_mean * 0.8:
        messages.append(f"当前总量低于样本均值 {1 - current / hist_mean:.0%}。")
    route_rank = df[df["date"] == snap_date].groupby("城市路线", as_index=False)["count"].sum().sort_values("count", ascending=False)
    if not route_rank.empty:
        top_route = route_rank.iloc[0]
        messages.append(f"当日最活跃路线为 {top_route['城市路线']}，记录数 {int(top_route['count'])}。")
    return messages[:4]


def build_transport_rank(df: pd.DataFrame, group_col: str, target_date: pd.Timestamp, top_n: int = 10) -> tuple[pd.DataFrame, pd.Timestamp | None]:
    snap_date = get_snapshot_date(df, target_date)
    if snap_date is None:
        return pd.DataFrame(), None
    temp = df[df["date"] == snap_date].groupby(group_col, as_index=False)["count"].sum().sort_values("count", ascending=False).head(top_n)
    return temp, snap_date


def build_transport_trend_chart(df: pd.DataFrame, title: str) -> go.Figure:
    temp = df.copy()
    if "date" in temp.columns:
        temp = temp.sort_values("date")
    if "日期标签" not in temp.columns and "date" in temp.columns:
        temp["日期标签"] = temp["date"].map(format_date_cn)
    fig = px.line(temp, x="日期标签", y="count", title=title, markers=False, category_orders={"日期标签": temp["日期标签"].drop_duplicates().tolist()} if "日期标签" in temp.columns else None)
    fig.update_traces(hovertemplate="日期：%{x}<br>调运量：%{y}<extra></extra>")
    fig.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量")
    return fig


def render_transport_focus_section(df: pd.DataFrame, title: str, targets: list[dict[str, str]], target_date: pd.Timestamp, both_directions: bool = False) -> None:
    """省份排行下钻模式：左侧堆积趋势图，右侧当日省份排行柱状图，点击省份下钻城市。"""
    render_section_header(f"📊 {title}")
    snap_date = get_snapshot_date(df, target_date)
    if snap_date is None:
        st.info(f"{title}：暂无数据。")
        return

    for target in targets:
        label = target["label"]
        field = target["field"]
        in_field = "调入省份" if field == "province" else "调入城市"
        out_field = "调出省份" if field == "province" else "调出城市"

        st.markdown(f"**{label}**")
        col_left, col_right = st.columns([1, 1])

        # 左侧：调入趋势堆积图（半宽）
        with col_left:
            stack_raw = df[df[in_field] == label].groupby(["date", "调出省份"], as_index=False)["count"].sum().sort_values(["date", "count"], ascending=[True, False])
            if not stack_raw.empty:
                top_src = stack_raw.groupby("调出省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(8)["调出省份"].tolist()
                stack_raw["来源分组"] = stack_raw["调出省份"].apply(lambda x: x if x in top_src else "其他")
                stacked = stack_raw.groupby(["date", "来源分组"], as_index=False)["count"].sum()
                stacked = sort_date_label_frame(stacked)
                stacked["日期标签"] = stacked["date"].map(format_date_cn)
                fig_stack = px.bar(stacked, x="日期标签", y="count", color="来源分组",
                                   title=f"调入{label} — 分省份堆积趋势", barmode="stack", text=None, custom_data=["date"], category_orders={"日期标签": stacked["日期标签"].drop_duplicates().tolist()})
                fig_stack.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量", hovermode="x unified")
                sel_stack = st.plotly_chart(fig_stack, key=f"focus_stack_{title}_{label}", on_select="rerun", width='stretch')
                if sel_stack and sel_stack.get("selection") and sel_stack["selection"].get("points"):
                    pt_stack = sel_stack["selection"]["points"][0]
                    stack_date = pt_stack.get("customdata", [None])[0]
                    if stack_date:
                        st.session_state[f"focus_stack_date_{title}_{label}"] = pd.Timestamp(stack_date)
            else:
                st.info(f"暂无调入{label}的历史数据。")

        # 右侧：来源省份排行 + 旁边直接显示城市明细
        with col_right:
            current_focus_date = st.session_state.get(f"focus_stack_date_{title}_{label}", snap_date)
            day_df = df[(df["date"] == pd.Timestamp(current_focus_date)) & (df[in_field] == label)]
            prov_rank = day_df.groupby("调出省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
            if not prov_rank.empty:
                sub_col1, sub_col2 = st.columns([1, 1])
                with sub_col1:
                    fig_prov = px.bar(prov_rank.sort_values("count"), x="count", y="调出省份",
                                      orientation="h", title=f"{format_date_cn(pd.Timestamp(current_focus_date))} 调入{label} 来源省份",
                                      text="count", color="count", color_continuous_scale="Blues")
                    fig_prov.update_layout(template="plotly_white", coloraxis_showscale=False)
                    sel_prov = st.plotly_chart(fig_prov, key=f"focus_prov_{title}_{label}", on_select="rerun", width='stretch')
                    clicked_prov = None
                    if sel_prov and sel_prov.get("selection") and sel_prov["selection"].get("points"):
                        pt = sel_prov["selection"]["points"][0]
                        clicked_prov = pt.get("label") or pt.get("y") or pt.get("x")
                    if clicked_prov:
                        st.session_state[f"focus_drill_prov_{title}_{label}"] = clicked_prov
                with sub_col2:
                    drill_prov = st.session_state.get(f"focus_drill_prov_{title}_{label}")
                    if drill_prov:
                        city_df = day_df[day_df["调出省份"] == drill_prov].groupby("调出城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                        if not city_df.empty:
                            fig_city = px.bar(city_df, x="调出城市", y="count",
                                              title=f"{drill_prov} → {label} 城市分布", text="count",
                                              color="count", color_continuous_scale="Oranges")
                            fig_city.update_traces(textposition="outside")
                            fig_city.update_layout(template="plotly_white", coloraxis_showscale=False)
                            render_plotly(fig_city, "transport_focus", title, label, "city", drill_prov)
                        st.caption(f"当前来源省份：{drill_prov}（点击左侧省份图切换）")
                    else:
                        st.info("点击左侧来源省份排行后，城市分布会直接显示在右侧。")
            else:
                st.info(f"{format_date_cn(snap_date)} 暂无调入{label}的记录。")

        # 如果两个方向都要展示（云南神农），再展示调出方向
        if both_directions:
            st.markdown(f"**{label}（调出方向）**")
            col_l2, col_r2 = st.columns([1, 1])
            with col_l2:
                out_stack = df[df[out_field] == label].groupby(["date", "调入省份"], as_index=False)["count"].sum().sort_values(["date", "count"], ascending=[True, False])
                if not out_stack.empty:
                    top_dest = out_stack.groupby("调入省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(8)["调入省份"].tolist()
                    out_stack["去向分组"] = out_stack["调入省份"].apply(lambda x: x if x in top_dest else "其他")
                    stacked_out = out_stack.groupby(["date", "去向分组"], as_index=False)["count"].sum()
                    stacked_out = sort_date_label_frame(stacked_out)
                    stacked_out["日期标签"] = stacked_out["date"].map(format_date_cn)
                    fig_out_stack = px.bar(stacked_out, x="日期标签", y="count", color="去向分组", title=f"调出{label} — 分省份堆积趋势", barmode="stack", text=None, custom_data=["date"], category_orders={"日期标签": stacked_out["日期标签"].drop_duplicates().tolist()})
                    fig_out_stack.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量", hovermode="x unified")
                    sel_out_stack = st.plotly_chart(fig_out_stack, key=f"focus_out_stack_{title}_{label}", on_select="rerun", width='stretch')
                    if sel_out_stack and sel_out_stack.get("selection") and sel_out_stack["selection"].get("points"):
                        pt_out = sel_out_stack["selection"]["points"][0]
                        out_date = pt_out.get("customdata", [None])[0]
                        if out_date:
                            st.session_state[f"focus_out_stack_date_{title}_{label}"] = pd.Timestamp(out_date)
                else:
                    st.info(f"暂无调出{label}的历史数据。")
            with col_r2:
                current_out_focus_date = st.session_state.get(f"focus_out_stack_date_{title}_{label}", snap_date)
                day_out_df = df[(df["date"] == pd.Timestamp(current_out_focus_date)) & (df[out_field] == label)]
                dest_rank = day_out_df.groupby("调入省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
                if not dest_rank.empty:
                    sub_col1, sub_col2 = st.columns([1, 1])
                    with sub_col1:
                        fig_dest = px.bar(dest_rank.sort_values("count"), x="count", y="调入省份", orientation="h", title=f"{format_date_cn(pd.Timestamp(current_out_focus_date))} 调出{label} 去向省份", text="count", color="count", color_continuous_scale="Greens")
                        fig_dest.update_layout(template="plotly_white", coloraxis_showscale=False)
                        sel_dest = st.plotly_chart(fig_dest, key=f"focus_dest_{title}_{label}", on_select="rerun", width='stretch')
                        clicked_dest = None
                        if sel_dest and sel_dest.get("selection") and sel_dest["selection"].get("points"):
                            pt2 = sel_dest["selection"]["points"][0]
                            clicked_dest = pt2.get("label") or pt2.get("y") or pt2.get("x")
                        if clicked_dest:
                            st.session_state[f"focus_drill_dest_{title}_{label}"] = clicked_dest
                    with sub_col2:
                        drill_dest = st.session_state.get(f"focus_drill_dest_{title}_{label}")
                        if drill_dest:
                            city_dest_df = day_out_df[day_out_df["调入省份"] == drill_dest].groupby("调入城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                            if not city_dest_df.empty:
                                fig_city_dest = px.bar(city_dest_df, x="调入城市", y="count", title=f"{label} → {drill_dest} 城市分布", text="count", color="count", color_continuous_scale="Purples")
                                fig_city_dest.update_traces(textposition="outside")
                                fig_city_dest.update_layout(template="plotly_white", coloraxis_showscale=False)
                                render_plotly(fig_city_dest, "transport_focus", title, label, "out_city", drill_dest)
                            st.caption(f"当前去向省份：{drill_dest}（点击左侧省份图切换）")
                        else:
                            st.info("点击左侧调出去向省份排行后，城市分布会直接显示在右侧。")
                else:
                    st.info(f"{format_date_cn(snap_date)} 暂无调出{label}的记录。")


# -----------------------------
# 鲜冻品分析工具
# -----------------------------

# 产品业务属性映射：判断某产品属于哪个信号维度
_BONE_KEYWORDS = ["骨", "大骨", "棒骨", "排骨", "肋排", "扇骨", "脊骨", "龙骨", "筒骨"]
_BYPRODUCT_KEYWORDS = ["副产", "猪头", "猪脚", "猪蹄", "猪耳", "猪肝", "猪肺", "猪心", "猪肚", "猪肠", "猪血", "下水", "猪尾"]
_FAT_KEYWORDS = ["肥肉", "猪油", "板油", "肥膘"]
_LEAN_KEYWORDS = ["瘦肉", "前腿", "后腿", "里脊", "通脊", "梅肉", "腱子", "精肉", "猪柳"]
_FROZEN_PORK_KEYWORDS = ["冻猪", "进口", "冻品", "冷冻"]


def _product_signal_type(label: str) -> str:
    """返回产品的业务信号类型"""
    for kw in _BONE_KEYWORDS:
        if kw in label:
            return "骨类"
    for kw in _BYPRODUCT_KEYWORDS:
        if kw in label:
            return "副产品"
    for kw in _FAT_KEYWORDS:
        if kw in label:
            return "肥肉"
    for kw in _LEAN_KEYWORDS:
        if kw in label:
            return "瘦肉"
    return "其他"


def _trend_text(diffs: pd.Series) -> str:
    """将价格变动序列转为简洁趋势描述"""
    if diffs.empty:
        return ""
    pos = (diffs > 0).sum()
    neg = (diffs < 0).sum()
    total = len(diffs)
    if pos == total:
        return f"连续{total}期上涨"
    if neg == total:
        return f"连续{total}期下跌"
    if pos >= total * 0.7:
        return "整体偏强"
    if neg >= total * 0.7:
        return "整体偏弱"
    return "震荡"


def build_price_alert_messages(df: pd.DataFrame, target_date: pd.Timestamp, label: str) -> list[str]:
    messages: list[str] = []
    series = df.groupby("date", as_index=False)["value"].mean().sort_values("date")
    snap_date = get_snapshot_date(series, target_date)
    if snap_date is None:
        return messages
    current = series.loc[series["date"] == snap_date, "value"].mean()
    prev = series[series["date"] < snap_date].tail(1)["value"].mean()
    if pd.notna(prev):
        delta = current - prev
        messages.append(f"{label} 当前值 {format_number(current)}，较前一日 {describe_delta(delta)}。")
    tail = series[series["date"] <= snap_date].tail(5)
    if len(tail) >= 4:
        diffs = tail["value"].diff().dropna()
        trend = _trend_text(diffs)
        if trend:
            messages.append(f"{label} 近期走势：{trend}。")
    ma10 = series[series["date"] <= snap_date].tail(10)["value"].mean()
    if pd.notna(ma10) and ma10 != 0:
        diff_ratio = current / ma10 - 1
        if abs(diff_ratio) >= 0.08:
            messages.append(f"{label} 较10日均值{'偏高' if diff_ratio > 0 else '偏低'} {abs(diff_ratio):.0%}。")
    yoy_match = same_day_last_year_match(series["date"], snap_date)
    if yoy_match is not None:
        yoy_value = series.loc[series["date"] == yoy_match, "value"].mean()
        if pd.notna(yoy_value):
            messages.append(f"{label} 较去年同期 {describe_delta(current - yoy_value)}。")
    return messages[:4]


def build_fresh_frozen_market_signals(price_df: pd.DataFrame, target_date: pd.Timestamp) -> list[dict]:
    """
    综合扫描全部鲜冻品，生成面向生猪行情的业务预警信号。
    返回 list[{level, title, body}]，level: "danger" / "warning" / "info"
    """
    signals: list[dict] = []
    if price_df.empty:
        return signals

    snap_date = get_snapshot_date(price_df, pd.Timestamp(target_date))
    if snap_date is None:
        return signals

    # 按产品聚合当日和近5日均值
    day_df = price_df[price_df["date"] == snap_date].groupby("product", as_index=False)["value"].mean()
    recent = price_df[price_df["date"] <= snap_date].copy()

    def _get_trend(product: str) -> tuple[float | None, str]:
        """返回 (当日价, 趋势描述)"""
        sub = recent[recent["product"] == product].groupby("date", as_index=False)["value"].mean().sort_values("date")
        if sub.empty:
            return None, ""
        cur = sub["value"].iloc[-1]
        diffs = sub["value"].tail(5).diff().dropna()
        return cur, _trend_text(diffs)

    # 骨类信号
    bone_products = [p for p in day_df["product"].unique() if _product_signal_type(p) == "骨类"]
    bone_rising, bone_falling = [], []
    for p in bone_products:
        cur, trend = _get_trend(p)
        if cur is None:
            continue
        if "上涨" in trend or "偏强" in trend:
            bone_rising.append(f"{p}（{format_number(cur)}）")
        elif "下跌" in trend or "偏弱" in trend:
            bone_falling.append(f"{p}（{format_number(cur)}）")

    if bone_rising:
        signals.append({
            "level": "warning",
            "title": "🦴 骨类价格上涨 → 利好屠宰",
            "body": (
                f"当前骨类走强：{'、'.join(bone_rising[:4])}。"
                "骨价上涨通常带动屠宰场提高收购积极性，有望推升日屠宰量。"
                "建议关注屠宰量同步变化，若量价共振则信号更强。"
            ),
        })
    if bone_falling:
        signals.append({
            "level": "info",
            "title": "🦴 骨类价格走弱 → 屠宰积极性降低",
            "body": (
                f"骨类走弱：{'、'.join(bone_falling[:4])}。"
                "骨价下跌压缩屠宰利润，屠企收购意愿趋降，屠宰量可能阶段性承压。"
            ),
        })

    # 副产品信号
    by_products = [p for p in day_df["product"].unique() if _product_signal_type(p) == "副产品"]
    by_rising, by_falling = [], []
    for p in by_products:
        cur, trend = _get_trend(p)
        if cur is None:
            continue
        if "上涨" in trend or "偏强" in trend:
            by_rising.append(f"{p}（{format_number(cur)}）")
        elif "下跌" in trend or "偏弱" in trend:
            by_falling.append(f"{p}（{format_number(cur)}）")

    if by_rising:
        signals.append({
            "level": "warning",
            "title": "🫀 副产品走强 → 综合利润改善",
            "body": (
                f"副产品上涨：{'、'.join(by_rising[:4])}。"
                "副产价格走强提升屠宰综合利润，屠企开工意愿增强，屠宰量有望放量。"
                "历史规律：副产与骨类同步走强时，屠宰量往往出现阶段高点。"
            ),
        })
    if by_falling:
        signals.append({
            "level": "info",
            "title": "🫀 副产品走弱 → 关注利润侵蚀",
            "body": (
                f"副产走弱：{'、'.join(by_falling[:4])}。"
                "副产下跌压缩屠宰毛利，若骨类同步弱势则屠企减产动力上升，需警惕屠宰量收缩拖累猪价。"
            ),
        })

    # 骨+副产双强 / 双弱
    if bone_rising and by_rising:
        signals.insert(0, {
            "level": "danger",
            "title": "🔥 骨类+副产双强 → 强烈利好屠宰放量",
            "body": (
                "骨类与副产品同步走强，屠宰综合利润显著扩张。"
                "历史上此类共振通常伴随屠宰量阶段性放量，需同步跟踪猪价是否因供给增加承压。"
            ),
        })
    if bone_falling and by_falling:
        signals.insert(0, {
            "level": "danger",
            "title": "❄️ 骨类+副产双弱 → 屠宰利润收缩风险",
            "body": (
                "骨类与副产品同步走弱，屠宰综合利润承压。"
                "此类组合历史上常伴随屠宰量下滑，若猪价同步高位则屠企观望情绪加重，供给节奏可能放缓。"
            ),
        })

    # 瘦肉类信号（反映终端需求）
    lean_products = [p for p in day_df["product"].unique() if _product_signal_type(p) == "瘦肉"]
    lean_rising = [p for p in lean_products if "上涨" in _get_trend(p)[1] or "偏强" in _get_trend(p)[1]]
    lean_falling = [p for p in lean_products if "下跌" in _get_trend(p)[1] or "偏弱" in _get_trend(p)[1]]
    if lean_rising:
        signals.append({
            "level": "info",
            "title": "🥩 瘦肉走强 → 终端需求偏好",
            "body": (
                f"{'、'.join(lean_rising[:3])} 等瘦肉类走强，"
                "终端消费需求相对积极，对猪价有一定支撑。"
            ),
        })
    if lean_falling:
        signals.append({
            "level": "info",
            "title": "🥩 瘦肉走弱 → 终端需求偏淡",
            "body": (
                f"{'、'.join(lean_falling[:3])} 等瘦肉类走弱，"
                "终端消费偏淡，猪价上行阻力加大，关注屠宰量是否随之收缩。"
            ),
        })

    # 鲜冻价差异常信号（冻品溢价/折价）
    fresh_df = price_df[price_df["dataset_type"] == "鲜品"]
    frozen_df = price_df[price_df["dataset_type"] == "冻品"]
    common_products = set(fresh_df["product"].unique()) & set(frozen_df["product"].unique())
    high_spread_items, low_spread_items = [], []
    for prod in list(common_products)[:6]:
        f_cur = fresh_df[fresh_df["product"] == prod].groupby("date")["value"].mean()
        z_cur = frozen_df[frozen_df["product"] == prod].groupby("date")["value"].mean()
        merged = pd.concat([f_cur.rename("fresh"), z_cur.rename("frozen")], axis=1).dropna()
        if merged.empty or len(merged) < 10:
            continue
        merged["spread"] = merged["fresh"] - merged["frozen"]
        cur_spread = merged["spread"].iloc[-1]
        pct = (merged["spread"] <= cur_spread).mean()
        if pct >= 0.85:
            high_spread_items.append(f"{prod}（鲜冻差 {format_number(cur_spread)}）")
        elif pct <= 0.15:
            low_spread_items.append(f"{prod}（鲜冻差 {format_number(cur_spread)}）")
    if high_spread_items:
        signals.append({
            "level": "warning",
            "title": "📦 鲜冻价差偏高 → 冻品套利空间收窄",
            "body": (
                f"{'、'.join(high_spread_items[:3])} 鲜冻差处于历史高位，"
                "鲜品相对冻品溢价明显。市场或存在鲜转冻动力，后续关注冻品库存变化。"
            ),
        })
    if low_spread_items:
        signals.append({
            "level": "info",
            "title": "📦 鲜冻价差偏低 → 冻品具备出库价值",
            "body": (
                f"{'、'.join(low_spread_items[:3])} 鲜冻差处于历史低位，"
                "冻品价格相对鲜品偏高，冻品出库意愿增强，关注供给端压力。"
            ),
        })

    return signals[:8]  # 最多展示8条


def build_spread_alert_messages(spread_df: pd.DataFrame, target_date: pd.Timestamp, label: str) -> list[str]:
    messages: list[str] = []
    if spread_df.empty:
        return messages
    snap_date = get_snapshot_date(spread_df, target_date)
    if snap_date is None:
        return messages
    current = spread_df.loc[spread_df["date"] == snap_date, "value"].mean()
    history = spread_df[spread_df["date"] <= snap_date]["value"]
    position_text, percentile = percentile_position_text(current, history)
    messages.append(f"{label} 当前价差 {format_number(current)}，所处位置：{position_text}。")
    if percentile is not None:
        if percentile >= 0.8:
            messages.append(f"{label} 价差处于高分位，后续留意收敛。")
        elif percentile <= 0.2:
            messages.append(f"{label} 价差处于低分位，后续留意走阔。")
    tail = spread_df[spread_df["date"] <= snap_date].tail(5)
    if len(tail) >= 4:
        diffs = tail["value"].diff().dropna()
        trend = _trend_text(diffs)
        if trend:
            messages.append(f"{label} 近期走势：{trend}。")
    return messages[:4]


# -----------------------------
# 周度涌益数据解析
# -----------------------------
def _is_period_date_header(row0: Any, row1: Any) -> bool:
    """判断前两列是否为"开始日期 / 结束日期"模式。"""
    h0 = text_of(row0).replace(" ", "").replace("\n", "")
    h1 = text_of(row1).replace(" ", "").replace("\n", "")
    return any(k in h0 for k in ["开始", "日期", "起始"]) and any(k in h1 for k in ["结束", "日期"])


def _find_data_start_row(rows: list[list[Any]], max_scan: int = 6) -> int | None:
    """找到第一行实际数据行（第一列是 datetime）。"""
    for i, row in enumerate(rows[:max_scan + 10]):
        if not row:
            continue
        ts = parse_date(row[0] if len(row) > 0 else None)
        if ts is not None:
            return i
    return None


def _find_all_header_rows(rows: list[list[Any]], data_start: int) -> list[int]:
    """返回数据行之前所有非空行的索引（可能有多行 header）。"""
    header_idxs = []
    for i in range(data_start):
        row = rows[i]
        if row and any(text_of(c) for c in row):
            header_idxs.append(i)
    return header_idxs


def _merge_header_rows(rows: list[list[Any]], header_idxs: list[int], n_cols: int) -> list[str]:
    """将多行 header 合并为一行列名（用 '·' 连接非空非重复部分）。"""
    merged = [""] * n_cols
    for idx in header_idxs:
        row = rows[idx]
        for col in range(n_cols):
            cell = text_of(row[col] if col < len(row) else None)
            if cell and cell not in merged[col]:
                merged[col] = (merged[col] + "·" + cell).strip("·") if merged[col] else cell
    return merged


def parse_weekly_sheet_universal(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    """
    通用周/月混合 sheet 解析器：
    - 周区间表默认以结束日期作为主 date，并保留 start_date/end_date
    - 支持文本周期拆解
    - 跳过环比/同比/备注等辅助列
    - 尽量补齐标准字段
    """
    records: list[dict] = []
    if not rows or len(rows) < 2:
        return records

    freq_type = infer_weekly_freq_type(sheet_name)
    topic_hint = "价格端" if any(flag in sheet_name for flag in ["出栏价", "母猪", "仔猪", "结算价"]) else ("价差端" if "价差" in sheet_name else "周度专题")
    data_start = _find_data_start_row(rows, max_scan=8)
    if data_start is None:
        for idx, row in enumerate(rows[:12]):
            if not row:
                continue
            start_ts, end_ts = parse_weekly_text_period_cell(row[0] if len(row) > 0 else None)
            if end_ts is not None:
                data_start = idx
                break
    if data_start is None or data_start == 0:
        return records

    header_idxs = _find_all_header_rows(rows, data_start)
    if not header_idxs:
        return records

    header_row = rows[data_start - 1]
    n_cols = max(len(header_row), max((len(r) for r in rows[data_start:data_start + 3]), default=0))
    merged_header = standardize_weekly_headers(_merge_header_rows(rows, header_idxs, n_cols))

    h0 = text_of(merged_header[0] if len(merged_header) > 0 else None)
    h1 = text_of(merged_header[1] if len(merged_header) > 1 else None)
    h2 = text_of(merged_header[2] if len(merged_header) > 2 else None)
    first_data_row = rows[data_start] if data_start < len(rows) else []

    has_two_date_cells = (
        len(first_data_row) >= 2
        and (
            parse_date(first_data_row[0]) is not None
            or parse_weekly_text_period_cell(first_data_row[0])[1] is not None
        )
        and parse_date(first_data_row[1]) is not None
    )
    period_from_text = parse_weekly_text_period_cell(first_data_row[0])[1] is not None
    is_period = (_is_period_date_header(h0, h1) and has_two_date_cells) or period_from_text

    # 月度-生产指标2 和供给端sheet强制使用简单日期模式（每列是一个指标/地区）
    if sheet_name == "月度-生产指标2" or sheet_name in SUPPLY_SIDE_SHEETS:
        is_period = False

    if is_period:
        has_metric_col = (
            len(merged_header) > 2
            and h2
            and "指标" in h2
            or (
                len(first_data_row) > 2
                and parse_date(first_data_row[2]) is None
                and to_float(first_data_row[2]) is None
                and len(h2) <= 30
                and not is_weekly_auxiliary_column(h2)
            )
        )
        data_col_start = 3 if has_metric_col else 2
        col_names = [merged_header[i] for i in range(data_col_start, len(merged_header))]
        for row in rows[data_start:]:
            if not row:
                continue
            start_ts = parse_date(row[0] if len(row) > 0 else None)
            end_ts = parse_date(row[1] if len(row) > 1 else None)
            if start_ts is None and end_ts is None:
                text_start, text_end = parse_weekly_text_period_cell(row[0] if len(row) > 0 else None)
                start_ts, end_ts = text_start, text_end
            if end_ts is None:
                end_ts = start_ts
            if end_ts is None:
                continue
            metric_label = normalize_weekly_header_label(row[2] if has_metric_col and len(row) > 2 else None) or "数值"
            source_row_type = "normal"
            raw_text = text_of(row[0] if len(row) > 0 else None)
            for i, col_name in enumerate(col_names):
                col_idx = data_col_start + i
                if col_idx >= len(row) or not col_name or is_weekly_auxiliary_column(col_name):
                    continue
                record = make_weekly_period_record(
                    sheet_name=sheet_name,
                    start_date=start_ts,
                    end_date=end_ts,
                    metric=metric_label,
                    series_name=col_name,
                    value=row[col_idx],
                    freq_type=freq_type,
                    source_row_type=source_row_type,
                    raw_text=raw_text,
                    extra={"topic": topic_hint},
                )
                if record is not None:
                    records.append(record)
        return records

    col_names = [merged_header[i] for i in range(1, len(merged_header))]
    for row in rows[data_start:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for i, col_name in enumerate(col_names):
            col_idx = i + 1
            if col_idx >= len(row) or not col_name or is_weekly_auxiliary_column(col_name):
                continue
            record = make_weekly_period_record(
                sheet_name=sheet_name,
                start_date=date,
                end_date=date,
                metric="数值",
                series_name=col_name,
                value=row[col_idx],
                freq_type=freq_type,
                raw_text=text_of(row[0] if len(row) > 0 else None),
                extra={"topic": topic_hint},
            )
            if record is not None:
                records.append(record)
    return records


def normalize_frozen_dataset_type(label: Any) -> str:
    text = normalize_weekly_header_label(label)
    if "鲜品" in text:
        return "鲜品"
    if "冻品" in text:
        return "冻品"
    return text or "未分类"


def parse_weekly_east_frozen_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    """专门解析“华东冻品价格”横向日期面板。"""
    records: list[dict] = []
    if len(rows) < 4:
        return records
    date_row = rows[1] if len(rows) > 1 else []
    type_row = rows[2] if len(rows) > 2 else []
    dates_filled = filled_right([parse_date(x) for x in date_row])
    types_filled = [normalize_frozen_dataset_type(x) for x in type_row]
    current_category = ""
    for row in rows[3:]:
        if not row:
            continue
        category_raw = normalize_weekly_header_label(row[0] if len(row) > 0 else None)
        if category_raw:
            current_category = category_raw
        product_raw = normalize_weekly_header_label(row[1] if len(row) > 1 else None)
        if not product_raw:
            continue
        for col in range(2, len(row)):
            date = dates_filled[col] if col < len(dates_filled) else None
            if date is None:
                continue
            dataset_type = types_filled[col] if col < len(types_filled) else ""
            low, high, mid, raw = parse_price_band(row[col])
            number = mid if mid is not None else row[col]
            record = build_weekly_record(
                sheet=sheet_name,
                date=date,
                start_date=date,
                end_date=date,
                metric=product_raw,
                series_name=f"{product_raw}｜{dataset_type}" if dataset_type else product_raw,
                value=number,
                province="全部",
                region="华东",
                region_type="region",
                low_price=low,
                high_price=high,
                raw_range=raw if raw != text_of(row[col]) else "",
                raw_text=text_of(row[col]),
                freq_type="周度",
                extra={
                    "dataset_type": dataset_type,
                    "category": current_category or "未分组",
                    "product": product_raw,
                    "topic": "冻品/猪肉",
                },
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_fresh_ratio_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 4:
        return records
    region_row = filled_right([normalize_weekly_header_label(x) for x in (rows[1] if len(rows) > 1 else [])])
    province_row = [normalize_weekly_header_label(x) for x in (rows[2] if len(rows) > 2 else [])]
    last_anchor_date: pd.Timestamp | None = None
    for row in rows[3:]:
        if not row:
            continue
        label = text_of(row[0] if len(row) > 0 else None)
        text_start, text_end = parse_weekly_text_period_cell(label)
        start_ts = parse_date(row[0] if len(row) > 0 else None)
        end_ts = parse_date(row[1] if len(row) > 1 else None)
        source_row_type = "normal"
        if text_end is not None and parse_date(row[1] if len(row) > 1 else None) is None:
            start_ts, end_ts = text_start, text_end
            source_row_type = "holiday_window"
        elif start_ts is None and end_ts is None:
            if "较去年" in label:
                end_ts = last_anchor_date
                start_ts = end_ts
                source_row_type = "yoy_note"
            else:
                end_ts = last_anchor_date
                start_ts = end_ts
                source_row_type = "remark"
        else:
            end_ts = end_ts or start_ts
        if end_ts is None:
            continue
        last_anchor_date = end_ts
        for col in range(2, len(row)):
            province_label = province_row[col] if col < len(province_row) else ""
            if not province_label:
                continue
            region_label = region_row[col] if col < len(region_row) else ""
            record = build_weekly_record(
                sheet=sheet_name,
                date=end_ts,
                start_date=start_ts,
                end_date=end_ts,
                metric="鲜销率",
                series_name=province_label,
                value=row[col],
                province=province_label,
                region=region_label,
                region_type="province",
                raw_text=label,
                source_row_type=source_row_type,
                freq_type="周度",
                extra={"topic": "屠宰/鲜销率"},
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_second_fatten_house_utilization_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 3:
        return records
    header_row = rows[1] if len(rows) > 1 else []
    date_headers = [parse_date(x) for x in header_row]
    for row in rows[2:]:
        if not row or len(row) < 2:
            continue
        province_label = normalize_weekly_header_label(row[1] if len(row) > 1 else None)
        if not province_label:
            continue
        for col in range(4, min(len(row), len(date_headers))):
            end_ts = date_headers[col]
            if end_ts is None:
                continue
            record = build_weekly_record(
                sheet=sheet_name,
                date=end_ts,
                start_date=end_ts,
                end_date=end_ts,
                metric="栏舍利用率",
                series_name=province_label,
                value=row[col],
                province=province_label,
                freq_type="旬度",
                raw_text=province_label,
                extra={"topic": "供给端"},
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_high_frequency_pig_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 3:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[1]]
    for row in rows[2:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        metric_label = normalize_weekly_header_label(row[1] if len(row) > 1 else None)
        if date is None or not metric_label:
            continue
        for col in range(2, len(headers)):
            series_name = headers[col]
            if not series_name or col >= len(row):
                continue
            province = "全部" if series_name in {"平均", "全国"} else series_name
            display_name = "全国" if province == "全部" else series_name
            record = build_weekly_record(
                sheet=sheet_name,
                date=date,
                start_date=date,
                end_date=date,
                metric=metric_label,
                series_name=display_name,
                value=row[col],
                province=province,
                display_name=display_name,
                raw_text=metric_label,
                freq_type="周度",
                extra={"topic": "价格端"},
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_second_fatten_cost_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 4:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[2]]
    for row in rows[3:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        weight_segment = text_of(row[1] if len(row) > 1 else None)
        if date is None or not weight_segment:
            continue
        segment_name = f"{weight_segment}kg"
        for col in range(2, len(headers)):
            metric = headers[col]
            if not metric or col >= len(row):
                continue
            record = build_weekly_record(
                sheet=sheet_name,
                date=date,
                start_date=date,
                end_date=date,
                metric=metric,
                series_name=segment_name,
                value=row[col],
                province="全部",
                raw_text=segment_name,
                freq_type="周度",
                extra={
                    "metric_group": segment_name,
                    "topic": "成本/利润",
                    "measure_name": metric,
                },
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_weight_split_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 4:
        return records
    raw_top_headers = [normalize_weekly_header_label(x) for x in (rows[1] if len(rows) > 1 else [])]
    padded_top_headers = raw_top_headers + [""] * max(0, len(rows[2]) - len(raw_top_headers))
    top_headers = filled_right(padded_top_headers)
    headers = [normalize_weekly_header_label(x) for x in rows[2]]
    for row in rows[3:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for col in range(1, len(headers)):
            header = headers[col]
            if not header or col >= len(row):
                continue
            parent_header = top_headers[col] if col < len(top_headers) else ""
            if col == 1:
                metric = "全国均重"
            elif parent_header == "出栏权重":
                metric = "集团权重" if header == "集团" else "散户权重"
            else:
                metric = "集团均重" if header == "集团" else "散户均重"
            display_name = metric
            province_name = "全部"
            record = build_weekly_record(
                sheet=sheet_name,
                date=date,
                start_date=date,
                end_date=date,
                metric=metric,
                series_name=display_name,
                value=row[col],
                province=province_name,
                display_name=display_name,
                raw_text=metric,
                freq_type="周度",
                extra={
                    "view_group": "体重拆分",
                    "measure_name": metric,
                    "topic": "体重端",
                },
            )
            if record is not None:
                record["series_name"] = display_name
                record["display_name"] = display_name
                record["metric"] = metric
                records.append(record)
    return records


def parse_weekly_metric_panel_sheet(rows: list[list[Any]], sheet_name: str, topic: str = "综合面板", header_row_idx: int = 1, data_start_row: int = 2) -> list[dict]:
    records: list[dict] = []
    if len(rows) <= header_row_idx:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[header_row_idx]]
    for row in rows[data_start_row:]:
        if not row:
            continue
        date = parse_date(row[0] if len(row) > 0 else None)
        if date is None:
            continue
        for col in range(1, len(headers)):
            metric = headers[col]
            if not metric or col >= len(row):
                continue
            record = build_weekly_record(
                sheet=sheet_name,
                date=date,
                start_date=date,
                end_date=date,
                metric=metric,
                series_name=metric,
                value=row[col],
                province="全部",
                raw_text=metric,
                freq_type=infer_weekly_freq_type(sheet_name),
                extra={"topic": topic},
            )
            if record is not None:
                records.append(record)
    return records


def parse_monthly_original_farm_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    records: list[dict] = []
    if len(rows) < 4:
        return records
    for row in rows[3:]:
        if not row:
            continue
        left_date = parse_date(row[0] if len(row) > 0 else None)
        if left_date is not None:
            left_metrics = {
                "后备母猪销量": row[1] if len(row) > 1 else None,
                "后备母猪销量环比": row[2] if len(row) > 2 else None,
                "后备母猪销量同比": row[3] if len(row) > 3 else None,
                "能繁母猪淘汰胎龄": row[4] if len(row) > 4 else None,
                "生猪出栏日龄": row[5] if len(row) > 5 else None,
            }
            for metric, value in left_metrics.items():
                record = build_weekly_record(
                    sheet=sheet_name,
                    date=left_date,
                    start_date=left_date,
                    end_date=left_date,
                    metric=metric,
                    series_name=metric,
                    value=value,
                    province="全部",
                    freq_type="月度",
                    extra={"topic": "月度产能/生产指标", "panel": "原种场与养殖场"},
                )
                if record is not None:
                    records.append(record)
        right_date = parse_date(row[7] if len(row) > 7 else None)
        if right_date is not None:
            remark = text_of(row[11] if len(row) > 11 else None)
            right_metrics = {
                "二元母猪销量": row[8] if len(row) > 8 else None,
                "二元母猪销量环比": row[9] if len(row) > 9 else None,
                "二元母猪销量同比": row[10] if len(row) > 10 else None,
            }
            for metric, value in right_metrics.items():
                record = build_weekly_record(
                    sheet=sheet_name,
                    date=right_date,
                    start_date=right_date,
                    end_date=right_date,
                    metric=metric,
                    series_name=metric,
                    value=value,
                    province="全部",
                    freq_type="月度",
                    raw_text=remark,
                    extra={"topic": "月度产能/生产指标", "panel": "二元母猪销量"},
                )
                if record is not None:
                    records.append(record)
    return records


def _parse_weekly_profit_latest_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict]:
    """周度-养殖利润最新：开始时间+结束时间+项目(利润)+按规模/来源分列"""
    records: list[dict] = []
    if len(rows) < 4:
        return records
    # 找表头行：第一个「col0 和 col1 都是日期」的行的上一行即为表头
    header_row_idx = None
    for i, row in enumerate(rows[1:10], 1):
        if row and len(row) >= 2 and parse_date(row[0]) is not None and parse_date(row[1]) is not None:
            header_row_idx = i - 1
            break
    if header_row_idx is None:
        return records
    headers = [normalize_weekly_header_label(x) for x in rows[header_row_idx]]
    top_headers = [normalize_weekly_header_label(x) for x in (rows[header_row_idx - 1] if header_row_idx > 0 else [])]
    # 补齐表头长度到数据行最大列数（尾列可能因空值被截断）
    data_max_cols = max((len(r) for r in rows[header_row_idx + 1:]), default=0)
    while len(headers) < data_max_cols:
        headers.append("")
    while len(top_headers) < data_max_cols:
        top_headers.append("")
    # 合并二级表头：子表头为空时用上级表头
    merged_headers: list[str] = []
    for col in range(data_max_cols):
        sub = headers[col] if col < len(headers) else ""
        top = top_headers[col] if col < len(top_headers) else ""
        merged_headers.append(sub or top)
    col_metric_map: dict[int, str] = {}
    for col in range(3, len(merged_headers)):
        h = merged_headers[col]
        if not h or is_weekly_auxiliary_column(h):
            continue
        col_metric_map[col] = f"自繁自养利润-{h}"
    # 特殊列名覆盖：识别规模段和来源类型
    for col, name in list(col_metric_map.items()):
        n = name
        if "50头" in n:
            col_metric_map[col] = "自繁自养利润-母猪50头以下"
        elif "50-200" in n:
            col_metric_map[col] = "自繁自养利润-50-200头"
        elif "200-500" in n:
            col_metric_map[col] = "自繁自养利润-200-500头"
        elif "500-2000" in n:
            col_metric_map[col] = "自繁自养利润-500-2000头"
        elif "2000-5000" in n:
            col_metric_map[col] = "自繁自养利润-2000-5000头"
        elif "5000-10000" in n or "5000-" in n:
            col_metric_map[col] = "自繁自养利润-5000-10000头"
        elif "外购" in n:
            col_metric_map[col] = "外购仔猪育肥利润"
        elif "农户" in n or "不同农" in n:
            col_metric_map[col] = "不同农户育肥利润"
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        start_ts = parse_date(row[0] if len(row) > 0 else None)
        end_ts = parse_date(row[1] if len(row) > 1 else None)
        if end_ts is None:
            continue
        project_label = normalize_weekly_header_label(row[2] if len(row) > 2 else None)
        # 遇到非「利润」项目行（如成本子表）则停止
        if project_label and project_label not in {"利润", ""}:
            break
        for col, metric in col_metric_map.items():
            if col >= len(row):
                continue
            record = build_weekly_record(
                sheet=sheet_name, date=end_ts, start_date=start_ts, end_date=end_ts,
                metric=metric, series_name=metric, value=row[col],
                province="全国", freq_type="周度",
                extra={"topic": "利润端"},
            )
            if record is not None:
                records.append(record)
    return records


def parse_weekly_special_sheet(rows: list[list[Any]], sheet_name: str) -> tuple[list[dict], str] | None:
    if sheet_name == "二育栏舍利用率":
        return parse_weekly_second_fatten_house_utilization_sheet(rows, sheet_name), "旬度栏舍利用率"
    if sheet_name == "鲜销率":
        return parse_weekly_fresh_ratio_sheet(rows, sheet_name), "鲜销率专题"
    if sheet_name == "周度-体重":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "体重端"), "商品猪体重"
    if sheet_name == "周度-屠宰厂宰前活猪重":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "体重端"), "宰前活猪重"
    if sheet_name == "周度-各体重段价差":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "价格端"), "各体重段价差"
    if sheet_name == "高频仔猪、母猪":
        return parse_weekly_high_frequency_pig_sheet(rows, sheet_name), "高频仔猪母猪面板"
    if sheet_name == "二育成本":
        return parse_weekly_second_fatten_cost_sheet(rows, sheet_name), "二育成本测算"
    if sheet_name == "华东冻品价格":
        return parse_weekly_east_frozen_sheet(rows, sheet_name), "横向日期+鲜冻双类型"
    if sheet_name == "周度-体重拆分":
        return parse_weekly_weight_split_sheet(rows, sheet_name), "体重/权重拆分"
    if sheet_name == "仔猪与商品猪利润对比":
        return parse_weekly_profit_compare_sheet(rows, sheet_name), "文本周期利润对比"
    if sheet_name == "周度-养殖利润最新":
        return _parse_weekly_profit_latest_sheet(rows, sheet_name), "养殖利润（按规模）"
    if sheet_name == "周度-河南屠宰白条成本":
        return parse_weekly_metric_panel_sheet(rows, sheet_name, topic="屠宰端"), "河南白条成本面板"
    if sheet_name == "周度-毛白价差":
        return apply_topic_to_records(parse_weekly_metric_panel_sheet(rows, sheet_name, topic="价差端", header_row_idx=0, data_start_row=1), "价差端"), "毛白价差面板"
    if sheet_name == "周度-商品猪出栏价":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "价格端"), "商品猪出栏价"
    if sheet_name == "周度-宰后结算价":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "屠宰端"), "宰后结算价"
    if sheet_name == "周度-50公斤二元母猪价格":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "价格端"), "二元母猪价格"
    if sheet_name == "周度-规模场15公斤仔猪出栏价":
        return apply_topic_to_records(parse_weekly_sheet_universal(rows, sheet_name), "价格端"), "仔猪出栏价"
    if sheet_name == "月度-原种场二元后备母猪销量及出栏日龄":
        return parse_monthly_original_farm_sheet(rows, sheet_name), "月度双面板拆分"
    if sheet_name == "月度-二元三元能繁比例":
        return parse_monthly_breeding_ratio_sheet(rows, sheet_name), "文本月份/半年度解析"
    return None


def finalize_weekly_df(records: list[dict]) -> pd.DataFrame:
    cols = [
        "sheet", "freq_type", "date", "start_date", "end_date", "metric", "series_name", "display_name",
        "province", "city", "region", "region_type", "value", "low_price", "high_price",
        "price_range_text", "raw_range", "raw_text", "source_row_type", "metric_group",
        "measure_name", "view_group", "dataset_type", "category", "product", "topic",
        "year", "month", "week", "month_day", "lunar_label", "lunar_order", "is_leap_lunar", "lunar_year",
    ]
    if not records:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(records)
    for col in cols:
        if col not in df.columns:
            df[col] = None
    if df["freq_type"].isna().any():
        df.loc[df["freq_type"].isna(), "freq_type"] = df.loc[df["freq_type"].isna(), "sheet"].map(infer_weekly_freq_type)
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce").dt.normalize()
    df["end_date"] = pd.to_datetime(df["end_date"], errors="coerce").dt.normalize()
    df["display_name"] = df["display_name"].fillna(df["series_name"])
    df = df.dropna(subset=["date", "value"])
    df = enrich_date_features(df)
    return df.sort_values(["sheet", "freq_type", "metric", "series_name", "date"]).reset_index(drop=True)


def build_weekly_dataset_from_path(path_str: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """读取涌益周度 Excel。缓存自动感知文件更新时间。"""
    return _cached_build_weekly_dataset_from_path(_get_file_version(path_str))


@st.cache_data(show_spinner=False)
def _cached_build_weekly_dataset_from_path(versioned_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    读取涌益周度 Excel，自动跳过说明/目录/停更 sheet，
    识别最新更新日期，距离最新主更新超过 60 天的 sheet 标记为停更并隐藏。
    返回：(data_df, meta_df[sheet, mode, latest_date, stale])
    """
    real_path = versioned_path.split("@@v=")[0]
    sheets = read_workbook_rows_from_path(real_path)
    all_records: list[dict] = []
    meta_rows: list[dict] = []

    candidate_sheets = {k: v for k, v in sheets.items() if k not in WEEKLY_SKIP_SHEETS and "计划" not in k}

    latest_dates: dict[str, pd.Timestamp | None] = {
        sheet_name: parse_latest_date_from_rows(rows)
        for sheet_name, rows in candidate_sheets.items()
    }
    global_latest = max((d for d in latest_dates.values() if d is not None), default=None)
    stale_threshold = 60

    for sheet_name, rows in candidate_sheets.items():
        latest = latest_dates.get(sheet_name)
        special_result = parse_weekly_special_sheet(rows, sheet_name)
        is_special = special_result is not None
        is_stale = (
            latest is None
            or (global_latest is not None and (global_latest - latest).days > stale_threshold)
        )
        if is_special or sheet_name in WEEKLY_ALWAYS_SHOW_SHEETS:
            is_stale = False
        meta_row = {
            "sheet": sheet_name,
            "latest_date": latest,
            "stale": is_stale,
            "freq_type": infer_weekly_freq_type(sheet_name),
        }
        meta_rows.append(meta_row)
        if is_stale:
            meta_row["mode"] = "停更隐藏"
            meta_row["records"] = 0
            continue

        if special_result is not None:
            records, mode = special_result
        else:
            records = parse_weekly_sheet_universal(rows, sheet_name)
            mode = "通用自适应解析"

        records = apply_weekly_unit_conversion(records, sheet_name)

        # 针对特定 sheet 排除无关指标（如日期列名等）
        if sheet_name in WEEKLY_SHEET_METRIC_EXCLUDE:
            exclude_set = WEEKLY_SHEET_METRIC_EXCLUDE[sheet_name]
            records = [r for r in records if r.get("series_name", "") not in exclude_set and r.get("metric", "") not in exclude_set]

        # 针对特定 sheet 只包含指定指标
        if sheet_name in WEEKLY_SHEET_METRIC_INCLUDE:
            include_set = WEEKLY_SHEET_METRIC_INCLUDE[sheet_name]
            records = [r for r in records if r.get("series_name", "") in include_set or r.get("metric", "") in include_set]
        # 供给端sheet排除非瘟前数据,保留所有区域（包括全国）
        if sheet_name in SUPPLY_SIDE_SHEETS:
            records = [r for r in records
                  if "非瘟前" not in str(r.get("series_name", ""))
               and "非瘟前" not in str(r.get("metric", ""))]

        freq_type = infer_weekly_freq_type(sheet_name)
        for record in records:
            record["freq_type"] = record.get("freq_type") or freq_type

        if records:
            all_records.extend(records)
        meta_row["mode"] = mode
        meta_row["records"] = len(records)

    data_df = finalize_weekly_df(all_records)
    meta_df = pd.DataFrame(meta_rows)
    return data_df, meta_df


def build_weekly_sheet_summary(weekly_df: pd.DataFrame, target_date: pd.Timestamp, sheet_name: str, label: str, show_comparison: bool = False) -> tuple[list[dict[str, str]], list[str]]:
    subset = weekly_df[(weekly_df["sheet"] == sheet_name) & (weekly_df["source_row_type"].fillna("normal") == "normal")].copy()
    if subset.empty:
        return [], []
    snap_date = get_snapshot_date(subset, target_date)
    if snap_date is None:
        return [], []
    current_df = subset[subset["date"] == snap_date].copy()
    if current_df.empty:
        return [], []
    if "series_name" in current_df.columns:
        preferred = pick_preferred_scope_name(current_df["series_name"].dropna().unique().tolist())
        if preferred:
            preferred_df = current_df[current_df["series_name"] == preferred].copy()
            if not preferred_df.empty:
                current_df = preferred_df
    current = current_df["value"].mean()
    if not show_comparison:
        cards = [{"label": label, "value": format_number(current), "extra": ""}]
        lines = [f"{label} 当前值 {format_number(current)}。"]
        return cards, lines
    prev_dates = subset[subset["date"] < snap_date]["date"]
    prev_date = prev_dates.max() if not prev_dates.empty else None
    prev_value = np.nan
    if prev_date is not None:
        prev_df = subset[subset["date"] == prev_date].copy()
        if "series_name" in current_df.columns and not current_df["series_name"].dropna().empty:
            preferred_name = current_df["series_name"].iloc[0]
            preferred_prev = prev_df[prev_df["series_name"] == preferred_name].copy()
            if not preferred_prev.empty:
                prev_df = preferred_prev
        if not prev_df.empty:
            prev_value = prev_df["value"].mean()
    delta = current - prev_value if pd.notna(prev_value) else np.nan
    history = subset[subset["date"] <= snap_date]["value"]
    position_text, _ = percentile_position_text(current, history)
    cards = [{"label": label, "value": format_number(current), "extra": f"{describe_delta(delta, '较上周')}｜{position_text}"}]
    lines = [f"{label} 当前 {format_number(current)}，{describe_delta(delta, '较上周')}，所处位置：{position_text}。"]
    return cards, lines


def build_weekly_special_snapshot_cards(df: pd.DataFrame, target_date: pd.Timestamp, metric_labels: list[str], show_comparison: bool = False) -> list[dict[str, str]]:
    cards: list[dict[str, str]] = []
    if df.empty:
        return cards
    snap_date = get_snapshot_date(df, target_date)
    if snap_date is None:
        return cards
    current_df = df[df["date"] == snap_date].copy()
    if current_df.empty:
        return cards
    for metric in metric_labels:
        sub = current_df[current_df["metric"] == metric].copy()
        if sub.empty:
            continue
        current = sub["value"].mean()
        extra = ""
        if show_comparison:
            prev_dates = df[(df["metric"] == metric) & (df["date"] < snap_date)]["date"]
            prev_date = prev_dates.max() if not prev_dates.empty else None
            prev_value = df[(df["metric"] == metric) & (df["date"] == prev_date)]["value"].mean() if prev_date is not None else np.nan
            extra = describe_delta(current - prev_value if pd.notna(prev_value) else np.nan, "较上期")
        cards.append({
            "label": metric,
            "value": format_number(current),
            "extra": extra,
        })
    return cards


def build_weekly_metric_spread_frames(df: pd.DataFrame, pairs: list[tuple[str, str, str]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for left_metric, right_metric, spread_name in pairs:
        spread_df = build_metric_spread_df(df, left_metric, right_metric, spread_name)
        if spread_df.empty:
            continue
        for col in ["freq_type", "topic", "region", "region_type", "source_row_type"]:
            if col in df.columns and col not in spread_df.columns:
                spread_df[col] = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            elif col in df.columns and col in spread_df.columns and spread_df[col].isna().all():
                spread_df[col] = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
        frames.append(spread_df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=df.columns)


def build_weekly_dual_axis_topic_chart(main_df: pd.DataFrame, spread_df: pd.DataFrame, frequency: str, title: str, use_lunar: bool) -> go.Figure:
    prepared_main = prepare_metric_compare_df(main_df.copy()) if not main_df.empty else main_df
    prepared_spread = prepare_metric_compare_df(spread_df.copy()) if not spread_df.empty else spread_df
    if not prepared_main.empty and not prepared_spread.empty:
        return build_trend_line_chart_with_spread(prepared_main, prepared_spread, frequency, title, use_lunar=use_lunar)
    if not prepared_main.empty:
        return build_multi_series_line_chart(prepared_main, frequency, title, hover_title="数值", use_lunar=use_lunar)
    return build_multi_series_line_chart(prepared_spread, frequency, title, hover_title="价差", use_lunar=use_lunar)


def weekly_change_style(delta: float | None, context: str = "level") -> str:
    if delta is None or pd.isna(delta):
        return "暂无可比变化"
    if abs(delta) < 1e-12:
        return "基本持平"
    if context == "spread":
        return "走阔" if delta > 0 else "收窄"
    if context == "inventory":
        return "累库" if delta > 0 else "去库"
    if context == "profit":
        return "改善" if delta > 0 else "恶化"
    return "走强" if delta > 0 else "走弱"


def build_weekly_business_line(label: str, current: float, delta: float | None, yoy_delta: float | None, position_text: str, topic: str, show_comparison: bool = True) -> str:
    if not show_comparison:
        if position_text and position_text != "样本不足":
            return f"{label} 当前值 {format_number(current)}，所处位置：{position_text}。"
        return f"{label} 当前值 {format_number(current)}。"
    context = "level"
    if "价差" in label:
        context = "spread"
    elif any(flag in label for flag in ["库存", "库容"]):
        context = "inventory"
    elif any(flag in label for flag in ["利润", "成本"]):
        context = "profit"
    weekly_text = weekly_change_style(delta, context)
    yoy_text = weekly_change_style(yoy_delta, context)
    if topic == "价格端":
        return f"{label} 当前录得 {format_number(current)}，周度表现{weekly_text}，同比{yoy_text}，当前处于{position_text}。"
    if topic == "体重端":
        return f"{label} 当前为 {format_number(current)}，较上期{describe_delta(delta, '')}，体重结构整体{weekly_text}，当前处于{position_text}。"
    if topic == "成本/利润":
        return f"{label} 当前值 {format_number(current)}，环比{weekly_text}，同比{yoy_text}，反映成本利润格局处于{position_text}。"
    if topic == "屠宰/鲜销率":
        return f"{label} 当前值 {format_number(current)}，周内表现{weekly_text}，同比{yoy_text}，屠宰链条指标处于{position_text}。"
    if topic == "冻品/猪肉":
        return f"{label} 当前录得 {format_number(current)}，周度{weekly_text}，同比{yoy_text}，鲜冻/猪肉结构位于{position_text}。"
    return f"{label} 当前值 {format_number(current)}，{describe_delta(delta, '较上周')}，{describe_delta(yoy_delta, '较去年同期')}，所处位置：{position_text}。"


def build_weekly_topic_snapshot(weekly_df: pd.DataFrame, target_date: pd.Timestamp, topic: str, label: str, show_comparison: bool = True) -> dict[str, Any] | None:
    subset = weekly_df[weekly_df["topic"] == topic].copy() if "topic" in weekly_df.columns else pd.DataFrame()
    subset = subset[subset["source_row_type"].fillna("normal").isin(["normal", "holiday_window"])] if not subset.empty else subset
    if subset.empty:
        return None
    snap_date = get_snapshot_date(subset, target_date)
    if snap_date is None:
        return None
    current_df = subset[subset["date"] == snap_date].copy()
    if current_df.empty:
        return None
    current = current_df["value"].mean()
    delta = np.nan
    yoy_delta = np.nan
    if show_comparison:
        prev_dates = subset[subset["date"] < snap_date]["date"]
        prev_date = prev_dates.max() if not prev_dates.empty else None
        prev_value = subset[subset["date"] == prev_date]["value"].mean() if prev_date is not None else np.nan
        yoy_match = same_day_last_year_match(subset["date"], snap_date)
        yoy_value = subset[subset["date"] == yoy_match]["value"].mean() if yoy_match is not None else np.nan
        delta = current - prev_value if pd.notna(prev_value) else np.nan
        yoy_delta = current - yoy_value if pd.notna(yoy_value) else np.nan
    position_text, _ = percentile_position_text(current, subset[subset["date"] <= snap_date]["value"])
    return {
        "label": label,
        "date": snap_date,
        "current": current,
        "delta": delta,
        "yoy_delta": yoy_delta,
        "position": position_text,
    }


def build_weekly_global_summary(weekly_df: pd.DataFrame, target_date: pd.Timestamp, show_comparison: bool = False) -> dict[str, Any]:
    """生成周度综合摘要：按用户指定来源 sheet 汇总核心面板。"""
    if weekly_df.empty:
        return {"ok": False}

    target = pd.Timestamp(target_date).normalize()
    summary_sources = [
        ("周度-商品猪出栏价", "商品猪出栏价"),
        ("周度-体重", "出栏均重"),
        ("周度-屠宰企业日度屠宰量", "屠宰企业日度屠宰量"),
        ("周度-冻品库存", "冻品库存"),
        ("鲜销率", "鲜销率"),
    ]

    cards: list[dict[str, str]] = []
    lines: list[str] = []
    for sheet_name, label in summary_sources:
        source_df = weekly_df[weekly_df["sheet"] == sheet_name].copy()
        if source_df.empty:
            continue
        if sheet_name == "周度-商品猪出栏价":
            source_df = source_df[source_df["series_name"].isin(["全国", "全国均价"])].copy() if source_df["series_name"].isin(["全国", "全国均价"]).any() else source_df
        elif sheet_name == "周度-体重":
            source_df = source_df[source_df["metric"] == "均重"].copy() if (source_df["metric"] == "均重").any() else source_df
        elif sheet_name == "周度-屠宰企业日度屠宰量":
            source_df = source_df[source_df["series_name"].isin(["全国", "全国均价"])].copy() if source_df["series_name"].isin(["全国", "全国均价"]).any() else source_df
        elif sheet_name == "周度-冻品库存":
            source_df = source_df[source_df["series_name"].isin(["全国", "全国均价"])].copy() if source_df["series_name"].isin(["全国", "全国均价"]).any() else source_df
        cards_part, lines_part = build_weekly_sheet_summary(source_df, target, sheet_name, label, show_comparison=show_comparison)
        if cards_part:
            cards.extend(cards_part[:1])
        if lines_part:
            lines.extend(lines_part[:1])

    brief = " ".join(lines[:4]) if lines else "暂无周度汇总数据。"
    return {
        "ok": bool(lines),
        "date": target,
        "title": "本周综合摘要",
        "brief": brief,
        "cards": cards,
        "lines": lines,
    }


# -----------------------------
# 模块页面：涌益周度数据
# -----------------------------
def build_weekly_board_local_summary(weekly_df: pd.DataFrame, target_date: pd.Timestamp, board_name: str, show_comparison: bool = False) -> dict[str, Any]:
    target = pd.Timestamp(target_date).normalize()
    if board_name == "价格端":
        subset = weekly_df[weekly_df["sheet"].isin(["周度-商品猪出栏价", "周度-50公斤二元母猪价格", "周度-规模场15公斤仔猪出栏价", "高频仔猪、母猪"])].copy()
    elif board_name == "屠宰端":
        subset = weekly_df[weekly_df["sheet"].isin(["周度-毛白价差", "周度-河南屠宰白条成本", "周度-冻品库存", "周度-冻品库存多样本", "鲜销率", "周度-屠宰企业日度屠宰量", "周度-屠宰新2022.10.28", "月度-淘汰母猪屠宰厂宰杀量", "周度-宰后结算价"])].copy()
    elif board_name == "猪肉产品":
        subset = weekly_df[weekly_df["sheet"].isin(["周度-猪肉价（前三等级白条均价）", "周度-猪肉产品价格", "华东冻品价格", "国产冻品2-4号肉价格"])].copy()
    elif board_name == "其他":
        subset = weekly_df[weekly_df["sheet"].isin(list(WEEKLY_FEED_SHEETS))].copy()
    else:
        subset = weekly_df[weekly_df["sheet"].isin(WEEKLY_BOARD_SHEETS.get(board_name, []))].copy()
    if subset.empty:
        return {"ok": False}
    payload = build_summary_payload(subset, target, show_comparison=show_comparison)
    if not payload.get("ok"):
        return payload
    payload["title"] = f"{board_name}板块摘要"
    return payload


def render_weekly_linked_price_spread_view(weekly_df: pd.DataFrame, frequency: str, target_date: pd.Timestamp, use_lunar: bool) -> bool:
    out_df = weekly_df[weekly_df["sheet"] == "周度-商品猪出栏价"].copy()
    settle_df = weekly_df[weekly_df["sheet"] == "周度-宰后结算价"].copy()
    hair_white_df = weekly_df[weekly_df["sheet"] == "周度-毛白价差"].copy()
    sow_df = weekly_df[weekly_df["sheet"] == "周度-50公斤二元母猪价格"].copy()
    piglet_df = weekly_df[weekly_df["sheet"] == "周度-规模场15公斤仔猪出栏价"].copy()
    if out_df.empty or settle_df.empty:
        return False

    scope_candidates = sorted(set(out_df["series_name"].dropna().tolist()) & set(settle_df["series_name"].dropna().tolist()))
    preferred_scope = pick_preferred_scope_name(scope_candidates) if scope_candidates else None
    selected_scope = st.selectbox("联动专题地区", scope_candidates, index=scope_candidates.index(preferred_scope) if preferred_scope in scope_candidates else 0, key="weekly_linked_scope") if scope_candidates else ""

    def _scope(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or not selected_scope:
            return df.copy()
        scoped = df[df["series_name"] == selected_scope].copy()
        return scoped if not scoped.empty else df.copy()

    out_scope = _scope(out_df)
    settle_scope = _scope(settle_df)
    sow_scope = _scope(sow_df)
    piglet_scope = _scope(piglet_df)

    out_scope = out_scope.copy()
    out_scope["metric"] = "商品猪出栏价"
    settle_scope = settle_scope.copy()
    settle_scope["metric"] = "宰后结算价"
    sow_scope = sow_scope.copy()
    if not sow_scope.empty:
        sow_scope["metric"] = "50公斤二元母猪价格"
    piglet_scope = piglet_scope.copy()
    if not piglet_scope.empty:
        piglet_scope["metric"] = "15公斤仔猪出栏价"

    spread_cross = build_cross_sheet_spread_df(out_scope, settle_scope, "宰后-出栏价差")
    if not spread_cross.empty:
        spread_cross["topic"] = "价差端"
        spread_cross["series_name"] = selected_scope or "宰后-出栏价差"
        spread_cross["display_name"] = spread_cross["series_name"]

    linked_main = pd.concat([frame for frame in [out_scope, settle_scope, sow_scope, piglet_scope] if not frame.empty], ignore_index=True)
    render_section_header("🔗 价格端联动总览")
    render_plotly(build_weekly_dual_axis_topic_chart(linked_main, spread_cross, frequency, f"价格联动｜{selected_scope or '全国'}", False), "weekly", "linked_price_spread", selected_scope, frequency)

    if not hair_white_df.empty:
        hair_white_line = hair_white_df[hair_white_df["metric"].isin(["价差", "毛白价差"])].copy()
        if hair_white_line.empty:
            hair_white_line = hair_white_df.copy()
        hair_white_line = hair_white_line.copy()
        hair_white_line["series_name"] = hair_white_line["metric"]
        render_section_header("📐 毛白价差专题")
        render_plotly(build_multi_series_line_chart(hair_white_line, frequency, "毛白价差走势", hover_title="价差", use_lunar=False), "weekly", "hair_white_spread", frequency)

    cards: list[dict[str, str]] = []
    cards.extend(build_weekly_special_snapshot_cards(out_scope, target_date, ["商品猪出栏价"], show_comparison=should_show_weekly_comparison("周度-商品猪出栏价")))
    cards.extend(build_weekly_special_snapshot_cards(settle_scope, target_date, ["宰后结算价"], show_comparison=should_show_weekly_comparison("周度-宰后结算价")))
    cards.extend(build_weekly_special_snapshot_cards(sow_scope, target_date, ["50公斤二元母猪价格"], show_comparison=should_show_weekly_comparison("周度-50公斤二元母猪价格")))
    cards.extend(build_weekly_special_snapshot_cards(piglet_scope, target_date, ["15公斤仔猪出栏价"], show_comparison=should_show_weekly_comparison("周度-规模场15公斤仔猪出栏价")))
    if not spread_cross.empty:
        cards.extend(build_weekly_special_snapshot_cards(spread_cross, target_date, ["宰后-出栏价差"], show_comparison=False))
    if not hair_white_df.empty:
        cards.extend(build_weekly_special_snapshot_cards(hair_white_df, target_date, ["价差"], show_comparison=False))
    if cards:
        render_metric_cards(cards[:6])
    return True


def render_weekly_special_topic(df: pd.DataFrame, selected_sheet: str, frequency: str, target_date: pd.Timestamp, use_lunar: bool, show_seasonal: bool = False) -> bool:
    if df.empty:
        return False
    show_comparison = should_show_weekly_comparison(selected_sheet)
    if selected_sheet == "鲜销率":
        normal_df = df[df["source_row_type"].fillna("normal") == "normal"].copy()
        holiday_df = df[df["source_row_type"] == "holiday_window"].copy()
        metric_options = ["鲜销率"]
        selected_metric = st.selectbox("选择指标", metric_options, key="weekly_fresh_ratio_metric") if show_seasonal else "鲜销率"
        if not normal_df.empty:
            render_section_header("📈 鲜销率主趋势")
            render_plotly(build_multi_series_line_chart(normal_df, frequency, f"鲜销率主趋势｜{selected_metric}", hover_title="鲜销率", use_lunar=False), "weekly", "fresh_ratio", selected_metric, frequency)
            if show_seasonal:
                _render_seasonal_section(normal_df.copy(), use_lunar, "weekly_fresh_ratio", frequency, metric_label="鲜销率")
        if not holiday_df.empty:
            render_section_header("🎯 鲜销率春节窗口专题")
            holiday_view = prepare_metric_compare_df(holiday_df.copy())
            render_plotly(build_multi_series_line_chart(holiday_view, "月度数据（均值）", "春节窗口鲜销率", hover_title="鲜销率"), "weekly", "fresh_ratio", "holiday")
        cards = build_weekly_special_snapshot_cards(normal_df, target_date, ["鲜销率"], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "华东冻品价格":
        categories = sorted(df["category"].dropna().unique().tolist()) if "category" in df.columns else []
        dataset_types = sorted(df["dataset_type"].dropna().unique().tolist()) if "dataset_type" in df.columns else []
        products = sorted(df["product"].dropna().unique().tolist()) if "product" in df.columns else []
        col1, col2, col3 = st.columns(3)
        with col1:
            selected_category = st.selectbox("产品大类", ["全部"] + categories, key="weekly_frozen_category") if categories else "全部"
        scoped = df.copy()
        if selected_category != "全部":
            scoped = scoped[scoped["category"] == selected_category].copy()
        with col2:
            available_products = sorted(scoped["product"].dropna().unique().tolist()) if "product" in scoped.columns else products
            selected_product = st.selectbox("单品", available_products, key="weekly_frozen_product") if available_products else ""
        if selected_product:
            scoped = scoped[scoped["product"] == selected_product].copy()
        with col3:
            selected_type = st.selectbox("鲜/冻类型", ["全部"] + dataset_types, key="weekly_frozen_type") if dataset_types else "全部"
        if selected_type != "全部":
            scoped = scoped[scoped["dataset_type"] == selected_type].copy()
        if not scoped.empty:
            render_plotly(build_product_price_chart(scoped, frequency, "华东冻品价格走势"), "weekly", "frozen", selected_product, selected_type, frequency)
            if show_seasonal:
                _render_seasonal_section(scoped.copy(), use_lunar, "weekly_frozen", frequency, metric_label=selected_product, scope_label="华东")
            spread_pairs = scoped.pivot_table(index=["date", "product"], columns="dataset_type", values="value", aggfunc="mean").reset_index()
            if {"鲜品", "冻品"}.issubset(set(spread_pairs.columns)):
                spread_pairs["value"] = spread_pairs["鲜品"] - spread_pairs["冻品"]
                spread_pairs["metric"] = "鲜冻价差"
                spread_pairs["series_name"] = spread_pairs["product"]
                spread_pairs["province"] = "全部"
                spread_pairs["city"] = ""
                render_section_header("📦 华东冻品鲜冻价差")
                render_plotly(build_multi_series_line_chart(spread_pairs, frequency, "鲜冻价差", hover_title="价差"), "weekly", "frozen_spread", selected_product, frequency)
        cards = build_weekly_special_snapshot_cards(scoped, target_date, [selected_product] if selected_product else [], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "二育栏舍利用率":
        provinces = sorted(df["province"].dropna().unique().tolist())
        default_provinces = provinces if provinces else []
        ensure_multi_choice("weekly_eryu_provinces", provinces, default_provinces)
        selected_provinces = st.multiselect("选择省份", provinces, default=default_provinces, key="weekly_eryu_provinces")
        current = df[df["province"].isin(selected_provinces)].copy() if selected_provinces else df.copy()
        if not current.empty:
            render_plotly(build_multi_series_line_chart(current, frequency, "二育栏舍利用率", hover_title="利用率", use_lunar=False), "weekly", "eryu", frequency)
            if show_seasonal:
                _render_seasonal_section(current.copy(), use_lunar, "weekly_eryu", frequency, metric_label="栏舍利用率")
        cards = build_weekly_special_snapshot_cards(current, target_date, ["栏舍利用率"], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "周度-体重拆分":
        # 直接从解析数据中获取所有可用的指标名（避免硬编码汉字可能出现的编码差异）
        all_series_names = sorted(df["series_name"].dropna().unique().tolist())
        available_series = [n for n in all_series_names if n not in {"日期", "数据日期", "月份"}]
        default_series = available_series if available_series else []
        selected_series = st.multiselect("选择体重拆分指标", available_series, default=default_series, key="weekly_weight_split_series") if available_series else []
        if show_seasonal:
            st.info("🔍 季节性对比仅支持单一指标，已自动截取首个选中项。")
            if selected_series:
                selected_series = selected_series[:1]
        current = df[df["series_name"].isin(selected_series)].copy() if selected_series else df.copy()
        render_plotly(build_multi_series_line_chart(current, frequency, "体重拆分", hover_title="数值", use_lunar=False), "weekly", "weight_split", tuple(selected_series), frequency)
        if show_seasonal and not current.empty:
            series_label = selected_series[0] if selected_series else (available_series[0] if available_series else "")
            _render_seasonal_section(current.copy(), use_lunar, "weekly_weight_split", frequency, metric_label=series_label)
        cards = build_weekly_special_snapshot_cards(current, target_date, selected_series[:5] if selected_series else available_series[:5], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "二育成本":
        segments = sorted(df["metric_group"].dropna().unique().tolist()) if "metric_group" in df.columns else []
        metrics = sorted(df["metric"].dropna().unique().tolist())
        col1, col2 = st.columns(2)
        with col1:
            selected_segment = st.selectbox("采购体重段", segments, key="weekly_second_fatten_segment") if segments else ""
        with col2:
            selected_metric = st.selectbox("成本指标", metrics, key="weekly_second_fatten_metric") if metrics else ""
        current = df.copy()
        if selected_segment:
            current = current[current["metric_group"] == selected_segment].copy()
        if selected_metric:
            current = current[current["metric"] == selected_metric].copy()
        if not current.empty:
            render_plotly(build_multi_series_line_chart(current, frequency, f"二育成本｜{selected_metric}", hover_title="数值", use_lunar=False), "weekly", "second_fatten", selected_segment, selected_metric)
            if show_seasonal:
                _render_seasonal_section(current.copy(), use_lunar, "weekly_second_fatten", frequency, metric_label=selected_metric or "")
            cards = build_weekly_special_snapshot_cards(current, target_date, [selected_metric], show_comparison=show_comparison)
            if cards:
                render_metric_cards(cards)
        return True
    if selected_sheet == "周度-河南屠宰白条成本":
        metric_pairs = [
            ("白条成本元/kg", "生猪结算元/kg", "宰后-出栏价差"),
            ("6个等级白条均价（元/kg）", "白条成本元/kg", "毛白价差"),
        ]
        spread_df = build_weekly_metric_spread_frames(df, metric_pairs)
        metric_options = [m for m in ["生猪结算元/kg", "白条成本元/kg", "6个等级白条均价（元/kg）", "白条头均利润（元/头）"] if m in df["metric"].unique()]
        default_metrics = metric_options[:1] if show_seasonal else metric_options[:2]
        selected_metrics = st.multiselect("选择屠宰主指标", metric_options, default=default_metrics, key="weekly_slaughter_metrics")
        if show_seasonal:
            st.info("🔍 季节性对比仅支持单一指标，已自动截取首个选中项。")
            if selected_metrics:
                selected_metrics = selected_metrics[:1]
        main_df = df[df["metric"].isin(selected_metrics)].copy() if selected_metrics else df.copy()
        if not main_df.empty:
            render_section_header("🐷 屠宰成本/利润主图")
            render_plotly(build_weekly_dual_axis_topic_chart(main_df, spread_df, frequency, "屠宰成本与价差", False), "weekly", "slaughter_dual", frequency)
            if show_seasonal:
                _render_seasonal_section(main_df.copy(), use_lunar, "weekly_slaughter", frequency, metric_label=selected_metrics[0] if selected_metrics else "白条成本元/kg", scope_label="河南")
        if not spread_df.empty and not show_seasonal:
            render_section_header("📐 派生价差专题")
            render_plotly(build_multi_series_line_chart(spread_df, frequency, "宰后-出栏价差 / 毛白价差", hover_title="价差", use_lunar=False), "weekly", "slaughter_spread", frequency)
        cards = build_weekly_special_snapshot_cards(df, target_date, ["白条成本元/kg", "6个等级白条均价（元/kg）", "白条头均利润（元/头）"], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "周度-养殖利润最新":
        metric_options = sorted(df["metric"].dropna().unique().tolist())
        default_metric = metric_options[0] if metric_options else None
        ensure_single_choice("weekly_profit_latest_metric", metric_options, default_metric)
        selected_metric = st.selectbox("选择养殖利润指标", metric_options, key="weekly_profit_latest_metric")
        current = df[df["metric"] == selected_metric].copy() if selected_metric else df.copy()
        if not current.empty:
            render_plotly(build_multi_series_line_chart(current, frequency, f"养殖利润｜{selected_metric}", hover_title="利润（元/头）", use_lunar=False), "weekly", "profit_latest", selected_metric, frequency)
            if show_seasonal:
                _render_seasonal_section(current.copy(), use_lunar, "weekly_profit_latest", frequency, metric_label=selected_metric)
        cards = build_weekly_special_snapshot_cards(current, target_date, [selected_metric] if selected_metric else [], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True
    if selected_sheet == "仔猪与商品猪利润对比":
        profit_metrics = sorted(df["metric"].dropna().unique().tolist())
        piglet_profit = [m for m in profit_metrics if "仔猪" in m and "利润" in m]
        hog_profit = [m for m in profit_metrics if "商品猪" in m and "利润" in m]
        col1, col2 = st.columns(2)
        with col1:
            selected_piglet = st.multiselect("选择仔猪利润指标", piglet_profit, default=piglet_profit[:1] if piglet_profit else [], key="weekly_piglet_profit")
        with col2:
            selected_hog = st.multiselect("选择商品猪利润指标", hog_profit, default=hog_profit[:1] if hog_profit else [], key="weekly_hog_profit")
        show_spread = st.checkbox("显示价差(右轴)", value=False, key="weekly_profit_spread")
        selected_metrics = selected_piglet + selected_hog
        main_df = df[df["metric"].isin(selected_metrics)].copy() if selected_metrics else pd.DataFrame()
        spread_df = pd.DataFrame()
        if show_spread and len(selected_piglet) == 1 and len(selected_hog) == 1:
            piglet_data = df[df["metric"] == selected_piglet[0]].copy()
            hog_data = df[df["metric"] == selected_hog[0]].copy()
            if not piglet_data.empty and not hog_data.empty:
                merged = pd.merge(piglet_data[["date", "value"]].rename(columns={"value": "piglet_value"}), hog_data[["date", "value"]].rename(columns={"value": "hog_value"}), on="date", how="inner")
                merged["value"] = merged["hog_value"] - merged["piglet_value"]
                merged["metric"] = "利润价差"
                merged["series_name"] = "商品猪-仔猪"
                merged["province"] = "全部"
                spread_df = merged[["date", "metric", "series_name", "value", "province"]].copy()
        if not main_df.empty:
            render_plotly(build_weekly_dual_axis_topic_chart(main_df, spread_df, frequency, "利润对比", False), "weekly", "profit_compare", frequency)
            if show_seasonal:
                _render_seasonal_section(main_df.copy(), use_lunar, "weekly_profit_compare", frequency, metric_label=selected_metrics[0] if selected_metrics else "")
        cards = build_weekly_special_snapshot_cards(df, target_date, selected_metrics[:3] if selected_metrics else [], show_comparison=show_comparison)
        if cards:
            render_metric_cards(cards)
        return True

def render_weekly_section(weekly_df: pd.DataFrame, target_date_value) -> None:
    """渲染涌益周度数据子页内容。"""
    if weekly_df.empty:
        st.warning("周度数据为空，请检查文件路径。")
        return

    target_date = pd.Timestamp(target_date_value).normalize()
    weekly_summary = build_weekly_global_summary(weekly_df, target_date, show_comparison=False)
    if weekly_summary.get("ok"):
        render_metric_cards(weekly_summary["cards"])
        render_summary_box(weekly_summary["title"], format_date_cn(weekly_summary["date"]), weekly_summary["brief"])
        render_signal_messages("摘要明细", weekly_summary.get("lines", []), "暂无可拆解周度摘要。")

    render_section_header("🔍 板块选择与图表")
    all_sheets = sorted(weekly_df["sheet"].dropna().unique().tolist())
    board_map: dict[str, list[str]] = {
        board_name: [sheet for sheet in sheet_names if sheet in all_sheets]
        for board_name, sheet_names in WEEKLY_BOARD_SHEETS.items()
    }
    board_map = {k: v for k, v in board_map.items() if v}
    if not board_map:
        st.info("当前周度数据暂无可展示的分析板块。")
        return

    board_options = list(board_map.keys())
    ensure_single_choice("weekly_board", board_options, board_options[0])
    selected_board = st.selectbox("选择分析板块", board_options, key="weekly_board")
    available_sheets = board_map.get(selected_board, [])
    if not available_sheets:
        st.info("当前板块暂无可展示的 Sheet。")
        return
    ensure_single_choice("weekly_selected_sheet", available_sheets, available_sheets[0] if available_sheets else None)
    selected_sheet = st.selectbox("选择 Sheet", available_sheets, key="weekly_selected_sheet")
    sheet_df = weekly_df[weekly_df["sheet"] == selected_sheet].copy()
    if sheet_df.empty:
        st.info(f"「{selected_sheet}」没有可解析数据。")
        return

    show_comparison = should_show_weekly_comparison(selected_sheet)
    sheet_freq_type = sheet_df["freq_type"].iloc[0] if not sheet_df.empty else "周度"
    if sheet_freq_type == "月度":
        frequency_options = ["月度数据（均值）"]
    elif sheet_freq_type == "旬度":
        frequency_options = ["旬度数据", "月度数据（均值）"]
    else:
        frequency_options = ["周度数据（均值）", "月度数据（均值）"]
    frequency, show_seasonal, use_lunar = _render_standard_controls("weekly", frequency_options=frequency_options)
    _, _, sheet_df = build_cn_date_range_selector(sheet_df, "weekly_chart", "📅 周度图表日期范围")
    range_text = st.session_state.get("weekly_chart_range_text", "")

    raw_metric_options = sorted(sheet_df["metric"].dropna().unique().tolist())
    special_sheet_metrics = {"周度-河南屠宰白条成本", "华东冻品价格", "鲜销率", "周度-体重拆分", "二育成本", "仔猪与商品猪利润对比", "周度-养殖利润最新", "二育栏舍利用率"}
    if len(raw_metric_options) > 1 and selected_sheet not in special_sheet_metrics:
        if show_seasonal:
            st.info("🔍 季节性对比仅支持单一指标，请在下方选择。")
        default_metric = raw_metric_options[0] if raw_metric_options else None
        ensure_single_choice("weekly_metric", raw_metric_options, default_metric)
        selected_metric = st.selectbox("选择指标", raw_metric_options, key="weekly_metric")
        sheet_df = sheet_df[sheet_df["metric"] == selected_metric].copy()
    else:
        selected_metric = raw_metric_options[0] if raw_metric_options else ""

    dimension = get_series_dimension(sheet_df)
    if dimension == "province":
        region_options = ["全国"] + sorted([x for x in sheet_df["province"].dropna().unique() if x and x != "全部"])
        preferred_scope = "全国"
        if show_seasonal:
            st.info("🔍 季节性对比仅支持单一区域，请在下方选择。")
            ensure_single_choice("selected_region_single", region_options, preferred_scope)
            selected_scope = st.selectbox("选择地区", region_options, key="selected_region_single")
            if selected_scope == "全国":
                sheet_df = sheet_df[sheet_df["province"] == "全部"].copy() if (sheet_df["province"] == "全部").any() else sheet_df.copy()
            else:
                sheet_df = sheet_df[sheet_df["province"] == selected_scope].copy()
        elif len(region_options) > 1:
            # 如果数据中没有"全部"省份，则默认选择所有省份
            has_national = (sheet_df["province"] == "全部").any()
            if has_national:
                default_regions = [preferred_scope]
            else:
            # 对于没有全国数据的sheet（如二育栏舍利用率），默认选择所有省份
                default_regions = region_options
            ensure_multi_choice("selected_regions", region_options, default_regions)
            selected_regions = st.multiselect("选择地区（可多选）", region_options, default=default_regions, key="selected_regions")
            if selected_regions:
                province_filters = ["全部" if region == "全国" else region for region in selected_regions]
                sheet_df = sheet_df[sheet_df["province"].isin(province_filters)].copy()

    series_options = sorted(sheet_df["series_name"].dropna().unique().tolist())
    # 供给端 sheet 排除日期/月份等标识列，避免作为展示序列
    if selected_board == "供给端":
        series_options = [s for s in series_options if s not in {"日期", "日期（月/日）", "月份", "时间", "期数", "月度"}]
    if len(series_options) > 1 and dimension == "none" and selected_sheet not in special_sheet_metrics:
        preferred_series = pick_preferred_scope_name(series_options)
        default_series = [preferred_series] if preferred_series else series_options[:3]
        ensure_multi_choice("weekly_series", series_options, default_series)
        selected_series = st.multiselect("选择展示序列", series_options, key="weekly_series")
        if selected_series:
            sheet_df = sheet_df[sheet_df["series_name"].isin(selected_series)].copy()

    st.markdown(f"<div class='mini-note'>{build_chart_context_note(selected_sheet, selected_metric, selected_board, target_date, show_seasonal, use_lunar, range_text)}</div>", unsafe_allow_html=True)

    handled_special = render_weekly_special_topic(sheet_df.copy(), selected_sheet, frequency, target_date, use_lunar, show_seasonal)
    if not handled_special:
        chart_df = prepare_metric_compare_df(sheet_df.copy()) if selected_metric and sheet_df["metric"].nunique() > 1 else sheet_df.copy()
        has_range = chart_df.get("raw_range", pd.Series(dtype=str)).fillna("").ne("").any() or chart_df.get("price_range_text", pd.Series(dtype=str)).fillna("").ne("").any()
        if show_seasonal and selected_sheet == "周度-体重拆分" and 'selected_series' in locals() and selected_series:
            chart_df = chart_df[chart_df["series_name"] == selected_series[0]].copy()
        tab_trend, tab_region = st.tabs(["📈 趋势/季节性", "🗺️ 区域对比"])
        with tab_trend:
            render_plotly(
                build_product_price_chart(chart_df, frequency, f"{selected_sheet} 走势") if has_range else build_multi_series_line_chart(chart_df, frequency, f"{selected_sheet} 走势", hover_title="数值", use_lunar=False),
                "weekly", "trend", selected_sheet, frequency,
            )
            if show_seasonal and not chart_df.empty:
                scope_label = selected_scope if 'selected_scope' in dir() else "全国"
                _render_seasonal_section(chart_df, use_lunar, f"weekly_{selected_sheet}", frequency, metric_label=selected_metric, scope_label=scope_label)
        with tab_region:
            if get_group_field(sheet_df) is not None and not sheet_df.empty:
                fig_rank, note_rank = build_rank_bar_chart(sheet_df, target_date)
                if note_rank:
                    st.info(note_rank)
                render_plotly(fig_rank, "weekly", "rank", selected_sheet)
                fig_heat, note_heat = build_regular_heatmap(sheet_df, target_date)
                if note_heat:
                    st.info(note_heat)
                render_plotly(fig_heat, "weekly", "heatmap", selected_sheet)
                if get_series_dimension(sheet_df) == "province":
                    fig_map, note_map = build_map_chart(sheet_df, target_date)
                    if note_map:
                        st.info(note_map)
                    render_plotly(fig_map, "weekly", "map", selected_sheet)
            else:
                st.info("当前口径不适合地区对比展示。")

    with st.expander("📊 近期数据一览", expanded=False):
        snap_date = get_snapshot_date(sheet_df, target_date)
        if snap_date is not None:
            snap_df = sheet_df[sheet_df["date"] == snap_date].groupby("series_name", as_index=False)["value"].mean().sort_values("value", ascending=False)
            if not snap_df.empty:
                snap_df.columns = ["序列", "数值"]
                snap_df["数值"] = snap_df["数值"].map(format_number)
                st.dataframe(snap_df, hide_index=True, width='stretch')
        st.caption(f"数据最新日期：{format_date_cn(pd.Timestamp(sheet_df['date'].max()))} | 记录点：{len(sheet_df)}")


# -----------------------------
# 模块页面：涌益现货（日度 + 周度合并）
# -----------------------------
def _render_seasonal_section(df: pd.DataFrame, use_lunar: bool, key_prefix: str, frequency: str = "日度数据", metric_label: str = "", scope_label: str = "") -> None:
    """统一的季节性折线图区。若数据含多个指标/地区，先让用户选择单个对象再展示。"""
    seasonal_df = df.copy()
    if "metric" in seasonal_df.columns:
        metric_options = [v for v in seasonal_df["metric"].dropna().unique() if v]
        if len(metric_options) > 1:
            st.info(f"⚠️ 季节性图仅支持单一指标展示（当前共 {len(metric_options)} 个指标），请在下方选择一个。")
            metric_label = st.selectbox("季节性：选择指标", metric_options, key=f"{key_prefix}_seasonal_metric")
            seasonal_df = seasonal_df[seasonal_df["metric"] == metric_label].copy()
        elif len(metric_options) == 1:
            if not metric_label:
                metric_label = str(metric_options[0])
    group_field = get_group_field(seasonal_df)
    if group_field and group_field in seasonal_df.columns:
        unique_vals = [v for v in seasonal_df[group_field].dropna().unique() if v and v != "全部"]
        if len(unique_vals) > 1:
            st.info(f"⚠️ 季节性图仅支持单一{'省份' if group_field == 'province' else '城市'}展示（当前共 {len(unique_vals)} 个），请在下方选择一个。")
            scope_label = st.selectbox(
                f"季节性：选择{'省份' if group_field == 'province' else '城市'}",
                unique_vals,
                key=f"{key_prefix}_seasonal_dim",
            )
            seasonal_df = seasonal_df[seasonal_df[group_field] == scope_label].copy()
        elif len(unique_vals) == 1:
            if not scope_label:
                scope_label = str(unique_vals[0])
    # 显示当前季节性对比的指标和区域
    if metric_label or scope_label:
        subtitle_parts = [p for p in [metric_label, scope_label] if p]
        st.markdown(f"**📊 季节性对比：{'｜'.join(subtitle_parts)}**")
    col1, col2 = st.columns(2)
    with col1:
        fig, note = build_seasonal_line_chart(seasonal_df, use_lunar, frequency)
        if note:
            st.info(note)
        else:
            render_plotly(fig, key_prefix, "seasonal_line", use_lunar, frequency)
    with col2:
        fig2, note2 = build_seasonal_heatmap(seasonal_df, use_lunar, frequency)
        if note2:
            st.info(note2)
        else:
            render_plotly(fig2, key_prefix, "seasonal_heatmap", use_lunar, frequency)


def _render_standard_controls(key_prefix: str, has_region: bool = False, frequency_options: list[str] | None = None) -> tuple[str, bool, bool]:
    """统一的时间口径 / 季节性 / 农历控件，返回 (frequency, show_seasonal, use_lunar)。
    仅在开启季节性对比时显示并启用农历模式。"""
    cols = st.columns([1, 1, 1])
    options = frequency_options or FREQUENCY_OPTIONS
    with cols[0]:
        frequency = st.selectbox("时间口径", options, key=f"{key_prefix}_frequency")
    with cols[1]:
        show_seasonal = st.checkbox("季节性对比", value=False, key=f"{key_prefix}_seasonal")
    with cols[2]:
        lunar_selected = st.checkbox("农历模式", value=False, key=f"{key_prefix}_lunar", disabled=not show_seasonal)
    use_lunar = show_seasonal and lunar_selected
    return frequency, show_seasonal, use_lunar


def apply_scope_selection(df: pd.DataFrame, show_seasonal: bool, region_mode: str = "default") -> tuple[pd.DataFrame, str]:
    dimension = get_series_dimension(df)
    if dimension == "none":
        return df.copy(), "全国"
    if dimension == "province":
        preferred_scope = pick_preferred_scope_name(df["province"].dropna().unique().tolist())
        if show_seasonal or region_mode == "single":
            selected_region = st.session_state.get("selected_region_single", preferred_scope)
            scoped = df[df["province"] == selected_region].copy() if selected_region else df.copy()
            return scoped, selected_region or "全国"
        selected_regions = st.session_state.get("selected_regions", [preferred_scope] if preferred_scope else [])
        frames = [df[df["province"] == region].copy() for region in selected_regions]
        scoped = pd.concat(frames, ignore_index=True).drop_duplicates() if frames else df.copy()
        return scoped, "、".join(selected_regions) if selected_regions else (preferred_scope or "全国")
    selected_province = st.session_state.get("selected_city_province")
    province_city_df = df[df["province"] == selected_province].copy() if selected_province else df.iloc[0:0].copy()
    if show_seasonal or region_mode == "single":
        selected_city = st.session_state.get("selected_city_single")
        scoped = province_city_df[province_city_df["city"] == selected_city].copy() if selected_city else province_city_df.iloc[0:0].copy()
        return scoped, f"{selected_province}｜{selected_city}" if selected_city else (selected_province or "云南")
    selected_cities = st.session_state.get("selected_cities", get_available_group_options(province_city_df, "city"))
    scoped = province_city_df[province_city_df["city"].isin(selected_cities)].copy() if selected_cities else province_city_df.copy()
    return scoped, f"{selected_province}｜{'、'.join(selected_cities) if selected_cities else '全部城市'}" if selected_province else "云南"


def prepare_metric_compare_df(df: pd.DataFrame) -> pd.DataFrame:
    """将多指标 DataFrame 合并为可用于折线图的格式。
    若数据含多个省份，series_name 用「指标｜省份」区分，否则只用 metric。"""
    if df.empty:
        return df.copy()
    result = df.copy()
    provinces = result["province"].dropna().unique() if "province" in result.columns else []
    multi_province = len([p for p in provinces if p and p != "全部"]) > 1
    if multi_province:
        result["series_name"] = result.apply(
            lambda r: f"{r['metric']}｜{r['province']}" if pd.notna(r.get("province")) and r.get("province") and r.get("province") != "全部" else str(r["metric"]),
            axis=1,
        )
    else:
        result["series_name"] = result["metric"]
    result["display_name"] = result["series_name"]
    return result


def get_yongyi_sheet_options(numeric_df: pd.DataFrame) -> list[str]:
    available = [s for s in numeric_df["sheet"].dropna().unique() if s != "市场主流标猪肥猪均价方便作图"]
    ordered = [s for s in YONGYI_SHEET_ORDER if s in available]
    remaining = [s for s in available if s not in YONGYI_SHEET_ORDER]
    return ordered + sorted(remaining)


def get_yongyi_sheet_display_name(sheet: str) -> str:
    return YONGYI_SHEET_DISPLAY_NAMES.get(sheet, sheet)


def build_chart_context_note(sheet_label: str, metric_label: str, scope_text: str, analysis_date: pd.Timestamp | None, show_seasonal: bool, use_lunar: bool, range_text: str = "") -> str:
    parts = [f"当前视图：{sheet_label}"]
    if metric_label:
        parts.append(metric_label)
    if scope_text:
        parts.append(f"口径：{scope_text}")
    if range_text:
        parts.append(range_text)
    if analysis_date is not None and not pd.isna(analysis_date):
        parts.append(f"分析日期：{format_date_cn(pd.Timestamp(analysis_date))}")
    if show_seasonal:
        parts.append("季节性")
    if use_lunar:
        parts.append("农历模式")
    return "｜".join(parts)


def style_highlight_control(label: str, key_prefix: str) -> str:
    st.markdown(f"<div class='highlight-control-title'>🏷️ {label}</div>", unsafe_allow_html=True)
    return st.selectbox("突出显示序列", ["不突出显示"], key=f"{key_prefix}_highlight_placeholder")


def apply_series_highlight(fig: go.Figure, highlight_name: str | None) -> go.Figure:
    if not highlight_name or highlight_name == "不突出显示":
        return fig
    for trace in fig.data:
        name = getattr(trace, "name", "")
        if name == highlight_name:
            trace.update(opacity=1.0)
            if hasattr(trace, "line"):
                trace.line.width = max(getattr(trace.line, "width", 2) or 2, 3.5)
            if hasattr(trace, "marker"):
                trace.marker.size = max(getattr(trace.marker, "size", 6) or 6, 8)
        else:
            trace.update(opacity=0.22)
            if hasattr(trace, "line") and getattr(trace.line, "width", None):
                trace.line.width = max((trace.line.width or 2) * 0.8, 1)
    return fig


def render_highlight_selector(options: list[str], key_prefix: str) -> str:
    if not options:
        return ""
    st.markdown("<div class='highlight-control-box'>", unsafe_allow_html=True)
    st.markdown("<div class='highlight-control-title'>🏷️ 标签突出显示</div>", unsafe_allow_html=True)
    selected = st.selectbox("选择需要突出的序列", ["不突出显示"] + options, key=f"{key_prefix}_highlight")
    st.caption("选中后图表内对应序列会加粗，其余序列弱化。")
    st.markdown("</div>", unsafe_allow_html=True)
    return selected


def build_yongyi_daily_view(numeric_df: pd.DataFrame, sheet: str, compare_mode: str) -> tuple[pd.DataFrame, str, str]:
    sheet_df = numeric_df[numeric_df["sheet"] == sheet].copy()
    if sheet_df.empty:
        return sheet_df, "", ""

    if sheet == "价格+宰量":
        allowed = [metric for metric in YONGYI_PRICE_SLAUGHTER_ALLOWED_METRICS if metric in sheet_df["metric"].unique()]
        sheet_df = sheet_df[sheet_df["metric"].isin(allowed)].copy()
        return sheet_df, YONGYI_PRICE_METRIC, compare_mode if compare_mode in allowed else ""

    if sheet == "出栏价":
        if compare_mode == "规模场-小散户之差":
            spread_df = build_metric_spread_df(sheet_df, "规模场", "小散户", compare_mode)
            return spread_df, compare_mode, ""
        metric = compare_mode if compare_mode in OUT_PRICE_COMPARE_OPTIONS else "均价"
        return sheet_df[sheet_df["metric"] == metric].copy(), metric, ""

    if sheet == "市场主流标猪肥猪价格":
        if compare_mode == "130-140kg-90-100kg价差":
            spread_df = build_metric_spread_df(sheet_df, "130-140kg均价", "90-100kg均价", compare_mode)
            return spread_df, compare_mode, ""
        if compare_mode == "150kg以上-90-100kg价差":
            spread_df = build_metric_spread_df(sheet_df, "150kg左右均价", "90-100kg均价", compare_mode)
            return spread_df, compare_mode, ""
        metric = compare_mode if compare_mode in MAINSTREAM_PRICE_COMPARE_OPTIONS else (compare_mode or "")
        return sheet_df[sheet_df["metric"] == metric].copy(), metric, ""

    metric = compare_mode if compare_mode else (sheet_df["metric"].dropna().iloc[0] if not sheet_df["metric"].dropna().empty else "")
    return sheet_df[sheet_df["metric"] == metric].copy(), metric, ""


def configure_yongyi_compare_controls(sheet: str, sheet_df: pd.DataFrame) -> dict[str, Any]:
    result: dict[str, Any] = {
        "mode": "default",
        "current": sheet_df.copy(),
        "base_current": sheet_df.copy(),
        "metric_label": "",
        "secondary_label": "",
        "spread_df": pd.DataFrame(columns=sheet_df.columns),
        "seasonal_candidates": pd.DataFrame(columns=sheet_df.columns),
        "region_mode": "default",
        "region_message": "",
    }
    if sheet == "价格+宰量":
        volume_options = [m for m in YONGYI_PRICE_SLAUGHTER_ALLOWED_METRICS if m != YONGYI_PRICE_METRIC and m in sheet_df["metric"].unique()]
        with st.sidebar:
            ensure_single_choice("selected_metric", volume_options, volume_options[0] if volume_options else None)
            st.selectbox("选择宰量指标", volume_options, key="selected_metric")
        secondary_metric = st.session_state["selected_metric"] if volume_options else ""
        metric = YONGYI_PRICE_METRIC
        current = sheet_df[sheet_df["metric"].isin([metric, secondary_metric])].copy() if secondary_metric else sheet_df[sheet_df["metric"] == metric].copy()
        result.update({
            "current": current,
            "base_current": current.copy(),
            "metric_label": metric,
            "secondary_label": secondary_metric,
            "seasonal_candidates": current.copy(),
            "highlight_options": [x for x in current.get("series_name", pd.Series(dtype=object)).dropna().unique().tolist() if x],
        })
        return result

    if sheet in {"出栏价", "市场主流标猪肥猪价格", "散户标肥价差", "各省份均价"}:
        if sheet == "出栏价":
            left_metric_options = [m for m in OUT_PRICE_COMPARE_OPTIONS if m in sheet_df["metric"].unique()]
            default_left = [m for m in ["规模场", "小散户"] if m in left_metric_options] or left_metric_options[:1]
            left_source_df = sheet_df.copy()
            compare_key = "out"
        elif sheet == "市场主流标猪肥猪价格":
            left_metric_options = [m for m in MAINSTREAM_PRICE_COMPARE_OPTIONS if m in sheet_df["metric"].unique()]
            default_left = [m for m in ["90-100kg均价", "130-140kg均价", "150kg左右均价"] if m in left_metric_options] or left_metric_options[:1]
            left_source_df = sheet_df.copy()
            compare_key = "mainstream"
        elif sheet == "各省份均价":
            left_source_df = sheet_df.copy()
            left_metric_options = [m for m in sorted(sheet_df["metric"].dropna().unique()) if m]
            default_left = ["均价"] if "均价" in left_metric_options else left_metric_options[:1]
            compare_key = "province_avg"
        else:
            price_view_df, spread_view_df = build_scattered_fat_price_views(sheet_df)
            left_source_df = price_view_df.copy()
            left_metric_options = sorted(price_view_df["metric"].dropna().unique()) if not price_view_df.empty else []
            default_left = [m for m in ["市场散户标重猪", "150公斤猪价格", "175公斤猪价格"] if m in left_metric_options] or left_metric_options[:1]
            result["spread_source_df"] = spread_view_df.copy()
            compare_key = "fat"

        with st.sidebar:
            ensure_multi_choice(f"{compare_key}_left_metrics", left_metric_options, default_left)
            st.multiselect("选择左轴价格", left_metric_options, key=f"{compare_key}_left_metrics")
        chosen_metrics = st.session_state.get(f"{compare_key}_left_metrics", default_left)
        chosen_metrics = [m for m in chosen_metrics if m in left_metric_options]
        if not chosen_metrics and left_metric_options:
            chosen_metrics = left_metric_options[:1]
        left_metric_count = len(chosen_metrics)
        region_mode = "multi" if left_metric_count <= 1 else "single"
        if sheet == "各省份均价":
            region_message = "当前表单支持区域价差分析，可选择两个区域生成价差副轴。"
        else:
            region_message = "左轴价格选择 1 个时，区域可多选；选择 2 个及以上时，区域仅支持单选，价差也只展示该区域。"
        with st.sidebar:
            st.caption(region_message)
            st.checkbox("显示价差", value=False, key=f"{compare_key}_show_spread")
        show_spread = st.session_state.get(f"{compare_key}_show_spread", False)

        base_current = left_source_df[left_source_df["metric"].isin(chosen_metrics)].copy() if chosen_metrics else left_source_df.copy()
        current = base_current.copy()
        spread_df = pd.DataFrame(columns=left_source_df.columns)
        metric_label = "、".join(chosen_metrics)
        seasonal_candidates = current.copy()
        secondary_label = ""

        if show_spread:
            if left_metric_count <= 1:
                region_options = sorted([x for x in base_current["province"].dropna().unique() if x and x != "全部"])
                with st.sidebar:
                    ensure_single_choice(f"{compare_key}_spread_region_left", region_options, region_options[0] if region_options else None)
                    st.selectbox("价差区域1", region_options, key=f"{compare_key}_spread_region_left")
                    right_default = region_options[1] if len(region_options) > 1 else (region_options[0] if region_options else None)
                    ensure_single_choice(f"{compare_key}_spread_region_right", region_options, right_default)
                    st.selectbox("价差区域2", region_options, key=f"{compare_key}_spread_region_right")
                region_left = st.session_state.get(f"{compare_key}_spread_region_left", "")
                region_right = st.session_state.get(f"{compare_key}_spread_region_right", "")
                spread_name = f"{region_right}-{region_left}价差" if region_left and region_right else "区域价差"
                spread_df = build_region_spread_df(base_current, region_right, region_left, spread_name)
                secondary_label = spread_name if not spread_df.empty else ""
            else:
                selected_scope = pick_preferred_scope_name(base_current["province"].dropna().unique().tolist())
                if get_series_dimension(base_current) == "city":
                    selected_scope = st.session_state.get("selected_city_single", selected_scope)
                elif region_mode == "single":
                    selected_scope = st.session_state.get("selected_region_single", selected_scope)
                filtered_current = base_current[base_current["province"] == selected_scope].copy() if selected_scope else base_current.copy()
                with st.sidebar:
                    ensure_single_choice(f"{compare_key}_spread_metric_left", chosen_metrics, chosen_metrics[0] if chosen_metrics else None)
                    st.selectbox("价差左轴价格1", chosen_metrics, key=f"{compare_key}_spread_metric_left")
                    right_default = chosen_metrics[1] if len(chosen_metrics) > 1 else (chosen_metrics[0] if chosen_metrics else None)
                    ensure_single_choice(f"{compare_key}_spread_metric_right", chosen_metrics, right_default)
                    st.selectbox("价差左轴价格2", chosen_metrics, key=f"{compare_key}_spread_metric_right")
                metric_left = st.session_state.get(f"{compare_key}_spread_metric_left", "")
                metric_right = st.session_state.get(f"{compare_key}_spread_metric_right", "")
                spread_name = f"{selected_scope}｜{metric_right}-{metric_left}价差" if metric_left and metric_right and selected_scope else (f"{metric_right}-{metric_left}价差" if metric_left and metric_right else "指标价差")
                spread_df = build_metric_spread_df(filtered_current, metric_right, metric_left, spread_name)
                secondary_label = spread_name if not spread_df.empty else ""
        if not spread_df.empty:
            seasonal_candidates = pd.concat([current.copy(), spread_df.copy()], ignore_index=True)

        highlight_source = current.copy()
        if not spread_df.empty:
            highlight_source = pd.concat([highlight_source, spread_df.copy()], ignore_index=True)
        highlight_options = [x for x in highlight_source.get("series_name", pd.Series(dtype=object)).dropna().unique().tolist() if x]
        result.update({
            "mode": "compare",
            "current": current,
            "base_current": base_current,
            "metric_label": metric_label,
            "secondary_label": secondary_label,
            "spread_df": spread_df,
            "seasonal_candidates": seasonal_candidates,
            "region_mode": region_mode,
            "region_message": region_message,
            "highlight_options": highlight_options,
        })
        return result

    metric_options = sorted(sheet_df["metric"].dropna().unique())
    with st.sidebar:
        ensure_single_choice("selected_metric", metric_options, metric_options[0] if metric_options else None)
        st.selectbox("选择指标", metric_options, key="selected_metric")
    metric = st.session_state.get("selected_metric", metric_options[0] if metric_options else "")
    current = sheet_df[sheet_df["metric"] == metric].copy() if metric else sheet_df.copy()
    result.update({
        "current": current,
        "base_current": current.copy(),
        "metric_label": metric,
        "seasonal_candidates": current.copy(),
        "highlight_options": [x for x in current.get("series_name", pd.Series(dtype=object)).dropna().unique().tolist() if x],
    })
    return result


def _resolve_data_path(pattern: str, label: str) -> str:
    """按文件名模式匹配最新文件，优先使用平台数据目录（与 PLATFORM_DATA_DIR 优先级一致）"""
    # 优先搜索 PLATFORM_DATA_DIR（本地平台数据 > data/）
    f = _find_latest_file(pattern, PLATFORM_DATA_DIR)
    if f:
        return str(f)
    # 兜底：尝试另一个目录
    fallback_dir = _REPO_DATA if PLATFORM_DATA_DIR == _LOCAL_DATA else _LOCAL_DATA
    if fallback_dir.exists():
        f = _find_latest_file(pattern, fallback_dir)
        if f:
            return str(f)
    # 诊断：列出可用文件
    available = list(PLATFORM_DATA_DIR.glob("*.xlsx")) if PLATFORM_DATA_DIR.exists() else []
    if not available and _REPO_DATA.exists():
        available = list(_REPO_DATA.glob("*.xlsx"))
    if available:
        names = ", ".join(f.name[:30] for f in available[:10])
        st.sidebar.error(f"未找到{label}数据，可用文件: {names}")
    else:
        st.sidebar.error(f"数据目录为空，请检查平台数据路径")
    return ""


# 独立日度模块
def render_yongyi_daily_module() -> None:
    with st.sidebar:
        st.header("📂 涌益日度数据源")
        yongyi_path = _resolve_data_path(r"涌益.*日度|日度数据", "涌益日度")
    try:
        numeric_df, metadata_df, parse_log = build_yongyi_dataset_from_path(yongyi_path)
    except Exception as exc:
        st.error(f"日度数据载入失败：{exc}")
        return
    if numeric_df.empty:
        st.warning("日度数据为空，请检查文件路径。")
        return

    render_module_lead("📌 日度口径｜调研来源：涌益咨询。综合摘要优先使用全国均价，若无全国则取云南；出栏价口径仅保留 规模场 / 小散户 / 均价；交割地出栏价不区分体重段。")
    target_date = st.date_input("综合分析日期", value=numeric_df["date"].max().date(), min_value=numeric_df["date"].min().date(), max_value=numeric_df["date"].max().date(), key="yongyi_target_date")
    global_summary_payload = build_yongyi_global_summary(numeric_df, pd.Timestamp(target_date))
    if global_summary_payload.get("ok"):
        if global_summary_payload.get("cards"):
            render_metric_cards(global_summary_payload["cards"])
        render_summary_box(global_summary_payload["title"], format_date_cn(global_summary_payload["date"]), global_summary_payload["brief"])
        render_signal_messages("摘要明细", global_summary_payload["lines"], "暂无可拆解摘要。")

    render_section_header("🔍 数据筛选")
    with st.sidebar:
        st.header("日度筛选条件")
        sheet_options = get_yongyi_sheet_options(numeric_df)
        sheet_display_map = {get_yongyi_sheet_display_name(s): s for s in sheet_options}
        selected_sheet_label = st.selectbox("选择表单", list(sheet_display_map.keys()), key="selected_sheet_label")
    sheet = sheet_display_map[selected_sheet_label]
    sheet_label = get_yongyi_sheet_display_name(sheet)

    sheet_df = numeric_df[numeric_df["sheet"] == sheet].copy()
    selection_state = configure_yongyi_compare_controls(sheet, sheet_df)
    current = selection_state["current"]
    base_current = selection_state["base_current"]
    metric = selection_state.get("metric_label", "")
    secondary_metric = selection_state.get("secondary_label", "")
    spread_df_current = selection_state.get("spread_df", pd.DataFrame(columns=sheet_df.columns))
    seasonal_candidates_df = selection_state.get("seasonal_candidates", current.copy())
    highlight_options = selection_state.get("highlight_options", [])
    region_mode = selection_state.get("region_mode", "default")
    region_message = selection_state.get("region_message", "")

    dimension = get_series_dimension(base_current)

    frequency, show_seasonal, use_lunar = _render_standard_controls("yongyi_daily")
    _, _, base_current = build_cn_date_range_selector(base_current, "yongyi_daily_chart", "📅 数据日期范围")
    range_text = st.session_state.get("yongyi_daily_chart_range_text", "")
    range_start = base_current["date"].min() if not base_current.empty else pd.Timestamp.today()
    range_end = base_current["date"].max() if not base_current.empty else pd.Timestamp.today()

    with st.sidebar:
        if region_message:
            st.caption(region_message)
        if dimension == "none":
            st.caption("当前表单为全国单一口径，地区筛选已自动隐藏。")
        elif dimension == "province":
            preferred_scope = pick_preferred_scope_name(base_current["province"].dropna().unique().tolist())
            region_options = sorted([x for x in base_current["province"].dropna().unique() if x])
            if show_seasonal or region_mode == "single":
                ensure_single_choice("selected_region_single", region_options, preferred_scope)
                st.selectbox("选择区域", region_options, key="selected_region_single")
            else:
                default_regions = [preferred_scope] if preferred_scope else region_options[:2]
                ensure_multi_choice("selected_regions", region_options, default_regions)
                st.multiselect("选择区域（可多选）", region_options, key="selected_regions")
        else:
            province_options = sorted([x for x in base_current["province"].dropna().unique() if x and x != "全部"])
            preferred_city_province = "云南" if "云南" in province_options else (province_options[0] if province_options else None)
            ensure_single_choice("selected_city_province", province_options, preferred_city_province)
            if province_options:
                st.selectbox("选择省份", province_options, key="selected_city_province")
            city_province = st.session_state.get("selected_city_province")
            if city_province:
                province_city_df = base_current[base_current["province"] == city_province].copy()
                city_options = get_available_group_options(province_city_df, "city")
                if show_seasonal or region_mode == "single":
                    ensure_single_choice("selected_city_single", city_options, city_options[0] if city_options else None)
                    if city_options:
                        st.selectbox("选择城市", city_options, key="selected_city_single")
                else:
                    ensure_multi_choice("selected_cities", city_options, city_options)
                    if city_options:
                        st.multiselect("选择城市（可多选）", city_options, key="selected_cities")
                st.caption("交割地出栏价不区分体重段，按地市均价展示。")

    current, selected_scope_text = apply_scope_selection(base_current, show_seasonal, region_mode=region_mode)

    if current.empty:
        st.warning("当前筛选条件下没有数据，请重新选择。")
        return

    min_date = current["date"].min().date()
    max_date_cur = current["date"].max().date()
    sheet_target_date = st.date_input("视图分析日期", value=min(max_date_cur, pd.Timestamp(target_date).date()), min_value=min_date, max_value=max_date_cur, key="sheet_target_date")

    st.markdown(f"<div class='mini-note'>{build_chart_context_note(sheet_label, metric + ((f'｜副轴：{secondary_metric}') if secondary_metric else ''), selected_scope_text, pd.Timestamp(sheet_target_date), show_seasonal, use_lunar, range_text)}</div>", unsafe_allow_html=True)

    allow_region_charts = supports_region_charts(base_current, sheet)
    tab1, tab2, tab3 = st.tabs(["📈 趋势 / 季节性", "🗺️ 地域对比", "📋 日期深度总结"])
    with tab1:
        highlight_name = render_highlight_selector(highlight_options, f"yongyi_daily_{sheet}")
        if show_seasonal and not seasonal_candidates_df.empty:
            _render_seasonal_section(seasonal_candidates_df.copy(), use_lunar, f"yongyi_daily_{sheet}_{metric}", frequency)
        else:
            fig_main: go.Figure | None = None
            if sheet == "价格+宰量" and secondary_metric:
                price_all = filter_df_by_date_window(sheet_df[sheet_df["metric"] == YONGYI_PRICE_METRIC].copy(), range_start, range_end)
                national_price = price_all[price_all["province"].apply(lambda p: is_national_scope(str(p)) if pd.notna(p) else False)]
                price_df_for_chart = national_price if not national_price.empty else price_all
                volume_df = filter_df_by_date_window(sheet_df[sheet_df["metric"] == secondary_metric].copy(), range_start, range_end)
                fig_main = build_dual_axis_line_chart(price_df_for_chart, volume_df, frequency, YONGYI_PRICE_METRIC, secondary_metric, use_lunar=False)
            elif selection_state.get("mode") == "compare":
                main_df = prepare_metric_compare_df(current.copy()) if not current.empty else current
                spread_plot_df = filter_df_by_date_window(spread_df_current.copy(), range_start, range_end) if not spread_df_current.empty else spread_df_current
                if not main_df.empty and not spread_plot_df.empty:
                    fig_main = build_trend_line_chart_with_spread(main_df, prepare_metric_compare_df(spread_plot_df.copy()), frequency, selected_scope_text, use_lunar=False)
                elif not main_df.empty:
                    fig_main = build_trend_line_chart(main_df, frequency, selected_scope_text, use_lunar=False)
                elif not spread_plot_df.empty:
                    fig_main = build_multi_series_line_chart(prepare_metric_compare_df(spread_plot_df.copy()), frequency, f"{sheet_label}｜{selected_scope_text}", hover_title="价差", use_lunar=False)
            else:
                fig_main = build_trend_line_chart(current, frequency, selected_scope_text, use_lunar=False)
            if fig_main is not None:
                render_plotly(apply_series_highlight(fig_main, highlight_name), "yongyi_d", "line", sheet, metric, selected_scope_text, frequency, highlight_name)

            if allow_region_charts:
                col1, col2 = st.columns(2)
                with col1:
                    fig, note = build_rank_bar_chart(current, pd.Timestamp(sheet_target_date))
                    if note:
                        st.info(note)
                    render_plotly(fig, "yongyi_d", "rank", sheet, metric, selected_scope_text)
                with col2:
                    fig, note = build_regular_heatmap(current, pd.Timestamp(sheet_target_date))
                    if note:
                        st.info(note)
                    render_plotly(fig, "yongyi_d", "heatmap", sheet, metric, selected_scope_text)

    with tab2:
        current_dimension = get_series_dimension(current)
        if not allow_region_charts:
            st.info("当前表单不适用地区排行图、地区热力图或地图展示。")
        elif current_dimension == "province":
            col1, col2 = st.columns(2)
            with col1:
                fig, note = build_map_chart(current, pd.Timestamp(sheet_target_date))
                if note:
                    st.info(note)
                render_plotly(fig, "yongyi_d", "map", sheet, metric, selected_scope_text)
            with col2:
                fig, note = build_box_chart(current)
                if note:
                    st.info(note)
                render_plotly(fig, "yongyi_d", "box", sheet, metric, selected_scope_text)
            rank_fig, rank_note = build_rank_bar_chart(current, pd.Timestamp(sheet_target_date))
            if rank_note:
                st.info(rank_note)
            render_plotly(rank_fig, "yongyi_d", "rank_full", sheet, metric, selected_scope_text)
        elif current_dimension == "city":
            col1, col2 = st.columns(2)
            with col1:
                fig, note = build_regular_heatmap(current, pd.Timestamp(sheet_target_date))
                if note:
                    st.info(note)
                render_plotly(fig, "yongyi_d", "heatmap_city", sheet, metric, selected_scope_text)
            with col2:
                fig, note = build_box_chart(current)
                if note:
                    st.info(note)
                render_plotly(fig, "yongyi_d", "box_city", sheet, metric, selected_scope_text)
            rank_fig, rank_note = build_rank_bar_chart(current, pd.Timestamp(sheet_target_date))
            if rank_note:
                st.info(rank_note)
            render_plotly(rank_fig, "yongyi_d", "rank_city", sheet, metric, selected_scope_text)
        else:
            st.info("当前表单为全国单一口径，地域对比图表不适用。")

    with tab3:
        payload = build_summary_payload(current, pd.Timestamp(sheet_target_date))
        render_summary(payload)

    with st.expander("解析日志"):
        st.dataframe(parse_log, width='stretch')
        if not metadata_df.empty:
            st.caption("交割地附加元数据仅保留升贴水等信息，已忽略交易均重口径。")
            st.dataframe(metadata_df.head(30), width='stretch')


# 独立周度模块
def render_yongyi_weekly_module() -> None:
    with st.sidebar:
        st.header("📂 涌益周度数据源")
        weekly_path = _resolve_data_path(r"涌益.*周度|周度数据", "涌益周度")
    try:
        weekly_df, weekly_meta = build_weekly_dataset_from_path(weekly_path)
    except Exception as exc:
        st.error(f"周度数据载入失败：{exc}")
        return

    render_module_lead("📌 周度口径｜调研来源：涌益咨询周度数据库。已自动跳过停更超 60 天及说明类 sheet；区间报价取均值绘图，悬停展示原区间。")

    if weekly_df.empty:
        st.warning("周度数据为空，请检查文件路径。")
        return

    target_date_input = st.date_input(
        "周度分析日期",
        value=weekly_df["date"].max().date(),
        min_value=weekly_df["date"].min().date(),
        max_value=weekly_df["date"].max().date(),
        key="weekly_target_date",
    )
    render_weekly_section(weekly_df, target_date_input)

    with st.expander("周度数据解析元信息"):
        if not weekly_meta.empty:
            st.dataframe(weekly_meta, width='stretch')


def _render_futures_tab(series_df: pd.DataFrame, tab_key: str, title: str, hover_title: str, contracts: list[str], default_n: int = 4, allow_single_seasonal: bool = True) -> None:
    """统一期货 tab 渲染：时间口径 / 季节性 / 农历 + 趋势图 + 季节性折线图（无热力图）。"""
    if series_df.empty:
        st.info(f"当前没有可展示的{title}数据。")
        return
    all_contracts = _sorted_contracts(series_df["contract"].dropna().unique().tolist()) or contracts
    default_sel = all_contracts[:default_n] if len(all_contracts) >= default_n else all_contracts
    selected = st.multiselect(f"选择合约", all_contracts, default=default_sel, key=f"{tab_key}_contracts")
    frequency, show_seasonal, use_lunar = _render_standard_controls(tab_key)
    chart_df = series_df[series_df["contract"].isin(selected)].copy() if selected else series_df.copy()
    if chart_df.empty:
        st.info("所选合约暂无数据。")
        return
    render_plotly(build_multi_series_line_chart(chart_df, frequency, title, hover_title=hover_title), "futures", tab_key, tuple(selected), frequency)
    if show_seasonal and allow_single_seasonal:
        single = st.selectbox("选择单合约查看季节性", all_contracts, key=f"{tab_key}_single")
        single_df = series_df[series_df["contract"] == single].copy()
        if not single_df.empty:
            _render_seasonal_section(single_df, use_lunar, f"futures_{tab_key}_{single}")


def render_yongyi_basis_futures_tab(futures_df: pd.DataFrame, yongyi_numeric_df: pd.DataFrame, yongyi_meta_df: pd.DataFrame, target_date: pd.Timestamp) -> None:
    price_df = futures_series(futures_df, "收盘价", "收盘价")
    contracts = get_active_lh_contracts(price_df, target_date)
    delivery_df = yongyi_numeric_df[(yongyi_numeric_df["sheet"] == "交割地市出栏价") & (yongyi_numeric_df["metric"] == "交割地出栏价")].copy()
    if price_df.empty or delivery_df.empty:
        st.info("缺少期货收盘价或涌益交割地市数据，暂无法计算基差（涌益版）。")
        return
    selected_contract = ""
    col1, col2, col3 = st.columns(3)
    with col1:
        selected_contract = st.selectbox("选择LH合约", contracts, key="yongyi_basis_contract") if contracts else ""
    city_basis_df = build_yongyi_city_basis_all_df(futures_df, yongyi_numeric_df, yongyi_meta_df, selected_contract) if selected_contract else pd.DataFrame()
    province_options = sorted([x for x in city_basis_df["province"].dropna().unique().tolist() if x and x != "全部"]) if not city_basis_df.empty else []
    ensure_single_choice("yongyi_basis_province", province_options, province_options[0] if province_options else None)
    with col2:
        selected_province = st.selectbox("选择省份", province_options, key="yongyi_basis_province") if province_options else ""
    city_options = sorted(city_basis_df[city_basis_df["province"] == selected_province]["city"].dropna().unique().tolist()) if selected_province and not city_basis_df.empty else []
    ensure_single_choice("yongyi_basis_city", city_options, city_options[0] if city_options else None)
    with col3:
        selected_city = st.selectbox("选择交割地市", city_options, key="yongyi_basis_city") if city_options else ""

    if not selected_contract:
        st.info("当前分析日期下暂无可用 LH 合约。")
        return
    if city_basis_df.empty:
        st.info("当前合约暂无可计算的涌益区域基差数据。")
        return

    if selected_contract:
        render_section_header("📍 区域基差（涌益）")
        region_basis_df = build_yongyi_region_basis_df(futures_df, yongyi_numeric_df, yongyi_meta_df, selected_contract)
        if region_basis_df.empty:
            st.info("当前合约暂无可用的区域基差数据。")
        else:
            region_names = sorted(region_basis_df["series_name"].dropna().unique().tolist())
            default_regions = region_names[: min(6, len(region_names))]
            selected_regions = st.multiselect("选择区域", region_names, default=default_regions, key="yongyi_region_basis_series")
            region_chart_df = region_basis_df[region_basis_df["series_name"].isin(selected_regions)].copy() if selected_regions else region_basis_df.copy()
            region_frequency, region_show_seasonal, region_use_lunar = _render_standard_controls("yongyi_region_basis")
            if region_chart_df.empty:
                st.info("所选区域暂无数据。")
            else:
                render_plotly(build_multi_series_line_chart(region_chart_df, region_frequency, f"{selected_contract} 各区域基差走势", hover_title="区域基差", use_lunar=region_use_lunar), "futures", "yongyi_region_basis", selected_contract, tuple(selected_regions), region_frequency)
                region_rank_df, region_snap_date = build_rank_df(region_basis_df, pd.Timestamp(target_date))
                if not region_rank_df.empty:
                    top_regions = region_rank_df.head(5)
                    bottom_regions = region_rank_df.sort_values("value", ascending=True).head(5)
                    col_top, col_bottom = st.columns(2)
                    with col_top:
                        fig = px.bar(top_regions.sort_values("value", ascending=True), x="value", y="series_name", orientation="h", color="value", color_continuous_scale="Reds")
                        fig.update_traces(hovertemplate="区域：%{y}<br>基差：%{x:.2f}<extra></extra>")
                        fig.update_layout(title=f"高基差区域｜{format_date_cn(region_snap_date)}", xaxis_title="基差", yaxis_title="区域", coloraxis_showscale=False, template="plotly_white")
                        render_plotly(fig, "futures", "yongyi_region_basis_top", selected_contract, format_date_cn(region_snap_date))
                    with col_bottom:
                        fig = px.bar(bottom_regions.sort_values("value", ascending=False), x="value", y="series_name", orientation="h", color="value", color_continuous_scale="Greens")
                        fig.update_traces(hovertemplate="区域：%{y}<br>基差：%{x:.2f}<extra></extra>")
                        fig.update_layout(title=f"低基差区域｜{format_date_cn(region_snap_date)}", xaxis_title="基差", yaxis_title="区域", coloraxis_showscale=False, template="plotly_white")
                        render_plotly(fig, "futures", "yongyi_region_basis_bottom", selected_contract, format_date_cn(region_snap_date))
                render_summary(build_summary_payload(region_basis_df, pd.Timestamp(target_date)))
                if region_show_seasonal and region_names:
                    selected_region_one = st.selectbox("选择单一区域查看季节性", region_names, key="yongyi_region_basis_single")
                    seasonal_region_df = region_basis_df[region_basis_df["series_name"] == selected_region_one].copy()
                    if not seasonal_region_df.empty:
                        _render_seasonal_section(seasonal_region_df, region_use_lunar, f"futures_yongyi_region_basis_{selected_contract}_{selected_region_one}", region_frequency)

    if not selected_contract or not selected_province or not selected_city:
        st.info("请选择合约、省份和交割地市。")
        return
    snapshot = build_yongyi_basis_snapshot(futures_df, yongyi_numeric_df, yongyi_meta_df, selected_contract, selected_province, selected_city, pd.Timestamp(target_date))
    if not snapshot.get("ok"):
        st.info(snapshot.get("message", "暂无可用数据。"))
        return
    cards = [
        {"label": "收盘价", "value": format_number(snapshot.get("close_price")), "extra": f"{snapshot['contract']}｜元/吨"},
        {"label": "交割地市价格", "value": format_number(snapshot.get("spot_price")), "extra": f"{snapshot['province']} {snapshot['city']}｜元/公斤"},
        {"label": "升贴水", "value": format_number(snapshot.get("premium")), "extra": classify_lh_contract_bucket(snapshot['contract']) or "未匹配"},
        {"label": "基差（涌益版）", "value": format_number(snapshot.get("basis_value")), "extra": f"分析日：{format_date_cn(snapshot['date'])}"},
    ]
    render_metric_cards(cards)
    render_summary_box(
        "基差（涌益版）",
        f"{snapshot['contract']}｜{snapshot['province']} {snapshot['city']}｜{format_date_cn(snapshot['date'])}",
        f"公式：收盘价 {format_number(snapshot.get('close_price'))} - 交割地市价格 {format_number(snapshot.get('spot_price'))} × 1000 + 升贴水 {format_number(snapshot.get('premium'))} = {format_number(snapshot.get('basis_value'))}",
    )
    frequency, show_seasonal, use_lunar = _render_standard_controls("yongyi_basis")
    render_plotly(build_multi_series_line_chart(snapshot["basis_df"], frequency, "基差（涌益版）走势", hover_title="基差", use_lunar=False), "futures", "yongyi_basis", selected_contract, selected_city, frequency)
    if show_seasonal:
        _render_seasonal_section(snapshot["basis_df"], use_lunar, f"futures_yongyi_basis_{selected_contract}_{selected_city}", frequency)
    render_summary(build_summary_payload(snapshot["basis_df"], snapshot["date"]))


# -----------------------------
# 期货深度分析与建议生成
# -----------------------------
def _compute_ma(series: pd.Series, n: int) -> pd.Series:
    return series.rolling(n, min_periods=1).mean()


def _safe_pct_change(current: float | None, previous: float | None) -> float | None:
    if current is None or previous is None:
        return None
    if pd.isna(current) or pd.isna(previous) or previous == 0:
        return None
    return (float(current) / float(previous) - 1) * 100


def _latest_series_value(series_df: pd.DataFrame, snap_date: pd.Timestamp) -> float | None:
    if series_df.empty:
        return None
    temp = series_df[series_df["date"] <= snap_date].sort_values("date")
    if temp.empty:
        return None
    value = temp.iloc[-1]["value"]
    return None if pd.isna(value) else float(value)


def _compute_futures_signal_score(value: float, positive_threshold: float, negative_threshold: float, strong_threshold: float | None = None) -> int:
    if pd.isna(value):
        return 0
    if value >= positive_threshold:
        if strong_threshold is not None and value >= strong_threshold:
            return 2
        return 1
    if value <= negative_threshold:
        if strong_threshold is not None and value <= -abs(strong_threshold):
            return -2
        return -1
    return 0


def _score_to_bias(score: int) -> str:
    if score >= 5:
        return "偏多"
    if score >= 2:
        return "观望偏多"
    if score <= -5:
        return "偏空"
    if score <= -2:
        return "观望偏空"
    return "中性"


def _basis_percentile_snapshot(futures_df: pd.DataFrame, contract: str, snap_date: pd.Timestamp) -> tuple[float | None, float | None, str]:
    basis_df = futures_series(futures_df, "期现基差", "基差")
    if basis_df.empty:
        return None, None, "暂无基差数据"
    contract_basis = basis_df[basis_df["contract"] == contract].copy()
    if contract_basis.empty:
        contract_basis = basis_df[basis_df["series_name"].astype(str).str.contains(contract, na=False)].copy()
    if contract_basis.empty:
        return None, None, "暂无该合约基差数据"
    current = _latest_series_value(contract_basis, snap_date)
    if current is None:
        return None, None, "暂无该日期基差数据"
    _, pct = percentile_position_text(current, contract_basis[contract_basis["date"] <= snap_date]["value"])
    if pct is None:
        return current, None, "基差历史样本不足"
    if pct >= 0.8:
        return current, pct, "基差处于高位"
    if pct <= 0.2:
        return current, pct, "基差处于低位"
    return current, pct, "基差处于中性区间"


def _contract_spread_snapshot(price_df: pd.DataFrame, contract: str, snap_date: pd.Timestamp) -> tuple[str | None, float | None, float | None, str]:
    contracts = get_active_lh_contracts(price_df, snap_date)
    if contract not in contracts:
        return None, None, None, "暂无可比价差"
    idx = contracts.index(contract)
    pair: tuple[str, str] | None = None
    if idx + 1 < len(contracts):
        pair = (contract, contracts[idx + 1])
    elif idx - 1 >= 0:
        pair = (contracts[idx - 1], contract)
    if pair is None:
        return None, None, None, "暂无可比价差"
    spread_df = build_spread_df(price_df, pair[0], pair[1])
    if spread_df.empty:
        return f"{pair[0]}-{pair[1]}", None, None, "价差数据不足"
    current = _latest_series_value(spread_df, snap_date)
    if current is None:
        return f"{pair[0]}-{pair[1]}", None, None, "价差数据不足"
    _, pct = percentile_position_text(current, spread_df[spread_df["date"] <= snap_date]["value"])
    if pct is None:
        return f"{pair[0]}-{pair[1]}", current, None, "价差历史样本不足"
    if pct >= 0.8:
        note = "价差处于历史高位"
    elif pct <= 0.2:
        note = "价差处于历史低位"
    else:
        note = "价差处于中性区间"
    return f"{pair[0]}-{pair[1]}", current, pct, note


def build_futures_contract_analysis(
    futures_df: pd.DataFrame,
    contract: str,
    target_date: pd.Timestamp,
    yongyi_numeric_df: pd.DataFrame | None = None,
    yongyi_meta_df: pd.DataFrame | None = None,
) -> dict:
    """对指定合约做技术面 + 资金面 + 基差面综合分析，生成结构化建议。"""
    result: dict = {
        "contract": contract,
        "date": target_date,
        "ok": False,
        "cards": [],
        "signals": [],
        "advice": "",
        "strong_signals": [],
        "risk_signals": [],
        "watch_signals": [],
        "bias": "中性",
        "score": 0,
    }

    price_df = futures_series(futures_df, "收盘价", "收盘价")
    price_s = price_df[price_df["contract"] == contract].sort_values("date")
    if price_s.empty:
        result["advice"] = f"合约 {contract} 暂无收盘价数据，无法分析。"
        return result

    snap_date = get_snapshot_date(price_s, target_date)
    if snap_date is None:
        result["advice"] = "所选日期无数据。"
        return result

    hist = price_s[price_s["date"] <= snap_date].copy()
    current_price = hist.iloc[-1]["value"] if not hist.empty else np.nan
    if pd.isna(current_price):
        result["advice"] = "收盘价数据为空。"
        return result

    prices = hist["value"].reset_index(drop=True)
    ma5_series = _compute_ma(prices, 5)
    ma10_series = _compute_ma(prices, 10)
    ma20_series = _compute_ma(prices, 20)
    ma60_series = _compute_ma(prices, 60)
    ma5 = ma5_series.iloc[-1]
    ma10 = ma10_series.iloc[-1]
    ma20 = ma20_series.iloc[-1]
    ma60 = ma60_series.iloc[-1]
    ma20_prev = ma20_series.iloc[-6] if len(ma20_series) >= 6 else np.nan

    momentum_5d = current_price - prices.iloc[-6] if len(prices) >= 6 else np.nan
    change_5d_pct = _safe_pct_change(current_price, prices.iloc[-6] if len(prices) >= 6 else None)
    change_10d_pct = _safe_pct_change(current_price, prices.iloc[-11] if len(prices) >= 11 else None)
    breakout_high = bool(len(prices) >= 20 and current_price >= hist.tail(20)["value"].max())
    breakout_low = bool(len(prices) >= 20 and current_price <= hist.tail(20)["value"].min())
    returns = prices.pct_change().dropna()
    vol_20 = returns.tail(20).std() * 100 if len(returns) >= 5 else np.nan
    _, pct = percentile_position_text(current_price, hist["value"])
    pct_label = f"{pct:.0%}" if pct is not None else "N/A"

    if pd.notna(ma5) and pd.notna(ma20):
        if ma5 > ma20 and current_price > ma20:
            trend_label = "短期多头趋势"
            trend_signal = "多"
        elif ma5 < ma20 and current_price < ma20:
            trend_label = "短期空头趋势"
            trend_signal = "空"
        else:
            trend_label = "震荡整理"
            trend_signal = "中性"
    else:
        trend_label = "数据不足"
        trend_signal = "中性"

    oi_df = futures_series(futures_df, "持仓量", "持仓量")
    oi_s = oi_df[oi_df["contract"] == contract].sort_values("date")
    oi_current = np.nan
    oi_delta = np.nan
    oi_ma5 = np.nan
    if not oi_s.empty:
        oi_hist = oi_s[oi_s["date"] <= snap_date]
        if not oi_hist.empty:
            oi_current = oi_hist.iloc[-1]["value"]
            oi_ma5 = oi_hist.tail(5)["value"].mean()
            if len(oi_hist) >= 2:
                oi_delta = oi_current - oi_hist.iloc[-2]["value"]

    vol_df = futures_series(futures_df, "成交量", "成交量")
    vol_s = vol_df[vol_df["contract"] == contract].sort_values("date")
    vol_current = np.nan
    vol_ma5 = np.nan
    if not vol_s.empty:
        vol_hist = vol_s[vol_s["date"] <= snap_date]
        if not vol_hist.empty:
            vol_current = vol_hist.iloc[-1]["value"]
            vol_ma5 = vol_hist.tail(5)["value"].mean()

    net_df = futures_series(futures_df, "前20多空净持仓", "前20净持仓")
    net_s = net_df[net_df["contract"] == contract].sort_values("date")
    net_current = np.nan
    net_delta_5d = np.nan
    net_trend = "N/A"
    if not net_s.empty:
        net_hist = net_s[net_s["date"] <= snap_date]
        if not net_hist.empty:
            net_current = net_hist.iloc[-1]["value"]
            if len(net_hist) >= 2:
                net_prev = net_hist.iloc[-2]["value"]
                if pd.notna(net_current) and pd.notna(net_prev):
                    net_trend = "增多" if net_current > net_prev else ("减空" if net_current < net_prev else "持平")
            if len(net_hist) >= 5:
                base_net = net_hist.iloc[-5]["value"]
                if pd.notna(base_net) and pd.notna(net_current):
                    net_delta_5d = net_current - base_net

    basis_current, basis_pct, basis_note = _basis_percentile_snapshot(futures_df, contract, snap_date)
    spread_name, spread_current, spread_pct, spread_note = _contract_spread_snapshot(price_df, contract, snap_date)
    yongyi_basis_snapshot = build_yongyi_city_basis_snapshot(futures_df, yongyi_numeric_df, yongyi_meta_df, snap_date, contract)

    trend_score = 0
    if pd.notna(ma5) and pd.notna(ma20):
        trend_score += 2 if ma5 > ma20 else -2
    if pd.notna(ma20) and current_price > ma20:
        trend_score += 1
    elif pd.notna(ma20):
        trend_score -= 1
    if pd.notna(ma60) and current_price > ma60:
        trend_score += 1
    elif pd.notna(ma60):
        trend_score -= 1
    if pd.notna(ma20_prev) and pd.notna(ma20):
        if ma20 > ma20_prev:
            trend_score += 1
        elif ma20 < ma20_prev:
            trend_score -= 1

    momentum_score = 0
    if change_5d_pct is not None:
        momentum_score += _compute_futures_signal_score(change_5d_pct, 1.5, -1.5, 3.5)
    if change_10d_pct is not None:
        momentum_score += _compute_futures_signal_score(change_10d_pct, 3.0, -3.0, 6.0)
    if breakout_high:
        momentum_score += 1
    if breakout_low:
        momentum_score -= 1

    flow_score = 0
    vol_ratio = None if pd.isna(vol_current) or pd.isna(vol_ma5) or vol_ma5 == 0 else float(vol_current / vol_ma5)
    oi_ratio = None if pd.isna(oi_current) or pd.isna(oi_ma5) or oi_ma5 == 0 else float(oi_current / oi_ma5)
    if vol_ratio is not None:
        if vol_ratio >= 1.4 and pd.notna(momentum_5d) and momentum_5d > 0:
            flow_score += 2
        elif vol_ratio >= 1.4 and pd.notna(momentum_5d) and momentum_5d < 0:
            flow_score -= 2
        elif vol_ratio <= 0.7:
            flow_score -= 1 if trend_signal == "多" else 0
    if pd.notna(oi_delta):
        if oi_delta > 0 and pd.notna(momentum_5d) and momentum_5d > 0:
            flow_score += 2
        elif oi_delta > 0 and pd.notna(momentum_5d) and momentum_5d < 0:
            flow_score -= 2
        elif oi_delta < 0 and pd.notna(momentum_5d) and momentum_5d > 0:
            flow_score -= 1
        elif oi_delta < 0 and pd.notna(momentum_5d) and momentum_5d < 0:
            flow_score += 1
    if pd.notna(net_current):
        if net_current > 0:
            flow_score += 1
        elif net_current < 0:
            flow_score -= 1
    if pd.notna(net_delta_5d):
        if net_delta_5d > 0:
            flow_score += 1
        elif net_delta_5d < 0:
            flow_score -= 1

    valuation_score = 0
    risk_score = 0
    yongyi_score = 0
    if pct is not None:
        if pct >= 0.85:
            valuation_score -= 2
            risk_score += 2
        elif pct >= 0.7:
            valuation_score -= 1
            risk_score += 1
        elif pct <= 0.15:
            valuation_score += 2
        elif pct <= 0.3:
            valuation_score += 1
    if basis_pct is not None:
        if basis_pct >= 0.8:
            valuation_score -= 1
        elif basis_pct <= 0.2:
            valuation_score += 1
    if spread_pct is not None and (spread_pct >= 0.85 or spread_pct <= 0.15):
        risk_score += 1
    if pd.notna(vol_20) and vol_20 >= 3.0:
        risk_score += 1
    if yongyi_basis_snapshot:
        high_city = yongyi_basis_snapshot.get("high")
        low_city = yongyi_basis_snapshot.get("low")
        extreme_city = yongyi_basis_snapshot.get("extreme")
        if high_city and low_city:
            yongyi_score += 1
        elif extreme_city and extreme_city.get("percentile") is not None:
            if extreme_city["percentile"] <= 0.2:
                yongyi_score += 1
            elif extreme_city["percentile"] >= 0.8:
                yongyi_score -= 1

    total_score = trend_score + momentum_score + flow_score + valuation_score + yongyi_score - risk_score
    bias = _score_to_bias(total_score)

    strong_signals: list[str] = []
    risk_signals: list[str] = []
    watch_signals: list[str] = []
    signals: list[str] = []

    if pd.notna(ma5) and pd.notna(ma20):
        trend_line = f"均线结构：MA5({format_number(ma5)}) {'>' if ma5 > ma20 else '<'} MA20({format_number(ma20)})，当前属于{trend_label}。"
        signals.append(trend_line)
        (strong_signals if trend_score >= 3 else watch_signals).append(trend_line)
    if pd.notna(ma60):
        line = f"中期趋势：价格{'站上' if current_price > ma60 else '跌破'}60日均线（{format_number(ma60)}）。"
        signals.append(line)
        (strong_signals if current_price > ma60 else risk_signals).append(line)

    if change_5d_pct is not None:
        line = f"动量表现：近5日涨跌幅 {change_5d_pct:+.1f}%，近10日涨跌幅 {change_10d_pct:+.1f}%。" if change_10d_pct is not None else f"动量表现：近5日涨跌幅 {change_5d_pct:+.1f}%。"
        signals.append(line)
        if momentum_score >= 2:
            strong_signals.append(line)
        elif momentum_score <= -2:
            risk_signals.append(line)
        else:
            watch_signals.append(line)
    if breakout_high:
        strong_signals.append("价格已突破近20日高位，盘面有继续上冲的技术条件。")
        signals.append("价格已突破近20日高位，盘面有继续上冲的技术条件。")
    if breakout_low:
        risk_signals.append("价格已跌破近20日低位，短线仍需防范惯性下探。")
        signals.append("价格已跌破近20日低位，短线仍需防范惯性下探。")

    if vol_ratio is not None:
        line = f"量能状态：当日成交量为5日均量的 {vol_ratio:.1f} 倍。"
        signals.append(line)
        if vol_ratio >= 1.4:
            strong_signals.append(line)
        elif vol_ratio <= 0.7:
            watch_signals.append(line)
    if oi_ratio is not None:
        line = f"持仓状态：当前持仓为5日均持仓的 {oi_ratio:.1f} 倍，较前一日{describe_delta(oi_delta)}。"
        signals.append(line)
        if pd.notna(oi_delta) and oi_delta > 0 and pd.notna(momentum_5d) and momentum_5d > 0:
            strong_signals.append(line)
        elif pd.notna(oi_delta) and oi_delta > 0 and pd.notna(momentum_5d) and momentum_5d < 0:
            risk_signals.append(line)
        else:
            watch_signals.append(line)
    if pd.notna(net_current):
        line = f"资金倾向：前20大户当前{'净多' if net_current > 0 else '净空'} {format_number(net_current)} 手，近5日变化 {describe_delta(net_delta_5d)}。"
        signals.append(line)
        if net_current > 0 and (pd.isna(net_delta_5d) or net_delta_5d >= 0):
            strong_signals.append(line)
        elif net_current < 0 and (pd.isna(net_delta_5d) or net_delta_5d <= 0):
            risk_signals.append(line)
        else:
            watch_signals.append(line)

    hist_line = f"历史位置：当前价格位于历史 {pct_label} 分位。" if pct is not None else "历史位置：样本不足，无法判断。"
    signals.append(hist_line)
    if pct is not None and pct >= 0.8:
        risk_signals.append(hist_line)
    elif pct is not None and pct <= 0.2:
        strong_signals.append(hist_line)
    else:
        watch_signals.append(hist_line)

    if basis_current is not None:
        line = f"基差状态：当前基差 {format_number(basis_current)}，{basis_note}。"
        signals.append(line)
        if basis_pct is not None and basis_pct <= 0.2:
            strong_signals.append(line)
        elif basis_pct is not None and basis_pct >= 0.8:
            risk_signals.append(line)
        else:
            watch_signals.append(line)
    if spread_name and spread_current is not None:
        line = f"价差状态：{spread_name} 当前为 {format_number(spread_current)}，{spread_note}。"
        signals.append(line)
        if spread_pct is not None and (spread_pct >= 0.85 or spread_pct <= 0.15):
            watch_signals.append(line)

    if yongyi_basis_snapshot:
        line = "涌益市级基差：" + yongyi_basis_snapshot["message"]
        signals.append(line)
        high_city = yongyi_basis_snapshot.get("high")
        low_city = yongyi_basis_snapshot.get("low")
        extreme_city = yongyi_basis_snapshot.get("extreme")
        if high_city and low_city:
            watch_signals.append(line)
        elif extreme_city and extreme_city.get("percentile") is not None:
            if extreme_city["percentile"] <= 0.2:
                strong_signals.append(line)
            elif extreme_city["percentile"] >= 0.8:
                risk_signals.append(line)
            else:
                watch_signals.append(line)
        if high_city:
            high_line = f"高位城市：{high_city.get('name', '')}，基差 {format_number(high_city.get('value'))}，历史分位 {high_city.get('percentile', 0):.0%}。"
            signals.append(high_line)
            watch_signals.append(high_line)
        if low_city:
            low_line = f"低位城市：{low_city.get('name', '')}，基差 {format_number(low_city.get('value'))}，历史分位 {low_city.get('percentile', 0):.0%}。"
            signals.append(low_line)
            watch_signals.append(low_line)

    advice_parts = [f"{contract} 当前综合评分为 {total_score:+d}，判断为“{bias}”。"]
    if bias == "偏多":
        advice_parts.append("趋势、动量与资金配合度较好，可优先顺势跟踪主升逻辑。")
    elif bias == "观望偏多":
        advice_parts.append("盘面略偏强，但仍需等待量仓或基差进一步确认，适合逢回调观察。")
    elif bias == "偏空":
        advice_parts.append("趋势与资金信号整体偏弱，短线应以防守和等待止跌信号为主。")
    elif bias == "观望偏空":
        advice_parts.append("当前偏弱但未形成一致性下行共振，追空性价比一般。")
    else:
        advice_parts.append("多空信号分化，暂以区间思路看待，等待突破后再调整策略。")

    if strong_signals:
        advice_parts.append("当前最值得跟踪的信号是：" + "；".join(strong_signals[:2]))
    if risk_signals:
        advice_parts.append("需要重点防范：" + "；".join(risk_signals[:2]))

    yongyi_card_value = "N/A"
    yongyi_card_extra = "暂无市级基差信号"
    yongyi_top_bottom_extra = "暂无高低位城市"
    if yongyi_basis_snapshot:
        extreme_city = yongyi_basis_snapshot.get("extreme")
        high_city = yongyi_basis_snapshot.get("high")
        low_city = yongyi_basis_snapshot.get("low")
        if extreme_city:
            yongyi_card_value = format_number(extreme_city.get("value"))
            yongyi_card_extra = f"{extreme_city.get('name', '')}｜历史分位 {extreme_city.get('percentile', 0):.0%}"
        if high_city and low_city:
            yongyi_top_bottom_extra = f"高位 {high_city.get('name', '')}｜低位 {low_city.get('name', '')}"
        elif high_city:
            yongyi_top_bottom_extra = f"高位 {high_city.get('name', '')}"
        elif low_city:
            yongyi_top_bottom_extra = f"低位 {low_city.get('name', '')}"

    cards = [
        {"label": "综合评分", "value": f"{total_score:+d}", "extra": f"结论：{bias}"},
        {"label": "趋势维度", "value": trend_label, "extra": f"MA5/MA20/MA60：{format_number(ma5)} / {format_number(ma20)} / {format_number(ma60)}"},
        {"label": "动量维度", "value": f"{change_5d_pct:+.1f}%" if change_5d_pct is not None else "N/A", "extra": f"10日：{f'{change_10d_pct:+.1f}%' if change_10d_pct is not None else 'N/A'}"},
        {"label": "量仓维度", "value": f"量比{vol_ratio:.1f} / 仓比{oi_ratio:.1f}" if vol_ratio is not None and oi_ratio is not None else format_number(vol_current), "extra": f"持仓变化：{describe_delta(oi_delta)}"},
        {"label": "资金维度", "value": format_number(net_current), "extra": f"前20净持仓：{net_trend}"},
        {"label": "估值维度", "value": pct_label, "extra": basis_note},
        {"label": "市级基差", "value": yongyi_card_value, "extra": yongyi_card_extra},
        {"label": "高低位城市", "value": yongyi_top_bottom_extra, "extra": f"市级基差评分 {yongyi_score:+d}"},
        {"label": "风险维度", "value": f"{vol_20:.2f}%" if pd.notna(vol_20) else "N/A", "extra": f"20日波动率｜{spread_note}"},
    ]

    result.update({
        "ok": True,
        "cards": cards,
        "signals": signals[:10],
        "advice": " ".join(advice_parts),
        "strong_signals": strong_signals[:5],
        "risk_signals": risk_signals[:5],
        "watch_signals": watch_signals[:5],
        "trend_signal": trend_signal,
        "current_price": current_price,
        "snap_date": snap_date,
        "bias": bias,
        "score": total_score,
    })
    return result


def build_futures_overall_advice(
    futures_df: pd.DataFrame,
    target_date: pd.Timestamp,
    yongyi_numeric_df: pd.DataFrame | None = None,
    yongyi_meta_df: pd.DataFrame | None = None,
    focus_contract: str | None = None,
) -> str:
    """基于当前日期所有可用合约数据，生成一段综合性操盘建议。"""
    price_df = futures_series(futures_df, "收盘价", "收盘价")
    contracts = get_active_lh_contracts(price_df, target_date)
    if not contracts:
        return "暂无足够期货数据生成综合建议。"

    analyses = []
    for c in contracts[:6]:
        a = build_futures_contract_analysis(
            futures_df,
            c,
            target_date,
            yongyi_numeric_df=yongyi_numeric_df,
            yongyi_meta_df=yongyi_meta_df,
        )
        if a.get("ok"):
            analyses.append(a)

    if not analyses:
        return "当前日期下合约数据不足，暂无综合建议。"

    bull_contracts = [a for a in analyses if a.get("score", 0) >= 2]
    bear_contracts = [a for a in analyses if a.get("score", 0) <= -2]
    neutral_contracts = [a for a in analyses if -2 < a.get("score", 0) < 2]
    strongest_bull = max(analyses, key=lambda x: x.get("score", -999))
    strongest_bear = min(analyses, key=lambda x: x.get("score", 999))
    opp = build_futures_opportunity_messages(
        futures_df,
        target_date,
        yongyi_numeric_df=yongyi_numeric_df,
        yongyi_meta_df=yongyi_meta_df,
        focus_contract=focus_contract or strongest_bull.get("contract"),
        include_yongyi=True,
    )
    yongyi_signal = build_yongyi_city_basis_signal(
        futures_df,
        yongyi_numeric_df,
        yongyi_meta_df,
        target_date,
        focus_contract or strongest_bull.get("contract"),
    )

    parts = []
    if len(bull_contracts) > len(bear_contracts):
        parts.append(f"当前盘面整体偏强，强势合约数量 {len(bull_contracts)} 个，高于偏弱合约的 {len(bear_contracts)} 个。")
    elif len(bear_contracts) > len(bull_contracts):
        parts.append(f"当前盘面整体偏弱，偏空合约数量 {len(bear_contracts)} 个，高于偏强合约的 {len(bull_contracts)} 个。")
    else:
        parts.append(f"当前多空力量相对均衡，强弱合约分布接近，中性合约 {len(neutral_contracts)} 个。")

    parts.append(f"多头代表合约是 {strongest_bull['contract']}（评分 {strongest_bull['score']:+d}，{strongest_bull['bias']}），空头代表合约是 {strongest_bear['contract']}（评分 {strongest_bear['score']:+d}，{strongest_bear['bias']}）。")

    if strongest_bull.get("strong_signals"):
        parts.append(f"当前最强顺势信号来自 {strongest_bull['contract']}：{strongest_bull['strong_signals'][0]}")
    if strongest_bear.get("risk_signals"):
        parts.append(f"当前最大风险信号来自 {strongest_bear['contract']}：{strongest_bear['risk_signals'][0]}")
    if yongyi_signal:
        parts.append("涌益市级基差观察：" + yongyi_signal)
    if opp:
        parts.append("盘面补充观察：" + " ".join(opp[:2]))

    parts.append("策略上可优先盯住高评分主力合约的趋势延续，同时结合基差、价差、涌益市级基差与净持仓背离信号判断追价还是等待回撤。")
    return " ".join(parts)



def render_futures_module() -> None:
    with st.sidebar:
        st.header("📂 期货数据源")
        futures_path = st.text_input("期货文件路径", value=DEFAULT_FUTURES_PATH, key="futures_path")
    try:
        futures_df, futures_log = build_futures_dataset_from_path(futures_path)
    except Exception as exc:
        st.error(f"期货数据载入失败：{exc}")
        return
    if futures_df.empty:
        st.warning("期货数据为空。")
        return

    render_module_lead("📌 生猪期货（LH）｜来源：大连商品交易所。平台提供收盘价、持仓量、成交量、基差、价差、前20净持仓与仓单等七大维度，并基于所选日期自动输出交易机会观察。")
    max_date = futures_df["date"].max()
    min_date = futures_df["date"].min()

    # 图表日期范围筛选
    _, _, futures_chart_df = build_cn_date_range_selector(futures_df, "futures_chart", "📅 期货图表日期范围")
    if futures_chart_df.empty:
        st.warning("所选期货图表日期范围内没有数据。")
        return

    col_date, col_contract = st.columns([1, 1])
    with col_date:
        target_date = st.date_input("期货分析日期", value=max_date.date(), min_value=min_date.date(), max_value=max_date.date(), key="futures_target_date")
    with col_contract:
        price_df_all = futures_series(futures_chart_df, "收盘价", "收盘价")
        all_contracts = get_active_lh_contracts(price_df_all, pd.Timestamp(target_date))
        selected_contract = st.selectbox("选择重点分析合约", all_contracts if all_contracts else ["暂无"], key="futures_selected_contract")

    try:
        yongyi_numeric_df, yongyi_meta_df, _ = build_yongyi_dataset_from_path(DEFAULT_YONGYI_PATH)
    except Exception:
        yongyi_numeric_df, yongyi_meta_df = None, None

    opportunity_messages = build_futures_opportunity_messages(
        futures_chart_df,
        pd.Timestamp(target_date),
        yongyi_numeric_df=yongyi_numeric_df,
        yongyi_meta_df=yongyi_meta_df,
        focus_contract=selected_contract if selected_contract != "暂无" else None,
        include_yongyi=True,
    )

    # ========================
    # ① 顶部：综合总体建议
    # ========================
    render_section_header("📊 期货市场综合建议")
    overall_advice = build_futures_overall_advice(
        futures_chart_df,
        pd.Timestamp(target_date),
        yongyi_numeric_df=yongyi_numeric_df,
        yongyi_meta_df=yongyi_meta_df,
        focus_contract=selected_contract if selected_contract != "暂无" else None,
    )
    st.markdown(
        f"<div class='summary-box'>"
        f"<div class='summary-title'>综合操盘建议</div>"
        f"<div class='summary-subtitle'>{format_date_cn(pd.Timestamp(target_date))} 基于所有合约多空格局与技术指标自动生成</div>"
        f"<div style='font-size:15px;color:#1e293b;line-height:1.8'>{overall_advice}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )
    render_signal_messages(
        "🎯 交易机会观察",
        opportunity_messages,
        "当前主要指标处于中性区间，暂无明显交易机会提示。",
    )

    # ========================
    # ② 合约深度分析
    # ========================
    if selected_contract and selected_contract != "暂无":
        render_section_header(f"🔬 {selected_contract} 深度技术分析")
        analysis = build_futures_contract_analysis(
            futures_chart_df,
            selected_contract,
            pd.Timestamp(target_date),
            yongyi_numeric_df=yongyi_numeric_df,
            yongyi_meta_df=yongyi_meta_df,
        )
        if analysis.get("ok"):
            render_metric_cards(analysis["cards"])
            render_summary_box(
                f"{selected_contract} 综合判断",
                f"{format_date_cn(analysis['snap_date'])}｜综合评分 {analysis['score']:+d}｜结论：{analysis['bias']}",
                analysis["advice"],
            )
            render_signal_messages(f"{selected_contract} 强信号", analysis.get("strong_signals", []), "当前暂无高置信度顺势信号。")
            render_signal_messages(f"{selected_contract} 风险点", analysis.get("risk_signals", []), "当前暂无突出的风险信号。")
            render_signal_messages(f"{selected_contract} 持续跟踪点", analysis.get("watch_signals", []), "当前暂无额外观察点。")
            render_signal_messages(f"{selected_contract} 技术信号明细", analysis["signals"], "暂无明显技术信号。")
        else:
            st.info(analysis.get("advice", "暂无分析数据。"))

    st.markdown("---")

    tabs = st.tabs(["收盘价", "持仓量", "成交量", "基差", "涌益区域基差", "价差", "前20多空净持仓", "仓单与交割量"])

    with tabs[0]:
        price_df = futures_series(futures_chart_df, "收盘价", "收盘价")
        frequency, show_seasonal_close, use_lunar_close = _render_standard_controls("futures_close")
        price_fig = build_multi_series_line_chart(price_df, frequency, "各合约收盘价", hover_title="收盘价（元/吨）", use_lunar=False)
        render_plotly(price_fig, "futures", "close", frequency, False)
        if show_seasonal_close and not price_df.empty:
            close_contract_options = get_active_lh_contracts(price_df, pd.Timestamp(target_date))
            single_close = st.selectbox("选择单合约查看季节性", close_contract_options if close_contract_options else sorted(price_df["contract"].dropna().unique().tolist()), key="futures_close_single")
            close_single_df = price_df[price_df["contract"] == single_close].copy()
            if not close_single_df.empty:
                _render_seasonal_section(close_single_df, use_lunar_close, f"futures_close_{single_close}")
        if selected_contract and selected_contract != "暂无":
            selected_contract_price = price_df[price_df["contract"] == selected_contract].copy()
            if not selected_contract_price.empty:
                export_fig = build_multi_series_line_chart(selected_contract_price, frequency, f"{selected_contract} 收盘价", hover_title="收盘价（元/吨）", use_lunar=False)
                analysis_export = build_futures_contract_analysis(
                    futures_chart_df,
                    selected_contract,
                    pd.Timestamp(target_date),
                    yongyi_numeric_df=yongyi_numeric_df,
                    yongyi_meta_df=yongyi_meta_df,
                )
                if analysis_export.get("ok"):
                    try:
                        export_bytes = build_futures_export_image(export_fig, selected_contract, pd.Timestamp(target_date), analysis_export.get("advice", ""), analysis_export.get("signals", []))
                        st.download_button("下载图表+技术分析总结图片", data=export_bytes, file_name=f"futures_{selected_contract}_{pd.Timestamp(target_date).strftime('%Y%m%d')}.png", mime="image/png", key="futures_export_image")
                    except Exception as exc:
                        st.info(f"当前环境暂不支持导出图片：{exc}")

    with tabs[1]:
        oi_df = futures_series(futures_chart_df, "持仓量", "持仓量")
        _render_futures_tab(oi_df, "oi", "各合约持仓量", "持仓量（手）", [])

    with tabs[2]:
        vol_df = futures_series(futures_chart_df, "成交量", "成交量")
        _render_futures_tab(vol_df, "vol", "各合约成交量", "成交量（手）", [])

    with tabs[3]:
        basis_df = futures_series(futures_chart_df, "期现基差")
        basis_names = sorted(basis_df["series_name"].dropna().unique().tolist())
        selected = st.multiselect("选择基差序列", basis_names, default=basis_names[:4], key="basis_series")
        frequency, show_seasonal_basis, use_lunar_basis = _render_standard_controls("basis")
        chart_df = basis_df[basis_df["series_name"].isin(selected)].copy()
        if chart_df.empty:
            st.info("当前没有可展示的基差数据。")
        else:
            render_plotly(build_multi_series_line_chart(chart_df, frequency, "期现基差与相关序列", hover_title="数值"), "futures", "basis", tuple(selected), frequency)
            render_signal_messages(
                "基差异常提示",
                opportunity_messages[:3],
                "暂无额外基差信号。",
            )
            if show_seasonal_basis:
                basis_only = chart_df[chart_df["metric"] == "基差"].copy()
                if not basis_only.empty:
                    selected_one = st.selectbox("选择单序列查看季节性", sorted(basis_only["series_name"].unique()), key="basis_single")
                    seasonal_df = basis_only[basis_only["series_name"] == selected_one].copy()
                    _render_seasonal_section(seasonal_df, use_lunar_basis, f"futures_basis_{selected_one}")

    with tabs[4]:
        st.info("涌益区域基差板块已暂时隐藏。")

    with tabs[5]:
        price_df_spread = futures_series(futures_chart_df, "收盘价", "收盘价")
        contracts_spread = get_active_lh_contracts(price_df_spread, pd.Timestamp(target_date))
        if len(contracts_spread) < 2:
            st.info("收盘价合约不足，暂无法计算价差。")
        else:
            col1, col2 = st.columns(2)
            with col1:
                contract_a = st.selectbox("合约A", contracts_spread, index=0, key="spread_a")
            with col2:
                contract_b = st.selectbox("合约B", [x for x in contracts_spread if x != contract_a], index=0, key="spread_b")
            frequency_spread, show_seasonal_spread, use_lunar_spread = _render_standard_controls("spread")
            spread_df = build_spread_df(price_df_spread, contract_a, contract_b)
            if spread_df.empty:
                st.info("当前组合没有重叠日期。")
            else:
                render_plotly(build_multi_series_line_chart(spread_df, frequency_spread, f"{contract_a}-{contract_b} 价差", hover_title="价差"), "futures", "spread", contract_a, contract_b, frequency_spread)
                render_signal_messages("价差交易提示", build_spread_alert_messages(spread_df, pd.Timestamp(target_date), f"{contract_a}-{contract_b}"), "暂无明显价差信号。")
                if show_seasonal_spread:
                    _render_seasonal_section(spread_df, use_lunar_spread, f"futures_spread_{contract_a}_{contract_b}")

    with tabs[6]:
        net_df = futures_series(futures_df, "前20多空净持仓", "前20净持仓")
        _render_futures_tab(net_df, "net", "前20多空净持仓", "净持仓（手）", [])
        render_signal_messages(
            "净持仓观察",
            opportunity_messages[1:4],
            "暂无明显净持仓背离。",
        )

    with tabs[7]:
        wh_df = futures_series(futures_chart_df, "仓单、虚实盘、交割量")
        available_metrics = sorted(wh_df["metric"].dropna().unique().tolist())
        st.caption("该页按原始文件真实字段展示；文件中没有明确虚实盘字段时不额外臆造。")
        if not available_metrics:
            st.info("当前没有可展示的数据。")
        else:
            selected_metrics = st.multiselect("选择指标", available_metrics, default=available_metrics[:4], key="warehouse_metrics")
            frequency_wh, show_seasonal_wh, use_lunar_wh = _render_standard_controls("warehouse")
            chart_df_wh = wh_df[wh_df["metric"].isin(selected_metrics)].copy()
            if not chart_df_wh.empty:
                chart_df_wh["series_name"] = chart_df_wh["metric"] + "｜" + chart_df_wh["series_name"]
                render_plotly(build_multi_series_line_chart(chart_df_wh, frequency_wh, "仓单 / 收盘价 / 交割量等指标", hover_title="数值"), "futures", "warehouse", tuple(selected_metrics), frequency_wh)
            seasonal_candidates = wh_df[wh_df["series_type"] == "seasonal"].copy()
            if not seasonal_candidates.empty and show_seasonal_wh:
                metric_one = st.selectbox("选择季节性指标", sorted(seasonal_candidates["metric"].unique()), key="warehouse_seasonal_metric")
                _render_seasonal_section(seasonal_candidates[seasonal_candidates["metric"] == metric_one].copy(), use_lunar_wh, f"futures_wh_{metric_one}")

    with st.expander("查看期货解析日志"):
        st.dataframe(futures_log, width='stretch')


# -----------------------------
# 模块页面：调运分析
# -----------------------------
def render_transport_module() -> None:
    with st.sidebar:
        st.header("📂 调运数据源")
        transport_path = _resolve_data_path(r"猪只调运|调运分析", "调运")
    try:
        transport_df = build_transport_dataset_from_path(transport_path)
    except Exception as exc:
        st.error(f"调运数据载入失败：{exc}")
        return
    if transport_df.empty:
        st.warning("调运数据为空。")
        return

    render_module_lead("📌 猪只调运监测｜来源：微信物流群信息提取（已二次去重）。提供重点区域定向监控与自定义路线分析，并自动识别放量/缩量异常。")
    page = st.radio("选择页面", ["重点监控看板", "自定义分析"], horizontal=True, key="transport_page")
    max_date = transport_df["date"].max()
    min_date = transport_df["date"].min()
    target_date = st.date_input("调运分析日期", value=max_date.date(), min_value=min_date.date(), max_value=max_date.date(), key="transport_target_date")

    _, _, transport_chart_df = build_cn_date_range_selector(transport_df, "transport_chart", "📅 调运图表日期范围")
    if transport_chart_df.empty:
        st.warning("所选调运图表日期范围内没有数据。")
        return

    if page == "重点监控看板":
        total_count = int(transport_chart_df["count"].sum())
        snap_date = get_snapshot_date(transport_chart_df, pd.Timestamp(target_date))
        day_count = int(transport_chart_df.loc[transport_chart_df["date"] == snap_date, "count"].sum()) if snap_date is not None else 0
        top_in_df, _ = build_transport_rank(transport_chart_df, "调入省份", pd.Timestamp(target_date), 10)
        top_out_df, _ = build_transport_rank(transport_chart_df, "调出省份", pd.Timestamp(target_date), 10)
        top_route_df, _ = build_transport_rank(transport_chart_df, "城市路线", pd.Timestamp(target_date), 10)
        render_metric_cards([
            {"label": "样本总量", "value": str(total_count), "extra": f"时间区间：{format_date_cn(min_date)} - {format_date_cn(max_date)}"},
            {"label": "所选日期调运量", "value": str(day_count), "extra": format_date_cn(snap_date)},
            {"label": "调入最多省份", "value": top_in_df.iloc[0]["调入省份"] if not top_in_df.empty else "暂无", "extra": f"记录数 {int(top_in_df.iloc[0]['count']) if not top_in_df.empty else 0}"},
            {"label": "调出最多省份", "value": top_out_df.iloc[0]["调出省份"] if not top_out_df.empty else "暂无", "extra": f"记录数 {int(top_out_df.iloc[0]['count']) if not top_out_df.empty else 0}"},
        ])
        render_signal_messages("调运异常提示", build_transport_anomaly_messages(transport_chart_df, pd.Timestamp(target_date)), "当前未发现明显异常波动。")

        # ---- 调入/调出省份排行图（点击省份直接下钻城市）----
        render_section_header("📊 省份调运排行 — 点击省份柱子可下钻城市")
        st.caption("点击下方任意省份柱子，页面将自动展示该省份的城市来源/去向分布，无需手动选择。")
        col1, col2 = st.columns(2)
        with col1:
            fig_in = build_bar_chart(top_in_df, "调入省份", "count", f"{format_date_cn(snap_date)} 调入最多的前10省份", text_col="count", color_col="count")
            selected_in = st.plotly_chart(fig_in, key="transport_top_in", on_select="rerun", width='stretch')
            # 读取点击事件
            clicked_in_province = None
            if selected_in and selected_in.get("selection") and selected_in["selection"].get("points"):
                pt = selected_in["selection"]["points"][0]
                clicked_in_province = pt.get("label") or pt.get("y") or pt.get("x")
        with col2:
            fig_out = build_bar_chart(top_out_df, "调出省份", "count", f"{format_date_cn(snap_date)} 调出最多的前10省份", text_col="count", color_col="count")
            selected_out = st.plotly_chart(fig_out, key="transport_top_out", on_select="rerun", width='stretch')
            clicked_out_province = None
            if selected_out and selected_out.get("selection") and selected_out["selection"].get("points"):
                pt = selected_out["selection"]["points"][0]
                clicked_out_province = pt.get("label") or pt.get("y") or pt.get("x")

        render_plotly(build_bar_chart(top_route_df.head(10), "城市路线", "count", f"{format_date_cn(snap_date)} 调运最多的路线", text_col="count", color_col="count"), "transport", "top_route", target_date)

        # ---- 点击下钻：调入省份 → 来源城市 ----
        if clicked_in_province:
            st.session_state["transport_drill_in_province"] = clicked_in_province
        if clicked_out_province:
            st.session_state["transport_drill_out_province"] = clicked_out_province

        # 调入下钻
        drill_in_prov = st.session_state.get("transport_drill_in_province")
        if drill_in_prov:
            render_section_header(f"🔍 调入「{drill_in_prov}」来源省份 & 城市分布")
            src_provinces_in = sorted(transport_chart_df[transport_chart_df["调入省份"] == drill_in_prov]["调出省份"].dropna().unique().tolist())
            if src_provinces_in:
                col_a, col_b = st.columns(2)
                with col_a:
                    # 来源省份排行
                    prov_rank_in = transport_df[(transport_df["调入省份"] == drill_in_prov) & (transport_df["date"] == snap_date)].groupby("调出省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
                    if not prov_rank_in.empty:
                        fig_prov_in = px.bar(prov_rank_in.sort_values("count"), x="count", y="调出省份", orientation="h", title=f"{format_date_cn(snap_date)} 调入{drill_in_prov} 来源省份", text="count", color="count", color_continuous_scale="Blues")
                        fig_prov_in.update_layout(template="plotly_white", coloraxis_showscale=False)
                        sel_prov_in = st.plotly_chart(fig_prov_in, key=f"drill_in_prov_{drill_in_prov}", on_select="rerun", width='stretch')
                        if sel_prov_in and sel_prov_in.get("selection") and sel_prov_in["selection"].get("points"):
                            pt2 = sel_prov_in["selection"]["points"][0]
                            clicked_src = pt2.get("label") or pt2.get("y") or pt2.get("x")
                            if clicked_src:
                                st.session_state["transport_drill_in_src_prov"] = clicked_src
                with col_b:
                    # 城市分布（基于所选来源省份）
                    src_prov = st.session_state.get("transport_drill_in_src_prov", src_provinces_in[0])
                    if src_prov not in src_provinces_in:
                        src_prov = src_provinces_in[0]
                    city_df_in = transport_df[
                        (transport_df["调入省份"] == drill_in_prov) &
                        (transport_df["调出省份"] == src_prov) &
                        (transport_df["date"] == snap_date)
                    ].groupby("调出城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                    if not city_df_in.empty:
                        fig_city_in = px.bar(city_df_in, x="调出城市", y="count", title=f"{src_prov} → {drill_in_prov} 城市分布（{format_date_cn(snap_date)}）", text="count", color="count", color_continuous_scale="Oranges")
                        fig_city_in.update_traces(textposition="outside")
                        fig_city_in.update_layout(template="plotly_white", coloraxis_showscale=False)
                        render_plotly(fig_city_in, "transport", "city_in", drill_in_prov, src_prov, snap_date)
                    st.caption(f"当前来源省份：{src_prov}（可点击左侧省份图切换）")
            else:
                st.info(f"{format_date_cn(snap_date)} 暂无调入{drill_in_prov}的记录。")

        # 调出下钻
        drill_out_prov = st.session_state.get("transport_drill_out_province")
        if drill_out_prov:
            render_section_header(f"🔍 调出「{drill_out_prov}」去向省份 & 城市分布")
            dest_provinces_out = sorted(transport_chart_df[transport_chart_df["调出省份"] == drill_out_prov]["调入省份"].dropna().unique().tolist())
            if dest_provinces_out:
                col_a, col_b = st.columns(2)
                with col_a:
                    prov_rank_out = transport_df[(transport_df["调出省份"] == drill_out_prov) & (transport_df["date"] == snap_date)].groupby("调入省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
                    if not prov_rank_out.empty:
                        fig_prov_out = px.bar(prov_rank_out.sort_values("count"), x="count", y="调入省份", orientation="h", title=f"{format_date_cn(snap_date)} {drill_out_prov} 调出去向省份", text="count", color="count", color_continuous_scale="Greens")
                        fig_prov_out.update_layout(template="plotly_white", coloraxis_showscale=False)
                        sel_prov_out = st.plotly_chart(fig_prov_out, key=f"drill_out_prov_{drill_out_prov}", on_select="rerun", width='stretch')
                        if sel_prov_out and sel_prov_out.get("selection") and sel_prov_out["selection"].get("points"):
                            pt3 = sel_prov_out["selection"]["points"][0]
                            clicked_dest = pt3.get("label") or pt3.get("y") or pt3.get("x")
                            if clicked_dest:
                                st.session_state["transport_drill_out_dest_prov"] = clicked_dest
                    else:
                        st.info("当前没有可展示的去向省份排行。")
                with col_b:
                    dest_prov = st.session_state.get("transport_drill_out_dest_prov", dest_provinces_out[0])
                    if dest_prov not in dest_provinces_out:
                        dest_prov = dest_provinces_out[0]
                    city_df_out = transport_df[
                        (transport_df["调出省份"] == drill_out_prov) &
                        (transport_df["调入省份"] == dest_prov) &
                        (transport_df["date"] == snap_date)
                    ].groupby("调入城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                    if not city_df_out.empty:
                        fig_city_out = px.bar(city_df_out, x="调入城市", y="count", title=f"{drill_out_prov} → {dest_prov} 城市分布（{format_date_cn(snap_date)}）", text="count", color="count", color_continuous_scale="Purples")
                        fig_city_out.update_traces(textposition="outside")
                        fig_city_out.update_layout(template="plotly_white", coloraxis_showscale=False)
                        render_plotly(fig_city_out, "transport", "city_out", drill_out_prov, dest_prov, snap_date)
                    else:
                        st.info("点击左侧调出省份排行后，城市分布会直接显示在右侧。")
                    st.caption(f"当前去向省份：{dest_prov}（可点击左侧省份图切换）")

        # 每日分区域调入/调出堆积图（针对指定区域）
        render_section_header("📊 每日分区域来源堆积分析")
        st.caption("点击堆积图中的某个日期柱子，可在下方查看该日的调入/调出省份明细与城市下钻（布局与下方重点板块一致）。")
        all_in_provinces = sorted(transport_chart_df["调入省份"].dropna().unique().tolist())
        all_out_provinces = sorted(transport_chart_df["调出省份"].dropna().unique().tolist())
        default_stack_idx = all_in_provinces.index(drill_in_prov) if drill_in_prov and drill_in_prov in all_in_provinces else (all_in_provinces.index("云南") if "云南" in all_in_provinces else 0)
        default_out_idx = all_out_provinces.index(drill_out_prov) if drill_out_prov and drill_out_prov in all_out_provinces else (all_out_provinces.index("云南") if "云南" in all_out_provinces else 0)
        col_stack_in, col_stack_out = st.columns(2)
        with col_stack_in:
            target_in = st.selectbox("查看调入省份（趋势图）", all_in_provinces, index=default_stack_idx, key="transport_stack_target")
            stack_df = transport_chart_df[transport_chart_df["调入省份"] == target_in].groupby(["date", "调出省份"], as_index=False)["count"].sum().sort_values(["date", "count"], ascending=[True, False])
            if not stack_df.empty:
                top_sources = stack_df.groupby("调出省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(8)["调出省份"].tolist()
                stack_df["来源分组"] = stack_df["调出省份"].apply(lambda x: x if x in top_sources else "其他")
                stacked = stack_df.groupby(["date", "来源分组"], as_index=False)["count"].sum()
                stacked = sort_date_label_frame(stacked)
                stacked["日期标签"] = stacked["date"].map(format_date_cn)
                fig_stack = px.bar(stacked, x="日期标签", y="count", color="来源分组", title=f"每日调入{target_in} — 分来源省份堆积图", text=None, barmode="stack", custom_data=["date"], category_orders={"日期标签": stacked["日期标签"].drop_duplicates().tolist()})
                fig_stack.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量", hovermode="x unified")
                sel_stack = st.plotly_chart(fig_stack, key="transport_stack_in_chart", on_select="rerun", width='stretch')
                if sel_stack and sel_stack.get("selection") and sel_stack["selection"].get("points"):
                    pt_stack = sel_stack["selection"]["points"][0]
                    picked_date = pt_stack.get("customdata", [None])[0]
                    if picked_date:
                        st.session_state["transport_stack_in_date"] = pd.Timestamp(picked_date)
                current_in_date = st.session_state.get("transport_stack_in_date", snap_date)
                # 来源省份排行（可点击下钻城市）
                rank_in_day = transport_chart_df[(transport_chart_df["调入省份"] == target_in) & (transport_chart_df["date"] == pd.Timestamp(current_in_date))].groupby("调出省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
                if not rank_in_day.empty:
                    st.caption(f"📅 点击日期后联动：{format_date_cn(pd.Timestamp(current_in_date))} 调入 {target_in} 来源省份排行（点击省份可下钻城市）")
                    col_rk_a, col_rk_b = st.columns(2)
                    with col_rk_a:
                        fig_rin = px.bar(rank_in_day.sort_values("count"), x="count", y="调出省份", orientation="h",
                                         title=f"{format_date_cn(pd.Timestamp(current_in_date))} 调入{target_in} 来源省份", text="count",
                                         color="count", color_continuous_scale="Blues")
                        fig_rin.update_layout(template="plotly_white", coloraxis_showscale=False)
                        sel_rin = st.plotly_chart(fig_rin, key=f"stack_in_rank_{target_in}", on_select="rerun", width='stretch')
                        if sel_rin and sel_rin.get("selection") and sel_rin["selection"].get("points"):
                            pt_rin = sel_rin["selection"]["points"][0]
                            clicked_src_prov = pt_rin.get("label") or pt_rin.get("y") or pt_rin.get("x")
                            if clicked_src_prov:
                                st.session_state[f"stack_in_drill_prov_{target_in}"] = clicked_src_prov
                    with col_rk_b:
                        drill_src = st.session_state.get(f"stack_in_drill_prov_{target_in}")
                        if drill_src:
                            city_in_day = transport_chart_df[(transport_chart_df["调入省份"] == target_in) & (transport_chart_df["调出省份"] == drill_src) & (transport_chart_df["date"] == pd.Timestamp(current_in_date))].groupby("调出城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                            if not city_in_day.empty:
                                fig_city_rin = px.bar(city_in_day, x="调出城市", y="count", title=f"{drill_src} → {target_in} 城市分布", text="count", color="count", color_continuous_scale="Oranges")
                                fig_city_rin.update_traces(textposition="outside")
                                fig_city_rin.update_layout(template="plotly_white", coloraxis_showscale=False)
                                render_plotly(fig_city_rin, "transport", "stack_in_city", target_in, drill_src, current_in_date)
                            st.caption(f"当前来源省份：{drill_src}（点击左侧省份图切换）")
                        else:
                            st.info("点击左侧来源省份排行可下钻城市分布。")
        with col_stack_out:
            target_out = st.selectbox("查看调出省份（趋势图）", all_out_provinces, index=default_out_idx, key="transport_out_stack_target")
            out_stack_df = transport_chart_df[transport_chart_df["调出省份"] == target_out].groupby(["date", "调入省份"], as_index=False)["count"].sum().sort_values(["date", "count"], ascending=[True, False])
            if not out_stack_df.empty:
                top_dests = out_stack_df.groupby("调入省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(8)["调入省份"].tolist()
                out_stack_df["去向分组"] = out_stack_df["调入省份"].apply(lambda x: x if x in top_dests else "其他")
                stacked_out = out_stack_df.groupby(["date", "去向分组"], as_index=False)["count"].sum()
                stacked_out = sort_date_label_frame(stacked_out)
                stacked_out["日期标签"] = stacked_out["date"].map(format_date_cn)
                fig_out_stack = px.bar(stacked_out, x="日期标签", y="count", color="去向分组", title=f"每日调出{target_out} — 分去向省份堆积图", text=None, barmode="stack", custom_data=["date"], category_orders={"日期标签": stacked_out["日期标签"].drop_duplicates().tolist()})
                fig_out_stack.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量", hovermode="x unified")
                sel_out_stack = st.plotly_chart(fig_out_stack, key="transport_stack_out_chart", on_select="rerun", width='stretch')
                if sel_out_stack and sel_out_stack.get("selection") and sel_out_stack["selection"].get("points"):
                    pt_out_stack = sel_out_stack["selection"]["points"][0]
                    picked_out_date = pt_out_stack.get("customdata", [None])[0]
                    if picked_out_date:
                        st.session_state["transport_stack_out_date"] = pd.Timestamp(picked_out_date)
                current_out_date = st.session_state.get("transport_stack_out_date", snap_date)
                # 去向省份排行（可点击下钻城市）
                rank_out_day = transport_chart_df[(transport_chart_df["调出省份"] == target_out) & (transport_chart_df["date"] == pd.Timestamp(current_out_date))].groupby("调入省份", as_index=False)["count"].sum().sort_values("count", ascending=False).head(10)
                if not rank_out_day.empty:
                    st.caption(f"📅 点击日期后联动：{format_date_cn(pd.Timestamp(current_out_date))} 调出 {target_out} 去向省份排行（点击省份可下钻城市）")
                    col_rk_c, col_rk_d = st.columns(2)
                    with col_rk_c:
                        fig_rout = px.bar(rank_out_day.sort_values("count"), x="count", y="调入省份", orientation="h",
                                          title=f"{format_date_cn(pd.Timestamp(current_out_date))} 调出{target_out} 去向省份", text="count",
                                          color="count", color_continuous_scale="Greens")
                        fig_rout.update_layout(template="plotly_white", coloraxis_showscale=False)
                        sel_rout = st.plotly_chart(fig_rout, key=f"stack_out_rank_{target_out}", on_select="rerun", width='stretch')
                        if sel_rout and sel_rout.get("selection") and sel_rout["selection"].get("points"):
                            pt_rout = sel_rout["selection"]["points"][0]
                            clicked_dest_prov = pt_rout.get("label") or pt_rout.get("y") or pt_rout.get("x")
                            if clicked_dest_prov:
                                st.session_state[f"stack_out_drill_prov_{target_out}"] = clicked_dest_prov
                    with col_rk_d:
                        drill_dest = st.session_state.get(f"stack_out_drill_prov_{target_out}")
                        if drill_dest:
                            city_out_day = transport_chart_df[(transport_chart_df["调出省份"] == target_out) & (transport_chart_df["调入省份"] == drill_dest) & (transport_chart_df["date"] == pd.Timestamp(current_out_date))].groupby("调入城市", as_index=False)["count"].sum().sort_values("count", ascending=False).head(15)
                            if not city_out_day.empty:
                                fig_city_rout = px.bar(city_out_day, x="调入城市", y="count", title=f"{target_out} → {drill_dest} 城市分布", text="count", color="count", color_continuous_scale="Purples")
                                fig_city_rout.update_traces(textposition="outside")
                                fig_city_rout.update_layout(template="plotly_white", coloraxis_showscale=False)
                                render_plotly(fig_city_rout, "transport", "stack_out_city", target_out, drill_dest, current_out_date)
                            st.caption(f"当前去向省份：{drill_dest}（点击左侧省份图切换）")
                        else:
                            st.info("点击左侧去向省份排行可下钻城市分布。")

        render_transport_focus_section(
            transport_chart_df,
            "云南神农销售区域",
            [{"label": "云南", "field": "province"}, {"label": "广东", "field": "province"}, {"label": "广西", "field": "province"}],
            pd.Timestamp(target_date),
            both_directions=True,
        )
        render_transport_focus_section(
            transport_chart_df,
            "疫情猪调入区域",
            [{"label": "临沂", "field": "city"}, {"label": "商丘", "field": "city"}],
            pd.Timestamp(target_date),
        )
        render_transport_focus_section(
            transport_chart_df,
            "肥大猪调入区域",
            [{"label": "四川", "field": "province"}, {"label": "云南", "field": "province"}, {"label": "贵州", "field": "province"}],
            pd.Timestamp(target_date),
        )
        render_transport_focus_section(
            transport_chart_df,
            "瘦长条大猪调入区域",
            [{"label": "广东", "field": "province"}, {"label": "浙江", "field": "province"}, {"label": "深圳", "field": "city"}],
            pd.Timestamp(target_date),
        )
    else:
        render_section_header("🔍 自定义路线分析")
        col1, col2 = st.columns(2)
        with col1:
            date_range = st.date_input("选择日期区间", value=(min_date.date(), max_date.date()), min_value=min_date.date(), max_value=max_date.date(), key="transport_range")
            start_date, end_date = (date_range[0], date_range[1]) if len(date_range) == 2 else (min_date.date(), max_date.date())
            in_provinces = st.multiselect("调入省份", sorted(transport_df["调入省份"].unique()), default=[], key="transport_in_province")
            out_provinces = st.multiselect("调出省份", sorted(transport_df["调出省份"].unique()), default=[], key="transport_out_province")
        with col2:
            in_cities = st.multiselect("调入城市", sorted(transport_df["调入城市"].unique()), default=[], key="transport_in_city")
            out_cities = st.multiselect("调出城市", sorted(transport_df["调出城市"].unique()), default=[], key="transport_out_city")
            chart_type = st.selectbox("图表类型", ["折线图", "柱状图"], key="transport_chart_type")
            group_by = st.selectbox("统计维度", ["城市路线", "调入省份", "调出省份", "调入城市", "调出城市"], key="transport_group_by")

        frequency_t, show_seasonal_t, use_lunar_t = _render_standard_controls("transport_custom")

        temp = transport_df.copy()
        temp = temp[(temp["date"] >= pd.Timestamp(start_date)) & (temp["date"] <= pd.Timestamp(end_date))].copy()
        if in_provinces:
            temp = temp[temp["调入省份"].isin(in_provinces)]
        if out_provinces:
            temp = temp[temp["调出省份"].isin(out_provinces)]
        if in_cities:
            temp = temp[temp["调入城市"].isin(in_cities)]
        if out_cities:
            temp = temp[temp["调出城市"].isin(out_cities)]
        if temp.empty:
            st.warning("当前自定义筛选下没有数据。")
            return

        trend = temp.groupby("date", as_index=False)["count"].sum().sort_values("date")
        # 构建成标准格式以复用季节性函数
        trend_std = trend.copy()
        trend_std["value"] = trend_std["count"]
        trend_std["series_name"] = "调运量"
        trend_std["province"] = "全部"
        trend_std["city"] = ""
        trend_std["sheet"] = "调运"
        trend_std["metric"] = "调运量"
        trend_std["display_name"] = "调运量"
        trend_std = enrich_date_features(trend_std)

        if show_seasonal_t:
            _render_seasonal_section(trend_std, use_lunar_t, "transport_custom_seasonal")
        else:
            if chart_type == "折线图":
                render_plotly(build_transport_trend_chart(trend, "筛选区间总调运趋势"), "transport_custom", "trend", frequency_t)
            else:
                fig = px.bar(trend, x="date", y="count", title="筛选区间总调运柱状图", text="count")
                fig.update_layout(template="plotly_white", xaxis_title="日期", yaxis_title="调运量")
                render_plotly(fig, "transport_custom", "bar", frequency_t)

        rank = temp.groupby(group_by, as_index=False)["count"].sum().sort_values("count", ascending=False).head(20)
        render_plotly(build_bar_chart(rank, group_by, "count", f"按{group_by}统计", text_col="count", color_col="count"), "transport_custom", "rank", group_by)


# -----------------------------
# 模块页面：鲜冻品数据库
# -----------------------------
def render_single_price_page(price_df: pd.DataFrame, page_key: str, page_title: str) -> None:
    """单品类价格分析页（鲜品 或 冻品）。
    支持产品/供应商选择、价差双轴图、季节性分析。"""
    # ── 空数据保护 ──
    if price_df.empty:
        st.warning(f"### ⚠️ 无{page_title}数据\n\n"
                   f"当前数据源中未找到{page_title}记录。\n\n"
                   "**可能原因**：\n"
                   "1. 数据文件中不包含该品类的价格日报 sheet\n"
                   "2. 文件名不匹配（鲜品需含\"鲜品\"或\"神农肉业\"，冻品需含\"冻品\"）\n"
                   "3. 日期格式无法解析\n\n"
                   "请展开下方「📊 原始 Excel 结构预览」排查。")
        return

    # ── 安全获取日期范围 ──
    date_min = price_df["date"].min()
    date_max = price_df["date"].max()
    if pd.isna(date_min) or pd.isna(date_max):
        st.warning(f"{page_title}数据日期无效，请检查数据源。")
        return
    default_date = min(date_max.date(), pd.Timestamp.today().date())
    min_date = date_min.date()
    max_date = date_max.date()

    # ── 产品 & 供应商选择 ──
    product_options = sorted(price_df["product"].dropna().unique().tolist())
    supplier_options = sorted(price_df["supplier"].dropna().unique().tolist()) if "supplier" in price_df.columns else []

    if not product_options:
        st.warning(f"{page_title}未找到任何产品，请检查数据。")
        return

    col_a, col_b, col_c = st.columns([1, 1, 1])
    with col_a:
        selected_product = st.selectbox(f"{page_title}产品", product_options, key=f"{page_key}_product")
    # 该产品可用的供应商
    product_suppliers = sorted(price_df[price_df["product"] == selected_product]["supplier"].dropna().unique().tolist()) if supplier_options else []
    with col_b:
        supplier_1 = st.selectbox(f"{page_title}系列 1", product_suppliers, key=f"{page_key}_sup1",
                                  index=0 if product_suppliers else 0)
    with col_c:
        default_s2 = product_suppliers[1] if len(product_suppliers) > 1 else (product_suppliers[0] if product_suppliers else "")
        try:
            idx2 = product_suppliers.index(default_s2)
        except (ValueError, IndexError):
            idx2 = 0
        supplier_2 = st.selectbox(f"{page_title}系列 2", product_suppliers, key=f"{page_key}_sup2", index=min(idx2, len(product_suppliers) - 1))

    # ── 价差开关 ──
    show_spread = st.checkbox("📏 显示价差（系列1 − 系列2，虚线·右轴）", value=False, key=f"{page_key}_show_spread")
    spread_type = "绝对值"
    if show_spread:
        spread_col1, spread_col2 = st.columns([1, 3])
        with spread_col1:
            spread_type = st.radio("价差类型", ["绝对值", "百分比"], horizontal=True, key=f"{page_key}_spread_type")

    target_date = st.date_input(f"{page_title}分析日期", value=default_date, min_value=min_date, max_value=max_date, key=f"{page_key}_date")
    frequency, show_seasonal, use_lunar = _render_standard_controls(page_key)
    _, _, filtered_price_df = build_cn_date_range_selector(price_df, f"{page_key}_chart", f"📅 {page_title}图表日期范围")

    # ── 提取两个系列 ──
    mask1 = (filtered_price_df["product"] == selected_product) & (filtered_price_df["supplier"] == supplier_1)
    s1 = filtered_price_df[mask1].copy()
    mask2 = (filtered_price_df["product"] == selected_product) & (filtered_price_df["supplier"] == supplier_2)
    s2 = filtered_price_df[mask2].copy()

    if s1.empty and s2.empty:
        st.warning(f"所选产品「{selected_product}」在日期范围内无数据。请检查产品/供应商选择或日期范围。")
        # 诊断信息
        with st.expander("🔍 数据诊断", expanded=True):
            st.markdown(f"- 产品列表: {product_options}")
            st.markdown(f"- 当前产品可用供应商: {product_suppliers}")
            st.markdown(f"- 日期范围: {filtered_price_df['date'].min()} ~ {filtered_price_df['date'].max()}")
            st.markdown(f"- 筛选后数据行数: {len(filtered_price_df)}")
        return

    # ── 构建合并数据用于价差计算 ──
    s1_agg = s1.groupby("date", as_index=False)["value"].mean().rename(columns={"value": "v1"})
    s2_agg = s2.groupby("date", as_index=False)["value"].mean().rename(columns={"value": "v2"})
    merged = s1_agg.merge(s2_agg, on="date", how="outer").sort_values("date")

    label1 = f"{selected_product} · {supplier_1}"
    label2 = f"{selected_product} · {supplier_2}"

    if show_spread:
        if spread_type == "百分比":
            merged["spread"] = np.where(
                merged["v2"].notna() & (merged["v2"] != 0),
                (merged["v1"] - merged["v2"]) / merged["v2"] * 100,
                np.nan,
            )
            spread_label = f"价差 %（{supplier_1} − {supplier_2}）"
            spread_axis_title = "价差 (%)"
        else:
            merged["spread"] = merged["v1"] - merged["v2"]
            spread_label = f"价差（{supplier_1} − {supplier_2}）"
            spread_axis_title = "价差 (元)"

    # ── 可视化 ──
    if show_seasonal:
        # 季节性图：使用合并后的主数据
        combined = pd.concat([
            s1.assign(series_name=label1),
            s2.assign(series_name=label2),
        ], ignore_index=True)
        _render_seasonal_section(combined, use_lunar, f"{page_key}_{selected_product}", frequency)
    elif show_spread and merged["spread"].notna().any():
        # 双轴图：左轴价格（实线），右轴价差（虚线）
        agg = aggregate_series_frequency(
            pd.concat([s1.assign(series_name=label1), s2.assign(series_name=label2)], ignore_index=True),
            frequency, use_lunar=use_lunar,
        )
        order = agg.sort_values("period")["period_label"].drop_duplicates().tolist()
        fig = go.Figure()
        # 鲜品红色系，冻品蓝色系，不同厂家深浅区分
        if "鲜品" in page_title:
            colors = ["#DC2626", "#EF4444", "#B91C1C", "#FCA5A5"]  # 红：标准、亮红、深红、浅红
        else:
            colors = ["#2563EB", "#3B82F6", "#1E40AF", "#93C5FD"]  # 蓝：标准、亮蓝、深蓝、浅蓝
        for i, sname in enumerate(sorted(agg["series_name"].dropna().unique())):
            sub = agg[agg["series_name"] == sname]
            fig.add_trace(go.Scatter(
                x=sub["period_label"], y=sub["value"],
                mode="lines+markers", name=sname,
                yaxis="y",
                line=dict(color=colors[i % len(colors)], width=2),
                marker=dict(size=5),
                hovertemplate=f"%{{x}}<br>{sname}：%{{y:.2f}}<extra></extra>",
            ))
        # 价差虚线（右轴）
        spread_agg = aggregate_series_frequency(
            merged[["date", "spread"]].rename(columns={"spread": "value"}).assign(series_name=spread_label),
            frequency, use_lunar=use_lunar,
        )
        spread_sub = spread_agg[spread_agg["series_name"] == spread_label]
        fig.add_trace(go.Scatter(
            x=spread_sub["period_label"], y=spread_sub["value"],
            mode="lines", name=spread_label,
            yaxis="y2",
            line=dict(color="#9333EA", width=2.5, dash="dash"),
            hovertemplate=f"%{{x}}<br>{spread_label}：%{{y:.2f}}<extra></extra>",
        ))
        fig.update_layout(
            title=f"{selected_product} 价格对比 · 价差分析（{frequency}）",
            xaxis=dict(title="日期", type="category", categoryorder="array", categoryarray=order),
            yaxis=dict(title="价格 (元)"),
            yaxis2=dict(title=spread_axis_title, overlaying="y", side="right", showgrid=False),
            template="plotly_white", hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=-0.35, xanchor="center", x=0.5),
        )
        render_plotly(fig, page_key, "spread_chart", selected_product, supplier_1, supplier_2, frequency)
    else:
        # 单轴图：也用红/蓝色系区分厂家
        combined = pd.concat([
            s1.assign(series_name=label1),
            s2.assign(series_name=label2),
        ], ignore_index=True)
        if "鲜品" in page_title:
            color_seq = ["#DC2626", "#EF4444"]  # 红色系
        else:
            color_seq = ["#2563EB", "#3B82F6"]  # 蓝色系
        fig = px.line(
            aggregate_series_frequency(combined, frequency),
            x="period_label", y="value", color="series_name",
            color_discrete_sequence=color_seq,
            markers=False,
        )
        fig.update_traces(hovertemplate="日期：%{x}<br>%{fullData.name}：%{y:.2f}<extra></extra>", line=dict(width=2))
        fig.update_layout(
            title=f"{selected_product} 价格走势（{frequency}）",
            xaxis_title="日期", yaxis_title="价格 (元)",
            template="plotly_white", hovermode="x unified",
        )
        render_plotly(fig, page_key, "price_chart", selected_product, frequency)

    # ── 异常提示 ──
    s1_for_alert = s1.copy()
    s1_for_alert["series_name"] = label1
    if not s1_for_alert.empty:
        render_signal_messages(f"{page_title}异常提示 · {label1}", build_price_alert_messages(s1_for_alert, pd.Timestamp(target_date), label1), "当前未发现明显异常。")


def render_fresh_frozen_module() -> None:
    """鲜冻品价格数据库模块。分别加载鲜品和冻品文件，冻品仅保留神农与牧原。"""

    # ── 数据源：分别定位鲜品和冻品文件 ──
    with st.sidebar:
        st.header("📂 鲜冻品数据源")
        fresh_file = _resolve_data_path(r"神农肉业.*鲜品|鲜品价格", "神农鲜品")
        # 冻品：精确匹配神农肉业-冻品价格，避免匹配到 5.鲜品冻品价格数据库.xlsx
        _FROZEN_LOCAL = Path(r"D:\CC\Desktop\平台数据\神农肉业-冻品价格.xlsx")
        _FROZEN_REPO = _REPO_DATA / "神农肉业-冻品价格.xlsx"
        if _FROZEN_LOCAL.is_file():
            frozen_file = str(_FROZEN_LOCAL)
        elif _FROZEN_REPO.is_file():
            frozen_file = str(_FROZEN_REPO)
        else:
            frozen_file = _resolve_data_path(r"神农肉业.*冻品价格\.xlsx", "神农冻品")
            if not frozen_file:
                frozen_file = _resolve_data_path(r"冻品", "神农冻品")
        if not fresh_file and not frozen_file:
            fresh_file = _resolve_data_path(r"鲜品|冻品", "鲜冻品")
            frozen_file = fresh_file

    # ── 分别加载鲜品和冻品 ──
    fresh_df = pd.DataFrame()
    frozen_df = pd.DataFrame()

    if fresh_file:
        try:
            fresh_df = build_fresh_frozen_dataset_from_path(fresh_file)
            fresh_df = fresh_df[fresh_df["dataset_type"] == "鲜品"].copy() if not fresh_df.empty else fresh_df
        except Exception as exc:
            st.error(f"鲜品加载失败：{exc}")

    frozen_path = frozen_file if frozen_file != fresh_file else fresh_file
    if frozen_path:
        try:
            frozen_df = build_fresh_frozen_dataset_from_path(frozen_path)
            if not frozen_df.empty:
                frozen_df = frozen_df[frozen_df["dataset_type"] == "冻品"].copy()
                # 冻品仅保留神农和牧原（仅当有供应商信息时过滤，宽表格式无供应商则保留全部）
                if "supplier" in frozen_df.columns and frozen_df["supplier"].notna().any():
                    has_supplier_data = frozen_df["supplier"].str.strip().str.len().sum() > 0
                    if has_supplier_data:
                        frozen_df = frozen_df[frozen_df["supplier"].str.contains("神农|牧原", na=False)].copy()
        except Exception as exc:
            st.error(f"冻品加载失败：{exc}")

    # 合并
    all_dfs = [df for df in [fresh_df, frozen_df] if not df.empty]
    price_df = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

    # ── 空数据 ──
    if price_df.empty:
        st.warning("### ⚠️ 鲜冻品数据为空，未能解析出任何价格记录")
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.markdown("**鲜品**")
            st.caption(f"文件：{fresh_file or '未找到'}")
            if not fresh_df.empty:
                st.success(f"✅ {len(fresh_df)} 条 · {fresh_df['product'].nunique()} 产品 · "
                          f"{fresh_df['supplier'].nunique()} 供应商")
            else:
                st.warning("无数据")
        with col_d2:
            st.markdown("**冻品**")
            st.caption(f"文件：{frozen_path or '未找到'}")
            if not frozen_df.empty:
                st.success(f"✅ {len(frozen_df)} 条 · {frozen_df['product'].nunique()} 产品 · "
                          f"{frozen_df['supplier'].nunique()} 供应商（仅神农/牧原）")
            else:
                st.warning("无数据")

        # 显示解析诊断
        with st.expander("🔍 详细解析诊断", expanded=True):
            t1, t2 = st.tabs(["鲜品解析日志", "冻品解析日志"])
            with t1:
                fresh_diag = fresh_df.attrs.get("parse_diag", []) if not fresh_df.empty else ["无数据"]
                st.code("\n".join(fresh_diag), language=None)
            with t2:
                frozen_diag = frozen_df.attrs.get("parse_diag", []) if not frozen_df.empty else ["无数据"]
                st.code("\n".join(frozen_diag), language=None)

        st.info("**调试建议**：确保 `data/` 目录下有对应 Excel 文件，内含「价格日报」sheet。"
                "鲜品日期为 M.DD 浮点（如 `5.16`），冻品为周区间（如 `\"5月18日-5月24日\"`）。"
                "展开上方「🔍 详细解析诊断」查看每个 sheet 的处理结果。")
        with st.expander("📊 原始 Excel 结构预览", expanded=False):
            ta, tb = st.tabs(["鲜品文件", "冻品文件"])
            with ta:
                _show_raw_excel_preview(fresh_file or "")
            with tb:
                _show_raw_excel_preview(frozen_path or "")
        return

    render_module_lead("📌 鲜品 / 冻品价格数据库｜区间报价取均值绘图。冻品仅展示神农与牧原品牌。")

    # ── 全局日期范围筛选 ──
    _, _, price_df = build_cn_date_range_selector(price_df, "fresh_frozen_global", "📅 鲜冻品图表日期范围")
    if price_df.empty:
        st.warning("所选日期范围内没有数据。")
        return

    # ── 综合行情预警 ──
    target_date_global = price_df["date"].max()
    if pd.notna(target_date_global):
        render_section_header("🚨 行情业务预警（基于全品类自动扫描）")
        signals = build_fresh_frozen_market_signals(price_df, target_date_global)
        render_market_signals(signals)

    st.markdown("---")
    sub_page = st.radio("选择子页", ["鲜品", "冻品", "鲜冻价差"], horizontal=True, key="price_page")

    fresh_view = price_df[price_df["dataset_type"] == "鲜品"].copy()
    frozen_view = price_df[price_df["dataset_type"] == "冻品"].copy()

    if sub_page == "鲜品":
        render_single_price_page(fresh_view, "fresh", "鲜品")
    elif sub_page == "冻品":
        if frozen_view.empty:
            st.warning("### ⚠️ 无冻品数据")
            st.info("冻品数据为空，可能是文件缺失、sheet 被跳过或日期格式无法解析。")
            with st.expander("🔍 冻品解析诊断", expanded=True):
                frozen_diag = frozen_df.attrs.get("parse_diag", []) if not frozen_df.empty else ["冻品 DataFrame 为空"]
                st.code("\n".join(frozen_diag), language=None)
            with st.expander("📊 冻品原始 Excel 预览", expanded=False):
                _show_raw_excel_preview(frozen_path or "")
        else:
            render_single_price_page(frozen_view, "frozen", "冻品")
    else:
        # ── 鲜冻价差：冻品周度价格展开到每日，与鲜品日度价格对比 ──
        if fresh_view.empty:
            st.warning("无鲜品数据，无法计算鲜冻价差。")
            return
        if frozen_view.empty:
            st.warning("无冻品数据，无法计算鲜冻价差。")
            return

        st.markdown("**冻品为周度价格（全周每日同一价格），鲜品为日度价格。选择产品与厂家进行对比。**")

        # 获取供应商列表
        fresh_suppliers_all = sorted(fresh_view["supplier"].dropna().unique().tolist()) if "supplier" in fresh_view.columns else []
        frozen_suppliers_all = sorted(frozen_view["supplier"].dropna().unique().tolist()) if "supplier" in frozen_view.columns else []

        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
        with col1:
            fresh_product = st.selectbox("鲜品产品", sorted(fresh_view["product"].unique()), key="spread_fresh_product")
        with col2:
            f_sups = sorted(fresh_view[fresh_view["product"] == fresh_product]["supplier"].dropna().unique().tolist()) if fresh_suppliers_all else []
            fresh_supplier = st.selectbox("鲜品厂家", f_sups, key="spread_fresh_supplier") if f_sups else None
        with col3:
            frozen_product = st.selectbox("冻品产品", sorted(frozen_view["product"].unique()), key="spread_frozen_product")
        with col4:
            z_sups = sorted(frozen_view[frozen_view["product"] == frozen_product]["supplier"].dropna().unique().tolist()) if frozen_suppliers_all else []
            frozen_supplier = st.selectbox("冻品厂家", z_sups, key="spread_frozen_supplier") if z_sups else None

        # 筛选鲜品日度数据
        f_mask = fresh_view["product"] == fresh_product
        if fresh_supplier:
            f_mask &= fresh_view["supplier"] == fresh_supplier
        fresh_daily = fresh_view[f_mask][["date", "value"]].copy()
        fresh_daily = fresh_daily.groupby("date", as_index=False)["value"].mean()
        fresh_daily = fresh_daily.rename(columns={"value": "fresh_value"})

        # 筛选冻品周度数据
        z_mask = frozen_view["product"] == frozen_product
        if frozen_supplier:
            z_mask &= frozen_view["supplier"] == frozen_supplier
        frozen_weekly = frozen_view[z_mask][["date", "value"]].copy()
        frozen_weekly = frozen_weekly.groupby("date", as_index=False)["value"].mean()
        frozen_weekly = frozen_weekly.rename(columns={"value": "frozen_value"})

        # 冻品周度价格展开到每日：根据相邻周期间隔确定覆盖天数
        frozen_weekly = frozen_weekly.sort_values("date").reset_index(drop=True)
        frozen_daily_rows = []
        for i, (_, row) in enumerate(frozen_weekly.iterrows()):
            week_start = row["date"]
            # 下一条记录的日期 - 1 天 = 本周覆盖的结束日期
            if i + 1 < len(frozen_weekly):
                next_start = frozen_weekly.iloc[i + 1]["date"]
                week_end = next_start - pd.Timedelta(days=1)
            else:
                week_end = week_start + pd.Timedelta(days=6)  # 最后一周默认7天
            days_in_week = (week_end - week_start).days + 1
            for d in range(days_in_week):
                frozen_daily_rows.append({
                    "date": week_start + pd.Timedelta(days=d),
                    "frozen_value": row["frozen_value"]
                })
        frozen_daily = pd.DataFrame(frozen_daily_rows)

        # 合并：inner join 确保只保留两者都有的日期
        spread_df = fresh_daily.merge(frozen_daily, on="date", how="inner")
        if spread_df.empty:
            f_dates = f"{fresh_daily['date'].min().date()} ~ {fresh_daily['date'].max().date()}" if not fresh_daily.empty else "无"
            z_dates = f"{frozen_daily['date'].min().date()} ~ {frozen_daily['date'].max().date()}" if not frozen_daily.empty else "无"
            z_weekly_dates = f"{frozen_weekly['date'].min().date()} ~ {frozen_weekly['date'].max().date()}" if not frozen_weekly.empty else "无"
            st.warning(f"当前鲜品和冻品没有可重叠日期。\n\n"
                       f"- 鲜品日度范围：{f_dates}（{len(fresh_daily)} 天）\n"
                       f"- 冻品周度范围：{z_weekly_dates}（{len(frozen_weekly)} 周）\n"
                       f"- 冻品展开后范围：{z_dates}（{len(frozen_daily)} 天）")
            return
        spread_df["value"] = spread_df["fresh_value"] - spread_df["frozen_value"]
        f_label = f"{fresh_product}" + (f" · {fresh_supplier}" if fresh_supplier else "")
        z_label = f"{frozen_product}" + (f" · {frozen_supplier}" if frozen_supplier else "")
        spread_df["series_name"] = f"{f_label} − {z_label}"
        spread_df["sheet"] = "鲜冻价差"
        spread_df["metric"] = "鲜冻价差"
        spread_df["province"] = "全部"
        spread_df["city"] = ""
        spread_df["display_name"] = spread_df["series_name"]

        # 双轴图：左轴鲜品+冻品价格（实线），右轴价差（虚线）
        fig = go.Figure()
        # 鲜品价格 实线
        fig.add_trace(go.Scatter(
            x=spread_df["date"], y=spread_df["fresh_value"],
            mode="lines+markers", name=f_label,
            yaxis="y",
            line=dict(color="#DC2626", width=2),
            marker=dict(size=4),
            hovertemplate="%{x|%Y-%m-%d}<br>" + f"{f_label}：%{{y:.2f}}<extra></extra>",
        ))
        # 冻品价格 实线（周度阶梯展开）
        fig.add_trace(go.Scatter(
            x=spread_df["date"], y=spread_df["frozen_value"],
            mode="lines+markers", name=z_label,
            yaxis="y",
            line=dict(color="#2563EB", width=2),
            marker=dict(size=4),
            hovertemplate="%{x|%Y-%m-%d}<br>" + f"{z_label}：%{{y:.2f}}<extra></extra>",
        ))
        # 价差 虚线
        fig.add_trace(go.Scatter(
            x=spread_df["date"], y=spread_df["value"],
            mode="lines", name=f"{f_label} − {z_label} 价差",
            yaxis="y2",
            line=dict(color="#9333EA", width=2.5, dash="dash"),
            hovertemplate="%{x|%Y-%m-%d}<br>价差：%{y:.2f}<extra></extra>",
        ))
        fig.update_layout(
            title=f"鲜冻价差走势 —— {f_label} − {z_label}",
            xaxis=dict(title="日期", tickformat="%m-%d", gridcolor="#E5E7EB"),
            yaxis=dict(title="价格 (元)", gridcolor="#E5E7EB"),
            yaxis2=dict(title="价差 (元)", overlaying="y", side="right", showgrid=False),
            template="plotly_white", hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=-0.35, xanchor="center", x=0.5),
            height=500,
        )
        render_plotly(fig, "fresh_frozen", "spread", fresh_product, frozen_product)

        # 价差统计
        st.caption(
            f"价差均值 {spread_df['value'].mean():.2f} 元 ｜ "
            f"最大 {spread_df['value'].max():.2f} 元 ｜ "
            f"最小 {spread_df['value'].min():.2f} 元 ｜ "
            f"共 {len(spread_df)} 个交易日"
        )


def _show_raw_excel_preview(file_path: str) -> None:
    """显示 Excel 文件的原始结构预览，用于调试数据解析问题。"""
    if not file_path or not Path(file_path).is_file():
        st.info(f"文件不存在：{file_path}")
        return
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            st.markdown(f"**Sheet: `{ws.title}`** ({ws.max_row} 行 × {ws.max_column} 列)")
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                values = list(row)
                while values and values[-1] is None:
                    values.pop()
                rows.append(values)
                if i >= 5:
                    break
            if rows:
                # 构建 DataFrame 展示
                max_len = max(len(r) for r in rows)
                cols = [f"Col{j}" for j in range(max_len)]
                data_rows = []
                for r in rows:
                    data_rows.append([str(v)[:30] if v is not None else "" for v in r] + [""] * (max_len - len(r)))
                preview_df = pd.DataFrame(data_rows, columns=cols)
                st.dataframe(preview_df, use_container_width=True)
        wb.close()
    except Exception as e:
        st.error(f"读取失败：{e}")


# -----------------------------
# 主界面
# -----------------------------
st.set_page_config(
    page_title="生猪现货-期货综合分析平台",
    layout="wide",
    menu_items={"About": "生猪市场研究平台 · 涌益咨询数据 · 内部使用"},
)
inject_css()

# 顶部导语
st.markdown(
    "<h1 style='font-size:28px;font-weight:900;color:#0f172a;letter-spacing:-0.02em;margin-bottom:4px'>"
    "⚑ 生猪现货-期货综合研究平台"
    "</h1>",
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='color:#64748b;font-size:14px;margin-top:0'>"
    "数据来源：涌益咨询日度 / 周度数据库、大连商品交易所期货行情、调运监测系统、鲜冻品价格数据库。"
    "｜ 平台特性：规则化综合摘要、异常/机会自动提示、季节性与农历分析。"
    "</p>",
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown(
        "<div style='font-size:15px;font-weight:700;color:#1e40af;margin:0 0 8px 0'>📐 模块导航</div>",
        unsafe_allow_html=True,
    )
    module = st.radio(
        "当前模块",
        ["涌益日度数据", "涌益周度数据", "调运分析", "鲜品和冻品数据库"],
        key="module",
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.caption("使用说明：在左侧选择模块后，对应数据源路径与筛选条件将出现在下方。各模块顶部会生成智能摘要。")
    if st.sidebar.button("🔄 清除缓存并刷新", use_container_width=True, help="数据文件更新后，点击此按钮清除缓存并重新加载最新数据"):
        st.cache_data.clear()
        st.rerun()

if module == "涌益日度数据":
    render_yongyi_daily_module()
elif module == "涌益周度数据":
    render_yongyi_weekly_module()
elif module == "调运分析":
    render_transport_module()
else:
    render_fresh_frozen_module()

